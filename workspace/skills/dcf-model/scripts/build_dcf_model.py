#!/usr/bin/env python3
"""
生成包含实时公式的 DCF (Discounted Cash Flow) Excel 模型。
准确提取富途API字段，包含完整的 WACC 和 FCF 计算公式链。
参考 get_financials_statements.py 的调用方式获取股价和总股本，包含详尽的调试日志。
"""
import argparse
import re
import zipfile
import tempfile
import os
import logging
import sys as _sys
import os as _os
from datetime import date
from pathlib import Path
from typing import Optional, Tuple

# 确保能导入 futuapi 的 common 模块 (与 get_financials_statements.py 保持一致)
_script_dir = _os.path.dirname(_os.path.abspath(__file__))
_potential_paths = [
    _os.path.join(_script_dir, "..", "futuapi", "scripts"),
    _os.path.join(_script_dir, "..", "..", "futuapi", "scripts"),
    _os.path.join(_script_dir, "futuapi", "scripts")
]
for p in _potential_paths:
    if _os.path.isdir(p):
        _sys.path.insert(0, _os.path.normpath(p))
        break

try:
    from common import create_quote_context, check_ret, safe_close, is_empty
except ImportError:
    logging.error("无法导入 common 模块，请确保 futuapi/scripts 目录在 sys.path 中。")
    # 定义空函数以防脚本完全崩溃
    def create_quote_context(*args, **kwargs): raise ImportError("common module not found")
    def check_ret(ret, data, ctx, msg): pass
    def safe_close(ctx): pass
    def is_empty(data): return data is None or len(data) == 0

# 配置日志输出
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
except ImportError as exc:
    raise ImportError("openpyxl required: pip install openpyxl") from exc

# ==================== Styles ====================
FONT_BLUE = Font(color="0000FF")            # 硬编码输入
FONT_BLACK = Font(color="000000")           # 公式计算
FONT_GREEN = Font(color="008000")           # 跨 Sheet 引用
FONT_PURPLE = Font(color="800080")          # 同 Sheet 直接引用 (无计算)
FONT_WHITE_BOLD = Font(color="FFFFFF", bold=True)
FONT_BOLD = Font(bold=True)
FONT_BLACK_BOLD = Font(color="000000", bold=True)
FILL_DARK_BLUE = PatternFill("solid", fgColor="1F4E79")
FILL_LIGHT_BLUE = PatternFill("solid", fgColor="D9E1F2")
FILL_MEDIUM_BLUE = PatternFill("solid", fgColor="BDD7EE")
FILL_INPUT_GREY = PatternFill("solid", fgColor="F2F2F2")
FILL_FORECAST_GREEN = PatternFill("solid", fgColor="E2F0D9")
FILL_VALUATION_ORANGE = PatternFill("solid", fgColor="FCE4D6")
FILL_VALUATION_ORANGE_DARK = PatternFill("solid", fgColor="F4B183")
BOTTOM_BORDER = Border(bottom=Side(style="thin", color="000000"))

# ==================== 数字格式 (对齐 schema.md) ====================
# 注意: 由于 skill 需支持多币种 (USD/CNY/HKD/…), 数字格式统一去掉货币符号 ($),
# 币种通过 A2 表头与行标签明示; 每股价格则用 FMT_PRICE (无货币符号)。
FMT_CURRENCY   = "#,##0;(#,##0);\"-\""      # 金额 (通用, 零值显示 -)
FMT_CURRENCY_M = "#,##0;(#,##0);\"-\""      # 百万单位金额
FMT_PRICE      = "#,##0.00"                 # 每股价格 (币种在标签中明示)
FMT_PERCENT    = "0.0%"                     # 百分比
FMT_MULTIPLE   = "0.0\"x\""                 # 倍数
FMT_SHARES     = "#,##0.00"                 # 股本 (百万股)
FMT_DECIMAL    = "0.0000"                   # 贴现因子
FMT_FX         = "0.0000"                   # FX 汇率 (4 位小数)

# ==================== 币种识别 & FX ====================
# 财报 Excel 第二列 "单位" 字符串 -> ISO 币种代码
# 对齐 generate_financial_excel.py:745 currency_mapping 的反向表
_UNIT_NAME_TO_CURRENCY = {
    "百万人民币": "CNY",
    "百万港元":   "HKD",
    "百万美元":   "USD",
    "百万欧元":   "EUR",
    "百万英镑":   "GBP",
    "百万日元":   "JPY",
    "百万新加坡元": "SGD",
    "百万澳元":   "AUD",
    "百万加元":   "CAD",
}

# 交易所前缀 -> 股价 (交易) 币种
_MARKET_PREFIX_TO_CURRENCY = {
    "US": "USD",
    "HK": "HKD",
    "SH": "CNY",
    "SZ": "CNY",
}

# Futu FX 快照代码候选 (顺序尝试): (from_ccy, to_ccy) -> [futu_code, ...]
# 说明: 富途外汇代码可能因权限/上线状态不同, 这里给出候选, 逐个探测。
_FX_FUTU_CODES = {
    ("USD", "CNY"): ["HK.USDCNH", "HK.USDCNY"],
    ("USD", "HKD"): ["HK.USDHKD"],
    ("HKD", "CNY"): ["HK.HKDCNH", "HK.HKDCNY"],
    ("EUR", "USD"): ["HK.EURUSD"],
    ("GBP", "USD"): ["HK.GBPUSD"],
    ("USD", "JPY"): ["HK.USDJPY"],
}

# 兜底汇率 (仅在 Futu FX 抓取失败时使用): 1 from_ccy = X to_ccy
# 请在 Excel 中用户覆盖以获得实时价格 — 这些仅作应急默认值
_FX_FALLBACKS = {
    ("USD", "CNY"): 7.20,
    ("USD", "HKD"): 7.80,
    ("HKD", "CNY"): 0.92,
    ("EUR", "USD"): 1.08,
    ("GBP", "USD"): 1.27,
    ("USD", "JPY"): 155.0,
}

# ==================== Beta 基准指数 & Rf/ERP 常量 ====================
# 按交易场所前缀选择大盘基准 (用于计算个股 5Y monthly beta)
_BENCHMARK_INDEX = {
    "US": "US.SPY",       # S&P 500 ETF
    "HK": "HK.800000",    # 恒生指数
    "SH": "SH.000300",    # 沪深 300
    "SZ": "SH.000300",    # A 股统一用沪深 300
}

# 按报表币种 (Reporting Currency) 决定 Rf 与 ERP:
# - Rf: 10Y 主权债券收益率
# - ERP: Damodaran country equity risk premium (年度更新, 2024/2025 数值)
# 用户可在 Excel WACC Sheet 的蓝色输入格中直接覆盖为最新值
_RF_ERP_BY_CURRENCY = {
    "USD": {"rf": 0.043, "erp": 0.055, "rf_src": "10Y US Treasury",  "erp_src": "Damodaran US ERP"},
    "HKD": {"rf": 0.040, "erp": 0.060, "rf_src": "HKGB 10Y",         "erp_src": "Damodaran HK country ERP"},
    "CNY": {"rf": 0.025, "erp": 0.065, "rf_src": "中国 10Y 国债",     "erp_src": "Damodaran CN country ERP"},
    "EUR": {"rf": 0.027, "erp": 0.050, "rf_src": "German Bund 10Y",  "erp_src": "Damodaran EU ERP"},
    "GBP": {"rf": 0.040, "erp": 0.055, "rf_src": "UK Gilt 10Y",      "erp_src": "Damodaran UK ERP"},
    "JPY": {"rf": 0.015, "erp": 0.055, "rf_src": "JGB 10Y",          "erp_src": "Damodaran JP ERP"},
}
_RF_ERP_DEFAULT = {"rf": 0.043, "erp": 0.055, "rf_src": "Fallback (USD)", "erp_src": "Fallback (USD)"}

# Beta 默认兜底值 (当 Futu 历史行情不足或调用失败时使用)
_BETA_FALLBACK = 1.20
_BETA_MIN_SAMPLES = 24   # 最少 24 个月度收益率样本 (=25 期月线) 才计算 beta

# ==================== Comment helper ====================
_COMMENT_AUTHOR = "DCF Builder"

def add_comment(cell, text: str, width: int = 260, height: int = 100):
    """给单元格附加气泡备注, 统一样式。"""
    c = Comment(text, _COMMENT_AUTHOR)
    c.width = width
    c.height = height
    cell.comment = c

def add_source_comment(cell, system: str, ref: str = "", extra: str = ""):
    """为硬编码输入添加 schema.md 规定格式的 Source 备注:
      Source: [System/Document], [Date], [Reference]
    """
    today = date.today().isoformat()
    parts = [f"Source: {system}", today]
    if ref: parts.append(ref)
    text = ", ".join(parts)
    if extra: text += f"\n{extra}"
    add_comment(cell, text)

# ==================== Utility ====================
def safe_divide(n, d): return n / d if d != 0 else 0.0

def _ensure_shared_strings(file_path: Path) -> Tuple[Path, bool]:
    file_path = Path(file_path)
    try:
        with zipfile.ZipFile(str(file_path), 'r') as zf:
            if 'xl/sharedStrings.xml' in zf.namelist(): return file_path, False
            ct = zf.read('[Content_Types].xml').decode('utf-8', errors='ignore')
            if 'sharedStrings' not in ct: return file_path, False
    except:
        return file_path, False
    fd, tmp = tempfile.mkstemp(suffix='.xlsx', dir=str(file_path.parent))
    os.close(fd)
    try:
        with zipfile.ZipFile(str(file_path), 'r') as src, zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as dst:
            for item in src.namelist(): dst.writestr(item, src.read(item))
            dst.writestr('xl/sharedStrings.xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"/>')
        return Path(tmp), True
    except:
        return file_path, False

def _read_excel_map(file_path: Path) -> Tuple[dict, Optional[str]]:
    """读取财报 Excel, 返回 (data_dict, currency_code)。

    currency_code 从第二列 "单位" 字符串反向解析 (如 "百万人民币"->"CNY")。
    若单位列缺失或为纯 "百万"/"%" 兜底则返回 None。
    """
    fp, is_tmp = _ensure_shared_strings(file_path)
    try:
        wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
        sheet = wb.active
        data = {}
        currency_code = None
        header_row = 1
        for r in range(1, 5):
            vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1) if sheet.cell(r, c).value]
            if "FY" in " ".join(str(v) for v in vals): header_row = r; break
        periods = [str(sheet.cell(header_row, c).value) for c in range(3, sheet.max_column + 1)]
        for r in range(header_row + 1, sheet.max_row + 1):
            ind = sheet.cell(r, 1).value
            if not ind: continue
            ind = str(ind).strip()
            # 从第二列单位字符串识别币种 (只取第一次成功匹配)
            if currency_code is None:
                unit_str = sheet.cell(r, 2).value
                if isinstance(unit_str, str):
                    mapped = _UNIT_NAME_TO_CURRENCY.get(unit_str.strip())
                    if mapped:
                        currency_code = mapped
            vals = {}
            for c, p in enumerate(periods):
                v = sheet.cell(r, c + 3).value
                num = 0.0
                if isinstance(v, (int, float)): num = float(v)
                elif isinstance(v, str) and v:
                    cl = v.replace(",", "").replace("-", "").strip()
                    if cl.replace(".", "", 1).isdigit(): num = float(cl) * (-1 if v.strip().startswith("-") else 1)
                vals[p] = num
            data[ind] = vals
        wb.close()
        return data, currency_code
    finally:
        if is_tmp and fp.exists(): os.remove(str(fp))

def find_local_file(excels_path: Path, ticker: str, suffix: str) -> Optional[Path]:
    pattern = re.compile(rf'^.*_{re.escape(ticker)}_{re.escape(suffix)}_.*\.(xlsx|xls)$', re.IGNORECASE)
    files = [f for f in excels_path.iterdir() if f.is_file() and pattern.match(f.name)]
    return sorted(files)[-1] if files else None

def normalize_stock_code(ticker: str) -> str:
    """将 ticker 转为富途格式: BABA -> US.BABA, 00700 -> HK.00700, 600519 -> SH.600519"""
    ticker = ticker.strip().upper()
    if '.' in ticker and ticker.split('.')[0] in ('US', 'HK', 'SH', 'SZ'):
        return ticker
    if ticker.isdigit():
        n = int(ticker)
        if n >= 600000: return f"SH.{ticker}"
        elif n >= 300000: return f"SZ.{ticker}"
        else: return f"HK.{ticker.zfill(5)}"
    return f"US.{ticker}"


def infer_trading_currency(stock_code: str) -> str:
    """从富途格式股票代码 (US.BABA / HK.00700 / SH.600519) 推断交易 (股价) 币种。

    - 交易场所决定报价币种, 这一映射永远正确 (US.*→USD, HK.*→HKD, SH./SZ.*→CNY)
    - 与财报币种可能不同 (ADR/H 股), 后者从财报 Excel 单位列读取
    """
    prefix = stock_code.split('.')[0] if '.' in stock_code else ""
    return _MARKET_PREFIX_TO_CURRENCY.get(prefix, "USD")

# ==================== Futu Market Data Fetching ====================
def fetch_market_data_from_futu(stock_code: str) -> dict:
    """
    参考 get_financials_statements.py 的调用方式，
    使用 common 模块的 create_quote_context, check_ret, safe_close 获取股票快照数据。
    """
    result = {"stock_price": None, "shares_outstanding": None, "shares_source": "Unknown"}
    ctx = None

    logger.info(f"正在通过 create_quote_context 连接 FutuOpenD 获取 {stock_code} 的数据...")

    try:
        # 1. 创建连接 (与 get_financials_statements.py 相同)
        ctx = create_quote_context()

        # 2. 调用 API 获取市场快照 (包含最新价格和总股本)
        logger.info(f"[API 调用] ctx.get_market_snapshot([{stock_code}])")
        ret, data = ctx.get_market_snapshot([stock_code])

        # 3. 检查返回值 (与 get_financials_statements.py 相同)
        check_ret(ret, data, ctx, "获取市场快照")

        if is_empty(data):
            logger.warning("get_market_snapshot 返回空数据，可能是股票代码错误或无权限")
        else:
            logger.debug(f"get_market_snapshot 返回列名: {data.columns.tolist()}")
            row = data.iloc[0]

            # 4. 提取价格
            if 'last_price' in data.columns and row['last_price'] is not None and float(row['last_price']) > 0:
                result["stock_price"] = float(row['last_price'])
                logger.info(f"  -> 成功获取价格 (last_price): {result['stock_price']}")
            else:
                logger.warning(f"  -> last_price 缺失或为 0，原始值: {row.get('last_price', 'N/A')}")

            # 5. 提取总股本
            if 'issued_shares' in data.columns and row['issued_shares'] is not None and float(row['issued_shares']) > 0:
                raw_shares = float(row['issued_shares'])
                result["shares_outstanding"] = raw_shares / 1_000_000  # 转换为百万股
                result["shares_source"] = "Futu get_market_snapshot (issued_shares)"
                logger.info(f"  -> 成功获取总股本 (issued_shares): {raw_shares:,.0f} (已转换为 {result['shares_outstanding']:,.2f}M)")
            else:
                logger.warning(f"  -> issued_shares 缺失或为 0，原始值: {row.get('issued_shares', 'N/A')}")

    except SystemExit:
        # check_ret 在失败时会抛出 SystemExit，这里捕获以防脚本直接退出
        logger.error("check_ret 触发了 SystemExit，可能是 OpenD 未连接、API 调用频率超限或权限不足。")
    except Exception as e:
        logger.exception(f"Futu API 调用发生未预期异常: {e}")
    finally:
        # 6. 关闭连接 (与 get_financials_statements.py 相同)
        if ctx:
            safe_close(ctx)
            logger.info("FutuOpenD 连接已关闭。")

    return result

# ==================== FX Rate Fetching ====================
def fetch_fx_rate_from_futu(from_ccy: str, to_ccy: str) -> Tuple[float, str]:
    """获取 1 from_ccy = X to_ccy 的汇率。

    优先尝试 Futu 外汇快照 (`get_market_snapshot(['HK.USDCNH', ...])`),
    失败时使用 _FX_FALLBACKS 中的常量兜底。

    Returns:
        (rate, source): source 是数据来源标识, 用于 Excel 备注
    """
    from_ccy, to_ccy = from_ccy.upper(), to_ccy.upper()
    if from_ccy == to_ccy:
        return 1.0, "Same currency (no conversion)"

    codes = _FX_FUTU_CODES.get((from_ccy, to_ccy), [])
    inverse_codes = _FX_FUTU_CODES.get((to_ccy, from_ccy), [])
    ctx = None
    try:
        if codes or inverse_codes:
            logger.info(f"正在通过 FutuOpenD 获取 {from_ccy}->{to_ccy} 汇率...")
            ctx = create_quote_context()
            # 正向候选码
            for fx_code in codes:
                try:
                    logger.info(f"[API 调用] ctx.get_market_snapshot([{fx_code}])")
                    ret, data = ctx.get_market_snapshot([fx_code])
                    if ret == 0 and not is_empty(data):
                        row = data.iloc[0]
                        px = row.get('last_price')
                        if px is not None and float(px) > 0:
                            logger.info(f"  -> 获取汇率成功 ({fx_code}): 1 {from_ccy} = {float(px):.4f} {to_ccy}")
                            return float(px), f"Futu get_market_snapshot ({fx_code})"
                    else:
                        logger.warning(f"  -> {fx_code} 无有效数据: ret={ret}")
                except Exception as e:
                    logger.warning(f"  -> {fx_code} 调用失败: {e}")
            # 反向候选码 (拿到 to->from, 求倒数)
            for fx_code in inverse_codes:
                try:
                    logger.info(f"[API 调用] 尝试反向汇率 ctx.get_market_snapshot([{fx_code}])")
                    ret, data = ctx.get_market_snapshot([fx_code])
                    if ret == 0 and not is_empty(data):
                        row = data.iloc[0]
                        px = row.get('last_price')
                        if px is not None and float(px) > 0:
                            rate = 1.0 / float(px)
                            logger.info(f"  -> 反向汇率成功 ({fx_code}={float(px):.4f}): 1 {from_ccy} = {rate:.4f} {to_ccy}")
                            return rate, f"Futu get_market_snapshot ({fx_code}, inverse)"
                except Exception as e:
                    logger.warning(f"  -> {fx_code} 调用失败: {e}")
    except Exception as e:
        logger.warning(f"FX 获取过程发生异常: {e}")
    finally:
        if ctx:
            safe_close(ctx)

    # Fallback
    if (from_ccy, to_ccy) in _FX_FALLBACKS:
        rate = _FX_FALLBACKS[(from_ccy, to_ccy)]
        logger.warning(f"FX API 失败, 使用回退常量 1 {from_ccy} = {rate} {to_ccy}")
        return rate, f"Fallback constant (please override in Excel)"
    if (to_ccy, from_ccy) in _FX_FALLBACKS:
        rate = 1.0 / _FX_FALLBACKS[(to_ccy, from_ccy)]
        logger.warning(f"FX API 失败, 使用反向回退常量: 1 {from_ccy} = {rate:.4f} {to_ccy}")
        return rate, f"Fallback constant (inverse, please override in Excel)"
    logger.warning(f"无 {from_ccy}->{to_ccy} 汇率映射, 使用 1.0 (可能不正确)")
    return 1.0, f"NO MAPPING — please override in Excel"

# ==================== Beta 计算 ====================
def get_benchmark_for(stock_code: str) -> str:
    """根据股票所属交易场所选择大盘基准指数, 用于计算个股 Beta。"""
    prefix = stock_code.split('.')[0] if '.' in stock_code else ""
    return _BENCHMARK_INDEX.get(prefix, "US.SPY")


def _fetch_monthly_returns(ctx, code: str, months: int = 60):
    """通过 Futu 获取月度前复权 K 线并计算月度收益率序列。

    Returns:
        list[float] 月度收益率 (可能为空); 失败时返回 []
    """
    try:
        from futu.common.constant import KLType, AuType   # 延迟导入, 避免顶部依赖
    except ImportError:
        logger.warning("futu-api 未安装, 无法计算 Beta")
        return []
    end = date.today().isoformat()
    # 多请求 ~1 年冗余, 避免遇到停牌导致样本不足
    from datetime import timedelta
    start = (date.today() - timedelta(days=int(365 * (months / 12 + 1)))).isoformat()
    try:
        ret, df, _ = ctx.request_history_kline(
            code, start=start, end=end, ktype=KLType.K_MON, autype=AuType.QFQ, max_count=1000
        )
    except Exception as e:
        logger.warning(f"  -> {code} request_history_kline 异常: {e}")
        return []
    if ret != 0 or df is None or df.empty:
        logger.warning(f"  -> {code} 月线数据不可用 (ret={ret})")
        return []
    closes = [float(v) for v in df['close'].tolist() if v is not None]
    if len(closes) < 2:
        return []
    return [(closes[i] - closes[i-1]) / closes[i-1] for i in range(1, len(closes)) if closes[i-1]]


def compute_beta_from_futu(stock_code: str, benchmark_code: Optional[str] = None,
                            months: int = 60) -> Tuple[Optional[float], str]:
    """通过 Futu `request_history_kline` 拉取 5Y monthly K 线, 计算个股 Beta。

    Beta = cov(stock_return, mkt_return) / var(mkt_return)

    Returns:
        (beta, source): beta 为 None 时表示计算失败, 调用方应用兜底值
    """
    benchmark = benchmark_code or get_benchmark_for(stock_code)
    logger.info(f"正在通过 FutuOpenD 计算 {stock_code} 5Y monthly Beta (基准: {benchmark})...")
    ctx = None
    try:
        ctx = create_quote_context()
        stock_r = _fetch_monthly_returns(ctx, stock_code, months)
        mkt_r = _fetch_monthly_returns(ctx, benchmark, months)
        n = min(len(stock_r), len(mkt_r))
        if n < _BETA_MIN_SAMPLES:
            logger.warning(f"  -> 样本不足 ({n} < {_BETA_MIN_SAMPLES}), 无法计算 Beta")
            return None, f"Insufficient history ({n} months)"
        stock_r = stock_r[-n:]
        mkt_r = mkt_r[-n:]
        mean_s = sum(stock_r) / n
        mean_m = sum(mkt_r) / n
        cov = sum((s - mean_s) * (m - mean_m) for s, m in zip(stock_r, mkt_r)) / n
        var_m = sum((m - mean_m) ** 2 for m in mkt_r) / n
        if var_m <= 0:
            logger.warning(f"  -> 基准方差为 0, 无法计算 Beta")
            return None, "Zero market variance"
        beta = cov / var_m
        logger.info(f"  -> 成功: Beta({stock_code} vs {benchmark}) = {beta:.4f}, n={n} months")
        return beta, f"Futu {n}M monthly kline (vs {benchmark}), computed {date.today().isoformat()}"
    except SystemExit:
        logger.error("Beta 计算触发 SystemExit (Futu 未连接?)")
        return None, "Futu API failure"
    except Exception as e:
        logger.exception(f"Beta 计算异常: {e}")
        return None, f"Exception: {e}"
    finally:
        if ctx:
            safe_close(ctx)
            logger.info("FutuOpenD 连接已关闭 (Beta 计算)。")

# ==================== Financial Data Extraction ====================
def extract_financial_data(workspace: Path, ticker: str) -> dict:
    excels_path = workspace / "excels"
    inc_file = find_local_file(excels_path, ticker, "income")
    bs_file = find_local_file(excels_path, ticker, "balance")
    cf_file = find_local_file(excels_path, ticker, "cashflow")
    inc_result = _read_excel_map(inc_file) if inc_file else ({}, None)
    bs_result = _read_excel_map(bs_file) if bs_file else ({}, None)
    cf_result = _read_excel_map(cf_file) if cf_file else ({}, None)
    inc_data, inc_ccy = inc_result
    bs_data, bs_ccy = bs_result
    cf_data, cf_ccy = cf_result

    # 报表币种: 以 income 为准, 三张表应一致, 不一致时警告
    reporting_currency = inc_ccy or bs_ccy or cf_ccy
    detected_ccys = [c for c in (inc_ccy, bs_ccy, cf_ccy) if c]
    if len(set(detected_ccys)) > 1:
        logger.warning(f"财报三张表币种不一致: income={inc_ccy}, balance={bs_ccy}, cashflow={cf_ccy} — 以 income 为准")

    all_fy = set()
    for d in [inc_data, bs_data, cf_data]:
        for k, v in d.items(): all_fy.update([p for p in v.keys() if "FY" in p])
    fy_cols = sorted(list(all_fy), reverse=True)
    if not fy_cols:
        raise RuntimeError(f"未找到 {ticker} 的历史财报数据，请先运行 futu-financial-report 生成 Excel")
    latest_fy, prev_fy = fy_cols[0], fy_cols[1] if len(fy_cols) > 1 else fy_cols[0]

    # 取最近 5 个财年，按时间正序排列 (oldest -> newest)；同时保留紧邻的更早一年用于计算最老年份的增长率
    hist_fys = list(reversed(fy_cols[:5]))
    prior_to_hist = fy_cols[5] if len(fy_cols) > 5 else None

    def gv(data, keys, col):
        for k in keys:
            if k in data and col in data[k]: return data[k][col]
        return 0.0

    # ---- 历史逐年抽取（按 hist_fys 正序）----
    def _series(data, keys):
        return [gv(data, keys, fy) for fy in hist_fys]

    revenue_series = _series(inc_data, ["总收入", "营业总收入"])
    ebit_series   = _series(inc_data, ["营业利润"])
    tax_series    = _series(inc_data, ["所得税"])
    ebt_series    = _series(inc_data, ["税前利润"])
    # CapEx: 严格来自现金流量表明细字段, 避免使用 income sheet 里含"投资活动净额"兜底的加工指标
    # (兜底口径会把长期股权投资/有价证券买卖计入, 显著高估 CapEx)
    capex_series  = [abs(v) for v in _series(cf_data, ["资本开支(CapEx明细)"])]
    if all(v == 0.0 for v in capex_series):
        # 旧版 cashflow Excel 无严格口径行 -> 直接尝试原始明细字段名
        # (若 futu-financial-report 已升级但用户尚未重新生成 cashflow Excel)
        capex_series = [abs(v) for v in _series(cf_data, [
            "购建固定资产及无形资产净额",                      # 美股 8046
            "购建固定资产",                                    # 港股 5071 (需另加 5073, 但至少不含兜底)
            "购建固定资产、无形资产和其他长期资产支付的现金",   # A 股 3043
        ])]
    if all(v == 0.0 for v in capex_series):
        # 最终回退: 使用 income sheet 的 CapEx (含投资活动净额兜底口径, 可能高估)
        logger.warning(
            "cashflow Excel 无 CapEx 明细字段, 回退到 income Excel 的 CapEx "
            "(该口径可能因投资活动净额兜底而高估 CapEx, 请重新生成 cashflow Excel 获得严格口径)"
        )
        capex_series = [abs(v) for v in _series(inc_data, ["资本开支(CapEx)", "资本开支"])]
    da_series     = _series(cf_data, ["折旧摊销及损耗"])
    if all(v == 0.0 for v in da_series):
        da_series = _series(inc_data, ["折旧摊销及损耗"])

    # 计算历史逐年比率
    def _pct_series(nums, denoms):
        return [safe_divide(n, d) if d else 0.0 for n, d in zip(nums, denoms)]

    ebit_margin_series = _pct_series(ebit_series, revenue_series)
    da_pct_series      = _pct_series(da_series, revenue_series)
    capex_pct_series   = _pct_series(capex_series, revenue_series)
    tax_rate_series    = [safe_divide(t, e) if e > 0 else 0.0 for t, e in zip(tax_series, ebt_series)]

    # 营收增长率：最老一年需用 prior_to_hist 计算，否则留空
    rev_growth_series = []
    prior_rev = gv(inc_data, ["总收入", "营业总收入"], prior_to_hist) if prior_to_hist else 0.0
    for i, rev in enumerate(revenue_series):
        prev = revenue_series[i-1] if i > 0 else prior_rev
        rev_growth_series.append(safe_divide(rev - prev, prev) if prev else None)

    # ---- 最新一年（用于返回给下游） ----
    revenue = revenue_series[-1]
    ebit = ebit_series[-1]
    tax = tax_series[-1]
    ebt = ebt_series[-1]
    capex = capex_series[-1]
    da = da_series[-1]

    # 合理性检查: CapEx% 超过 20% 通常是数据口径问题 (兜底到投资活动净额)
    capex_ratio_check = safe_divide(capex, revenue)
    if capex_ratio_check > 0.20:
        logger.warning(
            f"CapEx% = {capex_ratio_check:.1%} 超过 20%, 可能是数据源含投资活动净额, "
            f"请核对 cashflow Excel 中的「资本开支(CapEx明细)」行 (capex={capex:,.0f}, revenue={revenue:,.0f})"
        )

    rev_growth = rev_growth_series[-1] if rev_growth_series[-1] is not None else 0.0
    tax_rate = tax_rate_series[-1] if tax_rate_series[-1] > 0 else 0.25

    # ---- Net Debt 构成 ----
    # 债务 = 短期借款 (含融资租赁) + 长期借款 (含长期融资租赁)
    # 现金 = 现金及等价物 + 短期投资 + 定期存款 (流动 + 非流动)
    # 富途报表结构:
    #   美股: 短期借款与融资租赁负债 / -现金和现金等价物 / -短期投资 (无独立长期借款字段)
    #   港股: 银行贷款及透支 / 长期银行贷款 / 长期融资租赁负债 / 现金及等价物
    #         定期存款-流动资产 (fid 5005) + 定期存款-非流动资产 (fid 5054)
    #   A 股: 短期借款 / 长期借款 / 货币资金
    # 使用 "首项命中" 的加和法, 避免重复计入 (如"-现金和现金等价物"与"现金及等价物"只取其一)
    # 注: 受限制现金 (港股"已抵押存款" / 美股"受限制现金") 按投行惯例不计入 Net Cash
    def _sum_first_match(data, groups, col):
        """对 groups 中每组 keys, 只取首项命中的值, 汇总。避免同类字段重复相加。"""
        total = 0.0
        for keys in groups:
            for k in keys:
                if k in data and col in data[k]:
                    v = data[k][col]
                    if v:
                        total += v
                        break   # 该组已命中, 跳到下一组
        return total

    # ---- 定期存款: 需要同时累加流动 + 非流动两个字段, 单独处理 ----
    def _sum_all_matches(data, keys, col):
        """对 keys 里所有命中的字段累加 (用于同一"类"下有多个独立子项的场景)。"""
        total = 0.0
        for k in keys:
            if k in data and col in data[k]:
                v = data[k][col]
                if v:
                    total += v
        return total

    debt = _sum_first_match(bs_data, [
        # 短期借款 (港股 / 美股 / A 股 命名差异)
        ["短期借款与融资租赁负债", "-短期借款", "短期借款", "银行贷款及透支"],
        # 长期借款
        ["长期借款", "长期银行贷款"],
        # 长期融资租赁 (港股常见, 美股已含在"短期借款与融资租赁负债")
        ["长期融资租赁负债"],
    ], latest_fy)

    cash = _sum_first_match(bs_data, [
        # 现金及等价物
        ["-现金和现金等价物", "现金及现金等价物", "现金及等价物", "货币资金"],
        # 短期投资 (港股 fid 5005 若 display_name='短期投资' 时命中此组)
        ["-短期投资", "短期投资"],
    ], latest_fy)
    # 定期存款: 流动 + 非流动 独立累加 (港股常见,
    # 若 fid 5005 显示为"定期存款-流动资产"未被上面的"短期投资组"命中, 会在此组累加)
    cash += _sum_all_matches(bs_data, [
        "定期存款-流动资产",
        "定期存款-非流动资产",
        "长期定期存款",
        "短期存款",
        "定期存款",
    ], latest_fy)

    # ===== 从 FutuOpenD 实时获取股票价格和总股本 =====
    stock_code = normalize_stock_code(ticker)
    market_data = fetch_market_data_from_futu(stock_code)

    # 如果获取失败，使用回退默认值
    if market_data.get("stock_price") is None:
        logger.warning(f"未能从 API 获取 {ticker} 的价格，使用默认回退值 118.9")
        market_data["stock_price"] = 118.9
        market_data["shares_source"] = "Default Fallback (API Failed)"

    if market_data.get("shares_outstanding") is None:
        logger.warning(f"未能从 API 获取 {ticker} 的总股本，使用默认回退值 8922.51M")
        market_data["shares_outstanding"] = 8922.51
        market_data["shares_source"] = "Default Fallback (API Failed)"

    # ===== 币种一致性: 交易币种 vs 报表币种 =====
    trading_currency = infer_trading_currency(stock_code)
    if not reporting_currency:
        # 财报 Excel 未在单位列写入可识别的币种字符串 (兜底"百万"), 用交易币种猜测
        logger.warning(f"财报 Excel 未提供币种标识, 回退为按 stock_code 前缀推断: {trading_currency}")
        reporting_currency = trading_currency
    fx_rate, fx_source = fetch_fx_rate_from_futu(trading_currency, reporting_currency)
    logger.info(f"币种: trading={trading_currency}, reporting={reporting_currency}, "
                f"fx=1 {trading_currency} = {fx_rate:.4f} {reporting_currency}")

    # ===== WACC 输入个股化: Beta / Rf / ERP =====
    # Beta: Futu 5Y monthly 自算, 失败回退到常量
    benchmark = get_benchmark_for(stock_code)
    beta_calc, beta_source = compute_beta_from_futu(stock_code, benchmark)
    if beta_calc is None:
        logger.warning(f"Beta 计算失败, 使用兜底值 {_BETA_FALLBACK}")
        beta_calc = _BETA_FALLBACK
        beta_source = f"Fallback constant ({beta_source})"

    # Rf / ERP: 按报表币种查常量表 (用户可在 Excel WACC Sheet 覆盖)
    rf_erp = _RF_ERP_BY_CURRENCY.get(reporting_currency, _RF_ERP_DEFAULT)
    rf_rate = rf_erp["rf"]
    erp     = rf_erp["erp"]
    rf_src  = rf_erp["rf_src"]
    erp_src = rf_erp["erp_src"]
    logger.info(f"WACC 输入: beta={beta_calc:.4f} ({benchmark}), "
                f"Rf={rf_rate:.2%} ({rf_src}), ERP={erp:.2%} ({erp_src})")

    return {
        "ticker": ticker, "revenue": revenue, "ebit": ebit, "da": da,
        "capex": capex, "rev_growth": rev_growth, "tax_rate": tax_rate,
        "debt": debt, "cash": cash,
        # 市场数据
        "stock_price": market_data["stock_price"],
        "shares_outstanding": market_data["shares_outstanding"],
        "shares_source": market_data["shares_source"],
        # 币种 & FX
        "reporting_currency": reporting_currency,
        "trading_currency": trading_currency,
        "fx_rate": fx_rate,               # 1 trading_ccy = fx_rate reporting_ccy
        "fx_source": fx_source,
        # WACC 输入 (个股化)
        "beta": beta_calc,
        "beta_source": beta_source,
        "benchmark": benchmark,
        "rf_rate": rf_rate,
        "rf_source": rf_src,
        "erp": erp,
        "erp_source": erp_src,
        # 历史序列 (按 hist_fys 正序，最多 5 期)
        "hist_fys": hist_fys,
        "hist_revenue": revenue_series,
        "hist_ebit": ebit_series,
        "hist_da": da_series,
        "hist_capex": capex_series,
        "hist_tax": tax_series,
        "hist_ebit_margin": ebit_margin_series,
        "hist_da_pct": da_pct_series,
        "hist_capex_pct": capex_pct_series,
        "hist_tax_rate": tax_rate_series,
        "hist_rev_growth": rev_growth_series,
    }

# ==================== DCF Builder ====================
class DCFBuilder:
    """
    构建 DCF 模型 (严格对齐 references/schema.md).

    Sheet 1: DCF
      Section 1: Header (Row 1-2)
      Section 2: Case Selector (Row 4-5)
      Section 3: Market Data (Row 8-12)
      Section 4: Bear/Base/Bull Assumptions + Selected Case Consolidation (INDEX)
      Section 5: Historical & Projected Financials
      Section 6: Free Cash Flow Build
      Section 7: Discounting & Terminal Value
      Section 8: Valuation Summary + TV/EV Ratio Sanity Check
      Section 9: 3× 5×5 Sensitivity Tables (WACC×TGR, Growth×Margin, Beta×Rf → Implied Price)

    Sheet 2: WACC (CAPM + Capital Structure)
    """

    # 假设行顺序 (每个 scenario block 共 8 行)
    ASSUMPTION_LABELS = [
        ("Revenue Growth % -- 营收增长率",      "rev_growth",    FMT_PERCENT),
        ("EBIT Margin % -- 息税前利润率",        "ebit_margin",   FMT_PERCENT),
        ("D&A % of Revenue -- 折旧摊销占营收比", "da_pct",        FMT_PERCENT),
        ("CapEx % of Revenue -- 资本开支占营收比","capex_pct",    FMT_PERCENT),
        ("NWC % of Delta Revenue -- 净营运资本占营收变动比", "nwc_pct", FMT_PERCENT),
        ("Tax Rate -- 税率",                      "tax_rate",     FMT_PERCENT),
        ("Terminal Growth -- 永续增长率",         "terminal_growth", FMT_PERCENT),
        ("WACC -- 加权平均资本成本",              "wacc",          FMT_PERCENT),
    ]

    def __init__(self, d: dict):
        self.d = d
        self.years = 5
        self.wb = openpyxl.Workbook()
        # 场景默认值 (可根据 base 数据浮动 ±)
        base_growth = d["rev_growth"]
        base_margin = safe_divide(d["ebit"], d["revenue"])
        base_da     = safe_divide(d["da"], d["revenue"])
        base_capex  = safe_divide(d["capex"], d["revenue"])
        base_tax    = d["tax_rate"]

        # 三情景假设 (逐年 5 期; Terminal Growth 与 WACC 只用第一列)
        # Bear: 更低增长/毛利, 更高 WACC/CapEx
        # Bull: 更高增长/毛利, 更低 WACC
        def _decay(base, delta_start, decay_step):
            """生成 5 期递减序列: [base+delta, base+delta-step, ...]"""
            return [base + delta_start - decay_step * i for i in range(5)]

        self.scenarios = {
            "Bear": {
                "rev_growth":       _decay(base_growth, -0.03, 0.005),
                "ebit_margin":      _decay(base_margin, -0.03, 0.002),
                "da_pct":           [max(base_da, 0.03)] * 5,
                "capex_pct":        [max(base_capex, 0.05) + 0.01] * 5,
                "nwc_pct":          [0.02] * 5,
                "tax_rate":         [base_tax] * 5,
                "terminal_growth":  [0.020] + [None] * 4,
                "wacc":             [0.100] + [None] * 4,
            },
            "Base": {
                "rev_growth":       _decay(base_growth, 0.00, 0.003),
                "ebit_margin":      _decay(base_margin, 0.00, 0.000),
                "da_pct":           [max(base_da, 0.03)] * 5,
                "capex_pct":        [max(base_capex, 0.05)] * 5,
                "nwc_pct":          [0.01] * 5,
                "tax_rate":         [base_tax] * 5,
                "terminal_growth":  [0.025] + [None] * 4,
                "wacc":             [None] + [None] * 4,  # 由 WACC Sheet 引用
            },
            "Bull": {
                "rev_growth":       _decay(base_growth, 0.03, 0.005),
                "ebit_margin":      _decay(base_margin, 0.03, -0.002),  # 递增
                "da_pct":           [max(base_da, 0.03)] * 5,
                "capex_pct":        [max(base_capex, 0.05) - 0.005] * 5,
                "nwc_pct":          [0.005] * 5,
                "tax_rate":         [base_tax] * 5,
                "terminal_growth":  [0.030] + [None] * 4,
                "wacc":             [0.080] + [None] * 4,
            },
        }

        # WACC Sheet 默认输入 (来自个股化数据源, 均可在 Excel 中用户覆盖)
        self.wacc_inputs = {
            "risk_free_rate": d.get("rf_rate", 0.043),
            "beta":           d.get("beta", _BETA_FALLBACK),
            "erp":            d.get("erp", 0.055),
            "pre_tax_kd":     0.045,
        }

    # ---------- helpers ----------
    def _apply_input(self, cell, value, fmt=None, source_system: str = None, source_ref: str = ""):
        cell.value = value
        cell.font = FONT_BLUE
        cell.fill = FILL_INPUT_GREY
        if fmt:
            cell.number_format = fmt
        if source_system:
            add_source_comment(cell, source_system, source_ref)

    def build(self, output_path: Path):
        self._dcf()
        self._wacc()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(output_path))
        logger.info(f"DCF model saved: {output_path}")

    # ==================== Sheet 1: DCF ====================
    def _dcf(self):
        ws = self.wb.active; ws.title = "DCF"
        d = self.d
        hist_fys = d["hist_fys"]  # 正序 oldest -> newest, 最多 5 期
        n_hist = len(hist_fys)

        # --------- Section 1: Header (Row 1-2) ---------
        rep_ccy = d.get("reporting_currency") or "N/A"
        trd_ccy = d.get("trading_currency") or "N/A"
        c = ws.cell(1, 1, f"{d['ticker']} Corporation DCF Model / {d['ticker']} 贴现现金流估值模型")
        c.fill = FILL_DARK_BLUE; c.font = FONT_WHITE_BOLD
        ws.cell(2, 1,
                f"Ticker: {d['ticker']}  |  Date: {date.today().isoformat()}  |  Year End: FY  |  "
                f"Reporting Currency: Million {rep_ccy}  |  Trading Currency: {trd_ccy}")

        # --------- Section 2: Case Selector (Row 4-5) ---------
        ws.cell(4, 1, "Case Selector (1=Bear 悲观 / 2=Base 基准 / 3=Bull 乐观)")
        self._apply_input(ws.cell(4, 2), 2, source_system="User input")
        add_comment(ws.cell(4, 2), "Case Selector:\n  1 = Bear (悲观情景)\n  2 = Base (基准情景, 默认)\n  3 = Bull (乐观情景)\n\n所有下游预测通过 CHOOSE($B$4, Bear, Base, Bull) 引用选中情景.")
        ws.cell(5, 1, "Selected Case Name")
        c = ws.cell(5, 2, '=CHOOSE($B$4,"Bear","Base","Bull")'); c.font = FONT_BLACK
        add_comment(c, "公式: =CHOOSE($B$4, \"Bear\", \"Base\", \"Bull\")\n根据 B4 情景编号显示情景名。")

        # 数据来源标注
        c = ws.cell(6, 1,
                    f"Data Source 数据来源: {d['shares_source']}  |  "
                    f"FX: {d.get('fx_source', 'N/A')}  |  "
                    f"Beta: {d.get('beta_source', 'N/A')}")
        c.font = FONT_GREEN

        # --------- Section 3: Market Data ---------
        # 由于财报币种与股价币种可能不同 (ADR/H 股), 引入 FX Rate 把股价换算到报表币种,
        # 保证 EV → Equity → Implied Price 与 Current Price 全部在报表币种下比对。
        c = ws.cell(8, 1, "MARKET DATA -- 市场数据 (非情景依赖)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE

        # Row 9: Trading Price (原始股价, 交易币种)
        self._apply_input(ws.cell(9, 2), d["stock_price"], FMT_PRICE,
                          "Futu get_market_snapshot", "last_price")
        ws.cell(9, 1, f"Current Stock Price ({trd_ccy}) -- 当前股价(交易币种)")

        # Row 10: FX Rate (1 trading_ccy = X reporting_ccy)
        fx_lock = trd_ccy == rep_ccy
        fx_cell = ws.cell(10, 2, d.get("fx_rate", 1.0))
        fx_cell.font = FONT_BLUE if not fx_lock else FONT_BLACK
        fx_cell.fill = FILL_INPUT_GREY
        fx_cell.number_format = FMT_FX
        fx_comment_text = (
            f"FX 汇率: 1 {trd_ccy} = X {rep_ccy}\n"
            f"来源: {d.get('fx_source', 'N/A')}\n"
            f"日期: {date.today().isoformat()}\n"
            f"用户可覆盖此单元格以使用实时/自定义汇率。"
        )
        if fx_lock:
            fx_comment_text = (
                f"交易币种 = 报表币种 = {trd_ccy}, 无需汇率换算 (锁定为 1.0)。\n" + fx_comment_text
            )
        add_comment(fx_cell, fx_comment_text)
        ws.cell(10, 1, f"FX Rate: 1 {trd_ccy} = X {rep_ccy} -- 汇率")

        # Row 11: Reporting Stock Price = Trading Price × FX
        ws.cell(11, 1, f"Current Stock Price ({rep_ccy}) -- 当前股价(报表币种)")
        c = ws.cell(11, 2, "=B9*B10"); c.font = FONT_BLACK; c.number_format = FMT_PRICE
        add_comment(c, f"计算公式:\n  Stock Price ({rep_ccy}) = Stock Price ({trd_ccy}) × FX Rate\n  = B9 × B10")

        # Row 12: Shares Outstanding
        self._apply_input(ws.cell(12, 2), d["shares_outstanding"], FMT_SHARES,
                          "Futu get_market_snapshot", "issued_shares (÷1M)")
        ws.cell(12, 1, "Diluted Shares Outstanding (M) -- 稀释后总股本(百万)")

        # Row 13: Market Cap (in reporting currency)
        ws.cell(13, 1, f"Market Capitalization ({rep_ccy} M) -- 市值(报表币种,百万)")
        c = ws.cell(13, 2, "=B11*B12"); c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        add_comment(c, f"计算公式:\n  Market Cap ({rep_ccy} M) = Stock Price ({rep_ccy}) × Shares\n  = B11 × B12")

        # Row 14: Net Debt (already in reporting currency, from financial statements)
        # Net Debt = 短期借款 + 长期借款 (含融资租赁) − 现金及等价物 (含短期投资/定期存款流动+非流动)
        self._apply_input(ws.cell(14, 2), d["debt"] - d["cash"], FMT_CURRENCY_M,
                          f"富途「资产负债表」最新一年 ({rep_ccy})",
                          "短期借款 + 长期借款 − (现金及等价物 + 短期投资 + 定期存款(流动+非流动))")
        ws.cell(14, 1, f"Net Debt / (Net Cash) ({rep_ccy} M) -- 净债务/(净现金)")

        # 关键行号索引 (供后续区块引用) — 用命名字典让下游公式不必硬编码行号
        rows = {
            "case_selector": 4,
            "stock_price_trading": 9,      # 交易币种股价 (Futu 原始)
            "fx_rate": 10,                 # FX
            "stock_price": 11,             # 报表币种股价 (用于 Upside 比对)
            "shares": 12,
            "market_cap": 13,
            "net_debt": 14,
        }

        # --------- Section 4: 3 Scenario Blocks + Consolidation ---------
        # 列布局: A=标签; B..F = 5 期历史 (右对齐到 F, 不足 5 期时左侧留空); G..K = FY1..FY5 预测
        n_hist_capped = min(n_hist, 5)
        hist_start_col = 2 + (5 - n_hist_capped)              # 使最新历史右对齐到 F
        HIST_COLS = list(range(hist_start_col, hist_start_col + n_hist_capped))
        FCST_COLS = list(range(7, 12))                        # G..K
        # 供 section header 合并使用
        SEC_MERGE_END = 11                                    # A..K
        hist_series_offset = n_hist - n_hist_capped           # 若历史多于 5 期, 只取最近 5 期
        latest_hist_col = get_column_letter(HIST_COLS[-1]) if HIST_COLS else "B"

        row_ptr = 16   # Net Debt 位于 Row 14, 留空一行开始三情景块
        scenario_rows = {}  # {scenario_name: {assumption_key: excel_row}}

        # 历史列可展示的假设 key (Terminal Growth / WACC / NWC% 无历史数据)
        HIST_KEYS = {"rev_growth", "ebit_margin", "da_pct", "capex_pct", "tax_rate"}

        for scenario in ["Bear", "Base", "Bull"]:
            # section header
            c = ws.cell(row_ptr, 1, f"{scenario.upper()} CASE ASSUMPTIONS -- {'悲观' if scenario=='Bear' else ('基准' if scenario=='Base' else '乐观')}情景假设")
            c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
            ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=SEC_MERGE_END)
            row_ptr += 1

            # sub-header row: Assumption | Hist(fys) | FY1..FY5
            ws.cell(row_ptr, 1, "Assumption -- 假设项").font = FONT_BOLD
            ws.cell(row_ptr, 1).fill = FILL_LIGHT_BLUE
            for i, col in enumerate(HIST_COLS):
                fy_label = hist_fys[hist_series_offset + i]
                c = ws.cell(row_ptr, col, f"{fy_label} (A)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
            for i, col in enumerate(FCST_COLS):
                c = ws.cell(row_ptr, col, f"FY{i+1} (E)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
            row_ptr += 1

            # 8 assumption rows
            scenario_rows[scenario] = {}
            for label, key, fmt in self.ASSUMPTION_LABELS:
                ws.cell(row_ptr, 1, label)
                # 历史列: 展示实际比率 (若可用) —— 3 个情景块显示相同的历史实际值
                if key in HIST_KEYS:
                    hist_key_map = {
                        "rev_growth":  "hist_rev_growth",
                        "ebit_margin": "hist_ebit_margin",
                        "da_pct":      "hist_da_pct",
                        "capex_pct":   "hist_capex_pct",
                        "tax_rate":    "hist_tax_rate",
                    }
                    hist_data = d.get(hist_key_map[key], [])
                    for i, col in enumerate(HIST_COLS):
                        hv = hist_data[hist_series_offset + i] if hist_series_offset + i < len(hist_data) else None
                        if hv is None:
                            continue
                        c = ws.cell(row_ptr, col, hv); c.font = FONT_BLACK; c.number_format = fmt
                # 预测列: 情景假设输入 (蓝色)
                vals = self.scenarios[scenario][key]
                for i, col in enumerate(FCST_COLS):
                    v = vals[i]
                    if v is None:
                        # Base WACC 引用 WACC Sheet
                        if scenario == "Base" and key == "wacc" and i == 0:
                            c = ws.cell(row_ptr, col, "=WACC!B18"); c.font = FONT_GREEN
                            c.number_format = fmt
                            add_comment(c, "跨 Sheet 引用:\n  = WACC!B18\n(由 WACC 表 CAPM 计算得到)")
                        # 其他 None (Terminal Growth/WACC 只填第 1 列): 留空
                        continue
                    c = ws.cell(row_ptr, col, v); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY
                    c.number_format = fmt
                scenario_rows[scenario][key] = row_ptr
                row_ptr += 1
            row_ptr += 1  # 空行分隔

        # ---- Selected Case Consolidation Block ----
        c = ws.cell(row_ptr, 1, "SELECTED CASE ASSUMPTIONS -- 选中情景 (由 Case Selector 驱动)")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=SEC_MERGE_END)
        row_ptr += 1
        ws.cell(row_ptr, 1, "Assumption -- 假设项").font = FONT_BOLD
        ws.cell(row_ptr, 1).fill = FILL_LIGHT_BLUE
        for i, col in enumerate(HIST_COLS):
            fy_label = hist_fys[hist_series_offset + i]
            c = ws.cell(row_ptr, col, f"{fy_label} (A)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        for i, col in enumerate(FCST_COLS):
            c = ws.cell(row_ptr, col, f"FY{i+1} (E)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        row_ptr += 1

        # consolidation rows (历史列同上 Base 块引用; 预测列 CHOOSE)
        selected_rows = {}
        case_selector_ref = f"$B${rows['case_selector']}"
        for label, key, fmt in self.ASSUMPTION_LABELS:
            ws.cell(row_ptr, 1, label)
            bear_row = scenario_rows["Bear"][key]
            base_row = scenario_rows["Base"][key]
            bull_row = scenario_rows["Bull"][key]
            # 历史列: 直接引用 Base 块的历史值 (紫色, 3 情景历史相同)
            if key in HIST_KEYS:
                for col in HIST_COLS:
                    col_letter = get_column_letter(col)
                    src = f"{col_letter}{base_row}"
                    c = ws.cell(row_ptr, col, f"={src}"); c.font = FONT_PURPLE; c.number_format = fmt
            # 预测列: CHOOSE 聚合
            for i, col in enumerate(FCST_COLS):
                col_letter = get_column_letter(col)
                if key in ("wacc", "terminal_growth") and i > 0:
                    continue
                formula = f"=CHOOSE({case_selector_ref},{col_letter}{bear_row},{col_letter}{base_row},{col_letter}{bull_row})"
                c = ws.cell(row_ptr, col, formula); c.font = FONT_BLACK
                c.fill = FILL_FORECAST_GREEN; c.number_format = fmt
                add_comment(c, f"聚合公式:\n  =CHOOSE({case_selector_ref}, Bear, Base, Bull)\n  = CHOOSE({case_selector_ref}, {col_letter}{bear_row}, {col_letter}{base_row}, {col_letter}{bull_row})\n\n所有下游公式引用本行, 避免散落 IF 嵌套。")
            selected_rows[key] = row_ptr
            row_ptr += 1
        row_ptr += 1  # 空行

        # 保存关键引用行号
        rows["selected"] = selected_rows

        # --------- Section 5: Historical & Projected Financials ---------
        # 列布局: A=标签, B..F = 5 期历史 (最新历史在 F), G..K = FY1..FY5 预测
        c = ws.cell(row_ptr, 1, "HISTORICAL & PROJECTED FINANCIALS -- 历史与预测财务")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=SEC_MERGE_END)
        row_ptr += 1

        # sub-header
        ws.cell(row_ptr, 1, "Line Item (M) -- 项目").font = FONT_BOLD
        ws.cell(row_ptr, 1).fill = FILL_LIGHT_BLUE
        for i, col in enumerate(HIST_COLS):
            fy_label = hist_fys[hist_series_offset + i]
            c = ws.cell(row_ptr, col, f"{fy_label} (A)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        for i, col in enumerate(FCST_COLS):
            c = ws.cell(row_ptr, col, f"FY{i+1} (E)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        row_ptr += 1

        # Helper: 写入历史序列 (蓝色输入)
        def _write_hist_series(row, series):
            for i, col in enumerate(HIST_COLS):
                v = series[hist_series_offset + i] if hist_series_offset + i < len(series) else None
                if v is None: continue
                c = ws.cell(row, col, v); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY
                c.number_format = FMT_CURRENCY_M

        # ---- Row: Revenue ----
        rev_row = row_ptr
        ws.cell(rev_row, 1, "Revenue -- 营业收入")
        _write_hist_series(rev_row, d["hist_revenue"])
        # 首个预测年基于最后一个历史年
        prev_col_letter = latest_hist_col
        for i, col in enumerate(FCST_COLS):
            col_letter = get_column_letter(col)
            growth_ref = f"{col_letter}{selected_rows['rev_growth']}"
            c = ws.cell(rev_row, col, f"={prev_col_letter}{rev_row}*(1+{growth_ref})")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"计算公式:\n  Revenue(FY{i+1}) = Revenue(前年) × (1 + Growth%)\n  = {prev_col_letter}{rev_row} × (1 + {growth_ref})")
            prev_col_letter = col_letter
        row_ptr += 1

        # ---- Row: % Revenue Growth ----
        growth_disp_row = row_ptr
        ws.cell(growth_disp_row, 1, "  % Revenue Growth -- 营收增速")
        # 历史列: 用 (本年-上年)/上年 计算
        for i, col in enumerate(HIST_COLS):
            if i == 0:
                # 最老的历史年可能有上一年数据 (来自 prior_to_hist), 简化: 用 hist_rev_growth 数据
                hv = d["hist_rev_growth"][hist_series_offset + i]
                if hv is not None:
                    c = ws.cell(growth_disp_row, col, hv); c.font = FONT_BLACK; c.number_format = FMT_PERCENT
            else:
                prev_col = get_column_letter(HIST_COLS[i-1])
                cur_col = get_column_letter(col)
                c = ws.cell(growth_disp_row, col, f"={cur_col}{rev_row}/{prev_col}{rev_row}-1")
                c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        # 预测列: 引用 selected consolidation
        for i, col in enumerate(FCST_COLS):
            col_letter = get_column_letter(col)
            c = ws.cell(growth_disp_row, col, f"={col_letter}{selected_rows['rev_growth']}")
            c.font = FONT_PURPLE; c.number_format = FMT_PERCENT
            add_comment(c, f"同 Sheet 引用: 选中情景 Revenue Growth\n  = {col_letter}{selected_rows['rev_growth']}")
        row_ptr += 1

        # ---- Row: EBIT ----
        ebit_row = row_ptr
        ws.cell(ebit_row, 1, "EBIT -- 息税前利润")
        _write_hist_series(ebit_row, d["hist_ebit"])
        for i, col in enumerate(FCST_COLS):
            col_letter = get_column_letter(col)
            margin_ref = f"{col_letter}{selected_rows['ebit_margin']}"
            c = ws.cell(ebit_row, col, f"={col_letter}{rev_row}*{margin_ref}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"计算公式:\n  EBIT = Revenue × EBIT Margin\n  = {col_letter}{rev_row} × {margin_ref}")
        row_ptr += 1

        # ---- Row: % EBIT Margin (计算显示行, 全部用公式) ----
        margin_disp_row = row_ptr
        ws.cell(margin_disp_row, 1, "  % EBIT Margin -- 息税前利润率")
        for col in HIST_COLS + FCST_COLS:
            col_letter = get_column_letter(col)
            c = ws.cell(margin_disp_row, col, f"={col_letter}{ebit_row}/{col_letter}{rev_row}")
            c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        row_ptr += 1

        # ---- Row: Tax Rate (历史展示 + 预测引用) ----
        tax_disp_row = row_ptr
        ws.cell(tax_disp_row, 1, "Tax Rate -- 税率")
        for i, col in enumerate(HIST_COLS):
            hv = d["hist_tax_rate"][hist_series_offset + i]
            c = ws.cell(tax_disp_row, col, hv); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = FMT_PERCENT
        for i, col in enumerate(FCST_COLS):
            col_letter = get_column_letter(col)
            c = ws.cell(tax_disp_row, col, f"={col_letter}{selected_rows['tax_rate']}")
            c.font = FONT_PURPLE; c.number_format = FMT_PERCENT
        row_ptr += 1

        # ---- Row: Cash Taxes = -MAX(0, EBIT) × Tax Rate ----
        tax_row = row_ptr
        ws.cell(tax_row, 1, "Less: Cash Taxes -- 现金税额 (亏损不缴)")
        for col in HIST_COLS + FCST_COLS:
            col_letter = get_column_letter(col)
            c = ws.cell(tax_row, col, f"=-MAX(0,{col_letter}{ebit_row})*{col_letter}{tax_disp_row}")
            c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
            if col in FCST_COLS:
                c.fill = FILL_FORECAST_GREEN
            add_comment(c, f"计算公式:\n  Cash Taxes = -MAX(0, EBIT) × Tax Rate\n  = -MAX(0, {col_letter}{ebit_row}) × {col_letter}{tax_disp_row}\n  (EBIT 为负时不缴税)")
        row_ptr += 1

        # ---- Row: NOPAT = EBIT + Cash Taxes ----
        nopat_row = row_ptr
        ws.cell(nopat_row, 1, "NOPAT -- 税后净营业利润").font = FONT_BOLD
        for col in HIST_COLS + FCST_COLS:
            col_letter = get_column_letter(col)
            c = ws.cell(nopat_row, col, f"={col_letter}{ebit_row}+{col_letter}{tax_row}")
            c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
            if col in FCST_COLS:
                c.fill = FILL_FORECAST_GREEN
            add_comment(c, f"计算公式:\n  NOPAT = EBIT + Cash Taxes\n  = {col_letter}{ebit_row} + {col_letter}{tax_row}\n  (Cash Taxes 已带负号)")
        row_ptr += 2

        # --------- Section 6: Free Cash Flow Build ---------
        c = ws.cell(row_ptr, 1, "FREE CASH FLOW BUILD -- 自由现金流构建"); c.font = FONT_BOLD
        c.fill = FILL_LIGHT_BLUE
        row_ptr += 1

        # NOPAT reference row
        nopat_ref_row = row_ptr
        ws.cell(nopat_ref_row, 1, "NOPAT (from above)")
        for col in HIST_COLS + FCST_COLS:
            col_letter = get_column_letter(col)
            c = ws.cell(nopat_ref_row, col, f"={col_letter}{nopat_row}")
            c.font = FONT_PURPLE; c.number_format = FMT_CURRENCY_M
        row_ptr += 1

        # D&A (历史实际, 预测 = Revenue × D&A%)
        da_row = row_ptr
        ws.cell(da_row, 1, "Plus: D&A -- 折旧摊销 (加回)")
        _write_hist_series(da_row, d["hist_da"])
        for i, col in enumerate(FCST_COLS):
            col_letter = get_column_letter(col)
            da_pct_ref = f"{col_letter}{selected_rows['da_pct']}"
            c = ws.cell(da_row, col, f"={col_letter}{rev_row}*{da_pct_ref}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"计算公式:\n  D&A = Revenue × D&A%\n  = {col_letter}{rev_row} × {da_pct_ref}")
        row_ptr += 1

        # CapEx (历史用负号展示; 预测 = -Revenue × CapEx%)
        capex_row = row_ptr
        ws.cell(capex_row, 1, "Less: CapEx -- 资本开支")
        # 历史 CapEx: 富途数据已是正数, 这里取负展示为现金流出
        for i, col in enumerate(HIST_COLS):
            hv = d["hist_capex"][hist_series_offset + i]
            c = ws.cell(capex_row, col, -abs(hv)); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = FMT_CURRENCY_M
        for i, col in enumerate(FCST_COLS):
            col_letter = get_column_letter(col)
            capex_pct_ref = f"{col_letter}{selected_rows['capex_pct']}"
            c = ws.cell(capex_row, col, f"=-{col_letter}{rev_row}*{capex_pct_ref}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"计算公式:\n  CapEx = -Revenue × CapEx%\n  = -{col_letter}{rev_row} × {capex_pct_ref}\n  (负号表示现金流出)")
        row_ptr += 1

        # ΔNWC (历史第一年无 prior 置 0, 其余按 Δrev × NWC% 展示; 预测同)
        nwc_row = row_ptr
        ws.cell(nwc_row, 1, "Less: Δ Working Capital -- 营运资本变动")
        # 历史 ΔNWC: 用固定 NWC% (Base 情景 FY1 值) 近似
        base_nwc_pct = self.scenarios["Base"]["nwc_pct"][0]  # 1% 默认
        for i, col in enumerate(HIST_COLS):
            if i == 0:
                c = ws.cell(nwc_row, col, 0); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = FMT_CURRENCY_M
            else:
                cur_col = get_column_letter(col)
                prev_col = get_column_letter(HIST_COLS[i-1])
                c = ws.cell(nwc_row, col, f"=-({cur_col}{rev_row}-{prev_col}{rev_row})*{base_nwc_pct}")
                c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
                add_comment(c, f"计算公式 (历史近似):\n  ΔNWC = -(本年 Revenue - 上年 Revenue) × NWC% (Base 假设)\n  = -({cur_col}{rev_row} - {prev_col}{rev_row}) × {base_nwc_pct}")
        # 预测: 用 selected 的 NWC%
        prev_col_letter = latest_hist_col
        for i, col in enumerate(FCST_COLS):
            col_letter = get_column_letter(col)
            nwc_pct_ref = f"{col_letter}{selected_rows['nwc_pct']}"
            c = ws.cell(nwc_row, col, f"=-({col_letter}{rev_row}-{prev_col_letter}{rev_row})*{nwc_pct_ref}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"计算公式:\n  ΔNWC = -(本年 Revenue - 前年 Revenue) × NWC%\n  = -({col_letter}{rev_row} - {prev_col_letter}{rev_row}) × {nwc_pct_ref}")
            prev_col_letter = col_letter
        row_ptr += 1

        # Unlevered FCF = NOPAT + D&A + CapEx + ΔNWC (所有年份)
        fcf_row = row_ptr
        ws.cell(fcf_row, 1, "Unlevered FCF -- 无杠杆自由现金流").font = FONT_BOLD
        for col in HIST_COLS + FCST_COLS:
            col_letter = get_column_letter(col)
            c = ws.cell(fcf_row, col, f"={col_letter}{nopat_ref_row}+{col_letter}{da_row}+{col_letter}{capex_row}+{col_letter}{nwc_row}")
            c.font = FONT_BOLD; c.number_format = FMT_CURRENCY_M; c.border = BOTTOM_BORDER
            if col in FCST_COLS:
                c.fill = FILL_FORECAST_GREEN
            add_comment(c, f"计算公式:\n  Unlevered FCF = NOPAT + D&A + CapEx + ΔNWC\n  = {col_letter}{nopat_ref_row} + {col_letter}{da_row} + {col_letter}{capex_row} + {col_letter}{nwc_row}")
        row_ptr += 2

        # --------- Section 7: Discounting & Terminal Value ---------
        c = ws.cell(row_ptr, 1, "DISCOUNTING & TERMINAL VALUE -- 折现与终值")
        c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        row_ptr += 1

        first_fcst_letter = get_column_letter(FCST_COLS[0])

        # WACC & TGR references (显示行, 从 selected consolidation 取)
        wacc_ref_row = row_ptr
        ws.cell(wacc_ref_row, 1, "WACC (from Selected Case)")
        c = ws.cell(wacc_ref_row, 2, f"={first_fcst_letter}{selected_rows['wacc']}"); c.font = FONT_PURPLE; c.number_format = FMT_PERCENT
        add_comment(c, f"同 Sheet 引用: 选中情景 WACC (FY1 列)\n  = {first_fcst_letter}{selected_rows['wacc']}\n\n注: WACC 通常由 WACC Sheet 计算, 三情景中的 Base WACC 直接跨表引用 WACC!B18。")
        row_ptr += 1

        tgr_ref_row = row_ptr
        ws.cell(tgr_ref_row, 1, "Terminal Growth Rate (from Selected Case)")
        c = ws.cell(tgr_ref_row, 2, f"={first_fcst_letter}{selected_rows['terminal_growth']}"); c.font = FONT_PURPLE; c.number_format = FMT_PERCENT
        add_comment(c, f"同 Sheet 引用: 选中情景 Terminal Growth (FY1 列)\n  = {first_fcst_letter}{selected_rows['terminal_growth']}")
        row_ptr += 1

        # Discount Period (Mid-year convention)
        period_row = row_ptr
        ws.cell(period_row, 1, "Discount Period (Mid-year)")
        for i, col in enumerate(FCST_COLS):
            c = ws.cell(period_row, col, i + 0.5); c.font = FONT_BLUE; c.fill = FILL_FORECAST_GREEN
            c.number_format = FMT_DECIMAL
            if i == 0:
                add_comment(c, "半年惯例 (Mid-year convention):\n  FY1=0.5, FY2=1.5, ..., FY5=4.5\n  假设现金流均匀发生在会计年度中期。")
        row_ptr += 1

        # Discount Factor
        df_row = row_ptr
        ws.cell(df_row, 1, "Discount Factor")
        wacc_abs = f"$B${wacc_ref_row}"
        for i, col in enumerate(FCST_COLS):
            col_letter = get_column_letter(col)
            c = ws.cell(df_row, col, f"=1/(1+{wacc_abs})^{col_letter}{period_row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_DECIMAL
            add_comment(c, f"计算公式:\n  DF = 1 / (1+WACC)^period\n  = 1 / (1+{wacc_abs})^{col_letter}{period_row}")
        row_ptr += 1

        # PV of FCF
        pv_fcf_row = row_ptr
        ws.cell(pv_fcf_row, 1, "PV of FCF -- 自由现金流现值")
        for i, col in enumerate(FCST_COLS):
            col_letter = get_column_letter(col)
            c = ws.cell(pv_fcf_row, col, f"={col_letter}{fcf_row}*{col_letter}{df_row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"计算公式:\n  PV = FCF × Discount Factor\n  = {col_letter}{fcf_row} × {col_letter}{df_row}")
        row_ptr += 2

        # --------- Section 8: Valuation Summary ---------
        c = ws.cell(row_ptr, 1, "VALUATION SUMMARY -- 估值汇总")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=SEC_MERGE_END)
        row_ptr += 1
        val_start = row_ptr

        last_fcst = get_column_letter(FCST_COLS[-1])
        first_fcst = get_column_letter(FCST_COLS[0])

        # PV of Explicit FCFs
        pv_sum_row = row_ptr
        ws.cell(pv_sum_row, 1, "(+) PV of Explicit FCFs -- 显性期 FCF 现值合计")
        c = ws.cell(pv_sum_row, 2, f"=SUM({first_fcst}{pv_fcf_row}:{last_fcst}{pv_fcf_row})")
        c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        add_comment(c, f"计算公式:\n  = SUM({first_fcst}{pv_fcf_row}:{last_fcst}{pv_fcf_row})\n  (FY1..FY5 现值合计)")
        row_ptr += 1

        # Terminal FCF (Year 5+1)
        term_fcf_row = row_ptr
        ws.cell(term_fcf_row, 1, "Terminal FCF (Year 5+1)")
        c = ws.cell(term_fcf_row, 2, f"={last_fcst}{fcf_row}*(1+B{tgr_ref_row})")
        c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        add_comment(c, f"计算公式:\n  Terminal FCF = FY5 FCF × (1 + Terminal Growth)\n  = {last_fcst}{fcf_row} × (1 + B{tgr_ref_row})")
        row_ptr += 1

        # Terminal Value (Gordon Growth)
        tv_row = row_ptr
        ws.cell(tv_row, 1, "Terminal Value -- 终值 (Gordon Growth Model)")
        c = ws.cell(tv_row, 2, f"=B{term_fcf_row}/(B{wacc_ref_row}-B{tgr_ref_row})")
        c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        add_comment(c, f"计算公式 (Gordon Growth Model):\n  TV = Terminal FCF / (WACC - TGR)\n  = B{term_fcf_row} / (B{wacc_ref_row} - B{tgr_ref_row})\n\n约束: TGR 必须 < WACC, 否则分母为负/零导致无限值.")
        row_ptr += 1

        # PV of Terminal Value
        pv_tv_row = row_ptr
        ws.cell(pv_tv_row, 1, "(+) PV of Terminal Value -- 终值现值")
        c = ws.cell(pv_tv_row, 2, f"=B{tv_row}*{last_fcst}{df_row}")
        c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        add_comment(c, f"计算公式:\n  PV of TV = Terminal Value × DF(FY5)\n  = B{tv_row} × {last_fcst}{df_row}")
        row_ptr += 1

        # Enterprise Value
        ev_row = row_ptr
        ws.cell(ev_row, 1, f"  Enterprise Value ({rep_ccy} M) -- 企业价值 (=)").font = FONT_BOLD
        c = ws.cell(ev_row, 2, f"=B{pv_sum_row}+B{pv_tv_row}")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = FMT_CURRENCY_M
        add_comment(c, f"计算公式:\n  EV = PV of Explicit FCFs + PV of Terminal Value\n  = B{pv_sum_row} + B{pv_tv_row}")
        row_ptr += 1

        # TV / EV ratio (sanity check: 建议 50-70%)
        tv_ev_row = row_ptr
        ws.cell(tv_ev_row, 1, "  Terminal Value / EV -- 终值占比 (合理区间 50–70%)")
        c = ws.cell(tv_ev_row, 2, f"=B{pv_tv_row}/B{ev_row}")
        c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        add_comment(c, "合理性检验:\n  50–70%: 正常\n  >75%: 过度依赖远期假设 (红旗)\n  <40%: 终值假设过于保守")
        row_ptr += 1

        # Less: Net Debt
        nd_row = row_ptr
        ws.cell(nd_row, 1, "(-) Net Debt / (Net Cash)")
        c = ws.cell(nd_row, 2, f"=B{rows['net_debt']}"); c.font = FONT_PURPLE; c.number_format = FMT_CURRENCY_M
        add_comment(c, f"同 Sheet 引用: Market Data 净债务\n  = B{rows['net_debt']}")
        row_ptr += 1

        # Equity Value
        eq_row = row_ptr
        ws.cell(eq_row, 1, f"  Equity Value ({rep_ccy} M) -- 股权价值 (=)").font = FONT_BOLD
        c = ws.cell(eq_row, 2, f"=B{ev_row}-B{nd_row}")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = FMT_CURRENCY_M
        add_comment(c, f"计算公式:\n  Equity Value = EV - Net Debt\n  = B{ev_row} - B{nd_row}")
        row_ptr += 1

        # ÷ Shares
        shares_row = row_ptr
        ws.cell(shares_row, 1, "÷ Diluted Shares Outstanding (M)")
        c = ws.cell(shares_row, 2, f"=B{rows['shares']}"); c.font = FONT_PURPLE; c.number_format = FMT_SHARES
        add_comment(c, f"同 Sheet 引用: Market Data 稀释后总股本\n  = B{rows['shares']}")
        row_ptr += 1

        # Implied Price per Share
        implied_row = row_ptr
        ws.cell(implied_row, 1, f"  Implied Price per Share ({rep_ccy}) -- 每股内在价值 (=)").font = FONT_BOLD
        c = ws.cell(implied_row, 2, f"=B{eq_row}/B{shares_row}")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = FMT_PRICE
        add_comment(c, f"计算公式:\n  Implied Price = Equity Value / Shares\n  = B{eq_row} / B{shares_row}")
        row_ptr += 1

        # Current Price + Upside (使用报表币种口径以保证与 Implied Price 同币比对)
        cur_row = row_ptr
        ws.cell(cur_row, 1, f"Current Stock Price ({rep_ccy}) -- 当前股价(报表币种)")
        c = ws.cell(cur_row, 2, f"=B{rows['stock_price']}"); c.font = FONT_PURPLE; c.number_format = FMT_PRICE
        add_comment(c, f"同 Sheet 引用: 报表币种下的当前股价 (Trading Price × FX)\n  = B{rows['stock_price']}")
        row_ptr += 1

        upside_row = row_ptr
        ws.cell(upside_row, 1, "Implied Upside / (Downside) -- 隐含涨跌空间").font = FONT_BOLD
        c = ws.cell(upside_row, 2, f"=B{implied_row}/B{cur_row}-1")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = FMT_PERCENT
        add_comment(c, f"计算公式:\n  Upside% = Implied Price / Current Price - 1\n  = B{implied_row} / B{cur_row} - 1")
        row_ptr += 1

        # Valuation Summary 区块统一浅橙底纹
        for r in range(val_start, upside_row + 1):
            for col in (1, 2):
                if ws.cell(r, col).fill.fgColor is None or ws.cell(r, col).fill.patternType is None:
                    ws.cell(r, col).fill = FILL_VALUATION_ORANGE
        for r in (ev_row, eq_row, implied_row, upside_row):
            ws.cell(r, 2).fill = FILL_VALUATION_ORANGE_DARK

        row_ptr += 2

        # ===== 保存关键行号供敏感性表引用 =====
        self._dcf_refs = {
            "sheet": ws,
            "rows": rows,
            "selected_rows": selected_rows,
            "rev_row": rev_row,
            "fcf_row": fcf_row,
            "pv_fcf_row": pv_fcf_row,
            "df_row": df_row,
            "period_row": period_row,
            "wacc_ref_row": wacc_ref_row,
            "tgr_ref_row": tgr_ref_row,
            "implied_row": implied_row,
            "fcst_cols": FCST_COLS,
            "hist_cols": HIST_COLS,
            "latest_hist_col_letter": latest_hist_col,
        }

        # --------- Section 9: Sensitivity Analysis (3× 5×5 tables) ---------
        self._sensitivity_tables(ws, row_ptr)

        # 列宽
        ws.column_dimensions["A"].width = 48
        for col in HIST_COLS + FCST_COLS:
            ws.column_dimensions[get_column_letter(col)].width = 13

    # ==================== Sensitivity Tables ====================
    def _sensitivity_tables(self, ws, start_row: int):
        """在 DCF sheet 底部生成 3 张 5×5 敏感性表, 共 75 个完整重算公式。

        Table 1: WACC × Terminal Growth → Implied Price per Share
        Table 2: Revenue Growth × EBIT Margin → Implied Price per Share
        Table 3: Beta × Risk-Free Rate → Implied Price per Share

        每格用闭式公式重算隐含股价 (不依赖 Data Table 功能):
          Implied Price ≈ (PV_Explicit_FCFs + PV_TV - NetDebt) / Shares
        为保持公式独立性且中心格 = 主模型输出, 用如下近似:
          Explicit FCFs 视为常量 (不随敏感变量变化, 用 Base 值)
          仅 TV / WACC / TGR / DF(FY5) 或对应敏感变量重算
        """
        refs = self._dcf_refs
        selected_rows = refs["selected_rows"]
        rev_row = refs["rev_row"]
        fcf_row = refs["fcf_row"]
        pv_fcf_row = refs["pv_fcf_row"]
        wacc_ref_row = refs["wacc_ref_row"]
        tgr_ref_row = refs["tgr_ref_row"]
        rows = refs["rows"]
        FCST_COLS = refs["fcst_cols"]
        first_fcst = get_column_letter(FCST_COLS[0])
        last_fcst = get_column_letter(FCST_COLS[-1])

        row_ptr = start_row

        # ---- Table 1: WACC × Terminal Growth → Implied Price ----
        c = ws.cell(row_ptr, 1, "SENSITIVITY TABLE 1: WACC × Terminal Growth → Implied Share Price")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=7)
        row_ptr += 1

        # 轴值围绕 base 对称 ±2 步长
        # base wacc = B{wacc_ref_row}, base tgr = B{tgr_ref_row}
        # 我们用固定步长 0.5% (wacc) / 0.5% (tgr)
        # 行头 (WACC): 5 个值, 列 B
        # 列头 (TGR): 5 个值, 行 row_ptr
        wacc_step = 0.005
        tgr_step = 0.005
        row_ptr = self._draw_sens_table_wacc_tgr(ws, row_ptr, wacc_step, tgr_step)
        row_ptr += 2

        # ---- Table 2: Revenue Growth (FY1) × EBIT Margin (FY1) → Implied Price ----
        c = ws.cell(row_ptr, 1, "SENSITIVITY TABLE 2: Revenue Growth (FY1) × EBIT Margin (FY1) → Implied Share Price")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=7)
        row_ptr += 1
        row_ptr = self._draw_sens_table_growth_margin(ws, row_ptr, 0.02, 0.02)
        row_ptr += 2

        # ---- Table 3: Beta × Risk-Free Rate → Implied Price ----
        c = ws.cell(row_ptr, 1, "SENSITIVITY TABLE 3: Beta × Risk-Free Rate → Implied Share Price")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=7)
        row_ptr += 1
        row_ptr = self._draw_sens_table_beta_rf(ws, row_ptr, 0.15, 0.005)

    # ---- Sensitivity table 1 helper ----
    def _draw_sens_table_wacc_tgr(self, ws, start_row: int, wacc_step: float, tgr_step: float) -> int:
        """5×5: WACC × TGR → Implied Price.
        闭式公式:
          For each (wacc, tgr):
            DF_i = 1/(1+wacc)^period_i  (period = 0.5..4.5)
            PV_i = FCF_i × DF_i           (FCF_i 已定)
            PV_sum = Σ PV_i
            TV = FCF_5 × (1+tgr) / (wacc - tgr)
            PV_TV = TV × DF_5
            EV = PV_sum + PV_TV
            Equity = EV - NetDebt
            Implied = Equity / Shares
        中心格 = base WACC × base TGR = 主模型输出.
        """
        refs = self._dcf_refs
        rows = refs["rows"]
        fcf_row = refs["fcf_row"]
        wacc_ref_row = refs["wacc_ref_row"]
        tgr_ref_row = refs["tgr_ref_row"]
        FCST_COLS = refs["fcst_cols"]
        # 5 期 FCF 引用: C{fcf_row} .. G{fcf_row}
        fcf_cells = [f"{get_column_letter(c)}{fcf_row}" for c in FCST_COLS]
        periods = [0.5, 1.5, 2.5, 3.5, 4.5]
        net_debt = f"$B${rows['net_debt']}"
        shares = f"$B${rows['shares']}"
        base_wacc_cell = f"$B${wacc_ref_row}"
        base_tgr_cell = f"$B${tgr_ref_row}"

        # 表头行 (start_row):列 B 空, C..G = TGR 轴值
        hdr_row = start_row
        ws.cell(hdr_row, 1, "WACC \\ TGR").font = FONT_BOLD
        ws.cell(hdr_row, 1).fill = FILL_LIGHT_BLUE
        for j in range(5):
            offset = j - 2  # -2, -1, 0, +1, +2
            col = 2 + j  # B..F  (但 B 是行头, 我们让 TGR 从 C 开始)
        # 重新布局: 列头 TGR 在 C..G, 行头 WACC 在 B, 数据在 C..G × start_row+1..+5
        for j in range(5):
            offset = j - 2
            tgr_formula = f"={base_tgr_cell}+{offset}*{tgr_step}"
            c = ws.cell(hdr_row, 3 + j, tgr_formula); c.font = FONT_BLUE; c.fill = FILL_LIGHT_BLUE; c.number_format = FMT_PERCENT
        for i in range(5):
            offset = i - 2
            r = hdr_row + 1 + i
            wacc_formula = f"={base_wacc_cell}+{offset}*{wacc_step}"
            c = ws.cell(r, 2, wacc_formula); c.font = FONT_BLUE; c.fill = FILL_LIGHT_BLUE; c.number_format = FMT_PERCENT
            for j in range(5):
                cc = 3 + j
                col_letter = get_column_letter(cc)
                # WACC 取行头 B{r} (绝对列), TGR 取列头 col{hdr_row} (绝对行)
                wacc_ref = f"$B{r}"
                tgr_ref = f"{col_letter}${hdr_row}"
                # 5 期 DF
                df_terms = [f"({fcf_cells[k]}/(1+{wacc_ref})^{periods[k]})" for k in range(5)]
                pv_sum = "+".join(df_terms)
                # TV = FCF5 * (1+TGR) / (WACC - TGR)
                tv = f"{fcf_cells[-1]}*(1+{tgr_ref})/({wacc_ref}-{tgr_ref})"
                pv_tv = f"({tv})/(1+{wacc_ref})^{periods[-1]}"
                ev = f"({pv_sum})+({pv_tv})"
                equity = f"({ev})-{net_debt}"
                implied = f"IFERROR(({equity})/{shares},NA())"
                formula = f"={implied}"
                c = ws.cell(r, cc, formula)
                c.font = FONT_BLACK; c.number_format = FMT_PRICE
                if i == 2 and j == 2:
                    c.font = FONT_BLACK_BOLD; c.fill = FILL_MEDIUM_BLUE
                    add_comment(c, "★ 中心格 = Base case Implied Price\n  Base WACC × Base TGR\n  必须等于 Valuation Summary 输出值")
        return hdr_row + 5

    # ---- Sensitivity table 2 helper ----
    def _draw_sens_table_growth_margin(self, ws, start_row: int, growth_step: float, margin_step: float) -> int:
        """5×5: Revenue Growth (FY1) × EBIT Margin (FY1) → Implied Price.
        闭式近似:
          假设 FY1 Growth/Margin 变化, FY2..FY5 保持选中情景的原有序列.
          FCF_1 = Rev_LTM * (1+g_1) * m_1 * (1-tax) + D&A_1 - CapEx_1 - ΔNWC_1
          简化: 用 base FCF_2..FCF_5 作为常量, 仅 FCF_1 随 (g,m) 重算.
        """
        refs = self._dcf_refs
        rows = refs["rows"]
        rev_row = refs["rev_row"]
        fcf_row = refs["fcf_row"]
        wacc_ref_row = refs["wacc_ref_row"]
        tgr_ref_row = refs["tgr_ref_row"]
        selected_rows = refs["selected_rows"]
        FCST_COLS = refs["fcst_cols"]

        first_fcst_letter = get_column_letter(FCST_COLS[0])
        base_growth_cell = f"${first_fcst_letter}${selected_rows['rev_growth']}"      # FY1 Growth
        base_margin_cell = f"${first_fcst_letter}${selected_rows['ebit_margin']}"     # FY1 Margin
        # 最后一年历史 Revenue (LTM), 用于 FY1 重算
        rev_ltm = f"${refs['latest_hist_col_letter']}${rev_row}"
        base_wacc_cell = f"$B${wacc_ref_row}"
        base_tgr_cell = f"$B${tgr_ref_row}"
        net_debt = f"$B${rows['net_debt']}"
        shares = f"$B${rows['shares']}"
        # 用 selected 的 tax/da%/capex%/nwc% (FY1 列)
        tax_c = f"${first_fcst_letter}${selected_rows['tax_rate']}"
        da_c = f"${first_fcst_letter}${selected_rows['da_pct']}"
        capex_c = f"${first_fcst_letter}${selected_rows['capex_pct']}"
        nwc_c = f"${first_fcst_letter}${selected_rows['nwc_pct']}"
        # FCF_2..FCF_5 保持 base 值 (直接引用 D{fcf_row}..G{fcf_row})
        fcf_2_to_5 = [f"{get_column_letter(c)}${fcf_row}" for c in FCST_COLS[1:]]
        periods = [0.5, 1.5, 2.5, 3.5, 4.5]

        hdr_row = start_row
        ws.cell(hdr_row, 1, "Growth \\ Margin").font = FONT_BOLD
        ws.cell(hdr_row, 1).fill = FILL_LIGHT_BLUE
        for j in range(5):
            offset = j - 2
            m_formula = f"={base_margin_cell}+{offset}*{margin_step}"
            c = ws.cell(hdr_row, 3 + j, m_formula); c.font = FONT_BLUE; c.fill = FILL_LIGHT_BLUE; c.number_format = FMT_PERCENT
        for i in range(5):
            offset = i - 2
            r = hdr_row + 1 + i
            g_formula = f"={base_growth_cell}+{offset}*{growth_step}"
            c = ws.cell(r, 2, g_formula); c.font = FONT_BLUE; c.fill = FILL_LIGHT_BLUE; c.number_format = FMT_PERCENT
            for j in range(5):
                cc = 3 + j
                col_letter = get_column_letter(cc)
                g_ref = f"$B{r}"
                m_ref = f"{col_letter}${hdr_row}"
                # FY1 Revenue = LTM × (1 + g)
                rev1 = f"({rev_ltm}*(1+{g_ref}))"
                # FY1 EBIT = Rev1 × m
                ebit1 = f"({rev1}*{m_ref})"
                # Cash Tax = -MAX(0, EBIT1) × tax
                tax1 = f"(-MAX(0,{ebit1})*{tax_c})"
                # NOPAT = EBIT + Tax
                nopat1 = f"({ebit1}+{tax1})"
                # D&A = Rev1 × da%
                da1 = f"({rev1}*{da_c})"
                # CapEx = -Rev1 × capex%
                capex1 = f"(-{rev1}*{capex_c})"
                # ΔNWC = -(Rev1 - LTM) × nwc%
                nwc1 = f"(-({rev1}-{rev_ltm})*{nwc_c})"
                fcf1 = f"({nopat1}+{da1}+{capex1}+{nwc1})"
                # PV FCF1..5 (FY2-5 用 base 值)
                pv_terms = [f"({fcf1}/(1+{base_wacc_cell})^{periods[0]})"] + \
                           [f"({fcf_2_to_5[k]}/(1+{base_wacc_cell})^{periods[k+1]})" for k in range(4)]
                pv_sum = "+".join(pv_terms)
                # TV based on FY5 FCF (base 值)
                tv = f"{fcf_2_to_5[-1]}*(1+{base_tgr_cell})/({base_wacc_cell}-{base_tgr_cell})"
                pv_tv = f"({tv})/(1+{base_wacc_cell})^{periods[-1]}"
                ev = f"({pv_sum})+({pv_tv})"
                equity = f"({ev})-{net_debt}"
                implied = f"IFERROR(({equity})/{shares},NA())"
                c = ws.cell(r, cc, f"={implied}")
                c.font = FONT_BLACK; c.number_format = FMT_PRICE
                if i == 2 and j == 2:
                    c.font = FONT_BLACK_BOLD; c.fill = FILL_MEDIUM_BLUE
                    add_comment(c, "★ 中心格 ≈ Base case Implied Price\n  (FY1 Growth × FY1 Margin, 其余期用 base FCF)\n  与主模型可能有微小差异 (仅 FY1 参数变化)")
        return hdr_row + 5

    # ---- Sensitivity table 3 helper ----
    def _draw_sens_table_beta_rf(self, ws, start_row: int, beta_step: float, rf_step: float) -> int:
        """5×5: Beta × Risk-Free Rate → Implied Price.
        每格重算:
          Ke = Rf + β × ERP
          WACC = We × Ke + Wd × Kd*(1-t)
          再用 (WACC, TGR_base) 走完 DCF 得到 Implied Price
        为简化, 直接用 base WACC 输入 (股权成本变化传导至 WACC).
        权重 We/Wd 用 WACC Sheet 引用.
        """
        refs = self._dcf_refs
        rows = refs["rows"]
        fcf_row = refs["fcf_row"]
        wacc_ref_row = refs["wacc_ref_row"]
        tgr_ref_row = refs["tgr_ref_row"]
        FCST_COLS = refs["fcst_cols"]
        fcf_cells = [f"{get_column_letter(c)}${fcf_row}" for c in FCST_COLS]
        periods = [0.5, 1.5, 2.5, 3.5, 4.5]
        net_debt = f"$B${rows['net_debt']}"
        shares = f"$B${rows['shares']}"
        base_tgr_cell = f"$B${tgr_ref_row}"
        # 从 WACC sheet 引用
        we_ref = "WACC!$B$16"   # Equity Weight
        wd_ref = "WACC!$B$17"   # Debt Weight
        kd_after_tax = "WACC!$B$10"  # After-Tax Cost of Debt
        erp_ref = "WACC!$B$4"    # Equity Risk Premium
        base_beta_cell = "WACC!$B$3"
        base_rf_cell = "WACC!$B$2"

        hdr_row = start_row
        ws.cell(hdr_row, 1, "Beta \\ Rf").font = FONT_BOLD
        ws.cell(hdr_row, 1).fill = FILL_LIGHT_BLUE
        for j in range(5):
            offset = j - 2
            rf_formula = f"={base_rf_cell}+{offset}*{rf_step}"
            c = ws.cell(hdr_row, 3 + j, rf_formula); c.font = FONT_BLUE; c.fill = FILL_LIGHT_BLUE; c.number_format = FMT_PERCENT
        for i in range(5):
            offset = i - 2
            r = hdr_row + 1 + i
            beta_formula = f"={base_beta_cell}+{offset}*{beta_step}"
            c = ws.cell(r, 2, beta_formula); c.font = FONT_BLUE; c.fill = FILL_LIGHT_BLUE; c.number_format = "0.00"
            for j in range(5):
                cc = 3 + j
                col_letter = get_column_letter(cc)
                beta_ref = f"$B{r}"
                rf_ref = f"{col_letter}${hdr_row}"
                # Ke = Rf + β × ERP
                ke = f"({rf_ref}+{beta_ref}*{erp_ref})"
                # WACC = We × Ke + Wd × Kd(after-tax)
                wacc_formula = f"({we_ref}*{ke}+{wd_ref}*{kd_after_tax})"
                # DCF revaluation (与 Table 1 相同但 wacc 变量不同)
                df_terms = [f"({fcf_cells[k]}/(1+{wacc_formula})^{periods[k]})" for k in range(5)]
                pv_sum = "+".join(df_terms)
                tv = f"{fcf_cells[-1]}*(1+{base_tgr_cell})/({wacc_formula}-{base_tgr_cell})"
                pv_tv = f"({tv})/(1+{wacc_formula})^{periods[-1]}"
                ev = f"({pv_sum})+({pv_tv})"
                equity = f"({ev})-{net_debt}"
                implied = f"IFERROR(({equity})/{shares},NA())"
                c = ws.cell(r, cc, f"={implied}")
                c.font = FONT_BLACK; c.number_format = FMT_PRICE
                if i == 2 and j == 2:
                    c.font = FONT_BLACK_BOLD; c.fill = FILL_MEDIUM_BLUE
                    add_comment(c, "★ 中心格 = Base Beta × Base Rf\n  应等于 Valuation Summary 输出\n  (通过 WACC 传导重算)")
        return hdr_row + 5

    # ==================== Sheet 2: WACC ====================
    def _wacc(self):
        ws = self.wb.create_sheet("WACC")
        d = self.d

        # Cost of Equity (CAPM)
        c = ws.cell(1, 1, "COST OF EQUITY (CAPM) -- 股权成本"); c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

        rep_ccy = d.get("reporting_currency") or "N/A"
        rf_source = d.get("rf_source", "10Y sovereign yield")
        beta_source = d.get("beta_source", "5Y Monthly Beta")
        erp_source = d.get("erp_source", "Damodaran ERP")
        benchmark = d.get("benchmark", "N/A")

        self._apply_input(ws.cell(2, 2), self.wacc_inputs["risk_free_rate"], FMT_PERCENT,
                          f"{rf_source} ({rep_ccy})", "10Y 主权债券收益率")
        ws.cell(2, 1, f"Risk-Free Rate ({rep_ccy} 10Y) -- 无风险利率")

        self._apply_input(ws.cell(3, 2), self.wacc_inputs["beta"], "0.00",
                          beta_source, f"基准: {benchmark}")
        ws.cell(3, 1, f"Beta (5Y Monthly, vs {benchmark}) -- 贝塔系数")

        self._apply_input(ws.cell(4, 2), self.wacc_inputs["erp"], FMT_PERCENT,
                          f"{erp_source} ({rep_ccy})", "股权风险溢价 (国家/地区口径)")
        ws.cell(4, 1, f"Equity Risk Premium ({rep_ccy} country) -- 股权风险溢价")

        c = ws.cell(5, 1, "Cost of Equity -- 股权成本")
        c = ws.cell(5, 2, "=B2+B3*B4"); c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        add_comment(c, "计算公式 (CAPM):\n  Ke = Rf + β × ERP\n  = B2 + B3 × B4")

        # Cost of Debt
        c = ws.cell(7, 1, "COST OF DEBT -- 债务成本"); c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2)

        self._apply_input(ws.cell(8, 2), self.wacc_inputs["pre_tax_kd"], FMT_PERCENT,
                          "10-K / bond spread", "Pre-Tax Cost of Debt")
        ws.cell(8, 1, "Pre-Tax Cost of Debt -- 税前债务成本")

        self._apply_input(ws.cell(9, 2), d["tax_rate"], FMT_PERCENT,
                          "富途「利润表」", "所得税 / 税前利润")
        ws.cell(9, 1, "Tax Rate -- 税率")

        c = ws.cell(10, 1, "After-Tax Cost of Debt -- 税后债务成本")
        c = ws.cell(10, 2, "=B8*(1-B9)"); c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        add_comment(c, "计算公式:\n  Kd(after-tax) = Kd × (1 - Tax Rate)\n  = B8 × (1 - B9)")

        # Capital Structure
        c = ws.cell(12, 1, "CAPITAL STRUCTURE -- 资本结构"); c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=12, start_column=1, end_row=12, end_column=2)

        c = ws.cell(13, 1, "Market Capitalization (M) -- 市值")
        c = ws.cell(13, 2, "=DCF!B13"); c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
        add_comment(c, "跨 Sheet 引用:\n  = DCF!B13 (Market Cap, 报表币种)")

        c = ws.cell(14, 1, "Net Debt / (Net Cash) (M) -- 净债务")
        c = ws.cell(14, 2, "=DCF!B14"); c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
        add_comment(c, "跨 Sheet 引用:\n  = DCF!B14 (Net Debt, 报表币种)")

        c = ws.cell(15, 1, "Enterprise Capital (M) -- 企业资本")
        c = ws.cell(15, 2, "=B13+B14"); c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        add_comment(c, "计算公式:\n  Enterprise Capital = Market Cap + Net Debt\n  = B13 + B14")

        c = ws.cell(16, 1, "Equity Weight (We) -- 股权权重")
        c = ws.cell(16, 2, "=B13/B15"); c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        add_comment(c, "计算公式:\n  We = Market Cap / Enterprise Capital\n  = B13 / B15")

        c = ws.cell(17, 1, "Debt Weight (Wd) -- 债务权重")
        c = ws.cell(17, 2, "=B14/B15"); c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        add_comment(c, "计算公式:\n  Wd = Net Debt / Enterprise Capital\n  = B14 / B15")

        # WACC
        c = ws.cell(18, 1, "WACC -- 加权平均资本成本"); c.font = FONT_BOLD
        c = ws.cell(18, 2, "=B16*B5+B17*B10"); c.font = FONT_BLACK_BOLD; c.fill = FILL_VALUATION_ORANGE_DARK
        c.number_format = FMT_PERCENT
        add_comment(c, "计算公式:\n  WACC = We × Ke + Wd × Kd(after-tax)\n  = B16 × B5 + B17 × B10")

        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 16
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--ticker", required=True)
    parser.add_argument("--workspace", required=True)
    args = parser.parse_args()
    data = extract_financial_data(Path(args.workspace), args.ticker)
    output = Path(args.workspace) / "excels" / f"{args.ticker}_DCF_Model_{date.today().isoformat()}.xlsx"
    DCFBuilder(data).build(output)
