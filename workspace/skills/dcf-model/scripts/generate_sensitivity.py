#!/usr/bin/env python3
"""为 DCF 工作簿生成三张 5×5 实时公式敏感性表。"""

import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise ImportError("openpyxl required: pip install openpyxl") from exc


BASE_CELL_FILL = PatternFill("solid", fgColor="BDD7EE")
BASE_CELL_FONT = Font(bold=True)
TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
TITLE_FONT = Font(color="FFFFFF", bold=True)

TABLE_DEFINITIONS = (
    (
        "wacc_vs_terminal_growth",
        "WACC vs Terminal Growth - Implied Share Price",
        "WACC",
        "Terminal Growth",
        "wacc_tg",
        0.005,
        0.005,
    ),
    (
        "revenue_growth_vs_ebit_margin",
        "Revenue Growth vs EBIT Margin - Implied Share Price",
        "Revenue Growth",
        "EBIT Margin",
        "growth_margin",
        0.02,
        0.02,
    ),
    (
        "beta_vs_risk_free_rate",
        "Beta vs Risk-Free Rate - Implied Share Price",
        "Beta",
        "Risk-Free Rate",
        "beta_rf",
        0.15,
        0.005,
    ),
)


def generate_sensitivity_for_model(workbook, config, row_map: dict) -> int:
    """在 DCF 工作表底部生成全部 75 个敏感性公式。"""
    ws = workbook["DCF"]
    tables = (config.sensitivity or {}).get("tables", {})
    if int((config.sensitivity or {}).get("table_size", 5)) != 5:
        raise ValueError("sensitivity.table_size must be 5 because this model requires 75 formulas")

    base_wacc = _base_wacc(config)
    base_values = {
        "wacc_tg": (base_wacc, _base_terminal_growth(config)),
        "growth_margin": (
            float(config.scenario_series("base", "revenue_growth")[0]),
            float(config.scenario_series("base", "ebit_margin")[0]),
        ),
        "beta_rf": (
            float(config.wacc_inputs.get("beta", 1.0)),
            float(config.wacc_inputs.get("risk_free_rate", 0.043)),
        ),
    }

    formula_count = 0
    start_row = row_map["sensitivity_start"]
    for index, definition in enumerate(TABLE_DEFINITIONS):
        table_key, title, row_param, col_param, formula_type, default_row_step, default_col_step = definition
        settings = tables.get(table_key, {})
        if settings.get("enabled", True) is False:
            raise ValueError(f"sensitivity.tables.{table_key}.enabled must be true")
        row_step = float(settings.get("row_step", default_row_step))
        col_step = float(settings.get("col_step", default_col_step))
        row_base, col_base = base_values[formula_type]
        formula_count += _write_sensitivity_table(
            ws=ws,
            start_row=start_row + index * 10,
            title=title,
            row_param=row_param,
            col_param=col_param,
            row_base=row_base,
            col_base=col_base,
            row_step=row_step,
            col_step=col_step,
            row_map=row_map,
            config=config,
            formula_type=formula_type,
        )
    if formula_count != 75:
        raise ValueError(f"Expected 75 sensitivity formulas, generated {formula_count}")
    return formula_count


def _write_sensitivity_table(
    ws,
    start_row,
    title,
    row_param,
    col_param,
    row_base,
    col_base,
    row_step,
    col_step,
    row_map,
    config,
    formula_type,
):
    ws.cell(start_row, 1, title)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    for col in range(1, 7):
        ws.cell(start_row, col).fill = TITLE_FILL
        ws.cell(start_row, col).font = TITLE_FONT
    ws.cell(start_row + 1, 1, f"{row_param} / {col_param}")

    col_values = [col_base + offset * col_step for offset in (-2, -1, 0, 1, 2)]
    row_values = [row_base + offset * row_step for offset in (-2, -1, 0, 1, 2)]
    for index, value in enumerate(col_values):
        header = ws.cell(start_row + 1, 2 + index, value)
        header.number_format = "0.0x" if col_param == "Beta" else "0.0%"

    for row_index, value in enumerate(row_values):
        data_row = start_row + 2 + row_index
        header = ws.cell(data_row, 1, value)
        header.number_format = "0.0x" if row_param == "Beta" else "0.0%"
        for col_index in range(5):
            data_col = 2 + col_index
            row_header = f"$A{data_row}"
            col_header = f"{get_column_letter(data_col)}${start_row + 1}"
            formula = _build_sensitivity_formula(
                formula_type,
                row_header,
                col_header,
                row_map,
                config,
            )
            cell = ws.cell(data_row, data_col, formula)
            cell.number_format = config.output.get("number_format", {}).get("per_share", "$#,##0.00")
            if row_index == 2 and col_index == 2:
                cell.fill = BASE_CELL_FILL
                cell.font = BASE_CELL_FONT

    data_range = f"B{start_row + 2}:F{start_row + 6}"
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    return 25


def _build_sensitivity_formula(formula_type, row_ref, col_ref, row_map, config):
    """创建每个单元格的完整 DCF 重算公式。"""
    if formula_type == "wacc_tg":
        return _valuation_formula(config, row_map, row_ref, col_ref)

    if formula_type == "growth_margin":
        base_growth = float(config.scenario_series("base", "revenue_growth")[0])
        base_margin = float(config.scenario_series("base", "ebit_margin")[0])
        growth_delta = f"({row_ref}-{base_growth:g})"
        margin_delta = f"({col_ref}-{base_margin:g})"
        return _valuation_formula(
            config,
            row_map,
            f"$B${row_map['consol_wacc']}",
            f"$B${row_map['consol_terminal_g']}",
            growth_delta=growth_delta,
            margin_delta=margin_delta,
        )

    if formula_type == "beta_rf":
        equity_weight = "'WACC'!$B$18"
        debt_weight = "'WACC'!$B$19"
        erp = "'WACC'!$B$4"
        after_tax_debt = "'WACC'!$B$10"
        calculated_base_wacc = "'WACC'!$B$20"
        scenario_base_wacc = f"$B${row_map['consol_wacc']}"
        sensitivity_wacc = (
            f"(({col_ref}+{row_ref}*{erp})*{equity_weight}+"
            f"{after_tax_debt}*{debt_weight}+({scenario_base_wacc}-{calculated_base_wacc}))"
        )
        return _valuation_formula(
            config,
            row_map,
            sensitivity_wacc,
            f"$B${row_map['consol_terminal_g']}",
        )

    raise ValueError(f"Unknown sensitivity formula type: {formula_type}")


def _valuation_formula(
    config,
    row_map,
    wacc_ref,
    terminal_growth_ref,
    growth_delta=None,
    margin_delta=None,
):
    """重建预测现金流、终值和每股价值。"""
    rm = row_map
    history_end_col = get_column_letter(rm["historical_end_col"])
    revenue_expr = f"{history_end_col}{rm['revenue']}"
    pv_terms = []
    final_fcf = None
    final_revenue = None
    final_ebit = None

    for index in range(config.projection_years):
        assumption_col = get_column_letter(2 + index)
        growth = f"{assumption_col}${rm['consol_growth']}"
        if growth_delta:
            growth = f"({growth}+{growth_delta})"
        revenue_expr = f"({revenue_expr}*(1+{growth}))"
        margin = f"{assumption_col}${rm['consol_margin']}"
        if margin_delta:
            margin = f"({margin}+{margin_delta})"
        tax = f"{assumption_col}${rm['consol_tax']}"
        da_pct = f"{assumption_col}${rm['consol_da']}"
        capex_pct = f"{assumption_col}${rm['consol_capex']}"
        nwc_pct = f"{assumption_col}${rm['consol_nwc']}"
        if index == 0:
            prior_revenue = f"{history_end_col}{rm['revenue']}"
        else:
            prior_revenue = final_revenue
        ebit = f"({revenue_expr}*{margin})"
        nopat = f"({ebit}*(1-{tax}))"
        da = f"({revenue_expr}*{da_pct})"
        capex = f"({revenue_expr}*{capex_pct})"
        delta_nwc = f"(({revenue_expr}-{prior_revenue})*{nwc_pct})"
        fcf = f"({nopat}+{da}-{capex}-{delta_nwc})"
        period = index + 0.5 if config.mid_year_convention else index + 1.0
        pv_terms.append(f"{fcf}/(1+{wacc_ref})^{period:g}")
        final_revenue = revenue_expr
        final_ebit = ebit
        final_fcf = fcf

    terminal_method = config.terminal_value.get("method", "perpetuity_growth")
    terminal_period = config.projection_years - 0.5 if config.mid_year_convention else config.projection_years
    if terminal_method == "exit_multiple":
        exit_multiple = float(config.terminal_value["exit_multiple"])
        terminal_value = f"({final_ebit}*{exit_multiple:g})"
    else:
        terminal_value = (
            f"({final_fcf}*(1+{terminal_growth_ref})/"
            f"({wacc_ref}-{terminal_growth_ref}))"
        )
    pv_terminal = f"({terminal_value}/(1+{wacc_ref})^{terminal_period:g})"
    enterprise_value = "+".join(pv_terms + [pv_terminal])
    per_share_scale = 1000 if config.units == "B" else 1
    net_debt = f"$B${rm['net_debt']}"
    shares = f"$B${rm['shares']}"
    return f"=(({enterprise_value})-{net_debt})*{per_share_scale}/{shares}"


def _base_wacc(config) -> float:
    scenario_wacc = config.scenarios.get("base", {}).get("wacc")
    if scenario_wacc is not None:
        return float(scenario_wacc)
    rf = float(config.wacc_inputs.get("risk_free_rate", 0.043))
    beta = float(config.wacc_inputs.get("beta", 1.0))
    erp = float(config.wacc_inputs.get("equity_risk_premium", 0.055))
    after_tax_debt = float(config.wacc_inputs.get("pre_tax_cost_of_debt", 0.045)) * (
        1 - float(config.wacc_inputs.get("tax_rate", 0.21))
    )
    market_cap = float(config.stock_price) * float(config.shares_outstanding_m)
    enterprise_capital = market_cap + float(config.net_debt_m)
    if enterprise_capital == 0:
        return rf + beta * erp
    return (rf + beta * erp) * market_cap / enterprise_capital + after_tax_debt * float(
        config.net_debt_m
    ) / enterprise_capital


def _base_terminal_growth(config) -> float:
    return float(config.scenarios.get("base", {}).get("terminal_growth", config.terminal_growth))


def main() -> None:
    parser = argparse.ArgumentParser(description="Populate DCF sensitivity tables")
    parser.add_argument("--model", required=True, help="Input xlsx model")
    parser.add_argument("--config", required=True, help="DCF JSON config")
    parser.add_argument("--output", help="Output xlsx path; defaults to input model")
    args = parser.parse_args()

    from build_dcf_model import DCFConfig, DCFModelBuilder

    config = DCFConfig.from_json(args.config)
    workbook = openpyxl.load_workbook(args.model)
    builder = DCFModelBuilder(config)
    row_map = builder.row_map
    ws = workbook["DCF"]
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= row_map["sensitivity_start"]:
            ws.unmerge_cells(str(merged))
    for row in ws.iter_rows(min_row=row_map["sensitivity_start"]):
        for cell in row:
            cell.value = None
    generate_sensitivity_for_model(workbook, config, row_map)
    output = Path(args.output or args.model)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(f"Sensitivity tables saved to {output}")


if __name__ == "__main__":
    main()
