#!/usr/bin/env python3
"""验证 DCF 工作簿的结构、公式和关键估值约束。"""

import argparse
import json
from datetime import datetime
from pathlib import Path


class DCFModelValidator:
    """对脚本生成的两表 DCF 模型执行静态验证。"""

    REQUIRED_DCF_LABELS = (
        "Case (1=Bear, 2=Base, 3=Bull)",
        "BEAR CASE ASSUMPTIONS",
        "BASE CASE ASSUMPTIONS",
        "BULL CASE ASSUMPTIONS",
        "SELECTED CASE ASSUMPTIONS",
        "HISTORICAL & PROJECTED FREE CASH FLOW",
        "VALUATION SUMMARY",
        "Implied Price per Share",
    )
    SENSITIVITY_TITLES = (
        "WACC vs Terminal Growth - Implied Share Price",
        "Revenue Growth vs EBIT Margin - Implied Share Price",
        "Beta vs Risk-Free Rate - Implied Share Price",
    )
    EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")

    def __init__(self, excel_path: str | Path):
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl") from exc
        self.excel_path = str(excel_path)
        if not Path(excel_path).exists():
            raise FileNotFoundError(f"File not found: {excel_path}")
        self.workbook_formulas = openpyxl.load_workbook(excel_path, data_only=False)
        self.workbook_values = openpyxl.load_workbook(excel_path, data_only=True)
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.info: list[str] = []

    def validate_all(self) -> dict:
        self.check_sheet_structure()
        self.check_formula_errors()
        self.check_model_structure()
        self.check_dcf_logic()
        return {
            "file": self.excel_path,
            "validation_date": datetime.now().isoformat(),
            "status": "PASS" if not self.errors else "FAIL",
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
            "errors": self.errors,
            "warnings": self.warnings,
            "info": self.info,
        }

    def check_sheet_structure(self) -> None:
        sheet_names = self.workbook_formulas.sheetnames
        for sheet in ("DCF", "WACC"):
            if sheet not in sheet_names:
                self.errors.append(f"Required sheet missing: {sheet}")
            else:
                self.info.append(f"Found sheet: {sheet}")
        extras = [name for name in sheet_names if name not in ("DCF", "WACC")]
        if extras:
            self.warnings.append(f"Unexpected sheets present: {', '.join(extras)}")

    def check_formula_errors(self) -> None:
        total_formulas = 0
        for sheet_name in self.workbook_formulas.sheetnames:
            ws_formulas = self.workbook_formulas[sheet_name]
            ws_values = self.workbook_values[sheet_name]
            for row in ws_formulas.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        total_formulas += 1
                        for error in self.EXCEL_ERRORS:
                            if error in value:
                                self.errors.append(f"{error} in formula at {sheet_name}!{cell.coordinate}")
                    cached = ws_values[cell.coordinate].value
                    if isinstance(cached, str):
                        for error in self.EXCEL_ERRORS:
                            if error in cached:
                                self.errors.append(f"{error} cached at {sheet_name}!{cell.coordinate}")
                                break
        self.info.append(f"Total formulas: {total_formulas}")
        if total_formulas == 0:
            self.errors.append("Workbook contains no formulas")
        else:
            self.info.append("No literal Excel error tokens found in formulas")

    def check_model_structure(self) -> None:
        if "DCF" not in self.workbook_formulas.sheetnames or "WACC" not in self.workbook_formulas.sheetnames:
            return
        dcf = self.workbook_formulas["DCF"]
        wacc = self.workbook_formulas["WACC"]
        for label in self.REQUIRED_DCF_LABELS:
            if self._find_label(dcf, label) is None:
                self.errors.append(f"Required DCF section or label missing: {label}")
        if self._find_label(wacc, "WACC") is None:
            self.errors.append("WACC output is missing")

        sensitivity_formulas = 0
        for title in self.SENSITIVITY_TITLES:
            title_cell = self._find_label(dcf, title)
            if title_cell is None:
                self.errors.append(f"Sensitivity table missing: {title}")
                continue
            for row in dcf.iter_rows(
                min_row=title_cell.row + 2,
                max_row=title_cell.row + 6,
                min_col=2,
                max_col=6,
            ):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        sensitivity_formulas += 1
                    else:
                        self.errors.append(f"Sensitivity formula missing at DCF!{cell.coordinate}")
        if sensitivity_formulas == 75:
            self.info.append("Found all 75 sensitivity formulas")
        else:
            self.errors.append(f"Expected 75 sensitivity formulas, found {sensitivity_formulas}")

        selector = self._find_label(dcf, "Case (1=Bear, 2=Base, 3=Bull)")
        if selector is not None:
            selected_value = dcf.cell(selector.row, selector.column + 1).value
            if selected_value not in (1, 2, 3):
                self.errors.append("Case selector must contain 1, 2, or 3")

        consolidation = self._find_label(dcf, "SELECTED CASE ASSUMPTIONS")
        projection = self._find_label(dcf, "HISTORICAL & PROJECTED FREE CASH FLOW")
        if consolidation and projection:
            formulas = []
            for row in dcf.iter_rows(min_row=consolidation.row + 1, max_row=projection.row - 1):
                formulas.extend(
                    cell.value
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                )
            if not formulas or not all("CHOOSE(" in formula.upper() for formula in formulas):
                self.errors.append("Selected-case consolidation block must use CHOOSE formulas")

    def check_dcf_logic(self) -> None:
        if "DCF" not in self.workbook_formulas.sheetnames:
            return
        dcf = self.workbook_formulas["DCF"]
        for scenario in ("BEAR", "BASE", "BULL"):
            header = self._find_label(dcf, f"{scenario} CASE ASSUMPTIONS")
            if header is None:
                continue
            terminal_growth = self._value_below_label(dcf, header.row, "Terminal Growth", max_rows=10)
            wacc = self._value_below_label(dcf, header.row, "WACC", max_rows=10)
            if isinstance(terminal_growth, (int, float)) and isinstance(wacc, (int, float)):
                if terminal_growth >= wacc:
                    self.errors.append(
                        f"{scenario.title()} terminal growth ({terminal_growth:.2%}) must be below WACC ({wacc:.2%})"
                    )
                elif not 0.05 <= wacc <= 0.20:
                    self.warnings.append(f"{scenario.title()} WACC ({wacc:.2%}) is outside 5%-20%")
                else:
                    self.info.append(
                        f"{scenario.title()} terminal growth ({terminal_growth:.2%}) is below WACC ({wacc:.2%})"
                    )
            elif scenario == "BASE":
                self.warnings.append("Base WACC is formula-linked; numeric range requires Excel recalculation")

        values = self.workbook_values["DCF"]
        pv_terminal_cell = self._find_label(dcf, "PV of Terminal Value")
        enterprise_cell = self._find_label(dcf, "Enterprise Value")
        if pv_terminal_cell and enterprise_cell:
            pv_terminal = values.cell(pv_terminal_cell.row, pv_terminal_cell.column + 1).value
            enterprise_value = values.cell(enterprise_cell.row, enterprise_cell.column + 1).value
            if isinstance(pv_terminal, (int, float)) and isinstance(enterprise_value, (int, float)) and enterprise_value:
                proportion = pv_terminal / enterprise_value
                if proportion > 0.80 or proportion < 0.40:
                    self.warnings.append(f"Terminal value is {proportion:.1%} of enterprise value")
                else:
                    self.info.append(f"Terminal value is {proportion:.1%} of enterprise value")
            else:
                self.info.append("Numeric terminal-value proportion skipped because formulas have no cached results")

    @staticmethod
    def _find_label(ws, label: str):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == label:
                    return cell
        return None

    @staticmethod
    def _value_below_label(ws, start_row: int, label: str, max_rows: int):
        for row in range(start_row, start_row + max_rows + 1):
            if ws.cell(row, 1).value == label:
                return ws.cell(row, 2).value
        return None


def validate_dcf_model(excel_path: str | Path) -> dict:
    return DCFModelValidator(excel_path).validate_all()


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate a generated DCF model")
    parser.add_argument("excel_file", help="DCF xlsx file")
    parser.add_argument("output", nargs="?", help="Optional JSON report path")
    args = parser.parse_args()
    try:
        results = validate_dcf_model(args.excel_file)
        output = json.dumps(results, indent=2)
        print(output)
        if args.output:
            Path(args.output).write_text(output, encoding="utf-8")
        raise SystemExit(0 if results["status"] == "PASS" else 1)
    except SystemExit:
        raise
    except Exception as exc:
        result = {"file": args.excel_file, "status": "ERROR", "error": str(exc)}
        print(json.dumps(result, indent=2))
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()
