#!/usr/bin/env python3
"""
生成包含实时公式的 LBO (Leveraged Buyout) Excel 模型。

结构 (对齐 references/schema.md):
  Tab 1: Sources & Uses      -- 多档债务 + Equity Plug, Sources = Uses
  Tab 2: Operating Model     -- Closing + Year1-5 收入/EBITDA/EBIT/税/FCF
  Tab 3: Debt Schedule       -- Revolver/TLA/TLB/Senior Notes 多档 roll-forward,
                                Interest=期初余额×利率 (断循环), Cash Sweep 瀑布
  Tab 4: Returns Analysis    -- Exit EV/Equity, MOIC, IRR (基于现金流系列),
                                3 张 5×5 敏感性表 (Entry×Exit → IRR /
                                Entry×Leverage → MOIC / Growth×Margin → IRR)

字体颜色约定:
  蓝色 0000FF: 硬编码输入
  黑色 000000: 计算公式
  紫色 800080: 同 Sheet 引用
  绿色 008000: 跨 Sheet 引用

填充色:
  #1F4E79 深蓝 (Section header)
  #D9E1F2 浅蓝 (Column header)
  #F2F2F2 浅灰 (Input cell)
  #BDD7EE 中蓝 (Key output / Sensitivity center)
"""
import argparse
import re
import zipfile
import tempfile
import os
import logging
from datetime import date
from pathlib import Path
from typing import Optional, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
except ImportError as exc:
    raise ImportError("openpyxl required: pip install openpyxl") from exc

# ==================== Logging ====================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

# ==================== Styles ====================
FONT_BLUE = Font(color="0000FF")             # 硬编码输入
FONT_BLACK = Font(color="000000")            # 计算公式
FONT_PURPLE = Font(color="800080")           # 同 Sheet 引用
FONT_GREEN = Font(color="008000")            # 跨 Sheet 引用
FONT_WHITE_BOLD = Font(color="FFFFFF", bold=True)
FONT_BOLD = Font(bold=True)
FONT_BLACK_BOLD = Font(color="000000", bold=True)

FILL_DARK_BLUE = PatternFill("solid", fgColor="1F4E79")
FILL_LIGHT_BLUE = PatternFill("solid", fgColor="D9E1F2")
FILL_MEDIUM_BLUE = PatternFill("solid", fgColor="BDD7EE")
FILL_INPUT_GREY = PatternFill("solid", fgColor="F2F2F2")
FILL_NONE = PatternFill(fill_type=None)

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

ALIGN_RIGHT = Alignment(horizontal="right")

_COMMENT_AUTHOR = "LBO Builder"

def add_comment(cell, text: str, width: int = 280, height: int = 90):
    """给单元格附加气泡备注,备注计算公式。"""
    c = Comment(text, _COMMENT_AUTHOR)
    c.width = width
    c.height = height
    cell.comment = c

# ==================== Utility ====================
def safe_divide(n, d): return n / d if d != 0 else 0.0

def _ensure_shared_strings(file_path: Path) -> Tuple[Path, bool]:
    file_path = Path(file_path)
    try:
        with zipfile.ZipFile(str(file_path), 'r') as zf:
            if 'xl/sharedStrings.xml' in zf.namelist(): return file_path, False
            ct = zf.read('[Content_Types].xml').decode('utf-8', errors='ignore')
            if 'sharedStrings' not in ct: return file_path, False
    except Exception: return file_path, False
    fd, tmp = tempfile.mkstemp(suffix='.xlsx', dir=str(file_path.parent))
    os.close(fd)
    try:
        with zipfile.ZipFile(str(file_path), 'r') as src, zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as dst:
            for item in src.namelist(): dst.writestr(item, src.read(item))
            dst.writestr('xl/sharedStrings.xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"/>')
        return Path(tmp), True
    except Exception: return file_path, False

def _read_excel_map(file_path: Path) -> dict:
    fp, is_tmp = _ensure_shared_strings(file_path)
    try:
        wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
        sheet = wb.active
        data = {}
        header_row = 1
        for r in range(1, 5):
            vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1) if sheet.cell(r, c).value]
            if "FY" in " ".join(str(v) for v in vals): header_row = r; break
        periods = [str(sheet.cell(header_row, c).value) for c in range(3, sheet.max_column + 1)]
        for r in range(header_row + 1, sheet.max_row + 1):
            ind = sheet.cell(r, 1).value
            if not ind: continue
            ind = str(ind).strip()
            vals = {}
            for c, p in enumerate(periods):
                v = sheet.cell(r, c + 3).value
                num = 0.0
                if isinstance(v, (int, float)): num = float(v)
                elif isinstance(v, str) and v:
                    cl = v.replace(",", "").replace("-", "").strip()
                    if cl.replace(".", "", 1).isdigit(): num = float(cl) * (-1 if v.strip().startswith("-") else 1)
                vals[p] = num
            data[ind] = vals
        wb.close()
        return data
    finally:
        if is_tmp and fp.exists(): os.remove(str(fp))

def find_local_file(excels_path: Path, ticker: str, suffix: str) -> Optional[Path]:
    pattern = re.compile(rf'^.*_{re.escape(ticker)}_{re.escape(suffix)}_.*\.(xlsx|xls)$', re.IGNORECASE)
    files = [f for f in excels_path.iterdir() if f.is_file() and pattern.match(f.name)]
    return sorted(files)[-1] if files else None

# ==================== Data Extraction ====================
def extract_financial_data(workspace: Path, ticker: str) -> dict:
    """从富途生成的 Excel 中抽取 LTM 财务数据用于 LBO 模型。"""
    excels_path = workspace / "excels"
    inc_file = find_local_file(excels_path, ticker, "income")
    bs_file = find_local_file(excels_path, ticker, "balance")
    cf_file = find_local_file(excels_path, ticker, "cashflow")
    inc_data = _read_excel_map(inc_file) if inc_file else {}
    bs_data = _read_excel_map(bs_file) if bs_file else {}
    cf_data = _read_excel_map(cf_file) if cf_file else {}

    all_fy = set()
    for d in [inc_data, bs_data, cf_data]:
        for k, v in d.items(): all_fy.update([p for p in v.keys() if "FY" in p])
    fy_cols = sorted(list(all_fy), reverse=True)
    if not fy_cols:
        raise RuntimeError(f"未找到 {ticker} 的历史财报数据,请先运行 futu-financial-report 生成 Excel")
    latest_fy = fy_cols[0]
    prev_fy = fy_cols[1] if len(fy_cols) > 1 else latest_fy

    def gv(data, keys, col):
        for k in keys:
            if k in data and col in data[k]: return data[k][col]
        return 0.0

    revenue = gv(inc_data, ["总收入", "营业总收入"], latest_fy)
    prev_revenue = gv(inc_data, ["总收入", "营业总收入"], prev_fy)
    ebit = gv(inc_data, ["营业利润"], latest_fy)
    tax = gv(inc_data, ["所得税"], latest_fy)
    ebt = gv(inc_data, ["税前利润"], latest_fy)
    capex = abs(gv(inc_data, ["资本开支(CapEx)", "资本开支"], latest_fy))
    if capex == 0.0:
        capex = abs(gv(cf_data, ["资本开支(CapEx)", "资本开支", "购建固定资产、无形资产和其他长期资产支付的现金"], latest_fy))
    da = gv(cf_data, ["折旧摊销及损耗"], latest_fy)
    if da == 0.0: da = gv(inc_data, ["折旧摊销及损耗"], latest_fy)

    # EBITDA = EBIT + D&A, 若 D&A 缺失则以 CapEx 的 70% 估算
    if da == 0.0 and capex > 0:
        da = 0.7 * capex
        logger.warning("D&A 缺失,以 CapEx × 70%% 估算: %.2f", da)
    ebitda = ebit + da

    rev_growth = safe_divide(revenue - prev_revenue, prev_revenue)
    # 收敛异常增长率,避免极端 base case
    if rev_growth < -0.20 or rev_growth > 0.50:
        rev_growth = 0.05
    tax_rate = safe_divide(tax, ebt) if ebt > 0 else 0.25
    total_debt = gv(bs_data, ["短期借款与融资租赁负债", "短期借款"], latest_fy) + \
                 gv(bs_data, ["长期借款", "长期借款与融资租赁负债"], latest_fy)

    return {
        "ticker": ticker,
        "revenue": revenue,
        "ebit": ebit,
        "da": da,
        "ebitda": ebitda,
        "capex": capex if capex > 0 else 0.05 * revenue,  # 若缺失以 5% 营收估算
        "rev_growth": rev_growth,
        "tax_rate": tax_rate,
        "debt": total_debt,
    }

# ==================== LBO Builder ====================
class LBOBuilder:
    """
    构建 4-Tab LBO 模型 (Sources & Uses / Operating Model / Debt Schedule / Returns).

    主要假设:
      Entry Multiple : 默认 9.0x LTM EBITDA
      Exit Multiple  : 默认 9.0x Y5 EBITDA (与 Entry 一致 => Multiple Neutral)
      Debt Leverage  : Total Debt = 50% × Purchase Price
      债务结构 (占 Total Debt 比例, 加权平均利率):
        Revolver     : 0%  (预留额度, 初期不使用)
        Term Loan A  : 25% @ 5.5%, 每年强制摊销 10%
        Term Loan B  : 45% @ 6.5%, 到期一次还款 (Bullet)
        Senior Notes : 30% @ 8.0%, 到期一次还款
      Cash Sweep    : 100% 剩余 FCF 按 Revolver > TLA > TLB > Notes 顺序偿还
      Fees          : Transaction Fees 2% × EV, Financing Fees 2% × Debt
    """
    def __init__(self, d: dict, entry_multiple: float = 9.0, exit_multiple: float = 9.0):
        self.d = d
        self.entry_m = float(entry_multiple)
        self.exit_m = float(exit_multiple)
        self.years = 5
        self.wb = openpyxl.Workbook()

        # 派生比率
        self.ebitda_margin = safe_divide(d["ebitda"], d["revenue"])
        self.da_pct        = safe_divide(d["da"], d["revenue"])
        self.capex_pct     = safe_divide(d["capex"], d["revenue"])
        self.tax_rate      = d["tax_rate"]
        self.rev_growth    = d["rev_growth"]

        # LBO 结构性假设
        self.leverage_ratio       = 0.50  # Debt/Purchase Price
        self.transaction_fee_pct  = 0.02
        self.financing_fee_pct    = 0.02
        self.cash_to_bs_pct       = 0.00
        # 债务分档: (tranche_name, share_of_total_debt, interest_rate, mandatory_amort_pct/year)
        self.tranches = [
            ("Revolver",     0.00, 0.055, 0.00),
            ("Term Loan A",  0.25, 0.055, 0.10),
            ("Term Loan B",  0.45, 0.065, 0.00),
            ("Senior Notes", 0.30, 0.080, 0.00),
        ]
        self.nwc_pct_of_delta_rev = 0.10  # 净营运资本变动 = 10% × 营收变化

        # ---- 预计算行号布局 (跨 tab 引用必须先知道) ----
        # Sources & Uses (固定行)
        self.su_rows = {
            "ebitda": 4, "entry_multiple": 5, "ev": 6,
            "leverage_row": 10,
            "tranche_first": 14,
            "tranche_last": 14 + len(self.tranches) - 1,   # Notes 行 (=17)
            "sponsor_equity": 14 + len(self.tranches),     # 平衡项 (=18)
            "total_sources": 14 + len(self.tranches) + 1,  # (=19)
            "total_uses": 27,
        }

        # Debt Schedule 行布局:
        #   row 3 = header, row 4 = 第一档 subtitle, row 5 = 第一档 Beg
        #   每档 6 行 (subtitle / Beg / Interest / Amort / Sweep / Ending)
        rows_per_tranche = 6
        tranche_start_row = 5  # 第一档 Beginning Balance 行
        self.debt_tranche_positions = []  # (name, beg_row, share, rate, amort)
        for idx, (name, share, rate, amort) in enumerate(self.tranches):
            beg_row = tranche_start_row + idx * rows_per_tranche
            self.debt_tranche_positions.append((name, beg_row, share, rate, amort))
        # TOTALS 汇总区: 紧接最后一档之后
        last_beg = self.debt_tranche_positions[-1][1]  # 23
        last_ending = last_beg + 4                     # 27
        totals_header_row = last_ending + 1            # 28
        self.debt_rows = {
            "totals_header": totals_header_row,        # 28
            "total_beg":     totals_header_row + 1,    # 29
            "total_interest": totals_header_row + 2,   # 30
            "total_amort":   totals_header_row + 3,    # 31
            "total_sweep":   totals_header_row + 4,    # 32
            "total_ending":  totals_header_row + 5,    # 33
        }

    # ---------- helpers ----------
    def _sec_header(self, ws, row: int, col_start: int, col_end: int, text: str):
        ws.cell(row, col_start, text).font = FONT_WHITE_BOLD
        for c in range(col_start, col_end + 1):
            ws.cell(row, c).fill = FILL_DARK_BLUE

    def _col_headers(self, ws, row: int, labels: list, col_start: int = 2):
        for i, label in enumerate(labels):
            c = ws.cell(row, col_start + i, label)
            c.fill = FILL_LIGHT_BLUE; c.font = FONT_BOLD; c.alignment = Alignment(horizontal="center")

    # ---------- build ----------
    def build(self, output_path: Path):
        self._su()
        self._op()
        self._debt()
        self._ret()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(output_path))
        logger.info(f"LBO model saved: {output_path}")

    # ==================== Tab 1: Sources & Uses ====================
    def _su(self):
        ws = self.wb.active; ws.title = "Sources & Uses"
        self._sec_header(ws, 1, 1, 5, "SOURCES & USES -- 资金来源与用途")

        d = self.d
        # ---- 输入区 (row 3-8) ----
        ws.cell(3, 1, "TRANSACTION ASSUMPTIONS -- 交易假设").font = FONT_BOLD
        rows = [
            (4, "LTM EBITDA (M) -- 过去12个月EBITDA", d["ebitda"], "#,##0", None),
            (5, "Entry EBITDA Multiple (x) -- 入场倍数", self.entry_m, '0.0"x"', "入场时企业价值 / EBITDA 倍数"),
            (6, "Enterprise Value (M) -- 企业价值(购买价)", "=B4*B5", "#,##0", "EV = LTM EBITDA × Entry Multiple\n  = B4 × B5"),
            (7, "Transaction Fees % -- 交易费率", self.transaction_fee_pct, "0.0%", None),
            (8, "Financing Fees % -- 融资费率(占债务)", self.financing_fee_pct, "0.0%", None),
            (9, "Cash to Balance Sheet % of EV", self.cash_to_bs_pct, "0.0%", None),
            (10, "Leverage % (Debt / Purchase Price)", self.leverage_ratio, "0.0%", None),
        ]
        for r, label, val, fmt, note in rows:
            ws.cell(r, 1, label)
            c = ws.cell(r, 2, val)
            c.number_format = fmt
            if isinstance(val, str) and val.startswith("="):
                c.font = FONT_BLACK
            else:
                c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY
            if note:
                add_comment(c, note)

        # ---- SOURCES 区 (row 12-19) ----
        self._sec_header(ws, 12, 1, 5, "SOURCES -- 资金来源")
        self._col_headers(ws, 13, ["Amount (M)", "% of Cap"], col_start=2)
        ws.cell(13, 1, "Item").fill = FILL_LIGHT_BLUE; ws.cell(13, 1).font = FONT_BOLD

        # 债务分档
        # Total Debt = B10 (leverage) × B6 (EV); tranche = share × Total Debt
        total_debt_row = 18  # 稍后填 Total Debt 位置
        for i, (name, share, rate, amort) in enumerate(self.tranches):
            r = 14 + i
            ws.cell(r, 1, f"{name} ({share*100:.0f}%)")
            c = ws.cell(r, 2, f"=$B$10*$B$6*{share}"); c.font = FONT_BLACK; c.number_format = "#,##0"
            add_comment(c, f"计算公式:\n  {name} = Total Debt × 档位占比 ({share:.0%})\n  = Leverage × EV × {share}\n  利率 {rate:.1%}, 年强制摊销 {amort:.0%}")
            c2 = ws.cell(r, 3, f"=B{r}/$B$19"); c2.font = FONT_BLACK; c2.number_format = "0.0%"

        r_eq = 14 + len(self.tranches)  # Sponsor Equity 行
        ws.cell(r_eq, 1, "Sponsor Equity (Plug) -- 股东出资 (平衡项)")
        c = ws.cell(r_eq, 2, f"=B25-SUM(B14:B{r_eq-1})"); c.font = FONT_BLACK; c.number_format = "#,##0"
        add_comment(c, "Plug 平衡项:\n  Sponsor Equity = Total Uses - Total Debt\n  = B25 - SUM(B14:B{tail})\n(保证 Sources = Uses)".replace("{tail}", str(r_eq-1)))
        ws.cell(r_eq, 3, f"=B{r_eq}/$B$19").font = FONT_BLACK; ws.cell(r_eq, 3).number_format = "0.0%"

        r_total = r_eq + 1
        ws.cell(r_total, 1, "Total Sources").font = FONT_BOLD
        c = ws.cell(r_total, 2, f"=SUM(B14:B{r_eq})"); c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = "#,##0"
        add_comment(c, f"Total Sources = SUM(Revolver..Sponsor Equity)\n  = SUM(B14:B{r_eq})")
        ws.cell(r_total, 3, 1.0).font = FONT_BOLD
        ws.cell(r_total, 3).number_format = "0.0%"; ws.cell(r_total, 3).fill = FILL_MEDIUM_BLUE

        # ---- USES 区 (row 21-25) ----
        self._sec_header(ws, 21, 1, 5, "USES -- 资金用途")
        self._col_headers(ws, 22, ["Amount (M)", "% of Cap"], col_start=2)
        ws.cell(22, 1, "Item").fill = FILL_LIGHT_BLUE; ws.cell(22, 1).font = FONT_BOLD

        c = ws.cell(23, 1, "Purchase Enterprise Value -- 企业价值 (购买价)")
        c = ws.cell(23, 2, "=B6"); c.font = FONT_PURPLE; c.number_format = "#,##0"
        add_comment(c, "同 Sheet 引用:\n  = B6 (Enterprise Value = LTM EBITDA × Entry Multiple)")

        c = ws.cell(24, 1, "Transaction Fees -- 交易顾问费")
        c = ws.cell(24, 2, "=B6*B7"); c.font = FONT_BLACK; c.number_format = "#,##0"
        add_comment(c, "计算公式:\n  Transaction Fees = EV × Fee%\n  = B6 × B7")

        # Financing Fees = Total Debt (SUM B14..) × Financing Fee%
        # Cash to BS = EV × cash_pct
        c = ws.cell(25, 1, "Financing Fees -- 融资费用")
        total_debt_range = f"B14:B{r_eq-1}"
        c = ws.cell(25, 2, f"=SUM({total_debt_range})*B8"); c.font = FONT_BLACK; c.number_format = "#,##0"
        add_comment(c, f"计算公式:\n  Financing Fees = Total Debt × Financing Fee%\n  = SUM({total_debt_range}) × B8")

        c = ws.cell(26, 1, "Cash to Balance Sheet -- 保留在 BS 上的现金")
        c = ws.cell(26, 2, "=B6*B9"); c.font = FONT_BLACK; c.number_format = "#,##0"
        add_comment(c, "计算公式:\n  Cash to BS = EV × Cash %\n  = B6 × B9")

        # 需要 Total Uses 放在固定行 B25 之前 — 调整: 让 Purchase EV 从 row 23, Uses total = row 27
        ws.cell(27, 1, "Total Uses").font = FONT_BOLD
        c = ws.cell(27, 2, "=SUM(B23:B26)"); c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = "#,##0"
        add_comment(c, "Total Uses = Purchase EV + Transaction Fees + Financing Fees + Cash to BS\n  = SUM(B23:B26)")

        # 修正 Sponsor Equity 引用的 Total Uses 单元格 (原 B25 -> B27)
        c = ws.cell(r_eq, 2); c.value = f"=B27-SUM(B14:B{r_eq-1})"

        # ---- 平衡检查 (row 29) ----
        c = ws.cell(29, 1, "Check: Sources = Uses -- 勾稽检查 (应为 0)"); c.font = FONT_BOLD
        c = ws.cell(29, 2, f"=B{r_total}-B27"); c.font = FONT_BLACK; c.number_format = "#,##0.00"
        add_comment(c, "Sources - Uses (必须为 0)\n  = Total Sources - Total Uses")

        # 列宽
        ws.column_dimensions["A"].width = 46
        for col_letter, w in [("B", 16), ("C", 12), ("D", 4), ("E", 4)]:
            ws.column_dimensions[col_letter].width = w

    # ==================== Tab 2: Operating Model ====================
    def _op(self):
        """经营模型 (对齐 schema.md):
          - EBITDA Margin 直接作为驱动 (蓝色输入 / 公式)
          - EBIT = EBITDA - D&A
          - Net Income = (EBIT - Interest) × (1 - Tax Rate)
          - FCF = Net Income + D&A - CapEx - ΔNWC (可用于偿债的 Levered FCF)
        (schema.md 未要求 Gross Margin / SG&A 拆分,故省略以保持与 EBITDA 数据一致)
        """
        ws = self.wb.create_sheet("Operating Model")
        self._sec_header(ws, 1, 1, 7, "OPERATING MODEL -- 经营模型")

        # 列布局: A=Label, B=Closing (LTM), C..G = Year 1..5
        self._col_headers(ws, 3, ["Closing (LTM)", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"], col_start=2)
        ws.cell(3, 1, "Line Item (M)").fill = FILL_LIGHT_BLUE; ws.cell(3, 1).font = FONT_BOLD

        d = self.d

        # Row 4: Revenue
        ws.cell(4, 1, "Revenue -- 营业收入")
        c = ws.cell(4, 2, d["revenue"]); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = "#,##0"
        add_comment(c, "LTM Revenue: 从富途财报最新一年提取")
        for i in range(1, 6):
            col, prev = get_column_letter(2+i), get_column_letter(1+i)
            c = ws.cell(4, 2+i, f"={prev}4*(1+{col}5)"); c.font = FONT_BLACK; c.number_format = "#,##0"
            add_comment(c, f"计算公式:\n  Revenue(Y{i}) = Revenue(Y{i-1}) × (1 + Growth%)\n  = {prev}4 × (1 + {col}5)")

        # Row 5: Revenue Growth %
        ws.cell(5, 1, "Revenue Growth % -- 营收增长率")
        for i in range(1, 6):
            c = ws.cell(5, 2+i, self.rev_growth); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = "0.0%"

        # Row 6: EBITDA Margin % (直接驱动)
        ws.cell(6, 1, "EBITDA Margin % -- EBITDA 利润率")
        for i in range(0, 6):
            c = ws.cell(6, 2+i, self.ebitda_margin); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = "0.0%"
        add_comment(ws.cell(6, 7), "EBITDA Margin:\n  可逐年调整以体现 PE 经营改善假设\n  (蓝色输入, 用户可修改)")

        # Row 7: EBITDA = Revenue × Margin
        ws.cell(7, 1, "EBITDA").font = FONT_BOLD
        for i in range(0, 6):
            col = get_column_letter(2+i)
            c = ws.cell(7, 2+i, f"={col}4*{col}6"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE; c.number_format = "#,##0"
            add_comment(c, f"计算公式:\n  EBITDA = Revenue × EBITDA Margin\n  = {col}4 × {col}6")

        # Row 8: D&A %
        ws.cell(8, 1, "D&A % of Revenue -- 折旧摊销占营收比")
        for i in range(0, 6):
            c = ws.cell(8, 2+i, max(self.da_pct, 0.03)); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = "0.0%"

        # Row 9: D&A (负值)
        ws.cell(9, 1, "Less: D&A -- 折旧与摊销 (负值)")
        for i in range(0, 6):
            col = get_column_letter(2+i)
            c = ws.cell(9, 2+i, f"=-{col}4*{col}8"); c.font = FONT_BLACK; c.number_format = "#,##0;(#,##0)"
            add_comment(c, f"计算公式:\n  D&A = -Revenue × D&A%\n  = -{col}4 × {col}8")

        # Row 10: EBIT
        ws.cell(10, 1, "EBIT -- 息税前利润").font = FONT_BOLD
        for i in range(0, 6):
            col = get_column_letter(2+i)
            c = ws.cell(10, 2+i, f"={col}7+{col}9"); c.font = FONT_BOLD; c.number_format = "#,##0"
            add_comment(c, f"计算公式:\n  EBIT = EBITDA + D&A = {col}7 + {col}9\n  (D&A 已带负号)")

        # Row 11: Interest Expense (来自 Debt Schedule Total Interest)
        ws.cell(11, 1, "Less: Interest Expense -- 利息支出 (跨 Sheet)")
        ws.cell(11, 2, 0).number_format = "#,##0"
        interest_row = self.debt_rows["total_interest"]
        for i in range(1, 6):
            col = get_column_letter(2+i)
            c = ws.cell(11, 2+i, f"=-'Debt Schedule'!{col}{interest_row}"); c.font = FONT_GREEN; c.number_format = "#,##0;(#,##0)"
            add_comment(c, f"跨 Sheet 引用 (加负号):\n  = -Debt Schedule!{col}{interest_row}\n  (Total Interest, 期初余额×利率)")

        # Row 12: EBT
        ws.cell(12, 1, "EBT -- 税前利润")
        for i in range(0, 6):
            col = get_column_letter(2+i)
            c = ws.cell(12, 2+i, f"={col}10+{col}11"); c.font = FONT_BLACK; c.number_format = "#,##0"
            add_comment(c, f"EBT = EBIT + Interest = {col}10 + {col}11\n(Interest 已带负号)")

        # Row 13: Tax Rate
        ws.cell(13, 1, "Tax Rate -- 税率")
        for i in range(0, 6):
            c = ws.cell(13, 2+i, self.tax_rate); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = "0.0%"

        # Row 14: Taxes (负值, 亏损不缴税)
        ws.cell(14, 1, "Less: Taxes -- 所得税 (负值)")
        for i in range(0, 6):
            col = get_column_letter(2+i)
            c = ws.cell(14, 2+i, f"=-MAX(0,{col}12)*{col}13"); c.font = FONT_BLACK; c.number_format = "#,##0;(#,##0)"
            add_comment(c, f"计算公式:\n  Taxes = -MAX(0, EBT) × Tax Rate\n  = -MAX(0, {col}12) × {col}13\n  (亏损年度不缴税)")

        # Row 15: Net Income
        ws.cell(15, 1, "Net Income -- 净利润").font = FONT_BOLD
        for i in range(0, 6):
            col = get_column_letter(2+i)
            c = ws.cell(15, 2+i, f"={col}12+{col}14"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE; c.number_format = "#,##0"
            add_comment(c, f"Net Income = EBT + Taxes = {col}12 + {col}14\n(Taxes 已带负号)")

        # Row 16: Add-back D&A (非现金)
        ws.cell(16, 1, "Plus: D&A (Add back non-cash)")
        for i in range(0, 6):
            col = get_column_letter(2+i)
            c = ws.cell(16, 2+i, f"=-{col}9"); c.font = FONT_BLACK; c.number_format = "#,##0"
            add_comment(c, f"加回非现金项:\n  = -{col}9 (D&A 原为负号, 加回)")

        # Row 17: CapEx %
        ws.cell(17, 1, "CapEx % of Revenue")
        for i in range(0, 6):
            c = ws.cell(17, 2+i, max(self.capex_pct, 0.03)); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = "0.0%"

        # Row 18: CapEx (负值)
        ws.cell(18, 1, "Less: CapEx -- 资本开支 (负值)")
        for i in range(0, 6):
            col = get_column_letter(2+i)
            c = ws.cell(18, 2+i, f"=-{col}4*{col}17"); c.font = FONT_BLACK; c.number_format = "#,##0;(#,##0)"
            add_comment(c, f"计算公式:\n  CapEx = -Revenue × CapEx%\n  = -{col}4 × {col}17")

        # Row 19: NWC %
        ws.cell(19, 1, "NWC % of Δ Revenue")
        c = ws.cell(19, 2, self.nwc_pct_of_delta_rev); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = "0.0%"

        # Row 20: ΔNWC
        ws.cell(20, 1, "Less: Δ Working Capital -- 营运资本变动")
        ws.cell(20, 2, 0).number_format = "#,##0"
        for i in range(1, 6):
            col = get_column_letter(2+i); prev = get_column_letter(1+i)
            c = ws.cell(20, 2+i, f"=-({col}4-{prev}4)*$B$19"); c.font = FONT_BLACK; c.number_format = "#,##0;(#,##0)"
            add_comment(c, f"计算公式:\n  ΔNWC = -(本年营收 - 上年营收) × NWC%\n  = -({col}4 - {prev}4) × B19")

        # Row 21: Unlevered FCF (加回利息)
        ws.cell(21, 1, "Unlevered Free Cash Flow -- 无杠杆自由现金流").font = FONT_BOLD
        for i in range(0, 6):
            col = get_column_letter(2+i)
            c = ws.cell(21, 2+i, f"={col}15+{col}16+{col}18+{col}20-{col}11"); c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = "#,##0"
            add_comment(c, f"计算公式:\n  Unlevered FCF = Net Income + D&A + CapEx + ΔNWC - Interest\n  = {col}15 + {col}16 + {col}18 + {col}20 - {col}11\n  (加回利息还原为 unlevered)")

        # Row 22: Levered FCF (Available for debt repayment)
        ws.cell(22, 1, "Levered FCF (Available for Debt Repayment)").font = FONT_BOLD
        for i in range(0, 6):
            col = get_column_letter(2+i)
            c = ws.cell(22, 2+i, f"={col}15+{col}16+{col}18+{col}20"); c.font = FONT_BOLD; c.number_format = "#,##0"
            add_comment(c, f"计算公式:\n  Levered FCF = Net Income + D&A + CapEx + ΔNWC\n  = {col}15 + {col}16 + {col}18 + {col}20\n  (可用于偿还债务的现金)")

        # 列宽
        ws.column_dimensions["A"].width = 46
        for i in range(2, 8):
            ws.column_dimensions[get_column_letter(i)].width = 14

        # 保存关键行号供 Debt Schedule / Returns 引用
        self.op_rows = {
            "revenue":      4,
            "ebitda":       7,
            "ebit":        10,
            "interest":    11,
            "net_income":  15,
            "levered_fcf": 22,
        }

    # ==================== Tab 3: Debt Schedule ====================
    def _debt(self):
        ws = self.wb.create_sheet("Debt Schedule")
        self._sec_header(ws, 1, 1, 7, "DEBT SCHEDULE -- 债务偿还计划")

        # 列: A=Label, B=Closing (Day 1 余额), C..G = Year 1..5
        self._col_headers(ws, 3, ["Closing", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"], col_start=2)
        ws.cell(3, 1, "Line Item (M)").fill = FILL_LIGHT_BLUE; ws.cell(3, 1).font = FONT_BOLD

        # 每档 6 行 (subtitle / Beg / Interest / Mandatory Amort / Cash Sweep / Ending),
        # 行号布局已在 __init__ 中预先计算 (self.debt_tranche_positions)
        for idx, (name, base, share, rate, amort) in enumerate(self.debt_tranche_positions):
            # Section subtitle (base - 1)
            c = ws.cell(base - 1, 1, f"{name} @ {rate:.1%}  |  Mandatory Amort {amort:.0%}"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE

            # Beginning Balance (base)
            ws.cell(base, 1, "Beginning Balance -- 期初余额")
            # Closing = Sources & Uses initial debt
            su_row = self.su_rows["tranche_first"] + idx
            c = ws.cell(base, 2, f"='Sources & Uses'!B{su_row}"); c.font = FONT_GREEN; c.number_format = "#,##0"
            add_comment(c, f"跨 Sheet 引用:\n  = Sources & Uses!B{su_row}\n  ({name} 初始规模)")
            for i in range(1, 6):
                col, prev = get_column_letter(2+i), get_column_letter(1+i)
                c = ws.cell(base, 2+i, f"={prev}{base+4}"); c.font = FONT_BLACK; c.number_format = "#,##0"
                add_comment(c, f"Roll-forward:\n  Beg(Y{i}) = End(Y{i-1})\n  = {prev}{base+4}")

            # Interest Expense (base + 1) — 用期初余额, 断开循环引用
            ws.cell(base+1, 1, f"Interest Expense @ {rate:.1%} -- 利息支出")
            ws.cell(base+1, 2, 0).number_format = "#,##0"  # Closing 期无利息
            for i in range(1, 6):
                col = get_column_letter(2+i)
                c = ws.cell(base+1, 2+i, f"={col}{base}*{rate}"); c.font = FONT_BLACK; c.number_format = "#,##0"
                add_comment(c, f"计算公式:\n  Interest = Beginning Balance × Rate\n  = {col}{base} × {rate:.1%}\n  (用期初余额, 断开循环引用)")

            # Mandatory Amortization (base + 2)
            ws.cell(base+2, 1, "Mandatory Amortization -- 强制摊销")
            ws.cell(base+2, 2, 0).number_format = "#,##0;(#,##0)"
            for i in range(1, 6):
                col = get_column_letter(2+i)
                if amort > 0:
                    c = ws.cell(base+2, 2+i, f"=-MIN({col}{base},{amort}*'Sources & Uses'!B{su_row})")
                    c.font = FONT_BLACK
                    add_comment(c, f"计算公式:\n  Mandatory Amort = -MIN(期初余额, 摊销比例 × 原始债务)\n  = -MIN({col}{base}, {amort:.0%} × Sources & Uses!B{su_row})\n  (余额不能为负)")
                else:
                    c = ws.cell(base+2, 2+i, 0); c.font = FONT_BLACK
                    add_comment(c, "该档无强制摊销 (Bullet 到期一次还款)")
                c.number_format = "#,##0;(#,##0)"

            # Cash Sweep (base + 3) — 稍后统一填, 因为需要引用其他档 sweep 行
            ws.cell(base+3, 1, "Cash Sweep -- 现金瀑布优先偿还")
            for i in range(1, 6):
                ws.cell(base+3, 2+i, 0).number_format = "#,##0;(#,##0)"

            # Ending Balance (base + 4)
            ws.cell(base+4, 1, "Ending Balance -- 期末余额").font = FONT_BOLD
            for i in range(0, 6):
                col = get_column_letter(2+i)
                if i == 0:
                    c = ws.cell(base+4, 2+i, f"={col}{base}"); c.font = FONT_BLACK
                    add_comment(c, "Closing 期: Ending = Beginning")
                else:
                    c = ws.cell(base+4, 2+i, f"=MAX(0,{col}{base}+{col}{base+2}+{col}{base+3})"); c.font = FONT_BLACK
                    add_comment(c, f"Roll-forward:\n  End = MAX(0, Beg + Mandatory Amort + Cash Sweep)\n  = MAX(0, {col}{base} + {col}{base+2} + {col}{base+3})\n  (余额不能为负)")
                c.number_format = "#,##0"; c.fill = FILL_LIGHT_BLUE

        # ==== Cash Sweep 优先级瀑布 ====
        # 每档 Sweep = -MIN(该档剩余余额 [Beg+Amort], 剩余可用现金)
        # 剩余可用现金 = Levered FCF + Mandatory Amort Total (amort 已为负) + 已被更高优先级 sweep 的部分 (sweep 已为负)
        for i in range(1, 6):
            col = get_column_letter(2+i)
            for idx, (name, base, share, rate, amort) in enumerate(self.debt_tranche_positions):
                # 该档在本年扣除强制摊销后的剩余余额
                remain_bal = f"MAX(0,{col}{base}+{col}{base+2})"
                # 剩余可用现金 = Levered FCF + Σ(前面档的 Amort) + Σ(前面档的 Sweep, 已为负)
                # 但 Levered FCF 已经扣了 Interest (Net Income 里减了利息)
                # 用累计式: Levered FCF + Σ mandatory_amort_all_tranches + Σ sweep_higher_priority
                # 简化: 引用 Total Mandatory Amort + 已 sweep 的档
                prior_sweeps = "+".join(f"{col}{tp[1]+3}" for tp in self.debt_tranche_positions[:idx]) or "0"
                avail_expr = f"MAX(0,'Operating Model'!{col}{self.op_rows['levered_fcf']}+{col}{self.debt_rows['total_amort']}+({prior_sweeps}))"
                sweep_formula = f"=-MIN({remain_bal},{avail_expr})"
                c = ws.cell(base+3, 2+i, sweep_formula); c.font = FONT_BLACK; c.number_format = "#,##0;(#,##0)"
                add_comment(c, f"计算公式 (Cash Sweep 瀑布, 优先级 {idx+1}/{len(self.debt_tranche_positions)}):\n  剩余可用现金 = Levered FCF + Total Mandatory Amort + 已被更高优先级 sweep 的部分\n  本档 Sweep = -MIN(该档剩余余额, 剩余可用现金)\n  {name}")

        # ==== TOTALS 汇总区 ====
        totals_header = self.debt_rows["totals_header"]
        c = ws.cell(totals_header, 1, "TOTALS -- 汇总"); c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE

        beg_rows      = [pos[1]     for pos in self.debt_tranche_positions]
        interest_rows = [pos[1] + 1 for pos in self.debt_tranche_positions]
        amort_rows    = [pos[1] + 2 for pos in self.debt_tranche_positions]
        sweep_rows    = [pos[1] + 3 for pos in self.debt_tranche_positions]
        ending_rows   = [pos[1] + 4 for pos in self.debt_tranche_positions]

        ws.cell(self.debt_rows["total_beg"],      1, "Total Beginning Debt").font = FONT_BOLD
        ws.cell(self.debt_rows["total_interest"], 1, "Total Interest Expense").font = FONT_BOLD
        ws.cell(self.debt_rows["total_amort"],    1, "Total Mandatory Amort").font = FONT_BOLD
        ws.cell(self.debt_rows["total_sweep"],    1, "Total Cash Sweep").font = FONT_BOLD
        ws.cell(self.debt_rows["total_ending"],   1, "Total Ending Debt").font = FONT_BOLD

        for i in range(0, 6):
            col = get_column_letter(2+i)
            def _sum(rs): return "+".join(f"{col}{r}" for r in rs)
            ws.cell(self.debt_rows["total_beg"],      2+i, f"={_sum(beg_rows)}").font = FONT_BOLD
            ws.cell(self.debt_rows["total_interest"], 2+i, f"={_sum(interest_rows)}").font = FONT_BOLD
            ws.cell(self.debt_rows["total_amort"],    2+i, f"={_sum(amort_rows)}").font = FONT_BOLD
            ws.cell(self.debt_rows["total_sweep"],    2+i, f"={_sum(sweep_rows)}").font = FONT_BOLD
            c = ws.cell(self.debt_rows["total_ending"], 2+i, f"={_sum(ending_rows)}"); c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
            for k in ("total_beg", "total_interest", "total_amort", "total_sweep", "total_ending"):
                ws.cell(self.debt_rows[k], 2+i).number_format = "#,##0;(#,##0)"

        # 给 Total Interest / Ending 加公式备注
        for i in range(1, 6):
            col = get_column_letter(2+i)
            add_comment(ws.cell(self.debt_rows["total_interest"], 2+i),
                        f"Total Interest = Σ 各档利息\n  = {'+'.join(f'{col}{r}' for r in interest_rows)}")
            add_comment(ws.cell(self.debt_rows["total_ending"], 2+i),
                        f"Total Ending Debt = Σ 各档期末余额\n  = {'+'.join(f'{col}{r}' for r in ending_rows)}")

        # 列宽
        ws.column_dimensions["A"].width = 46
        for i in range(2, 8):
            ws.column_dimensions[get_column_letter(i)].width = 14

    # ==================== Tab 4: Returns Analysis ====================
    def _ret(self):
        ws = self.wb.create_sheet("Returns Analysis")
        self._sec_header(ws, 1, 1, 8, "RETURNS ANALYSIS -- 回报分析 + 敏感性表")

        # ---- 回报计算区 (row 3-16) ----
        ws.cell(3, 1, "RETURN CALCULATION -- 回报计算").font = FONT_BOLD; ws.cell(3, 1).fill = FILL_LIGHT_BLUE

        c = ws.cell(4, 1, "Initial Equity Investment (M) -- 初始股权投资")
        c = ws.cell(4, 2, f"='Sources & Uses'!B{self.su_rows['sponsor_equity']}"); c.font = FONT_GREEN; c.number_format = "#,##0"
        add_comment(c, f"跨 Sheet 引用:\n  = Sources & Uses!B{self.su_rows['sponsor_equity']}\n  (Sponsor Equity = Total Uses - Total Debt)")

        c = ws.cell(5, 1, "Exit EBITDA (Year 5) (M) -- 退出年 EBITDA")
        c = ws.cell(5, 2, f"='Operating Model'!G{self.op_rows['ebitda']}"); c.font = FONT_GREEN; c.number_format = "#,##0"
        add_comment(c, f"跨 Sheet 引用:\n  = Operating Model!G{self.op_rows['ebitda']}\n  (Year 5 EBITDA)")

        c = ws.cell(6, 1, "Exit Multiple (x) -- 退出倍数")
        c = ws.cell(6, 2, self.exit_m); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = '0.0"x"'

        c = ws.cell(7, 1, "Exit Enterprise Value -- 退出企业价值")
        c = ws.cell(7, 2, "=B5*B6"); c.font = FONT_BLACK; c.number_format = "#,##0"
        add_comment(c, "计算公式:\n  Exit EV = Exit EBITDA × Exit Multiple\n  = B5 × B6")

        c = ws.cell(8, 1, "Less: Net Debt at Exit (Year 5) -- 退出年净债务")
        end_row = self.debt_rows["total_ending"]
        c = ws.cell(8, 2, f"='Debt Schedule'!G{end_row}"); c.font = FONT_GREEN; c.number_format = "#,##0"
        add_comment(c, f"跨 Sheet 引用:\n  = Debt Schedule!G{end_row}\n  (Year 5 Total Ending Debt)")

        c = ws.cell(9, 1, "Exit Equity Value -- 退出股权价值").font = FONT_BOLD
        c = ws.cell(9, 2, "=B7-B8"); c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = "#,##0"
        add_comment(c, "计算公式:\n  Exit Equity = Exit EV - Net Debt\n  = B7 - B8")

        # 现金流系列 (用于 IRR): Year 0 = -Initial Equity, Year 1-4 = 0 (无分红), Year 5 = Exit Equity
        ws.cell(11, 1, "CASH FLOW SERIES for IRR -- IRR 现金流系列").font = FONT_BOLD; ws.cell(11, 1).fill = FILL_LIGHT_BLUE
        self._col_headers(ws, 12, ["Year 0", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"], col_start=2)
        ws.cell(12, 1, "Period").fill = FILL_LIGHT_BLUE; ws.cell(12, 1).font = FONT_BOLD

        ws.cell(13, 1, "Cash Flow (M)")
        c = ws.cell(13, 2, "=-B4"); c.font = FONT_BLACK; c.number_format = "#,##0;(#,##0)"
        add_comment(c, "Year 0 = -Initial Equity Investment (投入,负数)")
        for i in range(1, 5):
            col = get_column_letter(2+i)
            c = ws.cell(13, 2+i, 0); c.font = FONT_BLACK; c.number_format = "#,##0"
            add_comment(c, f"Year {i}: 假设无分红 (可根据实际填入 Interim Distributions)")
        c = ws.cell(13, 7, "=B9"); c.font = FONT_BLACK; c.number_format = "#,##0"
        add_comment(c, "Year 5 = +Exit Equity Value (退出,正数)")

        # MOIC & IRR
        ws.cell(15, 1, "MOIC (Money on Invested Capital) -- 倍数").font = FONT_BOLD
        c = ws.cell(15, 2, "=B9/B4"); c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = '0.00"x"'
        add_comment(c, "计算公式:\n  MOIC = Exit Equity / Initial Equity\n  = B9 / B4")

        ws.cell(16, 1, "IRR (Internal Rate of Return) -- 内部收益率").font = FONT_BOLD
        c = ws.cell(16, 2, "=IRR(B13:G13)"); c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = "0.0%"
        add_comment(c, "计算公式:\n  IRR = IRR(现金流系列 B13:G13)\n  Year 0 投入(负), Year 5 退出(正)")

        # ---- 敏感性表基础参数 (用于公式引用) ----
        # 保存基准值供敏感性表使用
        ws.cell(18, 1, "SENSITIVITY BASE PARAMETERS -- 敏感性表基准参数").font = FONT_BOLD; ws.cell(18, 1).fill = FILL_LIGHT_BLUE
        # 用于 Table 1/2/3 的基准变量
        ws.cell(19, 1, "Base LTM EBITDA")
        ws.cell(19, 2, f"='Sources & Uses'!B{self.su_rows['ebitda']}").font = FONT_GREEN
        ws.cell(20, 1, "Base Entry Multiple")
        ws.cell(20, 2, f"='Sources & Uses'!B{self.su_rows['entry_multiple']}").font = FONT_GREEN
        ws.cell(21, 1, "Base Exit Multiple")
        ws.cell(21, 2, "=B6").font = FONT_PURPLE
        ws.cell(22, 1, "Base Leverage Ratio")
        ws.cell(22, 2, "='Sources & Uses'!B10").font = FONT_GREEN
        ws.cell(23, 1, "Base Revenue Growth")
        ws.cell(23, 2, self.rev_growth).font = FONT_BLUE; ws.cell(23, 2).fill = FILL_INPUT_GREY; ws.cell(23, 2).number_format = "0.0%"
        ws.cell(24, 1, "Base EBITDA Margin")
        ws.cell(24, 2, self.ebitda_margin).font = FONT_BLUE; ws.cell(24, 2).fill = FILL_INPUT_GREY; ws.cell(24, 2).number_format = "0.0%"
        ws.cell(25, 1, "Base LTM Revenue")
        ws.cell(25, 2, f"='Operating Model'!B{self.op_rows['revenue']}").font = FONT_GREEN
        ws.cell(26, 1, "Base Total FCF over Hold (approximated as Debt paid down)")
        # Total FCF used to pay down debt ≈ Total Debt 初始 - Total Debt 期末
        end_row_debt = self.debt_rows["total_ending"]
        ws.cell(26, 2, f"=SUM('Sources & Uses'!B{self.su_rows['tranche_first']}:B{self.su_rows['tranche_last']})-'Debt Schedule'!G{end_row_debt}").font = FONT_BLACK
        ws.cell(26, 2).number_format = "#,##0"
        ws.cell(27, 1, "Base Net Debt at Exit")
        ws.cell(27, 2, "=B8").font = FONT_PURPLE
        for r in range(19, 28):
            ws.cell(r, 2).number_format = ws.cell(r, 2).number_format or "#,##0"

        # ---- 敏感性表 1: Entry × Exit Multiple → IRR (row 30-38) ----
        self._sensitivity_table_1_entry_exit_irr(ws, start_row=30)

        # ---- 敏感性表 2: Entry × Leverage → MOIC (row 42-50) ----
        self._sensitivity_table_2_entry_leverage_moic(ws, start_row=42)

        # ---- 敏感性表 3: Growth × Margin → IRR (row 54-62) ----
        self._sensitivity_table_3_growth_margin_irr(ws, start_row=54)

        # 列宽
        ws.column_dimensions["A"].width = 46
        for i in range(2, 8):
            ws.column_dimensions[get_column_letter(i)].width = 14

    # ---------- 敏感性表 helpers ----------
    def _sens_axis(self, base: float, step: float) -> list:
        """基于 base 生成 [base-2s, base-s, base, base+s, base+2s]"""
        return [base - 2*step, base - step, base, base + step, base + 2*step]

    def _write_sens_header(self, ws, row: int, title: str, row_axis_label: str, col_axis_label: str):
        c = ws.cell(row, 1, title); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.cell(row + 1, 1, f"{row_axis_label} \\ {col_axis_label}").font = FONT_BOLD
        ws.cell(row + 1, 1).fill = FILL_LIGHT_BLUE

    def _sensitivity_table_1_entry_exit_irr(self, ws, start_row: int):
        """Table 1: Entry Multiple (row 轴) × Exit Multiple (列 轴) → IRR
        闭式近似 (每格重算, 债务规模随 Entry 变化):
          Purchase_i        = EntryMult_i × EBITDA_ltm
          Debt0_i           = Leverage_base × Purchase_i
          InitEq_i          = Purchase_i - Debt0_i
          NetDebtExit_i     = MAX(0, Debt0_i - Base_FCF_over_hold)
          ExitEV_j          = ExitMult_j × ExitEBITDA_base
          ExitEq_ij         = ExitEV_j - NetDebtExit_i
          MOIC_ij           = ExitEq_ij / InitEq_i
          IRR_ij            = MOIC_ij ^ (1/5) - 1
        中心格 (base_entry, base_exit) = 模型 IRR (勾稽验证)
        """
        base_entry = self.entry_m
        base_exit = self.exit_m
        row_axis = self._sens_axis(base_entry, 1.0)  # Entry step = 1.0x
        col_axis = self._sens_axis(base_exit, 1.0)   # Exit step = 1.0x

        self._write_sens_header(ws, start_row, "TABLE 1: Entry × Exit Multiple → IRR", "Entry Multiple", "Exit Multiple")
        for j, v in enumerate(col_axis):
            c = ws.cell(start_row + 2, 3 + j, v); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE; c.number_format = '0.0"x"'
            c.alignment = Alignment(horizontal="center")
        for i, v in enumerate(row_axis):
            c = ws.cell(start_row + 3 + i, 2, v); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE; c.number_format = '0.0"x"'

        for i in range(5):
            for j in range(5):
                r = start_row + 3 + i
                c = 3 + j
                entry = f"$B{r}"
                exit_m = f"{get_column_letter(c)}${start_row + 2}"
                purchase = f"({entry}*$B$19)"                # EntryMult × Base LTM EBITDA
                debt0 = f"($B$22*{purchase})"                # Leverage_base × Purchase
                init_eq = f"({purchase}-{debt0})"
                net_debt_exit = f"MAX(0,{debt0}-$B$26)"      # Debt0 - Base FCF over hold
                exit_ev = f"({exit_m}*$B$5)"                 # ExitMult × Base Exit EBITDA
                exit_eq = f"({exit_ev}-{net_debt_exit})"
                moic = f"IFERROR(({exit_eq})/{init_eq},0)"
                formula = f"=IFERROR(IF({moic}>0,({moic})^(1/5)-1,-1),0)"
                cell = ws.cell(r, c, formula); cell.font = FONT_BLACK; cell.number_format = "0.0%"
                if i == 2 and j == 2:
                    cell.font = FONT_BLACK_BOLD; cell.fill = FILL_MEDIUM_BLUE
                    add_comment(cell, "★ 中心格 = Base Case\n  Entry = Base Entry Multiple\n  Exit = Base Exit Multiple\n  此处 IRR 应等于模型主计算区 B16 (取正 MOIC 时)")

    def _sensitivity_table_2_entry_leverage_moic(self, ws, start_row: int):
        """Table 2: Entry Multiple (row) × Leverage (col) → MOIC
        闭式近似:
          Purchase_i = EntryMult_i × EBITDA_ltm
          Debt0_ij   = Leverage_j × Purchase_i
          InitEq_ij  = Purchase_i - Debt0_ij
          NetDebtExit_ij = MAX(0, Debt0_ij - Base_Total_FCF)
          ExitEV     = Exit Mult(base) × Exit EBITDA(base)
          ExitEq_ij  = ExitEV - NetDebtExit_ij
          MOIC_ij    = ExitEq_ij / InitEq_ij
        中心格 (base_entry, base_leverage) = 模型 MOIC ✓
        """
        base_entry = self.entry_m
        base_leverage = self.leverage_ratio
        row_axis = self._sens_axis(base_entry, 1.0)
        col_axis = self._sens_axis(base_leverage, 0.10)  # ±10% leverage
        # 保证 leverage ∈ [0, 0.9]
        col_axis = [max(0.05, min(0.90, v)) for v in col_axis]

        self._write_sens_header(ws, start_row, "TABLE 2: Entry × Leverage → MOIC", "Entry Multiple", "Leverage %")
        for j, v in enumerate(col_axis):
            c = ws.cell(start_row + 2, 3 + j, v); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE; c.number_format = "0.0%"
        for i, v in enumerate(row_axis):
            c = ws.cell(start_row + 3 + i, 2, v); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE; c.number_format = '0.0"x"'

        for i in range(5):
            for j in range(5):
                r = start_row + 3 + i
                c = 3 + j
                entry = f"$B{r}"
                lev = f"{get_column_letter(c)}${start_row + 2}"
                purchase = f"({entry}*$B$19)"
                debt0 = f"({lev}*{purchase})"
                init_eq = f"({purchase}-{debt0})"
                # NetDebtExit = MAX(0, Debt0 - Base_Total_FCF)
                net_debt_exit = f"MAX(0,{debt0}-$B$26)"
                exit_ev = f"($B$21*$B$5)"
                exit_eq = f"({exit_ev}-{net_debt_exit})"
                moic = f"IFERROR(({exit_eq})/{init_eq},0)"
                cell = ws.cell(r, c, f"={moic}"); cell.font = FONT_BLACK; cell.number_format = '0.00"x"'
                if i == 2 and j == 2:
                    cell.font = FONT_BLACK_BOLD; cell.fill = FILL_MEDIUM_BLUE
                    add_comment(cell, "★ 中心格 = Base Case\n  Entry = Base Entry Multiple\n  Leverage = Base Leverage %\n  此处 MOIC 应等于模型主计算区 B15")

    def _sensitivity_table_3_growth_margin_irr(self, ws, start_row: int):
        """Table 3: Revenue Growth (row) × EBITDA Margin (col) → IRR
        闭式近似:
          Y5Revenue_i = LTM_Revenue × (1+growth_i)^5
          Y5EBITDA_ij = Y5Revenue_i × margin_j
          ExitEV_ij   = Exit Mult(base) × Y5EBITDA_ij
          ExitEq_ij   = ExitEV_ij - NetDebtExit_base
          MOIC_ij     = ExitEq_ij / InitEq_base
          IRR_ij      = MOIC_ij^(1/5) - 1
        中心格 (base_growth, base_margin) ≈ 模型 IRR ✓
        """
        base_growth = self.rev_growth
        base_margin = self.ebitda_margin
        row_axis = self._sens_axis(base_growth, 0.02)  # ±2%
        col_axis = self._sens_axis(base_margin, 0.02)  # ±2%
        col_axis = [max(0.02, v) for v in col_axis]

        self._write_sens_header(ws, start_row, "TABLE 3: Revenue Growth × EBITDA Margin → IRR", "Revenue Growth", "EBITDA Margin")
        for j, v in enumerate(col_axis):
            c = ws.cell(start_row + 2, 3 + j, v); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE; c.number_format = "0.0%"
        for i, v in enumerate(row_axis):
            c = ws.cell(start_row + 3 + i, 2, v); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE; c.number_format = "0.0%"

        for i in range(5):
            for j in range(5):
                r = start_row + 3 + i
                c = 3 + j
                growth = f"$B{r}"
                margin = f"{get_column_letter(c)}${start_row + 2}"
                y5_rev = f"($B$25*(1+{growth})^5)"
                y5_ebitda = f"({y5_rev}*{margin})"
                exit_ev = f"($B$21*{y5_ebitda})"
                exit_eq = f"({exit_ev}-$B$27)"
                init_eq = "$B$4"
                moic = f"IFERROR(({exit_eq})/{init_eq},0)"
                cell = ws.cell(r, c, f"=({moic})^(1/5)-1"); cell.font = FONT_BLACK; cell.number_format = "0.0%"
                # 用 IFERROR 保护, 若 MOIC 为负则显示 -100%
                cell.value = f"=IFERROR(IF({moic}>0,({moic})^(1/5)-1,-1),0)"
                if i == 2 and j == 2:
                    cell.font = FONT_BLACK_BOLD; cell.fill = FILL_MEDIUM_BLUE
                    add_comment(cell, "★ 中心格 = Base Case\n  Revenue Growth = Base\n  EBITDA Margin = Base\n  该表使用闭式近似:\n  Y5 EBITDA ≈ LTM_Revenue × (1+g)^5 × margin\n  中心格 IRR 应接近模型 B16")


# ==================== Main ====================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="生成 LBO Excel 模型 (4 Tab + 3 张敏感性表)")
    parser.add_argument("--ticker", required=True, help="股票代码,例如 BABA")
    parser.add_argument("--workspace", required=True, help="workspace 目录路径")
    parser.add_argument("--entry-multiple", type=float, default=9.0, help="入场 EBITDA 倍数 (默认 9.0x)")
    parser.add_argument("--exit-multiple", type=float, default=9.0, help="退出 EBITDA 倍数 (默认 9.0x)")
    args = parser.parse_args()

    data = extract_financial_data(Path(args.workspace), args.ticker)
    output = Path(args.workspace) / "excels" / f"{args.ticker}_LBO_Model_{date.today().isoformat()}.xlsx"
    LBOBuilder(data, entry_multiple=args.entry_multiple, exit_multiple=args.exit_multiple).build(output)
