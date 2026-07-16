#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
流式获取并生成财务报表Excel（支持三种报表类型 + 全市场适配 + 财务比率计算）

支持报表：
- 利润表 (statement_type=1) - 含毛利率/费用率/利润率等衍生指标
- 资产负债表 (statement_type=2)
- 现金流量表 (statement_type=3)

自动适配市场：
- A股 (SH./SZ.) -> field_id: 3xxx
- 美股/港股 (US./HK.) -> field_id: 8xxx

核心特性：
- 全程内存流式处理，无中间文件
- stdout 极简输出，最小化 token 消耗
- 动态字段映射，自动适配不同市场指标名称
- 财务比率自动计算（毛利率/费用占比/营业利润率/净利润率）
- YoY同比下降超过5%红色高亮显示

用法：
    python generate_financial_excel.py <stock_code> [--type T] [--num N] [--output <path>]

示例：
    python generate_financial_excel.py US.BABA --type income
    python generate_financial_excel.py HK.00700 --type balance
    python generate_financial_excel.py SH.600519 --type cashflow --num 20
"""

import glob
import json
import subprocess
import sys
import os
import re
import argparse
import logging
import time
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─────────────────────────────────────────────────────────────
# 路径配置（使用系统 Python 和 workspace 中的 futuapi）
# ─────────────────────────────────────────────────────────────

# 使用系统 Python
PYTHON_EXE = sys.executable

# futu API 脚本路径（workspace 中的 futuapi 副本）
skill_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FUTUAPI_SCRIPT = os.path.join(skill_dir, "..", "futuapi", "scripts", "quote", "get_financials_statements.py")
FUTUAPI_SCRIPT = os.path.normpath(FUTUAPI_SCRIPT)


# ─────────────────────────────────────────────────────────────
# 报表类型配置
# ─────────────────────────────────────────────────────────────

STATEMENT_TYPES = {
    'income':      (1, '利润表'),
    'balance':     (2, '资产负债表'),
    'cashflow':    (3, '现金流量表'),
}

# 颜色配置
COLORS = {
    'header':   'D9E1F2',      # 表头-浅蓝
    'revenue':  'DEEAF6',      # 收入-浅蓝
    'expense':  'FCE4D6',      # 费用-浅红
    'profit':   'E2EFDA',      # 利润-浅绿
    'asset':    'DBEEF3',      # 资产-浅蓝
    'liability': 'FCE4D6',     # 负债-浅红
    'equity':   'E2EFDA',      # 权益-浅绿
    'cash_in':  'C6EFCE',      # 现金流入-浅绿
    'cash_out': 'FCE4D6',      # 现金流出-浅红
    'core_ratio': '9BBB59',    # 核心利润率(毛利率/营业利润率/净利润率)-橄榄绿
    'expense_ratio': 'D8E4BC', # 费用率(营业/销售/管理/研发费用率)-豆绿色
    'ratio':    'FFF2CC',      # 其他一般比率-高亮黄
    'yoy':      None,          # YoY-无特殊背景
    'tax':      'FCE4D6',      # 税-浅红
    'derived':  'BDD7EE',      # 加工指标(FCFF等)-中蓝
    'fcf_ni':   'FFD9B3',      # 自由现金流净利润比(FCF/NI)-浅橙
    'default':  None,
}

# YoY阈值配置
YOY_DECLINE_THRESHOLD = -5.0   # 下降超过此阈值红色高亮
YOY_GROWTH_THRESHOLD = 30.0    # 增长超过此阈值绿色高亮

# 产物新鲜度：7天内的Excel直接复用，除非 --force
FRESH_TTL_SECONDS = 7 * 24 * 3600


def find_recent_excel(excels_dir, safe_code, statement_type, ttl_seconds=FRESH_TTL_SECONDS):
    """在 excels_dir 下查找匹配 {safe_code}_{statement_type}_*.xlsx 的最新文件。

    返回 (path, mtime) 或 None。仅返回 mtime 距今 <= ttl_seconds 的文件。
    """
    if not os.path.isdir(excels_dir):
        return None
    pattern = os.path.join(excels_dir, f"{safe_code}_{statement_type}_*.xlsx")
    candidates = glob.glob(pattern)
    if not candidates:
        return None
    now = time.time()
    best = None
    for path in candidates:
        try:
            mtime = os.path.getmtime(path)
        except OSError:
            continue
        if now - mtime > ttl_seconds:
            continue
        if best is None or mtime > best[1]:
            best = (path, mtime)
    return best

# ─────────────────────────────────────────────────────────────
# 财务比率配置 (分子field_id, 分母field_id, 基础名称, 是否紧跟父项后)
# ─────────────────────────────────────────────────────────────

# A股利润表比率配置 - (名称后缀, 分子field_id, 分母field_id, 父项field_id)
A_SHARE_RATIO_CONFIG = {
    'income': [
        ('毛利率', 3032-27, 3001, 3009),  # A股毛利=营收-成本，插入成本之后
        ('营业利润率', 3032, 3001, 3032),
        ('净利润率', 3043, 3001, 3043),
    ],
    'core_ratios': ['毛利率', '营业利润率', '净利润率'],  # 核心利润率，特殊着色
    'expense_ratios': [
        # (比率名称, 费用field_id, 收入field_id)
        ('营业费用率', 3005, 3001),  # 营业费用占比
        ('销售费用率', 3013, 3001),
        ('管理费用率', 3014, 3001),
        ('研发费用率', 3015, 3001),
    ],
}

# 美股利润表比率配置
US_RATIO_CONFIG = {
    'income': [
        ('毛利率', 8004, 8001, 8004),        # 毛利/总收入
        ('营业利润率', 8017, 8001, 8017),     # 营业利润/总收入
        ('净利润率', 8043, 8001, 8043),       # 归属于母公司净利润/总收入
    ],
    'core_ratios': ['毛利率', '营业利润率', '净利润率'],  # 核心利润率，特殊着色
    'expense_ratios': [
        # (比率名称, 费用field_id, 收入field_id)
        ('营业费用率', 8005, 8001),           # 营业费用占比
        ('销售和管理费用率', 8007, 8001),
        ('销售费用率', 8008, 8001),
        ('管理费用率', 8009, 8001),
        ('研发费用率', 8010, 8001),
    ],
}

# 港股利润表比率配置
# HK_RATIO_CONFIG update
HK_RATIO_CONFIG = {
    'income': [
        ('毛利率', 5010, 5001, 5010),        # 毛利/总收入
        ('营业利润率', 5034, 5001, 5034),     # 营业利润/总收入
        ('净利润率', 5051, 5001, 5051),       # 归属于母公司净利润/总收入
    ],
    'core_ratios': ['毛利率', '营业利润率', '净利润率'],  # 核心利润率，特殊着色
    'expense_ratios': [
        # (比率名称, 费用field_id, 收入field_id)
        ('销售成本率', 5008, 5001),           # 销售成本占比
        ('营业费用率', 5013, 5001),           # 营业费用占比
        ('销售及分销费用率', 5015, 5001),     # 销售及分销费用占比
        ('管理费用率', 5016, 5001),           # 管理费用占比
        ('研发费用率', 5017, 5001),           # 研发费用占比
    ],
}

# ─────────────────────────────────────────────────────────────
# 不同市场的核心字段映射 (市场前缀 -> {报表类型 -> 字段配置})
# ─────────────────────────────────────────────────────────────

# A股 (3xxx field_id)
A_SHARE_FIELDS = {
    'income': [
        (3001, '营业总收入',       'revenue'),
        (3009, '营业总成本',        'expense'),
        # 插入：毛利率 (由 (营收-成本)/营收 计算)
        (3013, '销售费用',        'expense'),
        (3014, '管理费用',        'expense'),
        (3015, '研发费用',        'expense'),
        (3032, '营业利润',        'profit'),
        # 插入：营业利润率
        (3038, '利润总额',        'profit'),
        (3039, '所得税费用',       'tax'),
        (3043, '净利润',         'profit'),
        # 插入：净利润率
        (3047, '归属母公司净利润',  'profit'),
    ],
    'balance': [
        # 资产
        (3001, '资产合计',       'asset'),
        (3002, '流动资产合计',     'asset'),
        (3003, '货币资金',        'asset'),
        (3004, '结算备付金',      'asset'),
        (3005, '拆出资金',        'asset'),
        (3006, '交易性金融资产',   'asset'),
        (3007, '衍生金融资产',     'asset'),
        (3008, '应收票据',        'asset'),
        (3009, '应收账款',        'asset'),
        (3010, '应收款项融资',     'asset'),
        (3011, '预付款项',        'asset'),
        (3012, '应收保费',        'asset'),
        (3013, '应收分保账款',    'asset'),
        (3014, '应收分保合同准备金', 'asset'),
        (3015, '其他应收款',      'asset'),
        (3016, '存货',          'asset'),
        (3017, '合同资产',        'asset'),
        (3018, '持有待售资产',    'asset'),
        (3019, '一年内到期的非流动资产', 'asset'),
        (3020, '其他流动资产',    'asset'),
        (3025, '非流动资产合计',   'asset'),
        (3026, '固定资产合计',     'asset'),
        (3027, '在建工程',        'asset'),
        (3028, '工程物资',        'asset'),
        (3029, '固定资产清理',    'asset'),
        (3030, '无形资产',        'asset'),
        (3031, '开发支出',        'asset'),
        (3032, '商誉',          'asset'),
        (3033, '长期待摊费用',    'asset'),
        (3034, '递延所得税资产',   'asset'),
        (3035, '其他非流动资产',  'asset'),
        # 负债
        (3055, '负债合计',       'liability'),
        (3056, '流动负债合计',    'liability'),
        (3057, '短期借款',        'liability'),
        (3058, '交易性金融负债',   'liability'),
        (3059, '衍生金融负债',     'liability'),
        (3060, '应付票据',        'liability'),
        (3061, '应付账款',        'liability'),
        (3062, '预收款项',        'liability'),
        (3063, '合同负债',        'liability'),
        (3064, '应付职工薪酬',    'liability'),
        (3065, '应交税费',        'liability'),
        (3066, '其他应付款',      'liability'),
        (3067, '一年内到期的非流动负债', 'liability'),
        (3079, '非流动负债合计',   'liability'),
        (3080, '长期借款',        'liability'),
        (3081, '应付债券',        'liability'),
        (3082, '长期应付款',      'liability'),
        (3083, '预计负债',        'liability'),
        (3084, '递延所得税负债',   'liability'),
        (3085, '递延收益',        'liability'),
        (3086, '其他非流动负债',  'liability'),
        # 权益
        (3097, '股东权益合计',     'equity'),
        (3098, '归属母公司所有者权益合计', 'equity'),
        (3099, '实收资本',        'equity'),
        (3100, '资本公积',        'equity'),
        (3101, '盈余公积',        'equity'),
        (3102, '专项储备',        'equity'),
        (3103, '未分配利润',      'equity'),
    ],
    'cashflow': [
        (3001, '经营活动现金流量净额',  'cash_in'),
        (3033, '投资活动现金流量净额',  'cash_out'),
        (3051, '融资活动现金流量净额',  'cash_out'),
        (3068, '期末现金及现金等价物余额', 'cash_in'),
    ],
}

# 美股 (8xxx field_id)
US_FIELDS = {
    'income': [
        (8001, '总收入',         'revenue'),    # 部分公司（如 AAPL）有此字段
        (8002, '营业总收入',      'revenue'),    # 部分公司（如 FUTU）用此字段作为收入
        (8003, '营业总成本',      'expense'),
        (8004, '毛利',           'profit'),
        (8005, '营业费用',        'expense'),
        (8017, '营业利润',        'profit'),
        (8034, '税前利润',        'profit'),
        (8035, '所得税',          'tax'),
        (8037, '净利润',         'profit'),
        (8043, '归属于母公司股东净利润', 'profit'),
    ],
    'balance': [
        (8001, '资产合计',       'asset'),
        (8002, '流动资产合计',     'asset'),
        (8003, '现金及现金等价物',  'asset'),
        (8004, '短期投资',        'asset'),
        (8005, '净应收账款',      'asset'),
        (8006, '应收票据',        'asset'),
        (8007, '其他应收款',      'asset'),
        (8008, '存货',          'asset'),
        (8009, '预付费用',        'asset'),
        (8010, '其他流动资产',    'asset'),
        (8022, '非流动资产合计',   'asset'),
        (8023, '长期投资',        'asset'),
        (8024, '固定资产净额',    'asset'),
        (8025, '土地和改进',      'asset'),
        (8026, '建筑物和改进',    'asset'),
        (8027, '设备',          'asset'),
        (8028, '在建工程',        'asset'),
        (8029, '租赁资产',        'asset'),
        (8030, '无形资产',        'asset'),
        (8031, '商誉',          'asset'),
        (8032, '长期待摊费用',    'asset'),
        (8033, '递延所得税资产',   'asset'),
        (8034, '其他非流动资产',  'asset'),
        (8055, '负债合计',       'liability'),
        (8056, '流动负债合计',    'liability'),
        (8057, '短期借款',        'liability'),
        (8058, '应付账款',        'liability'),
        (8059, '应付票据',        'liability'),
        (8060, '应计费用',        'liability'),
        (8061, '应付所得税',      'liability'),
        (8062, '其他流动负债',    'liability'),
        (8083, '非流动负债合计',   'liability'),
        (8084, '长期借款',        'liability'),
        (8085, '应付债券',        'liability'),
        (8086, '递延所得税负债',   'liability'),
        (8087, '其他非流动负债',  'liability'),
        (8100, '股东权益合计',     'equity'),
        (8101, '归属母公司所有者权益合计', 'equity'),
        (8102, '股本',          'equity'),
        (8103, '资本公积',        'equity'),
        (8104, '留存收益',        'equity'),
        (8105, '其他综合收益',    'equity'),
        (8106, '库藏股',        'equity'),
    ],
    'cashflow': [
        (8015, '经营活动现金流量净额',    'cash_in'),
        (8042, '投资活动现金流量净额',    'cash_out'),
        (8056, '融资活动现金流量净额',    'cash_out'),
        (8067, '现金及现金等价物期末余额', 'cash_in'),
        (8068, '现金及现金等价物净增加额', 'cash_in'),
        (8072, '自由现金流',             'cash_in'),
    ],
}

# 港股 (5xxx field_id) - 基于美团/腾讯实际字段
HK_FIELDS = {
    'income': [
        (5001, '营业总收入',       'revenue'),      # Operating revenue
        (5005, '营业总成本',        'expense'),      # Total operating costs
        (5008, '销售成本',        'expense'),      # Cost of sales
        (5010, '毛利',           'profit'),       # Gross profit
        (5013, '营业费用',        'expense'),      # Operating expenses
        (5015, '销售及分销成本',    'expense'),      # Selling and distribution
        (5016, '行政开支',        'expense'),      # Administrative expenses - 管理费用
        (5017, '研发费用',        'expense'),      # R&D costs
        (5034, '营业利润',        'profit'),       # Operating profit
        (5035, '财务收入',        'profit'),       # Finance income
        (5036, '财务成本',        'expense'),      # Finance costs
        (5037, '应占联营公司盈利', 'profit'),       # Share of profit of associates
        (5040, '税前利润',        'profit'),       # Profit before taxation
        (5043, '所得税',          'tax'),          # Income tax
        (5045, '本年利润',        'profit'),       # Profit/Loss for the year
        (5051, '归属于母公司股东净利润', 'profit'), # Profit for equity shareholders
    ],
    'balance': [
        (5001, '资产合计',       'asset'),
        (5002, '流动资产合计',     'asset'),
        (5003, '现金及现金等价物',  'asset'),
        (5004, '受限制现金',      'asset'),
        (5005, '短期投资',        'asset'),
        (5006, '应收账款',        'asset'),
        (5007, '其他应收款',      'asset'),
        (5008, '预付款项及按金',  'asset'),
        (5009, '存货',          'asset'),
        (5010, '金融资产',        'asset'),
        (5011, '其他流动资产',    'asset'),
        (5029, '非流动资产合计',   'asset'),
        (5030, '物业、厂房及设备', 'asset'),
        (5031, '在建工程',        'asset'),
        (5032, '土地使用权',      'asset'),
        (5033, '投资物业',        'asset'),
        (5034, '无形资产',        'asset'),
        (5035, '商誉',          'asset'),
        (5036, '长期投资',        'asset'),
        (5037, '递延所得税资产',   'asset'),
        (5038, '其他非流动资产',  'asset'),
        (5060, '负债合计',       'liability'),
        (5061, '流动负债合计',    'liability'),
        (5062, '短期借款',        'liability'),
        (5063, '应付账款',        'liability'),
        (5064, '应付票据',        'liability'),
        (5065, '其他应付款项',    'liability'),
        (5066, '合同负债',        'liability'),
        (5067, '应计费用',        'liability'),
        (5068, '应交税费',        'liability'),
        (5069, '一年内到期的长期负债', 'liability'),
        (5070, '其他流动负债',    'liability'),
        (5088, '非流动负债合计',   'liability'),
        (5089, '长期借款',        'liability'),
        (5090, '应付债券',        'liability'),
        (5091, '租赁负债',        'liability'),
        (5092, '递延所得税负债',   'liability'),
        (5093, '其他非流动负债',  'liability'),
        (5109, '股东权益合计',     'equity'),
        (5110, '归属于母公司股东权益合计', 'equity'),
        (5111, '股本',          'equity'),
        (5112, '储备',          'equity'),
        (5113, '留存收益',        'equity'),
        (5114, '其他权益',        'equity'),
    ],
    'cashflow': [
        (5058, '经营活动现金流量净额',    'cash_in'),
        (5076, '投资活动现金流量净额',    'cash_out'),
        (5086, '融资活动现金流量净额',    'cash_out'),
        (5100, '现金及现金等价物期末余额', 'cash_in'),
        (5101, '现金及现金等价物净增加额', 'cash_in'),
    ],
}

# 费用细分类目 (仅利润表) - (field_id: (名称, 是否取绝对值))
EXPENSE_ITEMS_MAP = {
    'us': {
        8007: ('销售和管理费用', False),
        8008: ('销售费用',       True),
        8009: ('管理费用',       True),
        8010: ('研发费用',       False),
    },
    'hk': {
        5015: ('销售及分销成本', False),
        5016: ('行政开支', False),
        5017: ('研发费用', False),
    },
    'a_share': {
        3013: ('销售费用',       False),
        3014: ('管理费用',       False),
        3015: ('研发费用',       False),
    },
}

# ─────────────────────────────────────────────────────────────
# FCFF (公司自由现金流) 计算配置
# ─────────────────────────────────────────────────────────────
# 公式（原始）: FCFF = 净利润 + 利息费用×(1-所得税率) + 折旧摊销 - 营运资本增加 - 资本支出
# 由现金流量表恒等式: OCF ≈ 净利润 + 折旧摊销 - 营运资本增加（忽略其他非现金项目）
# 因此可将 FCFF 重排为可跨市场稳定计算的等价形式:
#     FCFF = 经营活动现金流量净额 + 利息费用×(1-所得税率) - 资本支出
# 该形式仅依赖 3 个字段（OCF / 利息费用 / CapEx），A股、港股、美股均可拉到，比原公式
# 更稳定；并且当报表将 D&A 计入 OCF 加回项、把 ΔWC 计入 OCF 减项时，两式在数学上等价。
#
# 各市场的字段映射：{ 'ocf': 现金流表-经营活动现金流量净额, 'capex': 现金流表-资本支出（绝对值取正）,
#                    'interest_expense': 利润表-利息费用/财务费用（绝对值取正）,
#                    'net_profit': 利润表-净利润, 'ebt': 利润表-税前利润, 'tax': 利润表-所得税 }
# capex 字段有多个候选（不同报表用不同字段），按顺序取第一个非空值并累加。
FCFF_FIELD_MAP = {
    'us': {
        'ocf_fid':              8015,                # 经营活动现金流量净额
        'capex_fids':           [8046],              # 固定资产交易净额（含PPE和无形资产投资净额）
        # 部分公司季报（如 BABA 10-Q）不披露 8046 明细，仅有 8042 投资活动净额，回退使用
        'capex_fallback_fids':  [8042],
        'interest_expense_fid': 8020,                # 营业外利息费用
        'net_profit_fid':       8043,                # 归属于母公司股东净利润（部分公司无 8037）
        'ebt_fid':              8034,                # 税前利润
        'tax_fid':              8035,                # 所得税
    },
    'hk': {
        'ocf_fid':              5058,                # 经营活动现金流量净额（利润表 5058 与现金流表 5001 同义）
        # 港股现金流表：购买固定资产(5071) + 购买无形资产(5073) - 出售固定资产(5070)
        # 但 5058 是利润表口径，现金流表 OCF 为 5001（在利润表配置里 5058 不存在）
        'capex_fids':           [5071, 5073],        # 购买固定资产 + 购买无形资产（取绝对值累加）
        # Q1/Q9 累计季报港股 CF 不披露 5071/5073 明细，仅有 5069 投资活动净额，回退作为 CapEx 近似
        'capex_fallback_fids':  [5069],
        'interest_expense_fid': 5036,                # 财务成本（港股利润表口径）
        'net_profit_fid':       5051,                # 归属于母公司股东净利润
        'ebt_fid':              5040,                # 税前利润
        'tax_fid':              5043,                # 所得税
    },
    'a_share': {
        'ocf_fid':              3001,                # 现金流表：经营活动产生的现金流量净额
        'capex_fids':           [3043],              # 购建固定资产、无形资产和其他长期资产支付的现金
        'interest_expense_fid': 3016,                # 利润表：财务费用（利息费用为主要组成）
        'net_profit_fid':       3043,                # 利润表：净利润
        'ebt_fid':              3038,                # 利润表：利润总额
        'tax_fid':              3039,                # 利润表：所得税费用
    },
}
# 港股现金流表 OCF 使用 5001（与利润表 5058 不同数据来源）
FCFF_CF_OCF_FID = {
    'us':      8015,
    'hk':      5001,
    'a_share': 3001,
}
# 港股利润表中财务成本 5036 已计入营业利润前，直接使用（值多为负数表示费用）
# 各市场利息费用符号约定：均取绝对值作为"利息费用"

# ─────────────────────────────────────────────────────────────
# ROA / ROE 计算配置
# ─────────────────────────────────────────────────────────────
# 公式：
#   ROA = 年化净利润 / 资产合计 × 100%
#   ROE = 年化归属母公司净利润 / 归属母公司股东权益合计 × 100%
#
# 关键点：
#   - 分子取自利润表（年化后），分母取自资产负债表（时点数据，与利润表期次按 fiscal_year 对齐）
#   - 由于富途资产负债表默认仅按年度披露 (financial_type=7)，interim（H1/Q9/Q1 等）期次
#     若匹配不到同 fiscal_year 的年报数据，则使用最近一期的年报余额作为期末近似
#   - 部分公司资产负债表返回的"股东权益合计"字段 ID 与预置映射不同（如 MSFT 用 8081 而非 8100），
#     为提升适配性，同时尝试多个候选字段，取第一个非空值
#
# 字段候选（顺序 = 优先级）：
ROA_ROE_FIELD_MAP = {
    'us': {
        'total_assets_fids':   [8001],                # 资产合计
        'total_equity_fids':   [8085, 8081, 8100, 8101],  # 归属母公司/股东权益合计（顺序尝试）
        'net_profit_fid':      8043,                  # 归属母公司股东净利润（利润表）
        'total_ni_fid':        8037,                  # 净利润
    },
    'hk': {
        'total_assets_fids':   [5001],
        'total_equity_fids':   [5110, 5109],          # 归属母公司股东权益合计 / 股东权益合计
        'net_profit_fid':      5051,                  # 归属母公司股东净利润
        'total_ni_fid':        5045,                  # 本年利润
    },
    'a_share': {
        'total_assets_fids':   [3001],
        'total_equity_fids':   [3098, 3097],          # 归属母公司所有者权益合计 / 股东权益合计
        'net_profit_fid':      3047,                  # 归属母公司净利润
        'total_ni_fid':        3043,                  # 净利润
    },
}


def _pick_first_non_null(items_by_fid, candidate_fids):
    """按候选顺序返回第一个非空值；不同公司字段名不同时用于兜底"""
    for fid in candidate_fids:
        v = items_by_fid.get(fid)
        if v is not None:
            return float(v), fid
    return None, None


def _annualize_factor(period_text):
    """将 period_text 映射为年化倍数

    美股：Q1/Q2/Q3/Q4 均为单季，×4；H1 半年，×2；Q9 三季累计，×4/3；FY 年报 ×1
    港股/A股：Q1（累计3月）×4；H1/Q6（累计6月）×2；Q9（累计9月）×4/3；FY ×1
    """
    if not period_text:
        return 1.0
    pt = period_text.upper()
    if 'FY' in pt:
        return 1.0
    if 'Q9' in pt or '/9' in pt:
        return 4.0 / 3.0
    if 'H1' in pt or 'Q6' in pt:
        return 2.0
    if 'Q1' in pt or 'Q2' in pt or 'Q3' in pt or 'Q4' in pt:
        return 4.0
    return 1.0


def compute_roa_roe_values(income_reports, bs_reports, market_type, logger=None):
    """计算每期的 ROA 与 ROE

    Args:
        income_reports: 利润表期次列表
        bs_reports:     资产负债表期次列表（优先与利润表同频率——季度维度精准对齐，
                        比如 2026Q1 利润表 → 2026Q1 资产负债表 3/31 快照；仅年报数据时
                        interim 期次回退到最近一年报余额）

    Returns:
        (roa_pcts, roe_pcts): 两个 list，长度与 income_reports 一致，元素为百分数或 None
    """
    cfg = ROA_ROE_FIELD_MAP.get(market_type)
    if not cfg:
        return [None] * len(income_reports), [None] * len(income_reports)

    # 按 period_text 建索引（优先，季度级精准对齐）；同时按 fiscal_year 建回退索引
    bs_by_period = {}
    bs_by_year = {}
    bs_sorted = []  # 按 (fiscal_year, period_text) 排序的兜底列表
    for r in bs_reports:
        pt = r.get('period_text', '')
        fy = r.get('fiscal_year')
        items = {i['field_id']: i.get('data') for i in r.get('item_list', [])}
        if pt:
            bs_by_period[pt] = items
        # fiscal_year 索引记录该年最后出现的一条（迭代顺序通常是新→旧，覆盖后即最早一条）；
        # 对回退语义（无同期次时用同 fy 的年报数据）没有强绑定，够用即可
        if fy is not None:
            # 只有 FY 期次才作为"年报余额"回退候选（interim 快照不宜替代其他 interim）
            if 'FY' in pt.upper() or fy not in bs_by_year:
                bs_by_year[fy] = items
            bs_sorted.append((fy, pt, items))
    bs_sorted.sort(key=lambda x: (x[0], x[1]))

    def _pick_bs_items(period_text, fiscal_year):
        """按 period_text 精确匹配；缺失时回退到同 fy 年报，再兜底最近一年"""
        # 1) 优先按 period_text 精确匹配（季度对齐）
        if period_text and period_text in bs_by_period:
            return bs_by_period[period_text], period_text
        # 2) 回退：使用同 fiscal_year 的年报余额（近似期初/期末快照）
        if fiscal_year is not None and fiscal_year in bs_by_year:
            return bs_by_year[fiscal_year], f"FY{fiscal_year}(回退)"
        # 3) 再回退：<= 当前 fiscal_year 的最近一条 BS
        if fiscal_year is not None:
            candidates = [x for x in bs_sorted if x[0] <= fiscal_year]
            if candidates:
                _fy, _pt, items = candidates[-1]
                return items, f"{_pt}(邻近回退)"
        # 4) 最终兜底：取列表第一条
        if bs_sorted:
            _fy, _pt, items = bs_sorted[0]
            return items, f"{_pt}(首条兜底)"
        return None, None

    roa_values = []
    roe_values = []
    for rpt in income_reports:
        period = rpt.get('period_text', '')
        fy = rpt.get('fiscal_year')
        inc_items = {i['field_id']: i.get('data') for i in rpt.get('item_list', [])}

        bs_items, bs_period_tag = _pick_bs_items(period, fy)
        if not bs_items:
            roa_values.append(None)
            roe_values.append(None)
            if logger:
                logger.debug(f"ROA/ROE {period}: 无匹配资产负债表数据")
            continue

        total_assets, ta_fid = _pick_first_non_null(bs_items, cfg['total_assets_fids'])
        total_equity, te_fid = _pick_first_non_null(bs_items, cfg['total_equity_fids'])

        # 净利润：优先归母口径（与 ROE 匹配），若缺失回退到总净利润
        np_raw = inc_items.get(cfg['net_profit_fid'])
        total_ni_raw = inc_items.get(cfg['total_ni_fid'])
        net_profit_attributable = float(np_raw) if np_raw is not None else None
        total_net_income = float(total_ni_raw) if total_ni_raw is not None else net_profit_attributable

        # 年化系数
        factor = _annualize_factor(period)

        # ROA = 年化净利润 / 资产合计
        if total_net_income is not None and total_assets and total_assets != 0:
            roa = (total_net_income * factor) / total_assets * 100
            roa_values.append(round(roa, 2))
        else:
            roa_values.append(None)

        # ROE = 年化归属母公司净利润 / 归属母公司股东权益合计
        if net_profit_attributable is not None and total_equity and total_equity != 0:
            roe = (net_profit_attributable * factor) / total_equity * 100
            roe_values.append(round(roe, 2))
        else:
            roe_values.append(None)

        if logger:
            logger.debug(
                f"ROA/ROE {period} (fy={fy}, bs={bs_period_tag}, ann×{factor:.2f}): "
                f"assets[{ta_fid}]={total_assets}, equity[{te_fid}]={total_equity}, "
                f"ROA={roa_values[-1]}%, ROE={roe_values[-1]}%"
            )

    return roa_values, roe_values


# ─────────────────────────────────────────────────────────────
# 辅助函数
# ─────────────────────────────────────────────────────────────

def setup_logging(log_file):
    """设置日志系统"""
    if not os.path.exists(os.path.dirname(log_file)):
        os.makedirs(os.path.dirname(log_file))

    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s',
        filename=log_file,
        filemode='w',
        encoding='utf-8'
    )
    return logging.getLogger(__name__)


def log_stdout(msg):
    """输出到 stdout（极简）"""
    print(msg, flush=True)


def get_market_type(stock_code):
    """判断市场类型：a_share / hk / us / hk_us

    - A股 (SH./SZ.): 3xxx field_id
    - 港股 (HK.): 5xxx field_id
    - 美股 (US.): 8xxx field_id
    """
    if stock_code.startswith('SH.') or stock_code.startswith('SZ.'):
        return 'a_share'
    elif stock_code.startswith('HK.'):
        return 'hk'
    elif stock_code.startswith('US.'):
        return 'us'
    return 'hk_us'


def normalize_stock_code(stock_code):
    """标准化股票代码格式：
    - 港股（HK.）：代码补零到5位，如 HK.700 → HK.00700
    - A股（SH./SZ.）：代码补零到6位，如 SH.600 → SH.000600
    - 美股（US.）：保持原样
    """
    parts = stock_code.split('.', 1)
    if len(parts) != 2:
        return stock_code

    market, code = parts
    market_upper = market.upper()

    if market_upper == 'HK':
        # 港股5位数字
        code = code.zfill(5)
    elif market_upper in ['SH', 'SZ']:
        # A股6位数字
        code = code.zfill(6)

    return f"{market_upper}.{code}"


def get_unit_config(reports):
    """从财报数据中获取币种单位名称和百万转换除数

    Args:
        reports: API 返回的财报列表
    Returns:
        unit_name: 单位名称（如"百万人民币"、"百万港元"、"百万美元"等）
        divisor: 转换除数
    """
    # 币种代码到显示名称的映射
    currency_mapping = {
        'CNY': '百万人民币',
        'RMB': '百万人民币',
        'HKD': '百万港元',
        'HK$': '百万港元',
        'USD': '百万美元',
        'US$': '百万美元',
        'EUR': '百万欧元',
        'GBP': '百万英镑',
        'JPY': '百万日元',
        'SGD': '百万新加坡元',
        'AUD': '百万澳元',
        'CAD': '百万加元',
    }

    currency_code = None
    currency_info = None

    # 从第一条有币种信息的报表中获取
    if reports:
        for report in reports:
            currency_code = report.get('currency_code')
            currency_info = report.get('currency_info')
            if currency_code or currency_info:
                break

    # 优先使用 currency_code 映射，其次使用 currency_info
    if currency_code and currency_code in currency_mapping:
        unit_name = currency_mapping[currency_code]
    elif currency_info:
        # 使用返回的 currency_info，添加"百万"前缀
        unit_name = f"百万{currency_info}"
    else:
        # 默认兜底
        unit_name = "百万"

    # 所有币种原始单位都是元，转换到百万单位
    divisor = 1000000

    return unit_name, divisor, currency_code, currency_info


# ─────────────────────────────────────────────────────────────
# 错误类型
# ─────────────────────────────────────────────────────────────

class FetchError(Exception):
    """获取财务数据失败，带用户可读的错误信息和修复建议。"""
    def __init__(self, message, hint=None, log_detail=None):
        super().__init__(message)
        self.message = message
        self.hint = hint
        self.log_detail = log_detail

    def user_message(self):
        parts = [f"ERROR: {self.message}"]
        if self.hint:
            parts.append(f"提示: {self.hint}")
        return "\n".join(parts)


def validate_stock_code(stock_code):
    """校验股票代码格式，返回 (ok, error_message)。"""
    if not stock_code or not stock_code.strip():
        return False, "股票代码不能为空"
    code = stock_code.strip()
    if '.' not in code:
        return False, (
            f"股票代码格式错误: '{code}'。必须包含市场前缀和代码，用点号分隔。\n"
            f"正确格式示例：US.BABA（美股）、HK.00700（港股）、SH.600519（沪市A股）、SZ.000001（深市A股）\n"
            f"提示: 可用 stock-ticker skill 按公司名查股票代码，如 "
            f"`python workspace/skills/stock-ticker/scripts/search_ticker.py --company 腾讯`"
        )
    prefix, suffix = code.split('.', 1)
    prefix_upper = prefix.upper()
    if prefix_upper not in ('US', 'HK', 'SH', 'SZ'):
        return False, (
            f"不支持的市场前缀: '{prefix}'。支持的市场：US（美股）、HK（港股）、SH（沪市）、SZ（深市）。\n"
            f"正确格式示例：US.BABA / HK.00700 / SH.600519 / SZ.000001"
        )
    if not suffix or not suffix.strip():
        return False, f"股票代码缺失: '{code}'。点号后必须是股票代码，如 US.BABA 而非 US."
    if prefix_upper in ('HK',) and not suffix.isdigit():
        return False, f"港股代码必须是纯数字，如 HK.00700，收到 '{suffix}'"
    if prefix_upper in ('SH', 'SZ') and not suffix.isdigit():
        return False, f"A股代码必须是纯数字，如 SH.600519，收到 '{suffix}'"
    return True, None


def get_financial_data_with_structure(stock_code, statement_type, num=16, logger=None, financial_type_override=None):
    """从富途API获取财务数据及字段结构

    三张表统一按季度维度拉取，让利润表 / 资产负债表 / 现金流量表期次严格对齐：
    - 美股 (US.)：10 = 单季报+年报（美股按 10-Q/10-K 发布独立季度报表）
    - 港股 (HK.) / A 股 (SH./SZ.)：11 = 累计季报（Q1/Q6/Q9/年报）
      H 股/A 股企业只发布"一季报/半年报/三季报/年报"，不发布独立 Q2/Q3/Q4 单季报，
      因此用 11 才能拿到 Q6（半年报）与 Q9（三季报）；若继续用 10 会导致 Q2/Q3 全空。

    含义：
    - 利润表：区间累计（美股单季，HK/A股累计）
    - 现金流量表：与利润表同频，Q1/Q6/Q9/FY 为期初到该期末的累计口径；美股为单季
    - 资产负债表：季度末时点快照（3/31、6/30、9/30、12/31）

    对齐效果：利润表 2026Q1 ↔ 现金流量表 2026Q1（1–3 月累计）↔ 资产负债表 2026Q1
    （3-31 快照），三表期次可通过 period_text 直接对应。

    Args:
        financial_type_override: 显式指定 financial_type，覆盖默认逻辑（保留给未来定制拉取用；
                                 现在三张表的默认逻辑已一致，主流程一般不再需要 override）

    Raises:
        FetchError: 带用户可读信息的错误
    """
    st_code = STATEMENT_TYPES[statement_type][0]

    if financial_type_override is not None:
        financial_type = financial_type_override
    else:
        # 三张表默认统一按季度维度：港股 / A 股走累计季报 (Q1/Q6/Q9/FY)；美股走单季+年报
        market_type = get_market_type(stock_code)
        financial_type = 11 if market_type in ('hk', 'a_share') else 10

    if logger:
        msg = {7: '年报', 9: '单季报组合', 10: '单季报+年报', 11: '累计季报（Q1/Q6/Q9/年报）'}
        logger.info(f"使用 financial_type={financial_type} ({msg.get(financial_type, '未知')})")

    # 设置 PYTHONPATH 以找到 common 模块（在 futuapi/scripts 目录下）
    skill_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    futuapi_scripts_dir = os.path.normpath(os.path.join(skill_dir, "..", "futuapi", "scripts"))
    env = os.environ.copy()
    env['PYTHONPATH'] = futuapi_scripts_dir

    # 检查 futuapi 脚本是否存在
    if not os.path.isfile(FUTUAPI_SCRIPT):
        raise FetchError(
            f"futuapi 脚本不存在: {FUTUAPI_SCRIPT}",
            hint="请确认 workspace/skills/futuapi/scripts/quote/get_financials_statements.py 文件完整"
        )

    try:
        result = subprocess.run(
            [PYTHON_EXE, FUTUAPI_SCRIPT, stock_code,
             "--statement-type", str(st_code),
             "--financial-type", str(financial_type),
             "--num", str(num),
             "--json"],
            capture_output=True,
            timeout=120,
            env=env
        )
    except subprocess.TimeoutExpired:
        raise FetchError(
            f"获取 {stock_code} 财务数据超时（120秒）",
            hint="可能是 Futu OpenD 无响应或网络问题，请检查 OpenD 是否正常运行，或稍后重试",
        )
    except FileNotFoundError as e:
        raise FetchError(
            f"Python 解释器不存在或 futuapi 脚本无法执行",
            hint=f"错误详情: {e}"
        )

    stdout_text = result.stdout.decode('utf-8', errors='ignore')
    stderr_text = result.stderr.decode('utf-8', errors='ignore') if result.stderr else ""

    if logger:
        if stderr_text:
            logger.info(f"futuapi stderr: {stderr_text[:2000]}")
        logger.info(f"futuapi returncode={result.returncode}")

    # 检查子进程返回码
    if result.returncode != 0:
        # 尝试从 stderr 中提取关键信息
        stderr_lower = stderr_text.lower()
        if 'connect' in stderr_lower or 'refused' in stderr_lower or '11111' in stderr_text or 'opend' in stderr_lower:
            raise FetchError(
                f"无法连接 Futu OpenD（获取 {stock_code} 数据失败）",
                hint="请确认 Futu OpenD 已启动并登录，默认监听 127.0.0.1:11111。"
                     "可运行 `python workspace/skills/futuapi/check_env.py` 检查环境"
            )
        if 'no data' in stderr_lower or 'retcode' in stderr_lower and '0,' not in stderr_text:
            raise FetchError(
                f"股票 {stock_code} 无财务数据返回",
                hint=f"请检查代码是否正确，或该股票在富途是否有对应报表数据。stderr: {stderr_text[:500]}"
            )
        raise FetchError(
            f"获取 {stock_code} 财务数据失败（exit code {result.returncode}）",
            hint=f"错误详情见日志文件。stderr: {stderr_text[:300]}" if stderr_text else "可能是代码格式错误或 OpenD 未登录",
            log_detail=stderr_text
        )

    # 用bytes模式读取，避免Windows编码问题
    content = stdout_text

    # 如果 stdout 中包含错误信息但没有 JSON，给出具体提示
    json_start = content.find('{"code"')
    if json_start == -1:
        # 尝试从内容中识别错误
        content_lower = content.lower()
        if 'error' in content_lower or 'traceback' in content_lower:
            raise FetchError(
                f"获取 {stock_code} 财务数据返回异常",
                hint=f"API 返回了错误信息。stdout 摘要: {content[:500]}",
                log_detail=content
            )
        if not content.strip():
            raise FetchError(
                f"获取 {stock_code} 财务数据返回为空",
                hint="请检查 OpenD 是否正常运行、股票代码是否正确"
            )
        raise FetchError(
            f"无法解析 {stock_code} 的 API 返回数据（未找到 JSON 响应）",
            hint=f"返回内容前200字符: {content[:200]}"
        )

    json_end = content.rfind('}')
    if json_end == -1:
        raise FetchError(f"API 返回 JSON 不完整", hint="可能是网络中断或接口异常，请重试")

    json_str = content[json_start:json_end + 1]
    try:
        data = json.loads(json_str)
    except json.JSONDecodeError as e:
        raise FetchError(
            f"API 返回 JSON 解析失败: {e}",
            hint=f"JSON 片段: {json_str[:200]}"
        )

    # 检查错误：futuapi 脚本成功时返回 {"code": "<stock_code>", "data": {...}}，
    # code 字段是股票代码字符串而非数字返回码；失败时返回 {"error": "..."} 无 code/data。
    if 'error' in data and not data.get('data'):
        err_msg = str(data.get('error', ''))
        if 'no permission' in err_msg.lower() or 'permission' in err_msg.lower():
            raise FetchError(
                f"无权限获取 {stock_code} 的财务数据（富途返回: {err_msg}）",
                hint="请确认富途账号已登录且有该市场行情权限（港股/A股/美股可能需要相应权限包）"
            )
        raise FetchError(
            f"API 返回错误: {err_msg}",
            hint="可能是代码错误、停牌、或接口限流，请检查代码格式后重试"
        )
    ret_msg = data.get('msg', '')
    if ret_msg and not data.get('data'):
        raise FetchError(
            f"API 返回错误: {ret_msg}",
            hint="可能是代码错误、停牌、或接口限流，请检查代码格式后重试"
        )

    data_section = data.get('data', {})
    structure_list = data_section.get('structure_list')
    report_list = data_section.get('report_list')

    if not report_list:
        raise FetchError(
            f"股票 {stock_code} 无财报数据返回",
            hint=(
                "可能原因：\n"
                "  1. 股票代码错误或市场前缀错误（如把美股 BABA 写成 HK.BABA）\n"
                "  2. 该股票尚未发布任何财报（如新股/SPAC）\n"
                "  3. 请用 stock-ticker skill 确认代码："
                " `python workspace/skills/stock-ticker/scripts/search_ticker.py --company <公司名>`"
            )
        )

    if not structure_list:
        raise FetchError(
            f"股票 {stock_code} 的财报无字段结构信息",
            hint="该股票可能使用了不支持的报表格式，请检查是否为标准上市公司财报"
        )

    # 构建 field_id -> 显示名 的映射
    name_map = {e['field_id']: e['display_name'] or f"字段{e['field_id']}" for e in structure_list}

    return report_list, name_map


def convert_quarter_label(raw_label):
    """转换季度标签格式"""
    match = re.match(r'(\d{4}).*Q(\d)', raw_label)
    if match:
        return f"{match.group(1)}Q{match.group(2)}"
    return raw_label.replace('/', '').replace(' ', '')


def get_period_range(report):
    """计算财报周期的日期范围（资产负债表是时点数据，只需日期）"""
    end_date_str = report.get('date_time_str', '')
    return end_date_str


def format_value(val, divisor, is_pct=False):
    """格式化值"""
    if val is None or val == '' or val == 'N/A':
        return None
    if not isinstance(val, (int, float)):
        return val
    if is_pct:
        return round(val, 2)
    else:
        return round(val / divisor, 0)


def calculate_ratio(numerator_values, denominator_values, num_quarters, take_abs_numerator=False, logger=None):
    """计算比率（分子/分母 * 100），处理空值

    Args:
        numerator_values: 分子值数组
        denominator_values: 分母值数组
        num_quarters: 季度数
        take_abs_numerator: 是否对分子取绝对值（用于费用率计算，费用可能为负数）
    """
    ratio_values = []
    for i in range(num_quarters):
        n = numerator_values[i] if i < len(numerator_values) else None
        d = denominator_values[i] if i < len(denominator_values) else None
        if n is not None and d is not None and d != 0:
            # 对费用类分子取绝对值，避免费用率为负
            if take_abs_numerator:
                n = abs(n)
            ratio = round((n / d) * 100, 2)
            ratio_values.append(ratio)
        else:
            ratio_values.append(None)
    return ratio_values


def calculate_profit_margin_a_share(revenue_values, cost_values, num_quarters):
    """计算A股毛利率 (营收-成本)/营收"""
    margin_values = []
    for i in range(num_quarters):
        r = revenue_values[i] if i < len(revenue_values) else None
        c = cost_values[i] if i < len(cost_values) else None
        if r is not None and c is not None and r != 0:
            margin = round(((r - c) / r) * 100, 2)
            margin_values.append(margin)
        else:
            margin_values.append(None)
    return margin_values


def compute_fcff_values(income_reports, cf_by_period, market_type, logger=None):
    """计算 FCFF (公司自由现金流) 每期数值

    实现公式（等价形式，跨市场稳定）:
        FCFF = 经营活动现金流量净额 + 利息费用 × (1 - 所得税率) - 资本支出

    Args:
        income_reports:   利润表期次列表（顺序即列顺序）
        cf_by_period:     {period_text: item_map_by_field_id} 现金流量表按期查询
        market_type:      'us' / 'hk' / 'a_share'

    Returns:
        (fcff_list, capex_list): 两个 list，长度与 income_reports 一致，元素为原始金额（元）
                                  或 None（关键项缺失）。CapEx 已取绝对值（正数表示流出规模）
    """
    cfg = FCFF_FIELD_MAP.get(market_type)
    if not cfg:
        return [None] * len(income_reports), [None] * len(income_reports)

    cf_ocf_fid = FCFF_CF_OCF_FID.get(market_type, cfg.get('ocf_fid'))
    interest_fid = cfg['interest_expense_fid']
    net_profit_fid = cfg['net_profit_fid']
    ebt_fid = cfg['ebt_fid']
    tax_fid = cfg['tax_fid']
    capex_fids = cfg['capex_fids']

    results = []
    capex_out = []
    for rpt in income_reports:
        period = rpt.get('period_text', '')
        inc_items = {i['field_id']: i.get('data') for i in rpt.get('item_list', [])}
        cf_items = cf_by_period.get(period, {})

        # 1) OCF：来自现金流量表
        ocf = cf_items.get(cf_ocf_fid)

        # 2) 资本支出：现金流量表中多个字段（购建固定资产/无形资产等）累加取绝对值
        capex_total = 0.0
        capex_any = False
        for fid in capex_fids:
            v = cf_items.get(fid)
            if v is not None:
                capex_total += abs(float(v))
                capex_any = True
        # 回退：明细字段全缺失时使用总投资活动净额（用于 HK 累计季报 Q1/Q9）
        if not capex_any:
            for fid in cfg.get('capex_fallback_fids', []):
                v = cf_items.get(fid)
                if v is not None:
                    capex_total += abs(float(v))
                    capex_any = True
        capex = capex_total if capex_any else None

        # 3) 利息费用：来自利润表，取绝对值
        interest_raw = inc_items.get(interest_fid)
        interest = abs(float(interest_raw)) if interest_raw is not None else None

        # 4) 所得税率：由利润表 所得税 / 税前利润 推算；若无有效值则回退 25%
        ebt_raw = inc_items.get(ebt_fid)
        tax_raw = inc_items.get(tax_fid)
        tax_rate = None
        if ebt_raw is not None and tax_raw is not None:
            try:
                ebt_f = float(ebt_raw)
                tax_f = abs(float(tax_raw))
                if ebt_f != 0:
                    tr = tax_f / abs(ebt_f)
                    # 约束在合理区间 [0, 0.5]
                    if 0 <= tr <= 0.5:
                        tax_rate = tr
            except (TypeError, ValueError):
                pass
        if tax_rate is None:
            tax_rate = 0.25  # 默认名义税率兜底

        # 关键项缺失时该期无法计算
        if ocf is None or capex is None:
            results.append(None)
            capex_out.append(capex)  # capex 若单独可用，仍然输出（用于单独展示）
            if logger:
                logger.debug(f"FCFF {period}: 关键项缺失 ocf={ocf} capex={capex}")
            continue

        interest_component = interest * (1 - tax_rate) if interest is not None else 0.0
        fcff = float(ocf) + interest_component - capex
        results.append(fcff)
        capex_out.append(capex)

        if logger:
            logger.debug(
                f"FCFF {period}: ocf={ocf} + interest={interest}×(1-{tax_rate:.3f}) "
                f"- capex={capex} = {fcff:.0f}"
            )

    return results, capex_out


def build_rows_with_ratios(reports, name_map, field_config, expense_items, divisor, unit_name,
                            statement_type, market_type, ratio_config, logger=None,
                            fcff_values=None, capex_values=None,
                            roa_values=None, roe_values=None):
    """构建数据行，含财务比率计算（XX率紧跟XX后）

    Args:
        fcff_values:  FCFF 每期原始金额（未除以divisor），仅利润表使用
        capex_values: 资本支出 CapEx 每期原始金额（已取绝对值，未除以divisor），仅利润表使用
        roa_values:   ROA 百分数（已 ×100），仅利润表使用
        roe_values:   ROE 百分数（已 ×100），仅利润表使用
    """
    rows = []
    yoy_tracking = []  # 记录哪些行是YoY行，用于高亮：(row_idx, [value1, value2, ...])

    # 表头
    quarters = [convert_quarter_label(r.get('period_text', '')) for r in reports]
    num_quarters = len(quarters)
    header = ["财务指标", "单位"] + quarters
    rows.append(header)

    # 财报周期行
    period_row = ["财报日期", ""] + [get_period_range(r) for r in reports]
    rows.append(period_row)

    # 预提取所有字段值（原始数据，不除以divisor）
    field_values_raw = {}
    field_values_yoy_raw = {}
    for r in reports:
        items = {i['field_id']: i for i in r.get('item_list', [])}
        for fid in items:
            if fid not in field_values_raw:
                field_values_raw[fid] = []
                field_values_yoy_raw[fid] = []
            field_values_raw[fid].append(items[fid].get('data'))
            field_values_yoy_raw[fid].append(items[fid].get('yoy'))

    # 构建 field_id -> 行索引 的映射（用于插入比率）
    # 先构建基础字段位置映射
    base_field_positions = {}
    for idx, (fid, _, _) in enumerate(field_config):
        base_field_positions[fid] = idx

    # 确定插入位置 - 预先计算所有需要插入的比率位置
    insertions = {}  # {父项field_id: [(比率名称, 分子fid, 分母fid), ...]}

    # 确定收入字段 - 处理部分公司没有8001但有8002的情况
    if statement_type == 'income':
        if market_type == 'a_share':
            revenue_fid = 3001
        elif market_type == 'hk':
            revenue_fid = 5001
        else:  # us
            # 美股优先用8001（总收入），如果不存在则用8002（营业总收入）
            if 8001 in name_map:
                revenue_fid = 8001
            else:
                revenue_fid = 8002

        # 主指标比率
        for ratio_name, num_fid, den_fid, parent_fid in ratio_config.get('income', []):
            # 对于美股，如果原来的分母是8001但实际用的是8002，更新分母
            if market_type == 'us' and den_fid == 8001 and revenue_fid == 8002:
                den_fid = 8002
            if parent_fid not in insertions:
                insertions[parent_fid] = []
            insertions[parent_fid].append((ratio_name, num_fid, den_fid))

        # 费用比率 - 营业费用（注意：营业费用是主字段，不在expense_items子项中）
        if market_type == 'a_share':
            operating_expense_fid = 3005
        elif market_type == 'hk':
            operating_expense_fid = 5015
        else:  # us
            operating_expense_fid = 8005

        if operating_expense_fid in name_map:
            if operating_expense_fid not in insertions:
                insertions[operating_expense_fid] = []
            insertions[operating_expense_fid].append(('营业费用率', operating_expense_fid, revenue_fid))

        # 子项费用比率
        if expense_items:
            for expense_name, expense_fid, den_fid in ratio_config.get('expense_ratios', []):
                # 营业费用率已单独处理，跳过
                if '营业费用率' in expense_name:
                    continue
                # 对于美股，如果原来的分母是8001但实际用的是8002，更新分母
                if market_type == 'us' and den_fid == 8001 and revenue_fid == 8002:
                    den_fid = 8002
                if expense_fid in name_map and expense_fid in expense_items:
                    if expense_fid not in insertions:
                        insertions[expense_fid] = []
                    insertions[expense_fid].append((expense_name, expense_fid, den_fid))

    # 主循环 - 逐行构建
    for field_id, default_name, category in field_config:
        # 跳过不存在的字段
        if field_id not in name_map:
            continue

        actual_name = name_map.get(field_id, default_name)
        raw_values = field_values_raw.get(field_id, [None]*num_quarters)

        # 对费用、成本、支出类字段取绝对值（部分公司用负数表示费用）
        if category in ['expense', 'tax']:
            raw_values = [abs(v) if v is not None else None for v in raw_values]

        values = [format_value(v, divisor) for v in raw_values]
        rows.append([actual_name, unit_name] + values)

        # 添加 YoY 行（资产负债表除外）
        if statement_type != 'balance' and field_id in field_values_yoy_raw:
            yoy_values_raw = field_values_yoy_raw.get(field_id, [None]*num_quarters)
            # 注意：YoY 已经是百分比变化，不需要取绝对值
            yoy_values = [round(v, 2) if v is not None else None for v in yoy_values_raw]
            yoy_row_idx = len(rows)
            rows.append([f'{actual_name} YoY', '%'] + yoy_values)
            yoy_tracking.append((yoy_row_idx, yoy_values))  # 记录用于高亮

        # 检查是否需要在此字段后插入比率行
        if field_id in insertions:
            for ratio_name, num_fid, den_fid in insertions[field_id]:
                # 获取原始值计算比率
                num_raw = field_values_raw.get(num_fid, [None]*num_quarters)
                den_raw = field_values_raw.get(den_fid, [None]*num_quarters)

                # 特殊处理：A股毛利率 = (营收 - 成本)/营收
                if market_type == 'a_share' and ratio_name == '毛利率':
                    revenue_raw = field_values_raw.get(3001, [None]*num_quarters)
                    cost_raw = field_values_raw.get(3009, [None]*num_quarters)
                    ratio_values = calculate_profit_margin_a_share(revenue_raw, cost_raw, num_quarters)
                else:
                    # 费用率计算时，对分子（费用）取绝对值
                    take_abs = '费用率' in ratio_name or '成本率' in ratio_name
                    ratio_values = calculate_ratio(num_raw, den_raw, num_quarters, take_abs)

                rows.append([ratio_name, '%'] + ratio_values)

        # 特殊处理：A股营业费用（字段名可能为"销售费用"等）直接插入营业费用率
        # 检测字段名判断是否为营业费用相关
        if statement_type == 'income' and field_id in name_map:
            field_name = name_map[field_id]
            is_operating_expense = (field_id in [8005, 3005] or '营业费用' in field_name or '销售费用' in field_name and field_id not in [3013, 8008])
            if is_operating_expense and '营业费用率' not in str(rows):
                revenue_fid = 3001 if market_type == 'a_share' else 8001
                num_raw = field_values_raw.get(field_id, [None]*num_quarters)
                den_raw = field_values_raw.get(revenue_fid, [None]*num_quarters)
                # 如果 8001 不存在，尝试用 8002（营业总收入）
                if revenue_fid == 8001 and (not den_raw or all(v is None for v in den_raw)):
                    den_raw = field_values_raw.get(8002, [None]*num_quarters)
                ratio_values = calculate_ratio(num_raw, den_raw, num_quarters, take_abs_numerator=True)
                rows.append(['营业费用率', '%'] + ratio_values)

        # 处理费用细分类目及其比率
        if expense_items and statement_type == 'income' and field_id in [8005, 3005]:  # 营业费用字段
            for exp_fid, (exp_default_name, take_abs) in expense_items.items():
                if exp_fid in name_map and exp_fid in field_values_raw:
                    exp_actual_name = name_map.get(exp_fid, exp_default_name)
                    exp_raw_values = field_values_raw.get(exp_fid, [None]*num_quarters)
                    if take_abs:
                        exp_raw_values = [abs(v) if v is not None else None for v in exp_raw_values]
                    exp_values = [format_value(v, divisor) for v in exp_raw_values]
                    rows.append([exp_actual_name, unit_name] + exp_values)

                    # 费用 YoY 行
                    if exp_fid in field_values_yoy_raw:
                        exp_yoy_raw = field_values_yoy_raw.get(exp_fid, [None]*num_quarters)
                        exp_yoy_values = [round(v, 2) if v is not None else None for v in exp_yoy_raw]
                        exp_yoy_row_idx = len(rows)
                        rows.append([f'{exp_actual_name} YoY', '%'] + exp_yoy_values)
                        yoy_tracking.append((exp_yoy_row_idx, exp_yoy_values))

                    # 费用比率（如果配置了）
                    if exp_fid in insertions:
                        for ratio_name, num_f, den_f in insertions[exp_fid]:
                            rev_raw = field_values_raw.get(den_f, [None]*num_quarters)
                            # 如果 8001 不存在，尝试用 8002（营业总收入）
                            if den_f == 8001 and (not rev_raw or all(v is None for v in rev_raw)):
                                rev_raw = field_values_raw.get(8002, [None]*num_quarters)
                            # exp_raw_values 已经取过绝对值了，不需要再取
                            ratio_values = calculate_ratio(exp_raw_values, rev_raw, num_quarters, take_abs_numerator=False)
                            rows.append([ratio_name, '%'] + ratio_values)

    # ── 加工指标：公司自由现金流 FCFF (仅利润表) ──
    # 公式: FCFF = OCF + 利息费用 × (1 - 所得税率) - 资本支出
    # （等价于 净利润 + 利息费用×(1-税率) + 折旧摊销 - 营运资本增加 - 资本支出，
    #  但仅依赖现金流表 OCF/CapEx + 利润表 利息/税率，跨市场稳定）
    if statement_type == 'income' and fcff_values is not None:
        fcff_display = [format_value(v, divisor) for v in fcff_values]
        rows.append(['公司自由现金流(FCFF)', unit_name] + fcff_display)

        # ── 加工指标：自由现金流净利润比 FCF/NI (紧跟 FCFF 行) ──
        # 公式：FCF/NI = FCFF / 净利润 × 100%
        # 分母口径：与 FCFF 的"公司整体"口径对齐，优先使用总净利润 (US 8037 / HK 5045 / A股 3043)；
        # 若总净利润缺失则回退到归母净利润 (US 8043 / HK 5051 / A股 3047)
        # 语义：>100% 表示现金创造能力强于账面利润（如高折旧或客户预付款）；
        #      <100% 或负值多因大额 CapEx / 应收拉长 / 存货堆积等现金流质量问题
        ni_cfg = ROA_ROE_FIELD_MAP.get(market_type)
        fcf_ni_values = []
        for i in range(num_quarters):
            fcff_raw = fcff_values[i] if i < len(fcff_values) else None
            np_raw = None
            if ni_cfg:
                # 优先总净利润，缺失回退归母
                total_ni = field_values_raw.get(ni_cfg['total_ni_fid'], [None]*num_quarters)
                np_raw = total_ni[i] if i < len(total_ni) else None
                if np_raw is None:
                    attr_ni = field_values_raw.get(ni_cfg['net_profit_fid'], [None]*num_quarters)
                    np_raw = attr_ni[i] if i < len(attr_ni) else None
            if fcff_raw is not None and np_raw is not None and np_raw != 0:
                fcf_ni_values.append(round(float(fcff_raw) / float(np_raw) * 100, 2))
            else:
                fcf_ni_values.append(None)
        rows.append(['自由现金流净利润比(FCF/NI)', '%'] + fcf_ni_values)

    # ── 加工指标：资本支出 CapEx (仅利润表) ──
    # 数据来源：现金流量表（购建固定资产/无形资产等，取绝对值累加；无明细时回退到投资活动净额）
    # 紧跟 FCF/NI 行之后，便于对照 FCFF = OCF + 利息税盾 - CapEx 的计算过程
    if statement_type == 'income' and capex_values is not None:
        capex_display = [format_value(v, divisor) for v in capex_values]
        rows.append(['资本支出(CapEx)', unit_name] + capex_display)

    # ── 加工指标：ROA / ROE (仅利润表) ──
    # 公式：
    #   ROA = 年化净利润 / 资产合计 × 100%
    #   ROE = 年化归属母公司净利润 / 归属母公司股东权益合计 × 100%
    # 分子来自利润表（interim 期次年化），分母来自资产负债表（按 fiscal_year 匹配）
    if statement_type == 'income' and roa_values is not None:
        rows.append(['ROA(总资产收益率, 年化)', '%'] + list(roa_values))
    if statement_type == 'income' and roe_values is not None:
        rows.append(['ROE(净资产收益率, 年化)', '%'] + list(roe_values))

    if logger:
        logger.info(f"构建完成，共 {len(rows)} 行数据，{len(yoy_tracking)} 个YoY行需要高亮检查")

    return rows, quarters, yoy_tracking


def get_row_category(row_name):
    """根据行名称获取类别用于着色"""
    # 0. 加工指标（FCFF / FCF-NI / CapEx / ROA / ROE 等）优先匹配 - 中蓝背景
    # 注意：FCF/NI 单独分类为浅橙色，需在 FCFF 判断之前处理（否则"自由现金流"关键字会被 FCFF 抢先命中）
    if 'FCF/NI' in row_name or '自由现金流净利润比' in row_name:
        return 'fcf_ni'
    if 'FCFF' in row_name or '自由现金流' in row_name:
        return 'derived'
    if 'CapEx' in row_name or '资本支出' in row_name:
        return 'derived'
    if 'ROA' in row_name or 'ROE' in row_name or '资产收益率' in row_name or '净资产收益率' in row_name:
        return 'derived'

    # 1. 核心利润率优先匹配 - 橄榄绿背景
    core_ratios = ['毛利率', '营业利润率', '净利润率']
    for ratio in core_ratios:
        if ratio in row_name:
            return 'core_ratio'

    # 2. 费用率 - 豆绿色背景
    expense_ratios = ['营业费用率', '销售费用率', '管理费用率', '研发费用率', '销售和管理费用率']
    for ratio in expense_ratios:
        if ratio in row_name:
            return 'expense_ratio'

    # 3. 其他行按类别匹配
    category_markers = [
        ('收入', 'revenue'), ('营收', 'revenue'),
        ('成本', 'expense'), ('费用', 'expense'), ('支出', 'expense'),
        ('毛利', 'profit'), ('利润', 'profit'), ('收益', 'profit'), ('净利', 'profit'),
        ('资产', 'asset'), ('流动资产', 'asset'), ('非流动资产', 'asset'),
        ('负债', 'liability'), ('流动负债', 'liability'), ('非流动负债', 'liability'),
        ('权益', 'equity'), ('股东权益', 'equity'),
        ('经营', 'cash_in'), ('投资', 'cash_out'), ('融资', 'cash_out'),
        ('现金流', 'cash_in'), ('现金', 'cash_in'),
        ('税', 'tax'),
        ('率', 'ratio'),  # 其他比率 - 黄色背景
        ('YoY', 'yoy'),
    ]
    for keyword, category in category_markers:
        if keyword in row_name:
            return category
    return 'default'


def generate_excel_with_styling(rows, quarters, yoy_tracking, output_file, sheet_name, logger=None):
    """生成带样式的Excel（含YoY下降红色高亮和YoY增长绿色高亮）"""
    if not HAS_OPENPYXL:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = sheet_name

    # 基础样式
    header_font = Font(bold=True, name="微软雅黑")
    normal_font = Font(name="微软雅黑")
    red_font = Font(name="微软雅黑", color="FF0000", bold=True)    # 红色高亮字体（下降）
    green_font = Font(name="微软雅黑", color="00B050", bold=True)  # 绿色高亮字体（增长）
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 构建 YoY 行索引集合用于快速查找
    yoy_row_indices = {row_idx: values for row_idx, values in yoy_tracking}

    # 逐行写入
    for row_idx, row in enumerate(rows, 1):  # openpyxl 从1开始
        # 检查是否是 YoY 行（注意rows索引从0开始，row_idx从1开始，偏移量为1）
        is_yoy_row = (row_idx - 1) in yoy_row_indices
        yoy_values = yoy_row_indices.get(row_idx - 1, []) if is_yoy_row else []

        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = thin_border

            # 对齐方式
            if col_idx == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            # 表头样式
            if row_idx == 1:
                cell.font = header_font
                cell.fill = PatternFill(start_color=COLORS['header'],
                                        end_color=COLORS['header'],
                                        fill_type="solid")
            else:
                # 背景色（按类别）
                row_name = row[0] if row else ""
                category = get_row_category(row_name)
                bg_color = COLORS.get(category)
                if bg_color:
                    cell.fill = PatternFill(start_color=bg_color,
                                            end_color=bg_color,
                                            fill_type="solid")

                # YoY 行高亮：下降超过阈值=红色，增长超过阈值=绿色
                if is_yoy_row and col_idx >= 3:  # 第1列是名称，第2列是单位，第3列开始是数据
                    value_idx = col_idx - 3  # 转换为yoy_values的索引
                    if value_idx < len(yoy_values) and yoy_values[value_idx] is not None:
                        yoy_val = yoy_values[value_idx]
                        if isinstance(yoy_val, (int, float)):
                            if yoy_val <= YOY_DECLINE_THRESHOLD:
                                cell.font = red_font  # 红色高亮（大幅下降）
                            elif yoy_val >= YOY_GROWTH_THRESHOLD:
                                cell.font = green_font  # 绿色高亮（大幅增长）
                            else:
                                cell.font = normal_font
                        else:
                            cell.font = normal_font
                    else:
                        cell.font = normal_font
                else:
                    cell.font = normal_font

    # 调整列宽
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    for col in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # 冻结窗格和筛选
    ws.freeze_panes = 'C3'
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # 确保目录存在
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    wb.save(output_file)

    if logger:
        logger.info(f"Excel saved: {output_file}")
        logger.info(f"YoY highlight thresholds: decline {YOY_DECLINE_THRESHOLD}% (red), growth {YOY_GROWTH_THRESHOLD}% (green)")

    return output_file


def main():
    parser = argparse.ArgumentParser(description='流式生成财务报表Excel')
    parser.add_argument('stock_code', help='股票代码，如 US.BABA / HK.00700 / SH.600519')
    parser.add_argument('--type', '-t', default='income', choices=['income', 'balance', 'cashflow'],
                        help='报表类型：income(利润表), balance(资产负债表), cashflow(现金流量表)')
    parser.add_argument('--num', '-n', type=int, default=16, help='季度数量，默认16')
    parser.add_argument('--output', '-o', help='输出Excel文件路径')
    parser.add_argument('--force', action='store_true',
                        help='强制重新生成，忽略7天内的已有Excel产物')
    args = parser.parse_args()

    # 确保日志目录存在（相对于项目根目录）
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # workspace/skills/futu-financial-report/scripts/ -> 上4层到项目根目录
    project_root = os.path.abspath(os.path.join(script_dir, '..', '..', '..', '..'))
    logs_dir = os.path.join(project_root, 'workspace', 'excels', 'logs')
    if not os.path.exists(logs_dir):
        os.makedirs(logs_dir)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    statement_name = STATEMENT_TYPES[args.type][1]

    # 标准化股票代码格式（HK.700 → HK.00700, SH.600 → SH.000600）
    original_code = args.stock_code
    normalized_code = normalize_stock_code(original_code)

    # 缓存复用：默认输出路径下若已有 7 天内的 Excel 产物，直接返回复用（除非 --force）
    if not args.output and not args.force:
        excels_dir = os.path.join(project_root, 'workspace', 'excels')
        safe_code = normalized_code.replace('.', '_')
        cached = find_recent_excel(excels_dir, safe_code, args.type)
        if cached:
            cached_path, cached_mtime = cached
            age_days = (time.time() - cached_mtime) / 86400.0
            result = {
                "status": "ok",
                "stock_code": normalized_code,
                "statement_type": args.type,
                "statement_name": statement_name,
                "excel_path": os.path.abspath(cached_path),
                "cache_hit": True,
                "cache_age_days": round(age_days, 2),
                "message": f"Reused Excel generated {age_days:.1f} days ago; pass --force to regenerate",
            }
            log_stdout(f">>> Cache hit: reusing {cached_path} (age={age_days:.1f}d)")
            log_stdout("=" * 60)
            log_stdout(json.dumps(result, ensure_ascii=False))
            log_stdout("=" * 60)
            return 0

    log_file = os.path.join(logs_dir, f"financial_{normalized_code}_{args.type}_{timestamp}.log")
    logger = setup_logging(log_file)

    # ── Start processing ──
    try:
        if original_code != normalized_code:
            log_stdout(f">>> Stock code normalized: {original_code} → {normalized_code}")
        log_stdout(f">>> Processing {normalized_code} ({statement_name})")

        # 0. 前置校验：股票代码格式
        ok, err = validate_stock_code(normalized_code)
        if not ok:
            log_stdout(f"ERROR: {err}")
            return 1

        # 检查 openpyxl 依赖（提前报错，不要等数据拉完才说）
        if not HAS_OPENPYXL:
            log_stdout(
                "ERROR: 缺少 openpyxl 依赖，无法生成 Excel。\n"
                "提示: 请运行 `pip install openpyxl>=3.1.0` 安装，"
                "完整依赖: `pip install -r workspace/skills/futu-financial-report/scripts/requirements.txt`"
            )
            return 1

        # 1. 确定市场和字段配置
        market_type = get_market_type(normalized_code)
        logger.info(f"市场类型: {market_type}, 报表类型: {statement_name}")

        if market_type == 'hk_us':
            log_stdout(
                f"ERROR: 无法识别股票代码 '{normalized_code}' 所属市场。\n"
                f"提示: 代码必须以 US./HK./SH./SZ. 开头，例如 US.BABA、HK.00700、SH.600519。\n"
                f"可用 stock-ticker 查代码: `python workspace/skills/stock-ticker/scripts/search_ticker.py --company <公司名>`"
            )
            return 1

        # 根据市场类型选择字段配置
        if market_type == 'a_share':
            base_fields = A_SHARE_FIELDS.get(args.type, [])
            expense_items = EXPENSE_ITEMS_MAP['a_share'] if args.type == 'income' else None
            ratio_config = A_SHARE_RATIO_CONFIG
        elif market_type == 'hk':
            base_fields = HK_FIELDS.get(args.type, [])
            expense_items = EXPENSE_ITEMS_MAP['hk'] if args.type == 'income' else None
            ratio_config = HK_RATIO_CONFIG
        else:  # US
            base_fields = US_FIELDS.get(args.type, [])
            expense_items = EXPENSE_ITEMS_MAP['us'] if args.type == 'income' else None
            ratio_config = US_RATIO_CONFIG

        # 2. 获取数据（带字段结构）
        try:
            reports, name_map = get_financial_data_with_structure(normalized_code, args.type, args.num, logger)
        except FetchError as fe:
            logger.error(f"FetchError: {fe.message}\nhint={fe.hint}\ndetail={fe.log_detail}")
            log_stdout(fe.user_message())
            log_stdout(f"详细日志: {os.path.abspath(log_file)}")
            return 1

        # 从财报数据中获取真实的币种单位配置
        unit_name, divisor, currency_code, currency_info = get_unit_config(reports)
        logger.info(f"币种信息: code={currency_code}, info={currency_info}, unit={unit_name}")

        log_stdout(f"OK: Got {len(reports)} quarters data, {len(name_map)} fields")
        log_stdout(f"OK: Currency: {currency_code or currency_info or 'N/A'}, Unit: {unit_name}")

        # 3. 若为利润表：额外拉取现金流量表以计算 FCFF (公司自由现金流) 与 资本支出 CapEx
        fcff_values = None
        capex_values = None
        if args.type == 'income':
            try:
                # 利润表实际使用的 financial_type（复用 get_financial_data_with_structure 的默认逻辑）
                income_ftype = 11 if market_type in ('hk', 'a_share') else 10
                cf_reports, _cf_name_map = get_financial_data_with_structure(
                    normalized_code, 'cashflow', args.num, logger,
                    financial_type_override=income_ftype,
                )
                # 按 period_text 建索引: {period: {field_id: data}}
                cf_by_period = {}
                for r in cf_reports:
                    pt = r.get('period_text', '')
                    cf_by_period[pt] = {i['field_id']: i.get('data') for i in r.get('item_list', [])}
                logger.info(f"配套现金流数据 {len(cf_by_period)} 期已就绪，开始计算 FCFF / CapEx")
                fcff_values, capex_values = compute_fcff_values(reports, cf_by_period, market_type, logger)
                valid_cnt = sum(1 for v in fcff_values if v is not None)
                valid_capex = sum(1 for v in capex_values if v is not None)
                log_stdout(f"OK: FCFF computed for {valid_cnt}/{len(fcff_values)} periods, "
                           f"CapEx for {valid_capex}/{len(capex_values)} periods")
            except FetchError as fe:
                # 现金流拉取失败不影响利润表本体，仅跳过 FCFF/CapEx 加工指标
                logger.warning(f"跳过 FCFF/CapEx: 现金流数据获取失败: {fe.message}")
                log_stdout(f"WARN: 无法拉取现金流数据，跳过 FCFF/CapEx 加工指标（{fe.message}）")
                fcff_values = None
                capex_values = None
            except Exception as e:
                logger.warning(f"跳过 FCFF/CapEx: 未预期异常: {e}", exc_info=True)
                log_stdout(f"WARN: FCFF/CapEx 计算失败，跳过该加工指标: {e}")
                fcff_values = None
                capex_values = None

        # 3b. 若为利润表：额外拉取资产负债表以计算 ROA / ROE
        # 使用与利润表相同的 financial_type，让每一期利润表精确匹配到同期末资产负债表快照
        # （如 2026Q1 利润表 ↔ 2026-03-31 资产负债表），周期一致
        # 同时，若 workspace/excels/ 下无近期资产负债表 Excel，则同步生成一份（便于用户查看/复用）
        roa_values = None
        roe_values = None
        if args.type == 'income':
            try:
                # 与利润表同频（HK/A股=11 累计季报；美股=10 单季+年报），保证期次可精准对齐
                income_ftype_for_bs = 11 if market_type in ('hk', 'a_share') else 10
                bs_reports, bs_name_map = get_financial_data_with_structure(
                    normalized_code, 'balance', args.num, logger,
                    financial_type_override=income_ftype_for_bs,
                )
                logger.info(
                    f"配套资产负债表数据 {len(bs_reports)} 期已就绪（financial_type={income_ftype_for_bs}"
                    f"，与利润表同频），开始计算 ROA/ROE"
                )
                roa_values, roe_values = compute_roa_roe_values(reports, bs_reports, market_type, logger)
                valid_roa = sum(1 for v in roa_values if v is not None)
                valid_roe = sum(1 for v in roe_values if v is not None)
                log_stdout(f"OK: ROA computed for {valid_roa}/{len(roa_values)} periods, "
                           f"ROE computed for {valid_roe}/{len(roe_values)} periods")

                # 若无近期资产负债表 Excel（7天内），自动生成一份
                excels_dir = os.path.join(project_root, 'workspace', 'excels')
                safe_code = normalized_code.replace('.', '_')
                bs_cached = find_recent_excel(excels_dir, safe_code, 'balance')
                if bs_cached is None:
                    try:
                        # 复用刚才拉到的 bs_reports 直接生成 Excel，避免二次调用 API
                        bs_base_fields = (
                            A_SHARE_FIELDS['balance'] if market_type == 'a_share'
                            else HK_FIELDS['balance'] if market_type == 'hk'
                            else US_FIELDS['balance']
                        )
                        bs_unit_name, bs_divisor, _cc, _ci = get_unit_config(bs_reports)
                        bs_rows, bs_quarters, bs_yoy = build_rows_with_ratios(
                            bs_reports, bs_name_map, bs_base_fields, None,
                            bs_divisor, bs_unit_name, 'balance', market_type,
                            A_SHARE_RATIO_CONFIG if market_type == 'a_share'
                            else HK_RATIO_CONFIG if market_type == 'hk' else US_RATIO_CONFIG,
                            logger,
                        )
                        if not os.path.exists(excels_dir):
                            os.makedirs(excels_dir)
                        bs_output = os.path.join(
                            excels_dir, f"{safe_code}_balance_{timestamp}.xlsx"
                        )
                        generate_excel_with_styling(
                            bs_rows, bs_quarters, bs_yoy, bs_output,
                            STATEMENT_TYPES['balance'][1], logger,
                        )
                        log_stdout(f"OK: Balance sheet Excel generated (side-product): {bs_output}")
                    except Exception as e:
                        logger.warning(f"侧生成资产负债表 Excel 失败: {e}", exc_info=True)
                        log_stdout(f"WARN: 未能生成资产负债表 Excel（不影响利润表主输出）: {e}")
                else:
                    log_stdout(f"OK: Reusing existing balance sheet Excel: {bs_cached[0]}")
            except FetchError as fe:
                logger.warning(f"跳过 ROA/ROE: 资产负债表数据获取失败: {fe.message}")
                log_stdout(f"WARN: 无法拉取资产负债表数据，跳过 ROA/ROE 加工指标（{fe.message}）")
                roa_values = None
                roe_values = None
            except Exception as e:
                logger.warning(f"跳过 ROA/ROE: 未预期异常: {e}", exc_info=True)
                log_stdout(f"WARN: ROA/ROE 计算失败，跳过该加工指标: {e}")
                roa_values = None
                roe_values = None

        # 4. 构建数据行（含财务比率计算和YoY跟踪）
        rows, quarters, yoy_tracking = build_rows_with_ratios(
            reports, name_map, base_fields, expense_items,
            divisor, unit_name, args.type, market_type, ratio_config, logger,
            fcff_values=fcff_values,
            capex_values=capex_values,
            roa_values=roa_values,
            roe_values=roe_values,
        )

        log_stdout(f"OK: Built {len(rows)} data rows with {len(yoy_tracking)} YoY comparison rows")

        if len(rows) <= 2:  # 只有表头 + 日期行，无数据行
            log_stdout(
                f"ERROR: 构建后无有效数据行（可能是字段映射不匹配）\n"
                f"提示: 该公司的报表字段可能与默认映射不一致，检查日志查看 name_map。\n"
                f"详细日志: {os.path.abspath(log_file)}"
            )
            return 1

        # 4. 确定输出路径（相对于项目根目录）
        if args.output:
            output_file = args.output
        else:
            excels_dir = os.path.join(project_root, 'workspace', 'excels')
            if not os.path.exists(excels_dir):
                os.makedirs(excels_dir)
            safe_code = normalized_code.replace('.', '_')
            # 文件名格式：{代码}_{报表类型}_{时间}.xlsx
            output_file = os.path.join(excels_dir, f"{safe_code}_{args.type}_{timestamp}.xlsx")

        output_file = os.path.abspath(output_file)

        # 5. 生成 Excel（含YoY下降红色高亮）
        try:
            generate_excel_with_styling(rows, quarters, yoy_tracking, output_file, statement_name, logger)
        except PermissionError:
            log_stdout(
                f"ERROR: 无法写入 Excel 文件（文件可能被占用）: {output_file}\n"
                f"提示: 请关闭已打开的同名 Excel 文件后重试"
            )
            return 1
        except Exception as e:
            logger.exception("Excel 生成失败")
            log_stdout(f"ERROR: Excel 生成失败: {e}\n详细日志: {os.path.abspath(log_file)}")
            return 1

        # ── 最终输出 ──
        result = {
            "status": "ok",
            "stock_code": normalized_code,
            "statement_type": args.type,
            "statement_name": statement_name,
            "quarters": len(reports),
            "rows": len(rows),
            "yoy_highlight_threshold": f"decline {YOY_DECLINE_THRESHOLD}% (red), growth {YOY_GROWTH_THRESHOLD}% (green)",
            "excel_path": output_file,
            "log_path": os.path.abspath(log_file),
            "period_range": f"{quarters[0]} - {quarters[-1]}" if quarters else "",
        }

        log_stdout("=" * 60)
        log_stdout(json.dumps(result, ensure_ascii=False))
        log_stdout("=" * 60)

        return 0

    except KeyboardInterrupt:
        log_stdout("\nInterrupted by user")
        return 130
    except Exception as e:
        logger.exception("Unexpected error")
        log_stdout(f"ERROR: 未预期的错误: {e}")
        log_stdout(f"详细日志: {os.path.abspath(log_file)}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
