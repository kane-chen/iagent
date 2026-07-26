#!/usr/bin/env python3
"""
3-Statement Model - 三表联动财务模型 Excel 生成脚本

对齐 references/schema.md 结构:
  Tab 1  Assumptions          分区式假设 (HEADER / MARKET DATA / REVENUE / COST / BS / DEBT / DIVIDEND)
  Tab 2  Income Statement     5 期历史 + 5 期预测, 含 Margin% 展示行
  Tab 3  Balance Sheet        Days-driven, 含 Balance Check + Cash Tie-Out
  Tab 4  Cash Flow            OCF/CFI/CFF 三段, dNWC 严格符号规则
  Tab 5  D&A Schedule         PPE Beg -> CapEx -> Dep -> End
  Tab 6  Debt Schedule        Beg -> Issue -> Repay -> Sweep -> End, Interest = Beg x Rate
  Tab 7  Working Capital      AR Days / Inv Days / AP Days 驱动

关键设计:
  - 币种一致性: Reporting Currency (财报) vs Trading Currency (股价), FX 换算 (与 dcf-model 一致)
  - CapEx 严格口径: 从 cashflow Excel 的 "资本开支(CapEx明细)" 行读取
  - Interest = Beginning Debt x Rate (断开循环引用)
  - 历史期 Cash 从 BS 直读, 预测期 Cash = CF Ending Cash
  - Balance Check / Cash Tie-Out 红色条件格式
"""
import argparse
import re
import zipfile
import tempfile
import os
import logging
import sys as _sys
import os as _os
from datetime import date
from pathlib import Path
from typing import Optional, Tuple, List

# 确保能导入 futuapi 的 common 模块 (股价 / FX 获取)
_script_dir = _os.path.dirname(_os.path.abspath(__file__))
_potential_paths = [
    _os.path.join(_script_dir, "..", "futuapi", "scripts"),
    _os.path.join(_script_dir, "..", "..", "futuapi", "scripts"),
    _os.path.join(_script_dir, "futuapi", "scripts"),
]
for p in _potential_paths:
    if _os.path.isdir(p):
        _sys.path.insert(0, _os.path.normpath(p))
        break

try:
    from common import create_quote_context, check_ret, safe_close, is_empty
except ImportError:
    def create_quote_context(*args, **kwargs): raise ImportError("common module not found")
    def check_ret(ret, data, ctx, msg): pass
    def safe_close(ctx): pass
    def is_empty(data): return data is None or len(data) == 0

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%H:%M:%S')
logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
except ImportError as exc:
    raise ImportError("openpyxl required: pip install openpyxl") from exc

# ==================== Styles ====================
FONT_BLUE          = Font(color="0000FF")
FONT_BLACK         = Font(color="000000")
FONT_GREEN         = Font(color="008000")
FONT_PURPLE        = Font(color="800080")
FONT_WHITE_BOLD    = Font(color="FFFFFF", bold=True)
FONT_BOLD          = Font(bold=True)
FONT_BLACK_BOLD    = Font(color="000000", bold=True)
FONT_ITALIC_GREY   = Font(italic=True, color="595959", size=10)

FILL_DARK_BLUE     = PatternFill("solid", fgColor="1F4E79")
FILL_LIGHT_BLUE    = PatternFill("solid", fgColor="D9E1F2")
FILL_MEDIUM_BLUE   = PatternFill("solid", fgColor="BDD7EE")
FILL_INPUT_GREY    = PatternFill("solid", fgColor="F2F2F2")
FILL_FORECAST_GREEN = PatternFill("solid", fgColor="E2F0D9")
FILL_VALUATION_ORANGE      = PatternFill("solid", fgColor="FCE4D6")
FILL_VALUATION_ORANGE_DARK = PatternFill("solid", fgColor="F4B183")

BORDER_THIN_BOTTOM   = Border(bottom=Side(style="thin", color="000000"))
BORDER_MEDIUM_BOTTOM = Border(bottom=Side(style="medium", color="000000"))
BORDER_HAIR_BOTTOM   = Border(bottom=Side(style="hair", color="595959"))
BORDER_HIST_RIGHT    = Border(right=Side(style="thin", color="595959"))
BOTTOM_BORDER        = BORDER_THIN_BOTTOM  # DCF sheet alias

ALIGN_INDENT = Alignment(indent=2)

# ==================== 数字格式 (schema.md 标准) ====================
FMT_CURRENCY_M = '#,##0;(#,##0);"-"'
FMT_PRICE      = '#,##0.00'
FMT_PERCENT    = '0.0%'
FMT_MULTIPLE   = '0.0"x"'
FMT_SHARES     = '#,##0.00'
FMT_DECIMAL4   = '0.0000'
FMT_DAYS       = '0" days"'
FMT_CHECK      = '[Red][<>0]#,##0.00;[Red][<>0](#,##0.00);0'

# ==================== Comment helper ====================
_COMMENT_AUTHOR = "3-Statement Builder"

def add_comment(cell, text: str, width: int = 260, height: int = 100):
    c = Comment(text, _COMMENT_AUTHOR)
    c.width = width; c.height = height
    cell.comment = c

def add_source_comment(cell, system: str, ref: str = "", extra: str = ""):
    today = date.today().isoformat()
    parts = [f"Source: {system}", today]
    if ref: parts.append(ref)
    text = ", ".join(parts)
    if extra: text += "\n" + extra
    add_comment(cell, text)

def safe_divide(n, d):
    return n / d if d != 0 else 0.0

# ==================== 币种识别常量 ====================
_UNIT_NAME_TO_CURRENCY = {
    "百万人民币": "CNY", "百万港元": "HKD", "百万美元": "USD",
    "百万欧元": "EUR", "百万英镑": "GBP", "百万日元": "JPY",
    "百万新加坡元": "SGD", "百万澳元": "AUD", "百万加元": "CAD",
}
_MARKET_PREFIX_TO_CURRENCY = {"US": "USD", "HK": "HKD", "SH": "CNY", "SZ": "CNY"}
_FX_FUTU_CODES = {
    ("USD", "CNY"): ["HK.USDCNH", "HK.USDCNY"],
    ("USD", "HKD"): ["HK.USDHKD"],
    ("HKD", "CNY"): ["HK.HKDCNH", "HK.HKDCNY"],
}
_FX_FALLBACKS = {
    ("USD", "CNY"): 7.20, ("USD", "HKD"): 7.80, ("HKD", "CNY"): 0.92,
    ("EUR", "USD"): 1.08, ("GBP", "USD"): 1.27, ("USD", "JPY"): 155.0,
}

# ==================== Beta 基准指数 & Rf/ERP 常量 (用于 DCF/WACC sheet) ====================
_BENCHMARK_INDEX = {
    "US": "US.SPY",       # S&P 500 ETF (美国本土公司默认基准)
    "HK": "HK.800000",    # 恒生指数
    "SH": "SH.000300",    # 沪深 300
    "SZ": "SH.000300",    # A 股统一用沪深 300
}
# 中概股 ADR 专用基准 (美股上市但报表币种为 CNY → 业务在中国, 股价主要受中国宏观/监管驱动).
# KWEB (KraneShares CSI China Internet ETF) 覆盖 BABA/PDD/JD/TCOM/BIDU/NTES/美团/网易 等
# 主要中概互联网/科技股, 与 ADR 中概股相关性远高于 SPY (r 从 ~0.3 提升到 ~0.85), Beta 更贴近真实风险.
# 若 KWEB 数据不足 (< _BETA_MIN_SAMPLES 月样本, 例如 2013 年以前), 自动回退到 SPY 并记 warning.
_ADR_CN_BENCHMARK_PRIMARY  = "US.KWEB"    # 首选: 中概互联网 ETF (自 2013-08 上市, 覆盖 60+ 个月)
_ADR_CN_BENCHMARK_FALLBACK = "US.PGJ"     # 备选: Golden Dragon 广义中概 ETF
# 按报表币种查 10Y 主权债券收益率 与 Damodaran country ERP
_RF_ERP_BY_CURRENCY = {
    "USD": {"rf": 0.043, "erp": 0.055, "rf_src": "10Y US Treasury",  "erp_src": "Damodaran US ERP"},
    "HKD": {"rf": 0.040, "erp": 0.060, "rf_src": "HKGB 10Y",         "erp_src": "Damodaran HK country ERP"},
    "CNY": {"rf": 0.025, "erp": 0.065, "rf_src": "中国 10Y 国债",     "erp_src": "Damodaran CN country ERP"},
    "EUR": {"rf": 0.027, "erp": 0.050, "rf_src": "German Bund 10Y",  "erp_src": "Damodaran EU ERP"},
    "GBP": {"rf": 0.040, "erp": 0.055, "rf_src": "UK Gilt 10Y",      "erp_src": "Damodaran UK ERP"},
    "JPY": {"rf": 0.015, "erp": 0.055, "rf_src": "JGB 10Y",          "erp_src": "Damodaran JP ERP"},
}
_RF_ERP_DEFAULT = {"rf": 0.043, "erp": 0.055, "rf_src": "Fallback (USD)", "erp_src": "Fallback (USD)"}
_BETA_FALLBACK = 1.20
_BETA_MIN_SAMPLES = 24   # 最少 24 个月度收益率样本才计算 beta

# ==================== Utility - Fetch Financial Data from Futu API ====================
# 直接调用 ctx.get_financials_statements() 抽取三表数据 (替代读取本地 excel 文件)
# 参考: workspace/skills/futu-financial-report/scripts/get_financials_statements.py
_STATEMENT_TYPE_MAP = {
    "income":   1,  # 利润表
    "balance":  2,  # 资产负债表
    "cashflow": 3,  # 现金流量表
}
# financial_type=7 表示年报, 3-Statement 模型只需要 FY 数据
_FINANCIAL_TYPE_ANNUAL = 7


def _api_currency_to_reporting(currency_code: Optional[str]) -> Optional[str]:
    """API 返回的 currency_code (ISO 4217) 直接作为报告币种。空则返回 None。"""
    if not currency_code:
        return None
    c = str(currency_code).strip().upper()
    return c if c else None


def _fetch_statement_from_api(ctx, stock_code: str, statement_type: str,
                              max_pages: int = 10) -> Tuple[dict, Optional[str], dict]:
    """
    从 Futu API 拉取指定报表 (年报口径), 分页聚合后返回:
      data:      {display_name: {FY_key: value_in_millions}}  # 与原 _read_excel_map 兼容
      currency:  str (如 "CNY" / "USD" / "HKD" / None)
      fid_data:  {field_id: {FY_key: value_in_millions}}      # 用于按 field_id 精确访问 (跨市场消歧)

    - statement_type: "income" / "balance" / "cashflow"
    - 数值单位: API 返回的是原始货币单位 (元), 转换为百万单位 (÷ 1e6) 与 Excel 保持一致
    - FY_key 格式: "{fiscal_year}FY" (如 "2025FY"); 仅保留 fiscal_type=7 年报
    - 分页拉取: 每页 50 条, 直到 next_key == "-1" 或达到 max_pages
    - 跨市场消歧: 同一 display_name (如 "预付款项") 在美股 (fid 8016 流动) 与港股 (fid 5044 非流动)
      语义不同, 用 field_id 访问可避免误合并
    """
    if statement_type not in _STATEMENT_TYPE_MAP:
        raise ValueError(f"未知 statement_type={statement_type}, 可选: income/balance/cashflow")
    st_code = _STATEMENT_TYPE_MAP[statement_type]

    all_structure = {}          # {field_id: display_name}
    all_reports = []            # 累积所有页的 report 条目
    currency_code = None
    next_key = None
    for page in range(max_pages):
        ret, resp = ctx.get_financials_statements(
            stock_code,
            statement_type=st_code,
            financial_type=_FINANCIAL_TYPE_ANNUAL,
            currency_code=None,     # 不做币种转换, 保留原始货币
            next_key=next_key,
            num=50,
        )
        if ret != 0 or not isinstance(resp, dict):
            if page == 0:
                logger.warning(f"{stock_code} {statement_type} API 拉取失败: ret={ret}, data={resp}")
            break

        for entry in resp.get("structure_list", []) or []:
            fid = entry.get("field_id")
            name = entry.get("display_name") or f"字段{fid}"
            if fid is not None:
                all_structure[fid] = str(name).strip()

        page_reports = resp.get("report_list", []) or []
        all_reports.extend(page_reports)

        if currency_code is None:
            for r in page_reports:
                cc = _api_currency_to_reporting(r.get("currency_code"))
                if cc:
                    currency_code = cc
                    break

        nk = resp.get("next_key", "-1")
        if nk == "-1" or nk is None:
            break
        next_key = nk

    if not all_reports:
        return {}, currency_code, {}

    data: dict = {}       # display_name → {fy: value_m}
    fid_data: dict = {}   # field_id → {fy: value_m}
    for rpt in all_reports:
        ft = rpt.get("financial_type")
        if ft is not None and ft != _FINANCIAL_TYPE_ANNUAL:
            continue
        fy = rpt.get("fiscal_year")
        if fy is None:
            continue
        fy_key = f"{fy}FY"
        items = rpt.get("item_list", []) or []
        for item in items:
            fid = item.get("field_id")
            if fid is None:
                continue
            name = all_structure.get(fid)
            if not name:
                continue
            raw = item.get("data")
            if raw is None:
                continue
            try:
                val = float(raw)
            except (TypeError, ValueError):
                continue
            val_m = val / 1_000_000.0
            data.setdefault(name, {})[fy_key] = val_m
            fid_data.setdefault(fid, {})[fy_key] = val_m

    return data, currency_code, fid_data


# ==================== Utility - Excel Reading (legacy, kept for tools) ====================
def _ensure_shared_strings(file_path: Path) -> Tuple[Path, bool]:
    file_path = Path(file_path)
    try:
        with zipfile.ZipFile(str(file_path), 'r') as zf:
            if 'xl/sharedStrings.xml' in zf.namelist(): return file_path, False
            ct = zf.read('[Content_Types].xml').decode('utf-8', errors='ignore')
            if 'sharedStrings' not in ct: return file_path, False
    except:
        return file_path, False
    fd, tmp = tempfile.mkstemp(suffix='.xlsx', dir=str(file_path.parent))
    os.close(fd)
    try:
        with zipfile.ZipFile(str(file_path), 'r') as src, zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as dst:
            for item in src.namelist(): dst.writestr(item, src.read(item))
            dst.writestr('xl/sharedStrings.xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"/>')
        return Path(tmp), True
    except:
        return file_path, False


def _read_excel_map(file_path: Path) -> Tuple[dict, Optional[str]]:
    """读取财报 Excel, 返回 (data, currency_code)。"""
    fp, is_tmp = _ensure_shared_strings(file_path)
    try:
        wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
        sheet = wb.active
        data = {}
        currency_code = None
        header_row = 1
        for r in range(1, 5):
            vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1) if sheet.cell(r, c).value]
            if "FY" in " ".join(str(v) for v in vals):
                header_row = r
                break
        periods = [str(sheet.cell(header_row, c).value) for c in range(3, sheet.max_column + 1)]
        for r in range(header_row + 1, sheet.max_row + 1):
            ind = sheet.cell(r, 1).value
            if not ind: continue
            ind = str(ind).strip()
            if currency_code is None:
                unit_str = sheet.cell(r, 2).value
                if isinstance(unit_str, str):
                    mapped = _UNIT_NAME_TO_CURRENCY.get(unit_str.strip())
                    if mapped: currency_code = mapped
            vals = {}
            for c, p in enumerate(periods):
                v = sheet.cell(r, c + 3).value
                num = 0.0
                if isinstance(v, (int, float)):
                    num = float(v)
                elif isinstance(v, str) and v:
                    cl = v.replace(",", "").replace("-", "").strip()
                    if cl.replace(".", "", 1).isdigit():
                        num = float(cl) * (-1 if v.strip().startswith("-") else 1)
                vals[p] = num
            data[ind] = vals
        wb.close()
        return data, currency_code
    finally:
        if is_tmp and fp.exists(): os.remove(str(fp))


def find_local_file(excels_path: Path, ticker: str, suffix: str) -> Optional[Path]:
    pattern = re.compile(rf'^.*_{re.escape(ticker)}_{re.escape(suffix)}_.*\.(xlsx|xls)$', re.IGNORECASE)
    # 排除 Excel 临时锁文件 ~$xxx.xlsx
    files = [f for f in excels_path.iterdir()
             if f.is_file() and not f.name.startswith('~$') and pattern.match(f.name)]
    return sorted(files)[-1] if files else None


def normalize_stock_code(ticker: str) -> str:
    ticker = ticker.strip().upper()
    if '.' in ticker and ticker.split('.')[0] in ('US', 'HK', 'SH', 'SZ'): return ticker
    if ticker.isdigit():
        n = int(ticker)
        if n >= 600000: return f"SH.{ticker}"
        elif n >= 300000: return f"SZ.{ticker}"
        else: return f"HK.{ticker.zfill(5)}"
    return f"US.{ticker}"


def infer_trading_currency(stock_code: str) -> str:
    prefix = stock_code.split('.')[0] if '.' in stock_code else ""
    return _MARKET_PREFIX_TO_CURRENCY.get(prefix, "USD")


# ==================== Futu Market Data & FX ====================
def fetch_market_data_from_futu(stock_code: str) -> dict:
    result = {"stock_price": None, "shares_outstanding": None, "shares_source": "Unknown"}
    ctx = None
    logger.info(f"正在通过 FutuOpenD 获取 {stock_code} 的市场数据...")
    try:
        ctx = create_quote_context()
        ret, data = ctx.get_market_snapshot([stock_code])
        check_ret(ret, data, ctx, "获取市场快照")
        if not is_empty(data):
            row = data.iloc[0]
            if 'last_price' in data.columns and row['last_price'] and float(row['last_price']) > 0:
                result["stock_price"] = float(row['last_price'])
            if 'issued_shares' in data.columns and row['issued_shares'] and float(row['issued_shares']) > 0:
                raw = float(row['issued_shares'])
                result["shares_outstanding"] = raw / 1_000_000
                result["shares_source"] = "Futu get_market_snapshot (issued_shares)"
    except SystemExit:
        logger.error("Futu API 未连接")
    except Exception as e:
        logger.exception(f"Futu API 异常: {e}")
    finally:
        if ctx: safe_close(ctx)
    return result


def _sina_fx_quote(from_ccy: str, to_ccy: str) -> Tuple[Optional[float], str]:
    """通过新浪财经外汇快照获取汇率.

    新浪 API: https://hq.sinajs.cn/list=fx_s{from_lower}{to_lower}
    返回格式: var hq_str_fx_susdcny="YYYY-mm-dd HH:MM:SS,<last>,<bid>,<ask>,...";
    需要 Referer 头指向 sina.com.cn 白名单.
    """
    import urllib.request as _urlreq
    import re as _re
    code = f"fx_s{from_ccy.lower()}{to_ccy.lower()}"
    url = f"https://hq.sinajs.cn/list={code}"
    try:
        req = _urlreq.Request(url, headers={
            "Referer": "https://finance.sina.com.cn/",
            "User-Agent": "Mozilla/5.0",
        })
        with _urlreq.urlopen(req, timeout=5) as resp:
            raw = resp.read().decode("gbk", errors="ignore")
    except Exception as e:
        return None, f"Sina FX 请求失败 ({code}): {e}"
    m = _re.search(r'"([^"]*)"', raw)
    if not m or not m.group(1):
        return None, f"Sina FX 无数据 ({code})"
    parts = m.group(1).split(",")
    if len(parts) < 2:
        return None, f"Sina FX 响应格式异常 ({code})"
    # 优先取 index 1 (last), 兜底 bid/ask
    for idx in (1, 2, 3):
        if idx < len(parts):
            try:
                v = float(parts[idx])
            except (TypeError, ValueError):
                continue
            if 0.0001 < v < 100000:
                ts = parts[0] if parts else ""
                return v, f"Sina 新浪财经 ({code}, {ts})"
    return None, f"Sina FX 数值异常 ({code}): {parts[:4]}"


def fetch_fx_rate_from_sina(from_ccy: str, to_ccy: str) -> Tuple[Optional[float], str]:
    """新浪财经汇率获取, 支持正向 + 反向候选."""
    from_ccy, to_ccy = from_ccy.upper(), to_ccy.upper()
    if from_ccy == to_ccy: return 1.0, "Same currency (no conversion)"
    rate, msg = _sina_fx_quote(from_ccy, to_ccy)
    if rate is not None:
        return rate, msg
    inv_rate, inv_msg = _sina_fx_quote(to_ccy, from_ccy)
    if inv_rate is not None and inv_rate > 0:
        return 1.0 / inv_rate, f"{inv_msg} (inverse)"
    return None, f"Sina FX 双向均失败 (last: {inv_msg})"


def fetch_fx_rate_from_futu(from_ccy: str, to_ccy: str) -> Tuple[float, str]:
    """获取 1 from_ccy = X to_ccy 汇率。

    优先级:
      1) 新浪财经 hq.sinajs.cn 开放接口 (无需 FutuOpenD FX 权限, 主流货币对全覆盖)
      2) Futu 外汇快照 (`HK.USDCNH` 等, 需 FutuOpenD FX 权限)
      3) _FX_FALLBACKS 常量兜底
    """
    from_ccy, to_ccy = from_ccy.upper(), to_ccy.upper()
    if from_ccy == to_ccy: return 1.0, "Same currency (no conversion)"

    # ---- 优先: 新浪财经开放接口 ----
    sina_rate, sina_msg = fetch_fx_rate_from_sina(from_ccy, to_ccy)
    if sina_rate is not None:
        logger.info(f"Sina FX 成功: 1 {from_ccy} = {sina_rate:.4f} {to_ccy}")
        return sina_rate, sina_msg
    logger.warning(f"Sina FX 不可用 ({sina_msg}), 回退到 Futu FX")

    codes = _FX_FUTU_CODES.get((from_ccy, to_ccy), [])
    inverse_codes = _FX_FUTU_CODES.get((to_ccy, from_ccy), [])
    ctx = None
    try:
        if codes or inverse_codes:
            ctx = create_quote_context()
            for fx_code in codes:
                try:
                    ret, data = ctx.get_market_snapshot([fx_code])
                    if ret == 0 and not is_empty(data):
                        px = data.iloc[0].get('last_price')
                        if px and float(px) > 0:
                            return float(px), f"Futu get_market_snapshot ({fx_code})"
                except Exception: pass
            for fx_code in inverse_codes:
                try:
                    ret, data = ctx.get_market_snapshot([fx_code])
                    if ret == 0 and not is_empty(data):
                        px = data.iloc[0].get('last_price')
                        if px and float(px) > 0:
                            return 1.0 / float(px), f"Futu get_market_snapshot ({fx_code}, inverse)"
                except Exception: pass
    except Exception as e:
        logger.warning(f"FX 获取异常: {e}")
    finally:
        if ctx: safe_close(ctx)
    if (from_ccy, to_ccy) in _FX_FALLBACKS:
        return _FX_FALLBACKS[(from_ccy, to_ccy)], "Fallback constant"
    if (to_ccy, from_ccy) in _FX_FALLBACKS:
        return 1.0 / _FX_FALLBACKS[(to_ccy, from_ccy)], "Fallback constant (inverse)"
    return 1.0, "NO MAPPING"


# ==================== Beta 计算 (5Y monthly regression) ====================
def get_benchmark_for(stock_code: str, reporting_currency: Optional[str] = None) -> str:
    """根据股票所属交易场所 + 报表币种选择大盘基准指数, 用于计算个股 Beta.

    关键: **美股上市的中概股 ADR** (`US.*` 且报表币种 CNY, 如 BABA/PDD/TCOM/JD) 与 S&P 500 相关性
    极低 (r~0.3), 用 SPY 算 β 会严重低估 (~0.4). 业务/股价驱动因素在中国, 应改用中概股基准:
      - 首选 US.KWEB (KraneShares 中概互联网 ETF, 覆盖 BABA/PDD/JD/TCOM/BIDU/NTES 等)
      - 若 KWEB 样本不足 (< 24 月), 上游会自动回退到 SPY 并记 warning

    Args:
        stock_code: 富途格式股票代码 (US.BABA / HK.00700 / SH.600519)
        reporting_currency: 财报币种 (CNY / USD / HKD / ...)
    """
    prefix = stock_code.split('.')[0] if '.' in stock_code else ""
    # 美股上市中概股: 交易所=US 但报表币种=CNY → 用中概股基准
    if prefix == "US" and reporting_currency == "CNY":
        return _ADR_CN_BENCHMARK_PRIMARY
    return _BENCHMARK_INDEX.get(prefix, "US.SPY")


def _fetch_monthly_returns(ctx, code: str, months: int = 60) -> list:
    try:
        from futu.common.constant import KLType, AuType
    except ImportError:
        logger.warning("futu-api 未安装, 无法计算 Beta")
        return []
    from datetime import timedelta
    end = date.today().isoformat()
    start = (date.today() - timedelta(days=int(365 * (months / 12 + 1)))).isoformat()
    try:
        ret, df, _ = ctx.request_history_kline(
            code, start=start, end=end, ktype=KLType.K_MON, autype=AuType.QFQ, max_count=1000
        )
    except Exception as e:
        logger.warning(f"  -> {code} request_history_kline 异常: {e}")
        return []
    if ret != 0 or df is None or df.empty:
        return []
    closes = [float(v) for v in df['close'].tolist() if v is not None]
    if len(closes) < 2: return []
    return [(closes[i] - closes[i-1]) / closes[i-1] for i in range(1, len(closes)) if closes[i-1]]


def compute_beta_from_futu(stock_code: str, benchmark_code: Optional[str] = None,
                            months: int = 60,
                            reporting_currency: Optional[str] = None) -> Tuple[Optional[float], str]:
    """通过 Futu 拉取 60 个月月线, 计算 Beta = cov(stock, mkt) / var(mkt).

    自动 fallback: 若指定的基准 (如 US.KWEB 用于中概股) 月线样本不足 24 个月, 会:
      - ADR 中概股 (US.* + CNY): 尝试 KWEB → PGJ → SPY 递降
      - 其他情况: 保留原基准 (样本不足时返回 None 走兜底常量)
    """
    benchmark = benchmark_code or get_benchmark_for(stock_code, reporting_currency)
    logger.info(f"计算 Beta: {stock_code} vs {benchmark} (5Y monthly)...")

    # 构建基准候选链: 中概股 ADR 场景下 KWEB 若样本不足自动降级
    prefix = stock_code.split('.')[0] if '.' in stock_code else ""
    is_adr_cn = (prefix == "US" and reporting_currency == "CNY")
    if is_adr_cn:
        # 候选顺序: KWEB (首选) → PGJ (广义中概) → SPY (兜底大盘)
        candidates = [benchmark, _ADR_CN_BENCHMARK_FALLBACK, "US.SPY"]
        # 去重, 保持顺序
        seen = set(); benchmarks_to_try = []
        for b in candidates:
            if b and b not in seen:
                seen.add(b); benchmarks_to_try.append(b)
    else:
        benchmarks_to_try = [benchmark]

    ctx = None
    try:
        ctx = create_quote_context()
        stock_r = _fetch_monthly_returns(ctx, stock_code, months)
        if len(stock_r) < _BETA_MIN_SAMPLES:
            return None, f"Insufficient stock history ({len(stock_r)} months)"

        last_err = None
        for bm in benchmarks_to_try:
            mkt_r = _fetch_monthly_returns(ctx, bm, months)
            n = min(len(stock_r), len(mkt_r))
            if n < _BETA_MIN_SAMPLES:
                last_err = f"{bm} history insufficient ({n} months)"
                logger.warning(f"  -> {last_err}, 尝试下一基准")
                continue
            sr = stock_r[-n:]; mr = mkt_r[-n:]
            mean_s = sum(sr) / n; mean_m = sum(mr) / n
            cov = sum((s - mean_s) * (m - mean_m) for s, m in zip(sr, mr)) / n
            var_m = sum((m - mean_m) ** 2 for m in mr) / n
            if var_m <= 0:
                last_err = f"{bm} zero market variance"
                continue
            beta = cov / var_m
            # 相关系数用于诊断 (r 过低说明基准选错)
            var_s = sum((s - mean_s) ** 2 for s in sr) / n
            r_corr = cov / (var_s ** 0.5 * var_m ** 0.5) if var_s > 0 else 0.0
            note = "" if bm == benchmarks_to_try[0] else f" (fallback from {benchmarks_to_try[0]})"
            logger.info(f"  -> Beta({stock_code} vs {bm}){note} = {beta:.4f}, "
                        f"r={r_corr:.3f}, n={n} months")
            return beta, (f"Futu {n}M monthly kline (vs {bm}, r={r_corr:.2f}), "
                          f"{date.today().isoformat()}{note}")
        return None, f"All benchmarks failed ({last_err})"
    except SystemExit:
        return None, "Futu API failure"
    except Exception as e:
        logger.warning(f"Beta 计算异常: {e}")
        return None, f"Exception: {e}"
    finally:
        if ctx: safe_close(ctx)


# ==================== Financial Data Extraction ====================
def extract_financial_data(workspace: Path, ticker: str) -> dict:
    """从 Futu API 抽取三表历史财务数据 (替代原来的本地 Excel 读取)。

    workspace 参数保留是为了兼容, 但不再用作 Excel 数据源。
    Futu OpenD 必须已启动并登录, 否则抛错。
    """
    stock_code = normalize_stock_code(ticker)
    logger.info(f"通过 Futu API 拉取 {stock_code} 的三表历史数据 (年报口径)...")
    ctx = None
    try:
        ctx = create_quote_context()
        inc_data, inc_ccy, inc_fid = _fetch_statement_from_api(ctx, stock_code, "income")
        bs_data,  bs_ccy,  bs_fid  = _fetch_statement_from_api(ctx, stock_code, "balance")
        cf_data,  cf_ccy,  cf_fid  = _fetch_statement_from_api(ctx, stock_code, "cashflow")
    finally:
        if ctx: safe_close(ctx)

    if not inc_data and not bs_data and not cf_data:
        raise RuntimeError(
            f"Futu API 未返回 {stock_code} 的任何财务数据, 请检查:\n"
            f"  1) Futu OpenD 是否正常运行\n"
            f"  2) 股票代码是否正确\n"
            f"  3) 账号是否有对应市场行情权限"
        )

    reporting_currency = inc_ccy or bs_ccy or cf_ccy
    detected = [c for c in (inc_ccy, bs_ccy, cf_ccy) if c]
    if len(set(detected)) > 1:
        logger.warning(f"三张表币种不一致: inc={inc_ccy}, bs={bs_ccy}, cf={cf_ccy}")

    # 抽取所有 FY, 按时间倒序; 取最近 5 期正序
    # 关键: 只保留 BS 有数据的 FY 年份, 避免历史列大面积零值
    # (富途 API 对旧年度的 BS 覆盖常不足, 而 income/cashflow 可能覆盖更长历史)
    def _has_bs_coverage(fy: str) -> bool:
        """判定该 FY 是否有 BS 核心字段 (资产合计/负债合计/股东权益/现金 至少 2 项非零)"""
        core_keys = [
            "资产合计", "负债合计", "股东权益合计", "归属于母公司股东权益合计",
            # 现金字段跨市场兼容 (API 原始字段名, 与 Excel 加工字段不同)
            "-现金和现金等价物", "现金及现金等价物", "现金及现金等价物和短期投资",
            "现金及等价物", "货币资金",
        ]
        hits = 0
        for k in core_keys:
            if k in bs_data and fy in bs_data[k]:
                v = bs_data[k][fy]
                if v and abs(v) > 0:
                    hits += 1
        return hits >= 2

    all_fy = set()
    for d in [inc_data, bs_data, cf_data]:
        for k, v in d.items(): all_fy.update([p for p in v.keys() if "FY" in p])
    fy_cols_all = sorted(list(all_fy), reverse=True)
    if not fy_cols_all:
        raise RuntimeError(f"未找到 {ticker} 的历史财报数据, 请先运行 futu-financial-report")
    # 过滤: 只保留 BS 有覆盖的 FY, 避免 BS/CF 表出现大面积 0 值
    fy_cols_bs = [fy for fy in fy_cols_all if _has_bs_coverage(fy)]
    if not fy_cols_bs:
        logger.warning(f"{ticker}: BS 无任何 FY 覆盖, 回退到全部 income FY (BS 列可能大量 0)")
        fy_cols_bs = fy_cols_all
    n_bs = len(fy_cols_bs)
    if n_bs < len(fy_cols_all):
        logger.info(f"{ticker}: 全部 FY {len(fy_cols_all)} 期, BS 覆盖 {n_bs} 期; "
                    f"限制历史列到 {min(n_bs, 5)} 期以避免空值")
    fy_cols = fy_cols_bs
    hist_fys = list(reversed(fy_cols[:5]))
    prior_fy = fy_cols[5] if len(fy_cols) > 5 else None

    def gv(data, keys, col):
        for k in keys:
            if k in data and col in data[k]: return data[k][col]
        return 0.0

    def _series(data, keys):
        return [gv(data, keys, fy) for fy in hist_fys]

    def _sum_all_matches(data, keys, col):
        total = 0.0
        for k in keys:
            if k in data and col in data[k]:
                v = data[k][col]
                if v: total += v
        return total

    def _sum_first_match(data, groups, col):
        total = 0.0
        for keys in groups:
            for k in keys:
                if k in data and col in data[k]:
                    v = data[k][col]
                    if v: total += v; break
        return total

    # ---- 利润表历史序列 ----
    revenue_series = _series(inc_data, ["总收入", "营业总收入"])
    # 成本/费用类字段: 各市场符号约定不一 (港股 5005 可能为负, 美股 8003 为正), 统一取绝对值
    # 避免 WC schedule 的 Days 公式 (val × 365 / cost) 因负号导致 Days 为负
    cogs_series    = [abs(v) for v in _series(inc_data, ["营业总成本", "营业成本"])]
    opex_series    = [abs(v) for v in _series(inc_data, ["营业费用"])]
    ebit_series    = _series(inc_data, ["营业利润"])
    tax_series     = [abs(v) for v in _series(inc_data, ["所得税"])]
    ebt_series     = _series(inc_data, ["税前利润"])
    # NI: 归母口径优先, 精确匹配港股字段"归属母公司净利润"和美股字段"归属于母公司股东净利润"
    ni_series      = _series(inc_data, [
        "归属母公司净利润",           # 港股 5051 / 富途 HK 字段名
        "归属于母公司股东净利润",     # 美股 8043 富途字段名
        "归属于母公司股东的净利润",   # A 股常见字段名
        "净利润",                     # 合并口径 fallback (含少数股东损益)
    ])
    # 港股专属明细行 (5035/5036/5037), 美股/A 股返回 0 序列
    finance_income_series   = _series(inc_data, ["融资收入", "利息收入"])
    finance_cost_series     = _series(inc_data, ["融资成本", "利息费用", "财务费用"])
    equity_affiliate_series = _series(inc_data, ["应占联营公司利润", "应占联营公司盈利", "应占联营及合营公司损益", "投资收益"])

    def _series_first_nonzero(*sources):
        """按候选源顺序, 每年取第一个非零值 (逐年 fallback)。
        每个 source 是 (data_dict, keys_list) 二元组。"""
        out = []
        for fy in hist_fys:
            v = 0.0
            for data, keys in sources:
                candidate = gv(data, keys, fy)
                if candidate and abs(candidate) > 0:
                    v = candidate
                    break
            out.append(v)
        return out

    def _series_first_nonzero_sum(*sources):
        """按候选源顺序, 每年取第一个"任一 key 非零"的源, 并对该源所有 key 求和 (绝对值).
        用于 HK CapEx = |购买固定资产| + |购买无形资产| 等多字段合并。
        每个 source 是 (data_dict, keys_list) 二元组; keys_list 内所有 key 累加取绝对值."""
        out = []
        for fy in hist_fys:
            v_total = 0.0
            for data, keys in sources:
                s = 0.0
                any_hit = False
                for k in keys:
                    if k in data and fy in data[k]:
                        val = data[k][fy]
                        if val:
                            s += abs(val)
                            any_hit = True
                if any_hit and abs(s) > 0:
                    v_total = s
                    break
            out.append(v_total)
        return out

    # ---- CapEx 逐年优先级 fallback ----
    # API 原始字段跨市场差异:
    #   美股 (US): fid 8046 "固定资产交易净额" (净额, 通常为负)
    #   港股 (HK): fid 5071 "购买固定资产" + 5073 "购买无形资产" (两个字段, 都为负, 需累加取绝对值)
    #   A股 (CN): fid 3043 "购建固定资产、无形资产和其他长期资产支付的现金"
    # 同时保留 Excel 加工字段 "资本开支(CapEx)" / "资本开支(CapEx明细)" 兼容 (若数据源为 Excel 时)
    capex_series = _series_first_nonzero_sum(
        # HK: 购买固定资产 + 购买无形资产 (合并累加)
        (cf_data,  ["购买固定资产", "购买无形资产"]),
        # US: 固定资产交易净额 + 无形资产交易净额
        (cf_data,  ["固定资产交易净额", "无形资产交易净额"]),
        # A股: 购建固定资产、无形资产和其他长期资产支付的现金
        (cf_data,  ["购建固定资产、无形资产和其他长期资产支付的现金", "购建固定资产及无形资产净额", "购建固定资产"]),
        # Excel 加工字段兼容 (备用)
        (cf_data,  ["资本开支(CapEx明细)"]),
        (inc_data, ["资本开支(CapEx)", "资本开支"]),
    )
    if all(v == 0.0 for v in capex_series):
        logger.warning(f"{ticker}: 所有源均无 CapEx 数据 (income/cashflow 全空), CapEx 序列为 0")

    # ---- D&A: 两个口径分开抽取 ----
    # CF-side D&A (广口径, 用于 CF OCF 加回 与 PP&E 滚动):
    #   美股 fid 8019 "折旧摊销及损耗"
    #   港股 fid 5009 "折旧及摊销:" (注意末尾冒号!)
    #   A股 fid 3002 "折旧与摊销"
    # IS-side D&A (窄口径, 用于 IS EBITDA 计算, 与 EBIT 口径一致):
    #   美股 fid 8011 "折旧摊销及损耗" (仅固定资产折旧, 通常远小于 CF 广口径)
    #   港股 IS 无独立 D&A 字段 (回退到 CF 广口径)
    #   A股 fid 3020 "折旧与摊销"
    da_series_cf = _series_first_nonzero(
        # 加冒号版本 (港股 fid 5009 display_name)
        (cf_data,  ["折旧摊销及损耗", "折旧及摊销:", "折旧与摊销", "折旧及摊销"]),
        (inc_data, ["折旧摊销及损耗", "折旧及摊销:", "折旧与摊销", "折旧及摊销", "-折旧及摊销"]),
    )
    da_series_is = _series_first_nonzero(
        (inc_data, ["折旧摊销及损耗", "折旧及摊销:", "折旧与摊销", "折旧及摊销", "-折旧及摊销"]),
        (cf_data,  ["折旧摊销及损耗", "折旧及摊销:", "折旧与摊销", "折旧及摊销"]),
    )
    if all(v == 0.0 for v in da_series_cf) and all(v == 0.0 for v in da_series_is):
        logger.warning(f"{ticker}: 所有源均无 D&A 数据, EBITDA 可能明显低估")
    da_series = da_series_cf  # 默认导出 CF-side (保留原字段名, 供 D&A Schedule 使用)


    # ==========================================================================
    # BS 三层结构抽取 (L1 → L2 → L3) - 用 field_id 精确访问, 避免跨市场 display_name 歧义
    # L1: 资产/负债/权益 (3 大类, 已由 total_assets_series / total_liab_series / equity_series 覆盖)
    # L2: 流动资产/非流动资产 / 流动负债/非流动负债 / 归属母公司/少数股东 (6 项, 直接从 API 拿)
    # L3: 3-5 项 aggregated buckets, 按变现难度/偿还优先级排序, 相似指标合并
    # ==========================================================================

    # 市场判别 (需在 BS L3 之前, 用于 field_id 白名单选择)
    _prefix = stock_code.split('.')[0] if '.' in stock_code else ""
    if _prefix == "US":
        market_type = "us"
    elif _prefix == "HK":
        market_type = "hk"
    else:
        market_type = "cn"

    def _series_sum_keys(data, keys: List[str]) -> List[float]:
        """对每年遍历给定 keys 列表, 全部命中项累加 (0 值忽略). 用于 L3 合并"""
        out = []
        for fy in hist_fys:
            s = 0.0
            for k in keys:
                if k in data and fy in data[k]:
                    v = data[k][fy]
                    if v: s += v
            out.append(s)
        return out

    def _fid_series(fids: List[int]) -> List[float]:
        """按 field_id 列表累加, 每年独立求和; 0/缺失字段自动跳过"""
        out = []
        for fy in hist_fys:
            s = 0.0
            for fid in fids:
                if fid in bs_fid and fy in bs_fid[fid]:
                    v = bs_fid[fid][fy]
                    if v: s += v
            out.append(s)
        return out

    def _fid_get(fid: int) -> List[float]:
        """按单个 field_id 取一列, 缺失为 0"""
        out = []
        for fy in hist_fys:
            v = bs_fid.get(fid, {}).get(fy, 0.0)
            out.append(v or 0.0)
        return out

    # ---- Market-aware field_id 白名单 ----
    # 各市场 BS field_id 具有前缀语义: 美股 8xxx / 港股 5xxx / A股 3xxx
    # L2 (Level 2) fid 映射
    fid_l2_ca  = {"us": 8002, "hk": 5002, "cn": 3002}[market_type]
    fid_l2_nca = {"us": 8023, "hk": 5029, "cn": 3025}[market_type]
    fid_l2_cl  = {"us": 8049, "hk": 5061, "cn": 3051}[market_type]
    fid_l2_ncl = {"us": 8065, "hk": 5088, "cn": 3087}[market_type]
    fid_parent_eq = {"us": 8085, "hk": 5110, "cn": 3097}[market_type]
    fid_minority_eq = {"us": 8095, "hk": 5125, "cn": None}[market_type]
    fid_ta = {"us": 8001, "hk": 5001, "cn": 3001}[market_type]
    fid_tl = {"us": 8048, "hk": 5060, "cn": 3050}[market_type]
    fid_te = {"us": 8081, "hk": 5109, "cn": 3115}[market_type]

    # ---- L2: Balance Sheet 层级 2 直接从 API 抽取 ----
    ca_total_series  = _fid_get(fid_l2_ca)
    nca_total_series = _fid_get(fid_l2_nca)
    cl_total_series  = _fid_get(fid_l2_cl)
    ncl_total_series = _fid_get(fid_l2_ncl)
    parent_eq_series = _fid_get(fid_parent_eq)
    # Minority: 若无 (A 股) 用 股东权益合计 - 归属母公司近似
    if fid_minority_eq:
        minority_eq_series = _fid_get(fid_minority_eq)
    else:
        te_total = _fid_get(fid_te)
        minority_eq_series = [max(0.0, te - pe) for te, pe in zip(te_total, parent_eq_series)]
    # 若 API 未返回 minority 但有 股东权益合计 - 归属母公司差值, 用差值填充
    if all(v == 0.0 for v in minority_eq_series):
        te_total = _fid_get(fid_te)
        minority_eq_series = [max(0.0, te - pe) for te, pe in zip(te_total, parent_eq_series)]

    # ---- L3 Buckets: Market-aware field_id groups ----
    # 美股 8xxx BS fid 分布 (基于 BABA/AAPL/AMZN/PDD/GOOG/LI 观察):
    #   Current Assets (8002 total):
    #     - 8003 现金及现金等价物和短期投资 (父项, 已含 8004+8005)
    #     - 8004 -现金和现金等价物 (子项, 不再累加, 用父项 8003)
    #     - 8005 -短期投资 (子项, 已含在 8003)
    #     - 8006 应收款项 (父项, 已含 8007/8010-8013 子项)
    #     - 8016 预付款项, 8017 存货, 8018 受限制现金, 8019 递延资产, 8022 其他流动资产
    #   Non-Current Assets (8023 total):
    #     - 8024 固定资产净额 (父项, 含 8025/8026 子项)
    #     - 8028 总投资 (父项, 含 8031/8033 子项)
    #     - 8035 金融资产, 8037 长期应收票据, 8038 长期预付款项
    #     - 8039 商誉及其他无形资产 (父项, 含 8040/8041 子项)
    #     - 8045 非流动递延资产, 8047 其他非流动资产
    #   Current Liab (8049 total):
    #     - 8050 应付款项 (父项, 含 8051/8052/8054/8055 子项)
    #     - 8056 应计费用, 8057 短期借款与融资租赁负债 (父项含 8058/8060)
    #     - 8063 递延负债 (流动), 8064 其他流动负债
    #   Non-Current Liab (8065 total):
    #     - 8066 长期应付账款及其他应付款项, 8067 长期应计费用
    #     - 8068 长期借款与租赁负债 (父项含 8069/8070)
    #     - 8074 递延负债 (非流动), 8080 其他非流动负债
    #   Equity: 8085 归属母公司 (父项), 8086 股本 (含 8087/8088), 8090 资本公积
    #     8091 留存收益, 8092 库存股, 8093 不影响 RE 损益, 8094 其他股本权益, 8095 少数股东
    #
    # 港股 5xxx BS fid 分布 (基于 00700/83690 观察):
    #   Current Assets (5002 total):
    #     - 5003 现金及等价物, 5005 定期存款-流动, 5006 短期投资 (83690 有)
    #     - 5007 应收账款, 5014 预付款按金及其他应收款
    #     - 5017 按公平值入损益金融资产-流动 (类现金)
    #     - 5019 存货, 5022 已抵押存款, 5026 流动资产特殊项目
    #   Non-Current Assets (5029 total):
    #     - 5031 物业厂房及设备, 5032 在建工程, 5033 投资物业, 5034 土地使用权
    #     - 5036 可供出售金融资产-非流动 (00700 有), 5037 按公平值入损益金融资产-非流动
    #     - 5039 长期投资 (83690 有), 5044 预付款项 (非流动), 5046 无形资产
    #     - 5050 联营公司权益, 5053 合营公司权益, 5054 定期存款-非流动
    #     - 5056 递延税项资产, 5058 非流动资产特殊项目
    #   Current Liab (5061 total):
    #     - 5062 应付账款, 5063 应交税费, 5064 应付股利, 5066 应付票据
    #     - 5067 其他应付款及应计费用, 5068 预收款项 (83690)
    #     - 5070 银行贷款及透支, 5072 短期融资租赁负债, 5083 递延收入-流动
    #   Non-Current Liab (5088 total):
    #     - 5089 长期应付款, 5091 长期银行贷款, 5093 长期融资租赁负债
    #     - 5101 递延税项负债, 5102 递延收入-非流动, 5104 可转换票据及债券, 5106 其他非流动负债
    #   Equity: 5110 归属母公司, 5111 股本, 5112 股本溢价, 5115 保留溢利
    #     5121 -其他储备, 5123 股东权益特殊项目, 5125 少数股东

    if market_type == "us":
        # ─── 美股 L3 分桶 ───
        # 【流动资产】按变现速度: Cash → AR → Inventory → Other
        l3_cash_and_sti  = _fid_series([8003])                     # 父项已含 cash+STI
        l3_ar_and_prepaid = _fid_series([8006, 8016])              # 应收款项父项 + 预付款项
        l3_inventory      = _fid_series([8017])
        l3_other_ca       = _fid_series([8018, 8019, 8022])        # 受限制现金 + 递延资产(流) + 其他

        # 【非流动资产】按变现速度: PP&E → 投资 → 商誉 → 其他
        l3_ppe_and_property     = _fid_series([8024])              # 固定资产净额 (含累计折旧)
        l3_lt_investments       = _fid_series([8028, 8035])        # 总投资 + 金融资产
        l3_goodwill_intangibles = _fid_series([8039])              # 商誉及其他无形资产父项
        l3_other_nca            = _fid_series([8037, 8038, 8045, 8047])  # 长期应收+长期预付+递延+其他

        # 【流动负债】按偿还优先级: 债务 → 应付 → 税 → 其他
        l3_st_debt        = _fid_series([8057])                    # 短期借款与融资租赁负债父项
        # AP bucket 排除 -应交税费 (8052) 以避免与 taxes_payable 重复
        # 由于 8052 是 8050 的子项, 计算 AP bucket = 8050 - 8052
        us_ap = _fid_series([8050])
        us_taxes = _fid_series([8052])
        us_accrued = _fid_series([8056])
        l3_ap_and_accrued = [ap - tx + ac for ap, tx, ac in zip(us_ap, us_taxes, us_accrued)]
        l3_taxes_payable  = us_taxes
        l3_other_cl       = _fid_series([8063, 8064])              # 递延负债(流) + 其他流动负债

        # 【非流动负债】按偿还优先级: 长期债务 → 递延 → 其他
        l3_lt_debt       = _fid_series([8068])                     # 长期借款与租赁负债父项
        l3_deferred_liab = _fid_series([8074])                     # 递延负债 (非流动)
        l3_other_ncl     = _fid_series([8066, 8067, 8080])         # 长期应付 + 长期应计 + 其他非流动

        # 【权益细项】(归属母公司范围内)
        l3_common_stock_apic = _fid_series([8086, 8090])           # 股本 (含普通股/优先股) + 资本公积
        l3_retained_earnings = _fid_series([8091])                 # 留存收益
        l3_other_equity      = _fid_series([8092, 8093, 8094])     # 库存股 + 不影响RE损益 + 其他股本权益

    elif market_type == "hk":
        # ─── 港股 L3 分桶 ───
        l3_cash_and_sti  = _fid_series([5003, 5005, 5006, 5017])   # 现金 + 定期存款(流) + 短期投资 + FVTPL(流)
        l3_ar_and_prepaid = _fid_series([5007, 5014])              # 应收账款 + 预付款按金
        l3_inventory      = _fid_series([5019])
        l3_other_ca       = _fid_series([5022, 5026])              # 已抵押存款 + 流动资产特殊项目

        l3_ppe_and_property     = _fid_series([5031, 5032, 5033, 5034])  # 物业厂房 + 在建工程 + 投资物业 + 土地使用权
        l3_lt_investments       = _fid_series([5036, 5037, 5039, 5050, 5053, 5054])  # AFS + FVTPL(非流) + 长期投资 + 联营 + 合营 + 定期存款(非流)
        l3_goodwill_intangibles = _fid_series([5046])
        l3_other_nca            = _fid_series([5044, 5056, 5058])  # 预付款项(非流) + 递延税项资产 + 非流动特殊

        l3_st_debt        = _fid_series([5070, 5072])              # 银行贷款+透支 + 短期融资租赁
        l3_ap_and_accrued = _fid_series([5062, 5066, 5067, 5068])  # 应付账款 + 应付票据 + 其他应付+应计 + 预收
        l3_taxes_payable  = _fid_series([5063, 5064])              # 应交税费 + 应付股利
        l3_other_cl       = _fid_series([5083])                    # 递延收入-流动

        l3_lt_debt       = _fid_series([5091, 5093, 5104])         # 长期银行贷款 + 长期融资租赁 + 可转换票据
        l3_deferred_liab = _fid_series([5101, 5102])               # 递延税项 + 递延收入(非流)
        l3_other_ncl     = _fid_series([5089, 5106])               # 长期应付款 + 其他非流动负债

        l3_common_stock_apic = _fid_series([5111, 5112])           # 股本 + 股本溢价
        l3_retained_earnings = _fid_series([5115])                 # 保留溢利
        l3_other_equity      = _fid_series([5121, 5123])           # -其他储备 + 股东权益特殊项目

    else:  # A股 CN (3xxx)
        # 保守 fallback: 使用 display_name 匹配 (A 股样本不足, 先用简化映射)
        l3_cash_and_sti  = _series_sum_keys(bs_data, ["货币资金", "交易性金融资产"])
        l3_ar_and_prepaid = _series_sum_keys(bs_data, ["应收账款", "应收票据", "预付款项", "其他应收款"])
        l3_inventory      = _series(bs_data, ["存货"])
        l3_other_ca       = _series_sum_keys(bs_data, ["其他流动资产", "一年内到期的非流动资产"])
        l3_ppe_and_property     = _series_sum_keys(bs_data, ["固定资产合计", "在建工程", "工程物资", "固定资产清理"])
        l3_lt_investments       = _series_sum_keys(bs_data, ["长期股权投资", "其他非流动金融资产", "可供出售金融资产"])
        l3_goodwill_intangibles = _series_sum_keys(bs_data, ["无形资产", "商誉", "开发支出"])
        l3_other_nca            = _series_sum_keys(bs_data, ["其他非流动资产", "长期待摊费用", "递延所得税资产"])
        l3_st_debt        = _series_sum_keys(bs_data, ["短期借款", "一年内到期的非流动负债", "交易性金融负债"])
        l3_ap_and_accrued = _series_sum_keys(bs_data, ["应付账款", "应付票据", "应付职工薪酬", "预收款项", "合同负债"])
        l3_taxes_payable  = _series_sum_keys(bs_data, ["应交税费", "应付利息", "应付股利"])
        l3_other_cl       = _series_sum_keys(bs_data, ["其他流动负债", "其他应付款"])
        l3_lt_debt       = _series_sum_keys(bs_data, ["长期借款", "应付债券"])
        l3_deferred_liab = _series_sum_keys(bs_data, ["递延所得税负债", "递延收益"])
        l3_other_ncl     = _series_sum_keys(bs_data, ["其他非流动负债", "长期应付款"])
        l3_common_stock_apic = _series_sum_keys(bs_data, ["实收资本", "股本", "资本公积"])
        l3_retained_earnings = _series_sum_keys(bs_data, ["留存收益", "未分配利润"])
        l3_other_equity      = _series_sum_keys(bs_data, ["其他综合收益", "盈余公积", "库存股"])

    # 若 CS+APIC 全部为 0, 兜底 (Equity - RE - Other)
    if all(v == 0.0 for v in l3_common_stock_apic):
        l3_common_stock_apic = [max(0.0, pe - re - oe)
                                 for pe, re, oe in zip(parent_eq_series, l3_retained_earnings, l3_other_equity)]
    # 若 RE 全部为 0, 兜底: Equity - CS - Other
    if all(v == 0.0 for v in l3_retained_earnings):
        l3_retained_earnings = [max(0.0, pe - cs - oe)
                                 for pe, cs, oe in zip(parent_eq_series, l3_common_stock_apic, l3_other_equity)]

    # 校验并调整 "Other" plug 桶, 使 L3 sum 精确等于 L2 total
    # 原则: 前 N-1 桶保持从 API 直接抽取的值, 最后一个 "Other" 桶 = L2 - sum(前 N-1 桶)
    # 这样 Balance Check 在历史列一定为 0
    def _adjust_plug(l3_first_buckets: list, l2_total: List[float]) -> List[float]:
        """给定前 N-1 个 L3 桶 (每个是 series) 与 L2 总额, 返回作为 plug 的第 N 桶 series."""
        other = []
        for i in range(len(l2_total)):
            summed = sum(bucket[i] for bucket in l3_first_buckets)
            other.append(l2_total[i] - summed)  # 允许负值 (若 L3 前项已超过 L2, 表示重复)
        return other

    l3_other_ca  = _adjust_plug([l3_cash_and_sti, l3_ar_and_prepaid, l3_inventory], ca_total_series)
    l3_other_nca = _adjust_plug([l3_ppe_and_property, l3_lt_investments, l3_goodwill_intangibles], nca_total_series)
    l3_other_cl  = _adjust_plug([l3_st_debt, l3_ap_and_accrued, l3_taxes_payable], cl_total_series)
    l3_other_ncl = _adjust_plug([l3_lt_debt, l3_deferred_liab], ncl_total_series)
    l3_other_equity = _adjust_plug([l3_common_stock_apic, l3_retained_earnings], parent_eq_series)

    # ---- Legacy alias 抽取 (供 WC / D&A Schedule / IS 等下游使用) ----
    # 与 L3 aggregated 保持一致, 但用 legacy 名称 (避免大量 refactor)
    total_assets_series = _series(bs_data, ["资产合计"])
    total_liab_series   = _series(bs_data, ["负债合计"])
    equity_series = parent_eq_series
    re_series = l3_retained_earnings

    # ==========================================================================
    # Cash Flow L1/L2 抽取 (保持现有结构, 只改进映射)
    # L1 (OCF/CFI/CFF) 直接从 API 汇总项拉取
    # L2 (二级指标) 精确映射, 保证 L2 sum ≈ L1 (符号一致)
    # ==========================================================================
    def _cf_fid_series(fids: List[int]) -> List[float]:
        out = []
        for fy in hist_fys:
            s = 0.0
            for fid in fids:
                if fid in cf_fid and fy in cf_fid[fid]:
                    v = cf_fid[fid][fy]
                    if v: s += v
            out.append(s)
        return out

    def _cf_fid_get(fid: int) -> List[float]:
        out = []
        for fy in hist_fys:
            v = cf_fid.get(fid, {}).get(fy, 0.0)
            out.append(v or 0.0)
        return out

    if market_type == "us":
        # 美股 CF fid 分布:
        #   OCF 8015 (父项) = 8017 持续经营净收入 + 8019 D&A + 8026 递延所得税 + 8027 其他非现金 + 8028 营运资金变化
        #   CFI 8042 (父项) = 8046 固定资产 + 8047 无形资产 + 8049 业务交易 + 8051 投资产品 + 8054 其他
        #   CFF 8056 (父项) = 8058 债务发行/偿还净额 + 8059 普通股发行/回购净额 + 8061 现金股利 + 8065 其他融资
        ocf_total_series   = _cf_fid_get(8015)
        cfi_total_series   = _cf_fid_get(8042)
        cff_total_series   = _cf_fid_get(8056)
        cf_ni_series       = _cf_fid_get(8017)   # 持续经营净收入 (与 IS NI 对齐)
        cf_da_series       = _cf_fid_get(8019)   # 折旧摊销及损耗
        cf_dtx_series      = _cf_fid_get(8026)   # 递延所得税
        cf_other_op_series = _cf_fid_get(8027)   # 其他非现金
        cf_wc_change_series = _cf_fid_get(8028)  # 营运资金变化
        cf_capex_series    = [-abs(v) for v in _cf_fid_series([8046, 8047])]  # 固定资产+无形资产 (负值)
        cf_other_inv_series = _cf_fid_series([8049, 8051, 8054])  # 业务/投资产品/其他 (含符号)
        cf_debt_net_series = _cf_fid_get(8058)   # 债务发行/偿还净额 (含符号)
        cf_equity_net_series = _cf_fid_get(8059) # 普通股发行/回购净额
        cf_dividends_series  = _cf_fid_get(8061) # 现金股利 (负值)
        cf_other_fin_series  = _cf_fid_get(8065) # 其他融资
    elif market_type == "hk":
        # 港股 CF fid 分布:
        #   OCF 5001 = 5058 经营现金收入 + 5061 其他税项 + 5066 特殊项
        #   OCF 分解也可看: 5002 除税前经营利润+调整 → OCF
        #   CFI 5069 = 5070 出售固定资产 + 5071 购买固定资产 + 5073 购买无形资产 + 5074/5076/5077 投资交易
        #   CFF 5086 = 5087 新增借款 + 5088 偿还借款 + 5089 发行股份 + 5091 发行费用赎回 + 5093 已付利息 + 5094 已付股息
        ocf_total_series   = _cf_fid_get(5001)
        cfi_total_series   = _cf_fid_get(5069)
        cff_total_series   = _cf_fid_get(5086)
        # 港股 OCF 分解: 5003 除税前利润 (类 NI) + 5009 D&A + 5034 营运资金变动 + 其他调整
        cf_ni_series       = _cf_fid_get(5003)   # 除税前利润 (港股 CF 从税前利润起算, 非 NI)
        cf_da_series       = _cf_fid_get(5009)   # 折旧及摊销
        cf_dtx_series      = [0.0] * len(hist_fys)  # 港股无独立递延所得税
        # 其他非现金: 5004+5005+5006+5007+5013+5020+5024+5030+5032 (调整项)
        cf_other_op_series = _cf_fid_series([5004, 5005, 5006, 5007, 5013, 5020, 5024, 5030, 5032])
        cf_wc_change_series = _cf_fid_get(5034)  # 营运资金变动项目
        cf_capex_series    = [-abs(v) for v in _cf_fid_series([5071, 5073])]  # 购买固定资产 + 购买无形资产
        cf_other_inv_series = _cf_fid_series([5070, 5074, 5076, 5077, 5078, 5079, 5081, 5082, 5083, 5061, 5050, 5066])
        cf_debt_net_series = _cf_fid_series([5087, 5088])  # 新增借款 + 偿还借款
        cf_equity_net_series = _cf_fid_series([5089, 5091])  # 发行股份 + 发行费用赎回
        cf_dividends_series  = _cf_fid_get(5094)   # 已付股息-融资
        cf_other_fin_series  = _cf_fid_series([5090, 5093, 5096])  # 发行债券+已付利息+融资业务其他
    else:  # A 股 CN
        ocf_total_series   = _series(cf_data, ["经营活动产生的现金流量净额", "经营活动现金流量净额"])
        cfi_total_series   = _series(cf_data, ["投资活动产生的现金流量净额", "投资活动现金流量净额"])
        cff_total_series   = _series(cf_data, ["筹资活动产生的现金流量净额", "融资活动现金流量净额"])
        cf_ni_series       = [0.0] * len(hist_fys)
        cf_da_series       = _series(cf_data, ["折旧与摊销", "折旧及摊销"])
        cf_dtx_series      = [0.0] * len(hist_fys)
        cf_other_op_series = [0.0] * len(hist_fys)
        cf_wc_change_series = [0.0] * len(hist_fys)
        cf_capex_series    = [-abs(v) for v in _series(cf_data, [
            "购建固定资产、无形资产和其他长期资产支付的现金"])]
        cf_other_inv_series = [0.0] * len(hist_fys)
        cf_debt_net_series = [0.0] * len(hist_fys)
        cf_equity_net_series = [0.0] * len(hist_fys)
        cf_dividends_series  = [0.0] * len(hist_fys)
        cf_other_fin_series  = [0.0] * len(hist_fys)

    # 计算 OCF plug (其他非现金 = OCF Total - NI - D&A - DTax - WC Change - 已抽取其他)
    # 用于 "其他调整" plug 桶保证 L2 sum = L1
    _cf_ocf_l2_sum = [n+d+t+o+w for n,d,t,o,w in zip(
        cf_ni_series, cf_da_series, cf_dtx_series, cf_other_op_series, cf_wc_change_series)]
    cf_ocf_plug_series = [total - part for total, part in zip(ocf_total_series, _cf_ocf_l2_sum)]

    _cf_cfi_l2_sum = [c+o for c,o in zip(cf_capex_series, cf_other_inv_series)]
    cf_cfi_plug_series = [total - part for total, part in zip(cfi_total_series, _cf_cfi_l2_sum)]

    _cf_cff_l2_sum = [d+e+dv+o for d,e,dv,o in zip(
        cf_debt_net_series, cf_equity_net_series, cf_dividends_series, cf_other_fin_series)]
    cf_cff_plug_series = [total - part for total, part in zip(cff_total_series, _cf_cff_l2_sum)]


    # WC 用 (Days-driven schedule 需要 AR / Inv / AP 单项)
    ar_series = _series(bs_data, ["-应收账款净额", "应收账款净额", "应收账款", "应收款项"])
    ap_series = _series(bs_data, ["应付账款", "-应付账款"])
    inv_series = _series(bs_data, ["存货", "-存货"])
    # PPE (D&A Schedule 用)
    ppe_series = _series(bs_data, ["固定资产净额", "物业厂房及设备", "固定资产合计", "固定资产"])
    # Intangible (BS Level 3 用 l3_goodwill_intangibles; legacy alias 保留)
    intangible_series = l3_goodwill_intangibles
    # Cash (dcf-model 与 DCF Bridge 用)
    cash_series = l3_cash_and_sti[:]  # copy
    # Debt (total, WACC 与 Debt Schedule 用)
    debt_series = [st + lt for st, lt in zip(l3_st_debt, l3_lt_debt)]

    # ---- 假设默认值 (基于最新 FY) ----
    latest_idx = len(hist_fys) - 1
    lr = revenue_series[latest_idx] if revenue_series[latest_idx] else 1.0
    lc = cogs_series[latest_idx]
    prior_rev = gv(inc_data, ["总收入", "营业总收入"], prior_fy) if prior_fy else 0.0
    rev_growth_series = []
    for i, rev in enumerate(revenue_series):
        prev = revenue_series[i-1] if i > 0 else prior_rev
        rev_growth_series.append(safe_divide(rev - prev, prev) if prev else 0.0)
    latest_growth = rev_growth_series[-1] if rev_growth_series else 0.05

    cogs_pct  = safe_divide(cogs_series[latest_idx], lr)
    opex_pct  = safe_divide(opex_series[latest_idx], lr)
    da_pct    = safe_divide(da_series[latest_idx],   lr)
    capex_pct = safe_divide(capex_series[latest_idx], lr)
    tax_rate  = safe_divide(tax_series[latest_idx], ebt_series[latest_idx]) if ebt_series[latest_idx] > 0 else 0.25

    ar_days  = safe_divide(ar_series[latest_idx] * 365, lr)
    ap_days  = safe_divide(ap_series[latest_idx] * 365, lc) if lc else 0.0
    inv_days = safe_divide(inv_series[latest_idx] * 365, lc) if lc else 0.0

    # ---- NWC % of ΔRevenue: 用最近 3 年 ΔNWC / ΔRevenue 的均值, 避免单期噪声 ----
    # ΔNWC (CF 定义) = -(ΔAR + ΔInv - ΔAP) = (AR_prev - AR_curr) + (Inv_prev - Inv_curr) + (AP_curr - AP_prev)
    # 与 Working Capital sheet 的 dNWC 符号约定一致 (资产增加为现金流出, 负贡献 → NWC%_of_Δrev 为负)
    # 用 abs(ΔNWC / ΔRev) 取正号作为 DCF sheet 中 "NWC % of Delta Revenue" 假设
    _nwc_ratios = []
    for i in range(1, len(hist_fys)):
        d_ar  = ar_series[i-1] - ar_series[i]
        d_inv = inv_series[i-1] - inv_series[i]
        d_ap  = ap_series[i]   - ap_series[i-1]
        delta_nwc = d_ar + d_inv + d_ap
        delta_rev = revenue_series[i] - revenue_series[i-1]
        if delta_rev != 0:
            # WC = AR + Inv - AP (资产项), 常见口径为 ΔWC/ΔRev, 用绝对值让 DCF 假设为正百分比
            wc_change = -delta_nwc   # 转换: dNWC (现金视角) → ΔWC (资产项视角)
            ratio = wc_change / delta_rev
            _nwc_ratios.append(ratio)
    # 用最近 3 年均值 (若数据不足则用可得的均值); 缺失则回退 1%
    if _nwc_ratios:
        recent = _nwc_ratios[-3:] if len(_nwc_ratios) >= 3 else _nwc_ratios
        nwc_pct = sum(recent) / len(recent)
    else:
        nwc_pct = 0.01
    # 合理性约束: NWC% 通常在 -10% ~ +10%, 极端值 (ΔRev 极小导致比率失真) 收窄
    if abs(nwc_pct) > 0.10:
        logger.warning(f"NWC% = {nwc_pct:.1%} 极端值, 收窄到 ±10% 上限 (可能 ΔRev 太小导致比率失真)")
        nwc_pct = max(-0.10, min(0.10, nwc_pct))

    # ---- 市场数据 & 币种 ----
    stock_code = normalize_stock_code(ticker)
    market = fetch_market_data_from_futu(stock_code)
    if market.get("stock_price") is None:
        market["stock_price"] = 100.0
        market["shares_source"] = "Default Fallback"
    if market.get("shares_outstanding") is None:
        market["shares_outstanding"] = 1000.0
        market["shares_source"] = "Default Fallback"

    trading_currency = infer_trading_currency(stock_code)
    if not reporting_currency:
        logger.warning(f"财报未提供币种, 回退到交易币种 {trading_currency}")
        reporting_currency = trading_currency
    fx_rate, fx_source = fetch_fx_rate_from_futu(trading_currency, reporting_currency)
    logger.info(f"币种: trading={trading_currency}, reporting={reporting_currency}, "
                f"FX=1 {trading_currency} = {fx_rate:.4f} {reporting_currency}")

    # ---- WACC 输入个股化: Beta / Rf / ERP (用于新增的 DCF/WACC sheet) ----
    # 传入 reporting_currency 让 Beta 基准正确识别 ADR 中概股 (US.* + CNY → KWEB, 而非 SPY)
    benchmark = get_benchmark_for(stock_code, reporting_currency)
    beta_calc, beta_source = compute_beta_from_futu(
        stock_code, benchmark, reporting_currency=reporting_currency)
    if beta_calc is None:
        logger.warning(f"Beta 计算失败, 使用兜底 {_BETA_FALLBACK}: {beta_source}")
        beta_calc = _BETA_FALLBACK
        beta_source = f"Fallback constant ({beta_source})"
    rf_erp = _RF_ERP_BY_CURRENCY.get(reporting_currency, _RF_ERP_DEFAULT)
    rf_rate = rf_erp["rf"]; erp = rf_erp["erp"]
    rf_src = rf_erp["rf_src"]; erp_src = rf_erp["erp_src"]
    logger.info(f"WACC 输入: beta={beta_calc:.4f} ({benchmark}), "
                f"Rf={rf_rate:.2%} ({rf_src}), ERP={erp:.2%} ({erp_src})")

    # 市场判别: 已在 BS L3 抽取之前完成 (见上方 market_type 赋值)

    # EBIT Margin 与 Other Income 假设 (基于历史反算)
    ebit_margin_series = [safe_divide(e, r) for e, r in zip(ebit_series, revenue_series)]
    other_income_series = []
    for i, fy in enumerate(hist_fys):
        # 分市场计算 Other Income:
        # 港股: EBT - EBIT - FinInc + FinCost - EqAff (通常接近 0, 明细已拆细)
        # 其他: EBT - EBIT (吞并所有非营业项, 含 Interest)
        if market_type == "hk":
            oi = (ebt_series[i] - ebit_series[i] - finance_income_series[i]
                  + finance_cost_series[i] - equity_affiliate_series[i])
        else:
            oi = ebt_series[i] - ebit_series[i]
        other_income_series.append(oi)
    other_income_pct_series = [safe_divide(o, r) for o, r in zip(other_income_series, revenue_series)]
    latest_ebit_margin = ebit_margin_series[-1] if ebit_margin_series else 0.05
    latest_other_income_pct = other_income_pct_series[-1] if other_income_pct_series else 0.0

    return {
        "ticker": ticker, "stock_code": stock_code, "hist_fys": hist_fys,
        "market_type": market_type,
        "hist_revenue": revenue_series, "hist_cogs": cogs_series, "hist_opex": opex_series,
        "hist_ebit": ebit_series, "hist_tax": tax_series, "hist_ebt": ebt_series,
        "hist_ni": ni_series, "hist_da": da_series, "hist_capex": capex_series,
        "hist_da_is": da_series_is, "hist_da_cf": da_series_cf,
        "hist_cash": cash_series, "hist_ar": ar_series, "hist_ap": ap_series,
        "hist_inv": inv_series, "hist_ppe": ppe_series, "hist_intangible": intangible_series,
        "hist_equity": equity_series, "hist_re": re_series, "hist_debt": debt_series,
        "hist_total_assets": total_assets_series, "hist_total_liab": total_liab_series,
        "hist_rev_growth": rev_growth_series,
        # BS Level 2 (直接从 API 抽取)
        "hist_ca_total":  ca_total_series,
        "hist_nca_total": nca_total_series,
        "hist_cl_total":  cl_total_series,
        "hist_ncl_total": ncl_total_series,
        "hist_parent_eq": parent_eq_series,
        "hist_minority_eq": minority_eq_series,
        # BS Level 3 (aggregated buckets, 每组 3-5 项)
        "hist_l3_cash_sti":            l3_cash_and_sti,
        "hist_l3_ar_prepaid":          l3_ar_and_prepaid,
        "hist_l3_inventory":           l3_inventory,
        "hist_l3_other_ca":            l3_other_ca,
        "hist_l3_ppe_property":        l3_ppe_and_property,
        "hist_l3_lt_investments":      l3_lt_investments,
        "hist_l3_goodwill_intangibles": l3_goodwill_intangibles,
        "hist_l3_other_nca":           l3_other_nca,
        "hist_l3_st_debt":             l3_st_debt,
        "hist_l3_ap_accrued":          l3_ap_and_accrued,
        "hist_l3_taxes_payable":       l3_taxes_payable,
        "hist_l3_other_cl":            l3_other_cl,
        "hist_l3_lt_debt":             l3_lt_debt,
        "hist_l3_deferred_liab":       l3_deferred_liab,
        "hist_l3_other_ncl":           l3_other_ncl,
        "hist_l3_common_stock_apic":   l3_common_stock_apic,
        "hist_l3_retained_earnings":   l3_retained_earnings,
        "hist_l3_other_equity":        l3_other_equity,
        # Cash Flow L1/L2 (直接从 API 抽取, 与实际报告严格一致)
        "hist_ocf_total":              ocf_total_series,
        "hist_cfi_total":              cfi_total_series,
        "hist_cff_total":              cff_total_series,
        "hist_cf_ni":                  cf_ni_series,
        "hist_cf_da":                  cf_da_series,
        "hist_cf_dtx":                 cf_dtx_series,
        "hist_cf_other_op":            cf_other_op_series,
        "hist_cf_wc_change":           cf_wc_change_series,
        "hist_cf_ocf_plug":            cf_ocf_plug_series,
        "hist_cf_capex":               cf_capex_series,
        "hist_cf_other_inv":           cf_other_inv_series,
        "hist_cf_cfi_plug":            cf_cfi_plug_series,
        "hist_cf_debt_net":            cf_debt_net_series,
        "hist_cf_equity_net":          cf_equity_net_series,
        "hist_cf_dividends":           cf_dividends_series,
        "hist_cf_other_fin":           cf_other_fin_series,
        "hist_cf_cff_plug":            cf_cff_plug_series,
        # 港股专属明细行 (美股/A 股为 0 序列)
        "hist_finance_income": finance_income_series,
        "hist_finance_cost": finance_cost_series,
        "hist_equity_affiliate": equity_affiliate_series,
        "hist_other_income": other_income_series,
        "hist_ebit_margin": ebit_margin_series,
        "hist_other_income_pct": other_income_pct_series,
        "growth": latest_growth, "cogs_pct": cogs_pct, "opex_pct": opex_pct,
        "da_pct": da_pct, "capex_pct": capex_pct, "tax_rate": tax_rate,
        "nwc_pct": nwc_pct,   # ΔNWC/ΔRevenue 最近 3 年均值 (供 DCF sheet 使用, 有正负号)
        "ebit_margin": latest_ebit_margin, "other_income_pct": latest_other_income_pct,
        "ar_days": ar_days, "ap_days": ap_days, "inv_days": inv_days,
        "stock_price": market["stock_price"], "shares_outstanding": market["shares_outstanding"],
        "shares_source": market["shares_source"],
        "reporting_currency": reporting_currency, "trading_currency": trading_currency,
        "fx_rate": fx_rate, "fx_source": fx_source,
        # WACC 输入 (DCF/WACC sheet 用)
        "beta": beta_calc, "beta_source": beta_source, "benchmark": benchmark,
        "rf_rate": rf_rate, "rf_source": rf_src,
        "erp": erp, "erp_source": erp_src,
    }


# ==================== Model Builder ====================
class ThreeStatementBuilder:
    """构建 3-Statement 模型 (严格对齐 references/schema.md).

    Sheet 顺序:
      1) Assumptions       - 分区式假设 (HEADER / MARKET / REVENUE / COST / BS / DEBT / DIV)
      2) Income Statement  - 5A + 5E, 含 Margin% 展示行
      3) Balance Sheet     - Days-driven, Balance Check + Cash Tie-Out
      4) Cash Flow         - OCF/CFI/CFF, 三段
      5) D&A Schedule      - PP&E roll-forward
      6) Debt Schedule     - Beg -> Iss -> Repay -> Sweep -> End
      7) Working Capital   - AR/Inv/AP Days-driven

    列布局 (所有报表 Sheet):
      A = 标签, B = 单位
      C..G = 5 期历史 (最老 -> 最新), 若历史不足 5 期, 从右对齐 (G 为最新历史)
      H..L = 5 期预测 (FY1..FY5)
    """

    def __init__(self, d: dict):
        self.d = d
        self.hist_fys = d["hist_fys"]
        n_hist = len(self.hist_fys)
        # 历史列右对齐到 G (col 7)
        self.n_hist = min(n_hist, 5)
        self.hist_start_col = 3 + (5 - self.n_hist)   # 若不足 5 期, 左侧留空
        self.HIST_COLS = list(range(self.hist_start_col, self.hist_start_col + self.n_hist))
        self.FCST_COLS = list(range(8, 13))            # H..L
        self.ALL_COLS  = list(range(3, 13))            # C..L
        self.hist_offset = n_hist - self.n_hist         # 若历史多于 5 期, 只用最近 5 期

        self.wb = openpyxl.Workbook()
        # 存放各 Sheet 的行号索引 (供跨 Sheet 引用)
        self.rows = {"assump": {}, "is": {}, "bs": {}, "cf": {}, "da": {}, "debt": {}, "wc": {},
                     "dcf": {}, "wacc": {}}

    # ---------- 通用样式 helper ----------
    def _apply_input(self, cell, value, fmt=None, source_system=None, source_ref="", is_locked=False):
        cell.value = value
        cell.font = FONT_BLUE if not is_locked else FONT_BLACK
        cell.fill = FILL_INPUT_GREY
        if fmt: cell.number_format = fmt
        if source_system: add_source_comment(cell, source_system, source_ref)

    def _write_hist_input(self, ws, row, series, fmt=FMT_CURRENCY_M, source=None, ref=""):
        """将历史序列写入 HIST_COLS (蓝色输入, 带 Source comment)。"""
        for i, col in enumerate(self.HIST_COLS):
            v = series[self.hist_offset + i] if self.hist_offset + i < len(series) else None
            if v is None: continue
            c = ws.cell(row, col, v)
            c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = fmt
            if source: add_source_comment(c, source, ref)

    def _write_section_header(self, ws, row, label, span=12):
        c = ws.cell(row, 1, label)
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)

    def _write_column_headers(self, ws, row):
        """在 row 写入列头 (Item / Unit / 2020A..2024A / 2025E..2029E)。"""
        ws.cell(row, 1, "Line Item / 项目").font = FONT_BOLD
        ws.cell(row, 1).fill = FILL_LIGHT_BLUE
        ws.cell(row, 2, "Unit").font = FONT_BOLD
        ws.cell(row, 2).fill = FILL_LIGHT_BLUE
        for i, col in enumerate(self.HIST_COLS):
            fy_label = self.hist_fys[self.hist_offset + i]
            c = ws.cell(row, col, f"{fy_label} (A)")
            c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        for i, col in enumerate(self.FCST_COLS):
            c = ws.cell(row, col, f"FY{i+1} (E)")
            c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        # 历史列 vs 预测列 分隔线
        last_hist_col = self.HIST_COLS[-1] if self.HIST_COLS else 2
        for r_check in [row]:
            ws.cell(r_check, last_hist_col).border = BORDER_HIST_RIGHT

    def _fmt_column_widths(self, ws):
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 10
        for col in self.ALL_COLS:
            ws.column_dimensions[get_column_letter(col)].width = 13

    def build(self, output_path: Path):
        # 顺序: Assumptions -> Schedules -> IS -> CF -> BS -> DCF -> WACC
        # (下游 Sheet 引用上游行号, 但因为都是公式字符串, 顺序仅影响 rows 字典的填充时机)
        self._assumptions()
        self._wc_schedule()
        self._da_schedule()
        self._debt_schedule()
        self._income_statement()
        self._cash_flow()
        self._balance_sheet()
        self._wacc()   # 先建 WACC (DCF sheet 的 CAPM 引用 WACC!B18)
        self._dcf()    # 再建 DCF (含 3 张 5x5 敏感性表)
        # 按 schema.md Tab 顺序重排 Sheet
        desired = ["Assumptions", "Income Statement", "Balance Sheet", "Cash Flow",
                   "D&A Schedule", "Debt Schedule", "Working Capital", "DCF", "WACC"]
        # openpyxl 通过 _sheets 顺序控制
        self.wb._sheets = [self.wb[n] for n in desired if n in self.wb.sheetnames]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(output_path))
        logger.info(f"3-Statement + DCF model saved: {output_path}")

    # ==================== Tab 1: Assumptions ====================
    def _assumptions(self):
        ws = self.wb.active; ws.title = "Assumptions"
        d = self.d
        rep_ccy = d["reporting_currency"]; trd_ccy = d["trading_currency"]

        # --- HEADER 区 ---
        c = ws.cell(1, 1, f"{d['ticker']} 3-Statement Financial Model / {d['ticker']} 三表联动财务模型")
        c.fill = FILL_DARK_BLUE; c.font = FONT_WHITE_BOLD
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        c = ws.cell(2, 1, f"Ticker: {d['ticker']}  |  Date: {date.today().isoformat()}  |  Reporting: Million {rep_ccy}  |  Trading Ccy: {trd_ccy}")
        c.font = FONT_ITALIC_GREY
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        c = ws.cell(3, 1, f"Data Source: {d['shares_source']}  |  FX: {d['fx_source']}")
        c.font = FONT_GREEN
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)

        row = 5
        r = self.rows["assump"]

        # --- MARKET DATA 区 ---
        self._write_section_header(ws, row, "MARKET DATA -- 市场数据", span=3); row += 1
        ws.cell(row, 1, f"Current Stock Price ({trd_ccy})")
        self._apply_input(ws.cell(row, 2), d["stock_price"], FMT_PRICE,
                          "Futu get_market_snapshot", "last_price")
        r["stock_price"] = row; row += 1

        ws.cell(row, 1, f"FX Rate: 1 {trd_ccy} = X {rep_ccy}")
        fx_lock = (trd_ccy == rep_ccy)
        cell = ws.cell(row, 2, d["fx_rate"])
        cell.font = FONT_BLACK if fx_lock else FONT_BLUE
        cell.fill = FILL_INPUT_GREY; cell.number_format = FMT_DECIMAL4
        add_comment(cell, f"FX 汇率: 1 {trd_ccy} = X {rep_ccy}\n来源: {d['fx_source']}\n用户可覆盖")
        r["fx_rate"] = row; row += 1

        ws.cell(row, 1, f"Shares Outstanding (M)")
        self._apply_input(ws.cell(row, 2), d["shares_outstanding"], FMT_SHARES,
                          "Futu get_market_snapshot", "issued_shares (÷1M)")
        r["shares"] = row; row += 2

        # --- REVENUE ASSUMPTIONS 区 ---
        self._write_section_header(ws, row, "REVENUE ASSUMPTIONS -- 收入假设 (逐年输入)", span=12); row += 1
        # 列头 (FY1..FY5)
        for i, col in enumerate(self.FCST_COLS):
            c = ws.cell(row, col, f"FY{i+1}"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        row += 1
        # Revenue Growth % (5 期预测输入, 与历史比较)
        ws.cell(row, 1, "Revenue Growth % / 营收增长率")
        for i, col in enumerate(self.FCST_COLS):
            # 默认: 用最新历史增长率, 逐年递减 0.5%
            val = max(d["growth"] - 0.005 * i, 0.02)
            self._apply_input(ws.cell(row, col), val, FMT_PERCENT,
                              "Assumption", f"Base on {self.hist_fys[-1]} actual growth")
        r["rev_growth"] = row; row += 2

        # --- COST ASSUMPTIONS 区 ---
        self._write_section_header(ws, row, "COST ASSUMPTIONS -- 成本假设 (%/率, 逐年可变)", span=12); row += 1
        # 列头
        for i, col in enumerate(self.FCST_COLS):
            c = ws.cell(row, col, f"FY{i+1}"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        row += 1
        # COGS %
        ws.cell(row, 1, "COGS % of Revenue / 营收成本率")
        for col in self.FCST_COLS:
            self._apply_input(ws.cell(row, col), d["cogs_pct"], FMT_PERCENT,
                              "Assumption", f"Latest FY COGS/Revenue = {d['cogs_pct']:.2%}")
        r["cogs_pct"] = row; row += 1
        # OpEx %
        ws.cell(row, 1, "OpEx % of Revenue / 营业费用率 (S&M + G&A + R&D)")
        for col in self.FCST_COLS:
            self._apply_input(ws.cell(row, col), d["opex_pct"], FMT_PERCENT,
                              "Assumption", "Latest FY OpEx/Revenue")
        r["opex_pct"] = row; row += 1
        # D&A %
        ws.cell(row, 1, "D&A % of Revenue / 折旧摊销率")
        for col in self.FCST_COLS:
            # 使用最近一年实际值 (不加人为下限, 尊重公司真实结构)
            self._apply_input(ws.cell(row, col), d["da_pct"], FMT_PERCENT,
                              "Assumption", f"Latest FY D&A/Revenue = {d['da_pct']:.2%}")
        r["da_pct"] = row; row += 1
        # EBIT Margin % (直接驱动 EBIT 预测, 替代 GP - OpEx - D&A 组装)
        ws.cell(row, 1, "EBIT Margin % / 息税前利润率")
        for col in self.FCST_COLS:
            self._apply_input(ws.cell(row, col), d["ebit_margin"], FMT_PERCENT,
                              "Assumption", f"Latest FY EBIT/Revenue = {d['ebit_margin']:.2%}")
        r["ebit_margin"] = row; row += 1
        # Other Income % of Revenue (非营业净收益/损失, 港股主要含 应占联营/融资净收支 之外的残差; 美股/A股为 EBT − EBIT plug)
        ws.cell(row, 1, "Other Income % of Revenue / 其他非营业净收益率")
        for col in self.FCST_COLS:
            self._apply_input(ws.cell(row, col), d["other_income_pct"], FMT_PERCENT,
                              "Assumption", f"Latest FY (EBT-EBIT-fin adj)/Revenue = {d['other_income_pct']:.2%}")
        r["other_income_pct"] = row; row += 1
        # Tax Rate
        ws.cell(row, 1, "Tax Rate / 税率")
        for col in self.FCST_COLS:
            self._apply_input(ws.cell(row, col), d["tax_rate"], FMT_PERCENT,
                              "Assumption", "Latest FY 有效税率, 亏损默认 25%")
        r["tax_rate"] = row; row += 2

        # --- BS ASSUMPTIONS 区 ---
        self._write_section_header(ws, row, "BALANCE SHEET ASSUMPTIONS -- 资产负债表假设", span=12); row += 1
        for i, col in enumerate(self.FCST_COLS):
            c = ws.cell(row, col, f"FY{i+1}"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        row += 1
        # CapEx %
        ws.cell(row, 1, "CapEx % of Revenue / 资本开支率")
        for col in self.FCST_COLS:
            # 使用最近一年实际值 (不加人为下限, 尊重公司真实资本结构)
            self._apply_input(ws.cell(row, col), d["capex_pct"], FMT_PERCENT,
                              "Assumption", f"Latest FY CapEx (strict)/Revenue = {d['capex_pct']:.2%}")
        r["capex_pct"] = row; row += 1
        # AR Days
        ws.cell(row, 1, "AR Days / 应收账款周转天数")
        for col in self.FCST_COLS:
            self._apply_input(ws.cell(row, col), d["ar_days"], FMT_DAYS,
                              "Assumption", f"AR × 365 / Revenue = {d['ar_days']:.1f}")
        r["ar_days"] = row; row += 1
        # Inventory Days
        ws.cell(row, 1, "Inventory Days / 存货周转天数")
        for col in self.FCST_COLS:
            self._apply_input(ws.cell(row, col), d["inv_days"], FMT_DAYS,
                              "Assumption", f"Inv × 365 / COGS = {d['inv_days']:.1f}")
        r["inv_days"] = row; row += 1
        # AP Days
        ws.cell(row, 1, "AP Days / 应付账款周转天数")
        for col in self.FCST_COLS:
            self._apply_input(ws.cell(row, col), d["ap_days"], FMT_DAYS,
                              "Assumption", f"AP × 365 / COGS = {d['ap_days']:.1f}")
        r["ap_days"] = row; row += 2

        # --- DEBT ASSUMPTIONS 区 ---
        self._write_section_header(ws, row, "DEBT ASSUMPTIONS -- 债务假设", span=3); row += 1
        ws.cell(row, 1, "Interest Rate on Debt / 债务利率")
        self._apply_input(ws.cell(row, 2), 0.045, FMT_PERCENT,
                          "Assumption", "Default 4.5%, 10-K 债券利率或市场基准可覆盖")
        r["interest_rate"] = row; row += 1
        ws.cell(row, 1, "Mandatory Repayment / yr / 每年强制还款")
        self._apply_input(ws.cell(row, 2), 0.0, FMT_CURRENCY_M,
                          "Assumption", "每年强制偿还本金 (百万, 报表币种)")
        r["mandatory_repay"] = row; row += 1
        ws.cell(row, 1, "Cash Sweep % / 超额现金还款比例")
        self._apply_input(ws.cell(row, 2), 0.0, FMT_PERCENT,
                          "Assumption", "自由现金还款比例 (0=关闭)")
        r["cash_sweep"] = row; row += 2

        # --- DIVIDEND ASSUMPTIONS 区 ---
        self._write_section_header(ws, row, "DIVIDEND ASSUMPTIONS -- 股利/回购假设", span=3); row += 1
        ws.cell(row, 1, "Dividend Payout Ratio / 股利支付率")
        self._apply_input(ws.cell(row, 2), 0.0, FMT_PERCENT,
                          "Assumption", "Dividends / Net Income (0 = 无分红)")
        r["div_payout"] = row; row += 1
        ws.cell(row, 1, "Share Repurchases / yr / 每年回购金额")
        self._apply_input(ws.cell(row, 2), 0.0, FMT_CURRENCY_M,
                          "Assumption", "年度回购金额 (百万, 报表币种)")
        r["repurchase"] = row; row += 1

        # 列宽
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 14
        for col in self.FCST_COLS:
            ws.column_dimensions[get_column_letter(col)].width = 12

    # ==================== Tab 7: Working Capital Schedule ====================
    def _wc_schedule(self):
        """AR / Inventory / AP Days-driven; 生成 Balance 与 Δ 送 CF/BS。"""
        ws = self.wb.create_sheet("Working Capital")
        d = self.d; r = self.rows["wc"]
        assump = self.rows["assump"]

        self._write_section_header(ws, 1, "WORKING CAPITAL SCHEDULE -- 营运资本表", span=12)
        row = 2
        self._write_column_headers(ws, row); row += 1

        # 依赖: Revenue / COGS 从 IS 引用 (下面 IS 会填充 rows["is"]["revenue"], ["cogs"])
        # 但 WC 先于 IS 建立, 因此需要 pre-declare 目标行号。
        # 让我们约定 IS 行号 (稍后 IS 会遵守此布局):
        # IS 第 8 行 = Revenue, 第 10 行 = COGS (稍后 IS 严格按此)
        is_rev_row = 8
        is_cogs_row = 10
        self._is_layout = {"revenue": is_rev_row, "cogs": is_cogs_row}

        # ---- AR ----
        ws.cell(row, 1, "AR Days / 应收账款周转天数")
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_ar"][self.hist_offset + i]
            rev = self.d["hist_revenue"][self.hist_offset + i]
            days = safe_divide(v * 365, rev) if rev else 0.0
            c = ws.cell(row, col, days); c.font = FONT_BLACK; c.number_format = FMT_DAYS
        for i, col in enumerate(self.FCST_COLS):
            c = ws.cell(row, col, f"=Assumptions!${get_column_letter(col)}${assump['ar_days']}")
            c.font = FONT_GREEN; c.number_format = FMT_DAYS
        r["ar_days"] = row; row += 1

        ws.cell(row, 1, "AR Balance / 应收账款余额")
        ws.cell(row, 2, "M")
        self._write_hist_input(ws, row, self.d["hist_ar"],
                                source="富途 BS", ref="应收账款净额")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"='Income Statement'!{cL}{is_rev_row}*{cL}{r['ar_days']}/365"
            c = ws.cell(row, col, f); c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"AR = Revenue x AR Days / 365\n= 'Income Statement'!{cL}{is_rev_row} x {cL}{r['ar_days']} / 365")
        r["ar_bal"] = row; row += 1

        # Δ AR (Prior - Current, CF 影响: 资产增加 => -)
        ws.cell(row, 1, "Δ AR / 应收账款变动 (Prior − Current)")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col); prev_cL = get_column_letter(col - 1)
            if col == self.HIST_COLS[0]:
                ws.cell(row, col, 0).font = FONT_BLACK
            else:
                c = ws.cell(row, col, f"={prev_cL}{r['ar_bal']}-{cL}{r['ar_bal']}")
                c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
                if col in self.FCST_COLS: c.fill = FILL_FORECAST_GREEN
        r["ar_chg"] = row; row += 2

        # ---- Inventory ----
        ws.cell(row, 1, "Inventory Days / 存货周转天数")
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_inv"][self.hist_offset + i]
            cg = self.d["hist_cogs"][self.hist_offset + i]
            days = safe_divide(v * 365, cg) if cg else 0.0
            c = ws.cell(row, col, days); c.font = FONT_BLACK; c.number_format = FMT_DAYS
        for col in self.FCST_COLS:
            c = ws.cell(row, col, f"=Assumptions!${get_column_letter(col)}${assump['inv_days']}")
            c.font = FONT_GREEN; c.number_format = FMT_DAYS
        r["inv_days"] = row; row += 1

        ws.cell(row, 1, "Inventory Balance / 存货余额")
        ws.cell(row, 2, "M")
        self._write_hist_input(ws, row, self.d["hist_inv"], source="富途 BS", ref="存货")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            # 注意: IS COGS 在利润表中按投行惯例显示为**负数** (Less: COGS),
            # 而 Inventory Balance 应保持**正数** (与历史 blue input 口径一致).
            # 若直接 IS!cogs × InvDays / 365 会得到负数余额, 导致 Δ Inv 在 FY1 (Prior 正 − Current 负)
            # 出现虚高跳变 (数倍于后续年度). 用 ABS() 取绝对值消除符号误差.
            f = f"=ABS('Income Statement'!{cL}{is_cogs_row})*{cL}{r['inv_days']}/365"
            c = ws.cell(row, col, f); c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"Inventory = |COGS| × Inv Days / 365\n"
                           f"= ABS('Income Statement'!{cL}{is_cogs_row}) × {cL}{r['inv_days']} / 365\n"
                           "(取绝对值: IS COGS 按投行惯例存为负数, 但 Inventory 余额应保持正数)")
        r["inv_bal"] = row; row += 1

        ws.cell(row, 1, "Δ Inventory / 存货变动 (Prior − Current)")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col); prev_cL = get_column_letter(col - 1)
            if col == self.HIST_COLS[0]:
                ws.cell(row, col, 0).font = FONT_BLACK
            else:
                c = ws.cell(row, col, f"={prev_cL}{r['inv_bal']}-{cL}{r['inv_bal']}")
                c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
                if col in self.FCST_COLS: c.fill = FILL_FORECAST_GREEN
        r["inv_chg"] = row; row += 2

        # ---- AP ----
        ws.cell(row, 1, "AP Days / 应付账款周转天数")
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_ap"][self.hist_offset + i]
            cg = self.d["hist_cogs"][self.hist_offset + i]
            days = safe_divide(v * 365, cg) if cg else 0.0
            c = ws.cell(row, col, days); c.font = FONT_BLACK; c.number_format = FMT_DAYS
        for col in self.FCST_COLS:
            c = ws.cell(row, col, f"=Assumptions!${get_column_letter(col)}${assump['ap_days']}")
            c.font = FONT_GREEN; c.number_format = FMT_DAYS
        r["ap_days"] = row; row += 1

        ws.cell(row, 1, "AP Balance / 应付账款余额")
        ws.cell(row, 2, "M")
        self._write_hist_input(ws, row, self.d["hist_ap"], source="富途 BS", ref="应付账款")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            # 同 Inventory: IS COGS 按投行惯例存为负数, AP 余额应保持正数.
            # 用 ABS() 消除符号误差, 避免 Δ AP 在 FY1 出现虚高跳变.
            f = f"=ABS('Income Statement'!{cL}{is_cogs_row})*{cL}{r['ap_days']}/365"
            c = ws.cell(row, col, f); c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"AP = |COGS| × AP Days / 365\n"
                           f"= ABS('Income Statement'!{cL}{is_cogs_row}) × {cL}{r['ap_days']} / 365\n"
                           "(取绝对值: IS COGS 按投行惯例存为负数, 但 AP 余额应保持正数)")
        r["ap_bal"] = row; row += 1

        ws.cell(row, 1, "Δ AP / 应付账款变动 (Current − Prior)")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col); prev_cL = get_column_letter(col - 1)
            if col == self.HIST_COLS[0]:
                ws.cell(row, col, 0).font = FONT_BLACK
            else:
                c = ws.cell(row, col, f"={cL}{r['ap_bal']}-{prev_cL}{r['ap_bal']}")
                c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
                if col in self.FCST_COLS: c.fill = FILL_FORECAST_GREEN
        r["ap_chg"] = row; row += 2

        # ---- Total ΔNWC ----
        c = ws.cell(row, 1, "Total Δ NWC / 营运资本变动合计 (→ CF)"); c.font = FONT_BOLD
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['ar_chg']}+{cL}{r['inv_chg']}+{cL}{r['ap_chg']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_THIN_BOTTOM
            if col in self.FCST_COLS: cc.fill = FILL_FORECAST_GREEN
        r["dNWC"] = row; row += 1

        self._fmt_column_widths(ws)

    # ==================== Tab 5: D&A / PP&E Schedule ====================
    def _da_schedule(self):
        """PP&E Beg + CapEx - Dep + Other Non-Cash = Net End; 输出 Total D&A 送 IS/CF, PP&E Net 送 BS.

        采用 BS L3 "PP&E & Property" 广口径 (物业厂房+在建工程+投资物业+土地使用权),
        与 BS 保持严格一致 (D&A Schedule 输出的 PP&E End 直接是 BS L3 PP&E & Property).

        关键设计: 存量 (PP&E End) 是**窄口径** (仅物理固定资产 + 土地), 而流量 (CapEx / D&A) 是**广口径**
        (含无形资产/使用权摊销). 直接 Beg + CapEx - D&A = End 无法平衡, 会有非现金差额 (并购/减值/汇兑/
        无形净变动等). 参考 Debt Schedule 与 CF Schedule 的 plug 模式, 引入 "Other Non-Cash Adj" 行:
          - 历史列: Other = End - Beg - CapEx + D&A (反推 plug, 保证 rollforward 精确为 0)
          - 预测列: Other = 0 (假设; 用户可手工填入 M&A / 减值等一次性事项)
          - PP&E End 统一公式: Beg + CapEx - D&A + Other (历史列由 End 反推 Other, 预测列由 Other 加成 End)
        """
        ws = self.wb.create_sheet("D&A Schedule")
        d = self.d; r = self.rows["da"]
        assump = self.rows["assump"]
        is_layout = self._is_layout

        self._write_section_header(ws, 1, "D&A / PP&E SCHEDULE -- 折旧摊销与固定资产表", span=12)
        row = 2
        self._write_column_headers(ws, row); row += 1

        # 采用 hist_l3_ppe_property (BS L3 广口径, 与 BS 严格勾稽)
        ppe_hist = self.d.get("hist_l3_ppe_property") or self.d["hist_ppe"]

        # 提前锁定 5 行的行号 (Beg / CapEx / D&A / Other / End), 避免公式引用时行号未知
        beg_row_idx = row            # PP&E Beginning
        capex_row_idx = row + 1      # (+) CapEx
        da_row_idx = row + 2         # (-) D&A
        other_row_idx = row + 3      # (+/-) Other Non-Cash Adj (plug)
        end_row_idx = row + 4        # PP&E Ending
        check_row_idx = row + 5      # Rollforward Check

        # ---- PP&E Beginning Balance ----
        ws.cell(row, 1, "PP&E Beginning Balance -- 固定资产期初余额 (含物业/在建/土地)")
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            if i == 0:
                # 最老一期: Beg 无法从前期 End 引用, 用 End 反推 (假设 Other=0 于首期)
                # 即 Beg[first] = End[first] - CapEx[first] + D&A[first]
                capex = self.d["hist_capex"][self.hist_offset + i]
                da = self.d["hist_da"][self.hist_offset + i]
                v = ppe_hist[self.hist_offset + i] - capex + da
                c = ws.cell(row, col, v); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY
                c.number_format = FMT_CURRENCY_M
                add_source_comment(c, "反推 (End − CapEx + D&A)",
                                   "最老一期 Beg 由 End 反推 (假设 Other Non-Cash Adj = 0 于首期)")
            else:
                # Beg[i] = 前期 End (跨列同 sheet 引用)
                prev_cL = get_column_letter(col - 1)
                c = ws.cell(row, col, f"={prev_cL}{end_row_idx}")
                c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{end_row_idx}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["ppe_beg"] = row; row += 1

        # ---- CapEx ----
        ws.cell(row, 1, "(+) Capital Expenditure -- 加: 资本开支 (含固定资产+无形资产)")
        ws.cell(row, 2, "M")
        self._write_hist_input(ws, row, self.d["hist_capex"],
                                source="富途 CF", ref="固定资产+无形资产交易净额 (美股 8046+8047 / 港股 5071+5073)")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"='Income Statement'!{cL}{is_layout['revenue']}*Assumptions!{cL}{assump['capex_pct']}"
            c = ws.cell(row, col, f); c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["capex"] = row; row += 1

        # ---- Depreciation & Amortization ----
        # 采用 CF-side D&A (广口径): 涵盖固定资产折旧 + 无形资产摊销 + 使用权资产摊销 + 减值
        # 用途: CF 加回 (NI + D&A + ΔWC = OCF) 与 PP&E 滚动折算
        # IS "Less: D&A" 用 IS-side (窄口径, 仅固定资产折旧, 与 EBIT 口径一致)
        ws.cell(row, 1, "(-) Depreciation & Amortization -- 减: 折旧摊销 (CF 广口径, 含无形/使用权)")
        ws.cell(row, 2, "M")
        self._write_hist_input(ws, row, self.d["hist_da"],
                                source="富途现金流表", ref="折旧摊销及损耗 (广口径, 含无形/使用权/减值)")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"='Income Statement'!{cL}{is_layout['revenue']}*Assumptions!{cL}{assump['da_pct']}"
            c = ws.cell(row, col, f); c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["da"] = row; row += 1

        # ---- Other Non-Cash Adjustments (plug 行, 保证历史 rollforward 精确平衡) ----
        # 历史列: Other = End - Beg - CapEx + D&A (反推), 吸收 API 未捕获的 PP&E 非现金变动:
        #   - 无形资产净购置/剥离 (CF broad CapEx 含无形, BS 窄 PP&E 不含)
        #   - 使用权资产摊销与新增 (IFRS 16 影响)
        #   - 资产减值损失 / 汇兑重估 / 收购并入的 PP&E
        # 预测列: Other = 0 (默认无并购/减值假设), 用户可手工覆盖为并购/减值一次性事项
        # 首期历史 (i=0): 因 Beg 已由反推得到, 此期 Other 应为 0 (与首期 Beg 反推假设一致)
        ws.cell(row, 1, "(+/-) Other Non-Cash Adj -- 其他非现金变动 (plug: 并购/减值/无形/汇兑)")
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            cL = get_column_letter(col)
            if i == 0:
                # 首期: 与首期 Beg 反推假设一致, Other = 0
                c = ws.cell(row, col, 0); c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
                add_comment(c, "首期 Other = 0 (与首期 Beg 由 End 反推的假设一致)")
            else:
                # Other = End - Beg - CapEx + D&A (反推 plug, 保证等式精确)
                f = (f"={cL}{end_row_idx}-{cL}{beg_row_idx}"
                     f"-{cL}{capex_row_idx}+{cL}{da_row_idx}")
                c = ws.cell(row, col, f); c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
                add_comment(c, f"历史 plug: Other = End - Beg - CapEx + D&A\n"
                               f"= {cL}{end_row_idx} - {cL}{beg_row_idx} - {cL}{capex_row_idx} + {cL}{da_row_idx}\n"
                               "\n吸收 API 未捕获的 PP&E 非现金变动: 并购 / 减值 / 无形资产净变动 / 汇兑重估")
        for col in self.FCST_COLS:
            c = ws.cell(row, col, 0); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY
            c.number_format = FMT_CURRENCY_M
            add_source_comment(c, "Assumption", "预测期默认 0, 用户可手工填入并购/减值一次性事项")
        r["other"] = row; row += 1

        # ---- PP&E Ending Balance ----
        # 历史列: 直接引用 hist_l3_ppe_property (广口径, 与 BS L3 完全匹配, 蓝色 blue input)
        # 预测列: 公式 = Beg + CapEx - D&A + Other (含 plug 行, 允许用户手工填入调整)
        c = ws.cell(row, 1, "PP&E Ending Balance (Net) -- 固定资产期末净额 (含物业/在建/土地)")
        c.font = FONT_BOLD
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            v = ppe_hist[self.hist_offset + i]
            cc = ws.cell(row, col, v)
            cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "富途 BS", "L3 PP&E & Property (与 BS 严格匹配, 报告值)")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = (f"={cL}{beg_row_idx}+{cL}{capex_row_idx}-{cL}{da_row_idx}+{cL}{other_row_idx}")
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_THIN_BOTTOM
            cc.fill = FILL_FORECAST_GREEN
            add_comment(cc, "PP&E End = PP&E Beg + CapEx - D&A + Other Non-Cash Adj\n"
                            "(Other 默认 0, 用户可手工填入并购/减值一次性事项)")
        r["ppe_end"] = row; row += 1

        # ---- Rollforward Check ----
        # 历史列: End - Beg - CapEx + D&A - Other = 0 (由 Other plug 保证精确平衡)
        # 预测列: 同上, 也 = 0 (End 由公式生成)
        c = ws.cell(row, 1, "Rollforward Check -- 滚动勾稽 (Beg+CapEx−D&A+Other vs Actual End)")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = (f"={cL}{beg_row_idx}+{cL}{capex_row_idx}-{cL}{da_row_idx}"
                 f"+{cL}{other_row_idx}-{cL}{end_row_idx}")
            cc = ws.cell(row, col, f); cc.font = FONT_BOLD; cc.fill = FILL_MEDIUM_BLUE
            cc.number_format = FMT_CHECK
            add_comment(cc, "Beg + CapEx - D&A + Other - End (应为 0)\n"
                            "历史列 Other 为反推 plug 保证精确 = 0\n"
                            "预测列 End 由公式生成保证精确 = 0")
        r["rollforward_check"] = row; row += 1

        self._fmt_column_widths(ws)

    # ==================== Tab 6: Debt Schedule ====================
    def _debt_schedule(self):
        """LT Debt roll-forward: Beg + Issue - Mandatory - Sweep = End; Interest = Beg x Rate.

        采用 BS L3 "Long-term Debt" (长期借款) 作为 Debt Schedule 追踪对象,
        与 BS L3 严格勾稽 (BS L3 LT Debt End = Debt Schedule debt_end).
        短期借款 (ST Debt) 在 BS 中单独保持 prev year, 不进入 Debt Schedule 滚动。
        """
        ws = self.wb.create_sheet("Debt Schedule")
        d = self.d; r = self.rows["debt"]
        assump = self.rows["assump"]

        self._write_section_header(ws, 1, "DEBT SCHEDULE -- 长期债务表 (LT Debt Only)", span=12)
        row = 2
        self._write_column_headers(ws, row); row += 1

        # 采用 hist_l3_lt_debt (BS L3 长期借款广口径: 长期借款+长期融资租赁+可转换票据)
        debt_hist = self.d.get("hist_l3_lt_debt") or self.d["hist_debt"]

        # ---- Beginning Balance ----
        ws.cell(row, 1, "Beginning Balance -- 期初余额 (长期债务)")
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            if i == 0:
                # 反推 Beg (若无 prior 数据, 用当期 End 减去当期 Iss+Rep+Sweep 反推; 简化为 End)
                v = debt_hist[self.hist_offset + i]
                c = ws.cell(row, col, v); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY
                c.number_format = FMT_CURRENCY_M
                add_source_comment(c, "富途 BS L3", "最老一期用当期 LT Debt End 反推 Beg")
            else:
                prev_cL = get_column_letter(col - 1)
                c = ws.cell(row, col, f"={prev_cL}{row+4}")
                c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row+4}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["debt_beg"] = row; row += 1

        # ---- Issuance ----
        ws.cell(row, 1, "(+) Debt Issuance -- 加: 新增借款")
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS:
            i_ = self.HIST_COLS.index(col)
            if i_ == 0:
                ws.cell(row, col, 0).font = FONT_BLACK
            else:
                # End 在此行下方 3 行 (issuance→mandatory→sweep→debt_end), 用 row+3
                cL = get_column_letter(col)
                f = f"=MAX(0,{cL}{row+3}-{cL}{r['debt_beg']})"
                c = ws.cell(row, col, f); c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
                add_comment(c, "历史 Issuance = MAX(0, End − Beg) (End 增加则视为新增借款)")
        for col in self.FCST_COLS:
            c = ws.cell(row, col, 0); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY
            c.number_format = FMT_CURRENCY_M
            add_source_comment(c, "Assumption", "预测期无新增借款, 用户可修改")
        r["issuance"] = row; row += 1

        # ---- Mandatory Repayment ----
        ws.cell(row, 1, "(-) Mandatory Repayment -- 减: 强制还款")
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS:
            i_ = self.HIST_COLS.index(col)
            if i_ == 0:
                ws.cell(row, col, 0).font = FONT_BLACK
            else:
                # 历史: 若 End < Beg, 差额视为还款 (正数, 与预测期符号一致)
                # 注意: End 在此行下方 2 行 (mandatory→sweep→debt_end), 用 row+2
                cL = get_column_letter(col)
                f = f"=MAX(0,{cL}{r['debt_beg']}-{cL}{row+2})"
                c = ws.cell(row, col, f); c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
                add_comment(c, "历史 Repayment = MAX(0, Beg − End) (End 下降则视为还款)")
        for col in self.FCST_COLS:
            c = ws.cell(row, col, f"=Assumptions!$B${assump['mandatory_repay']}")
            c.font = FONT_GREEN; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["mandatory"] = row; row += 1

        # ---- Cash Sweep ----
        ws.cell(row, 1, "(-) Cash Sweep -- 减: 超额现金还款")
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS:
            ws.cell(row, col, 0).font = FONT_BLACK
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"=Assumptions!$B${assump['cash_sweep']}*{cL}{r['debt_beg']}"
            c = ws.cell(row, col, f); c.font = FONT_GREEN; c.fill = FILL_FORECAST_GREEN
            c.number_format = FMT_CURRENCY_M
        r["sweep"] = row; row += 1

        # ---- Ending Balance ----
        c = ws.cell(row, 1, "Ending Balance -- 期末余额 (长期债务)"); c.font = FONT_BOLD
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            v = debt_hist[self.hist_offset + i]
            cc = ws.cell(row, col, v); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "富途 BS L3", "长期借款 + 长期融资租赁 + 可转换票据 (与 BS L3 严格勾稽)")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['debt_beg']}+{cL}{r['issuance']}-{cL}{r['mandatory']}-{cL}{r['sweep']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.fill = FILL_FORECAST_GREEN
            cc.number_format = FMT_CURRENCY_M; cc.border = BORDER_THIN_BOTTOM
        r["debt_end"] = row; row += 1

        # ---- Interest Expense ----
        ws.cell(row, 1, "Interest Expense (= Beg x Rate) -- 利息费用")
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['debt_beg']}*Assumptions!$B${assump['interest_rate']}"
            c = ws.cell(row, col, f); c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
            if col in self.FCST_COLS: c.fill = FILL_FORECAST_GREEN
            add_comment(c, "Interest = Beginning Debt x Rate (期初余额×利率, 断循环引用)")
        r["interest"] = row; row += 1

        # ---- Rollforward Check ----
        c = ws.cell(row, 1, "Rollforward Check -- 滚动勾稽 (Beg+Iss−Rep−Sweep vs Actual End)")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['debt_beg']}+{cL}{r['issuance']}-{cL}{r['mandatory']}-{cL}{r['sweep']}-{cL}{r['debt_end']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BOLD; cc.fill = FILL_MEDIUM_BLUE
            cc.number_format = FMT_CHECK
            add_comment(cc, "非零通常来自 API 未捕获的债务变动 (如 debt-for-equity swap / 汇兑损益)")
        r["rollforward_check"] = row; row += 1

        self._fmt_column_widths(ws)

    # ==================== Tab 2: Income Statement ====================
    def _income_statement(self):
        """5A + 5E, Margin% 展示行 (indent italics), Interest 引用 Debt Schedule。"""
        ws = self.wb.create_sheet("Income Statement")
        d = self.d; r = self.rows["is"]
        assump = self.rows["assump"]
        debt = self.rows["debt"]

        self._write_section_header(ws, 1, "INCOME STATEMENT -- 利润表", span=12)
        row = 2
        self._write_column_headers(ws, row); row += 1

        # 严格布局: Revenue 必须落在 row 8, COGS 在 row 10 (WC/D&A 引用了此约定)
        # 用循环 pad 到 row 8 (安全, 不依赖手工计数)
        c = ws.cell(row, 1, "Currency: / 币种:")
        ws.cell(row, 2, f"Million {d['reporting_currency']}").font = FONT_ITALIC_GREY
        while row < 8:
            row += 1

        # ---- Revenue (row 8) ----
        assert row == 8, f"IS Layout error: Revenue must be row 8, got {row}"
        c = ws.cell(row, 1, "Revenue / 营业收入"); c.font = FONT_BOLD
        ws.cell(row, 2, "M")
        self._write_hist_input(ws, row, self.d["hist_revenue"],
                                source="富途利润表", ref="总收入/营业总收入")
        for i, col in enumerate(self.FCST_COLS):
            cL = get_column_letter(col); prev_cL = get_column_letter(col - 1)
            f = f"={prev_cL}{row}*(1+Assumptions!{cL}{assump['rev_growth']})"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.fill = FILL_FORECAST_GREEN
            cc.number_format = FMT_CURRENCY_M
            add_comment(cc, f"Revenue(FY{i+1}) = Prior x (1 + Growth%)")
        r["revenue"] = row; row += 1

        # ---- Revenue Growth % 展示行 (indent, italics) ----
        ws.cell(row, 1, "  Revenue Growth % / 营收增长率").font = FONT_ITALIC_GREY
        ws.cell(row, 1).alignment = ALIGN_INDENT
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col); prev_cL = get_column_letter(col - 1)
            if col == self.HIST_COLS[0]:
                ws.cell(row, col, 0).font = FONT_ITALIC_GREY
            else:
                c = ws.cell(row, col, f"={cL}{r['revenue']}/{prev_cL}{r['revenue']}-1")
                c.font = FONT_ITALIC_GREY; c.number_format = FMT_PERCENT
        r["rev_growth_disp"] = row; row += 1

        # ---- COGS (row 10) ----
        assert row == 10, f"IS Layout error: COGS must be row 10, got {row}"
        ws.cell(row, 1, "Less: COGS / 减: 营业成本"); ws.cell(row, 2, "M")
        # 历史用负数展示
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_cogs"][self.hist_offset + i]
            cc = ws.cell(row, col, -abs(v)); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "富途利润表", "营业总成本")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"=-{cL}{r['revenue']}*Assumptions!{cL}{assump['cogs_pct']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK; cc.fill = FILL_FORECAST_GREEN
            cc.number_format = FMT_CURRENCY_M
        r["cogs"] = row; row += 1

        # ---- Gross Profit ----
        c = ws.cell(row, 1, "Gross Profit / 毛利润"); c.font = FONT_BOLD
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['revenue']}+{cL}{r['cogs']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_THIN_BOTTOM
        r["gp"] = row; row += 1

        # ---- Gross Margin % (indent italics) ----
        ws.cell(row, 1, "  Gross Margin % / 毛利率").font = FONT_ITALIC_GREY
        ws.cell(row, 1).alignment = ALIGN_INDENT
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(row, col, f"={cL}{r['gp']}/{cL}{r['revenue']}")
            c.font = FONT_ITALIC_GREY; c.number_format = FMT_PERCENT
        row += 1

        # ---- OpEx ----
        ws.cell(row, 1, "Less: OpEx / 减: 营业费用 (S&M + G&A + R&D)"); ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_opex"][self.hist_offset + i]
            cc = ws.cell(row, col, -abs(v)); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "富途利润表", "营业费用")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"=-{cL}{r['revenue']}*Assumptions!{cL}{assump['opex_pct']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK; cc.fill = FILL_FORECAST_GREEN
            cc.number_format = FMT_CURRENCY_M
        r["opex"] = row; row += 1

        # ---- D&A ----
        # 历史列: 直接引用 IS-side D&A (窄口径, 仅固定资产折旧), 与 EBIT 口径一致
        #         保证 EBITDA = EBIT + |D&A_IS| 不重复扣除或高估
        # 预测列: 从 D&A Schedule 拉取 (Schedule 用 CF-side 广口径滚动 PP&E)
        # 注意: 历史 IS-D&A (5,079) 与 CF-D&A (47,118) 可能有较大差异
        #       差额 = 无形资产摊销 + 使用权资产摊销 + 减值 等, 已包含在 EBIT 但富途利润表未单列
        da_row = self.rows["da"]["da"]
        ws.cell(row, 1, "Less: D&A / 减: 折旧摊销 (IS 窄口径, 仅固定资产折旧)"); ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            v_is = self.d["hist_da_is"][self.hist_offset + i]
            v_cf = self.d["hist_da_cf"][self.hist_offset + i]
            cc = ws.cell(row, col, -abs(v_is) if v_is else 0)
            cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            gap = abs(v_cf) - abs(v_is)
            note = (f"IS-side D&A (富途利润表 '折旧摊销及损耗', 窄口径, 仅固定资产折旧). "
                    f"CF-side D&A = {v_cf:.0f} (广口径, 含无形/使用权/减值). "
                    f"差额 = {gap:.0f} 已包含在 EBIT 内但不单列; CF Statement 加回 D&A 用广口径")
            add_source_comment(cc, "富途利润表", note)
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(row, col, f"=-'D&A Schedule'!{cL}{da_row}")
            c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
            c.fill = FILL_FORECAST_GREEN
            add_comment(c, f"跨表引用: -'D&A Schedule'!{cL}{da_row} (预测期 IS-D&A 与 CF-D&A 假设一致)")
        r["da"] = row; row += 1

        # ---- EBIT ----
        # 历史列: 直接引用 hist_ebit (富途"营业利润") 避免 GP-OpEx-D&A 组装误差 (D&A 已含在成本/费用里)
        # 预测列: Revenue x EBIT Margin (从 Assumptions)
        c = ws.cell(row, 1, "EBIT / 息税前利润"); c.font = FONT_BOLD
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_ebit"][self.hist_offset + i]
            cc = ws.cell(row, col, v); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_THIN_BOTTOM
            add_source_comment(cc, "富途利润表", "营业利润 (直接读取, 避免 D&A 重复扣除)")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['revenue']}*Assumptions!{cL}{assump['ebit_margin']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.fill = FILL_FORECAST_GREEN
            cc.number_format = FMT_CURRENCY_M; cc.border = BORDER_THIN_BOTTOM
            add_comment(cc, "EBIT = Revenue x EBIT Margin (from Assumptions)")
        r["ebit"] = row; row += 1

        # ---- EBIT Margin % ----
        ws.cell(row, 1, "  EBIT Margin % / 息税前利润率").font = FONT_ITALIC_GREY
        ws.cell(row, 1).alignment = ALIGN_INDENT
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(row, col, f"={cL}{r['ebit']}/{cL}{r['revenue']}")
            c.font = FONT_ITALIC_GREY; c.number_format = FMT_PERCENT
        row += 1

        # ---- EBITDA (= EBIT + |D&A|; D&A 存为负) ----
        c = ws.cell(row, 1, "EBITDA / 息税折旧摊销前利润"); c.font = FONT_BOLD
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['ebit']}-{cL}{r['da']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.number_format = FMT_CURRENCY_M
        r["ebitda"] = row; row += 1

        # ---- 分市场展示: EBITDA → EBT 中间项 ----
        # 港股 (HK): 独立展示 5035 融资收入 / 5036 融资成本 / 5037 应占联营公司利润 + Other Income 残差
        # 美股/A股 (US/CN): 富途利润表无对应明细字段, 合并为单行"Non-Operating Items (Net)"
        #   = hist_ebt - hist_ebit (含利息费用/投资净收益/汇兑等所有营业外净额)
        #   预测列 = Revenue x Other Income % - Debt Schedule Interest (显式扣除利息)
        market_type = self.d.get("market_type", "us")

        if market_type == "hk":
            # ---- 港股: 4 行明细展示 ----
            # (+) Finance Income (港股 5035 融资收入)
            ws.cell(row, 1, "(+) Finance Income / 加: 融资收入"); ws.cell(row, 2, "M")
            for i, col in enumerate(self.HIST_COLS):
                v = self.d["hist_finance_income"][self.hist_offset + i]
                cc = ws.cell(row, col, v); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
                cc.number_format = FMT_CURRENCY_M
                add_source_comment(cc, "富途利润表", "融资收入 (港股 fid 5035)")
            for col in self.FCST_COLS:
                prev_cL = get_column_letter(col - 1)
                cc = ws.cell(row, col, f"={prev_cL}{row}"); cc.font = FONT_BLACK; cc.fill = FILL_FORECAST_GREEN
                cc.number_format = FMT_CURRENCY_M
            r["fin_income"] = row; row += 1

            # (-) Finance Cost / Interest Expense (港股 5036 融资成本)
            ws.cell(row, 1, "(-) Finance Cost / Interest Expense / 减: 融资成本 / 利息费用"); ws.cell(row, 2, "M")
            for i, col in enumerate(self.HIST_COLS):
                v = self.d["hist_finance_cost"][self.hist_offset + i]
                cc = ws.cell(row, col, -abs(v) if v else 0); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
                cc.number_format = FMT_CURRENCY_M
                add_source_comment(cc, "富途利润表", "融资成本 (港股 fid 5036)")
            for col in self.FCST_COLS:
                cL = get_column_letter(col)
                c = ws.cell(row, col, f"=-'Debt Schedule'!{cL}{debt['interest']}")
                c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
                c.fill = FILL_FORECAST_GREEN
                add_comment(c, "预测 Interest 从 Debt Schedule (Beg x Rate)")
            r["interest"] = row; row += 1

            # (+) Equity in Affiliates (港股 5037 应占联营公司利润)
            ws.cell(row, 1, "(+) Equity in Affiliates / 加: 应占联营公司利润"); ws.cell(row, 2, "M")
            for i, col in enumerate(self.HIST_COLS):
                v = self.d["hist_equity_affiliate"][self.hist_offset + i]
                cc = ws.cell(row, col, v); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
                cc.number_format = FMT_CURRENCY_M
                add_source_comment(cc, "富途利润表", "应占联营公司利润 (港股 fid 5037)")
            for col in self.FCST_COLS:
                prev_cL = get_column_letter(col - 1)
                cc = ws.cell(row, col, f"={prev_cL}{row}"); cc.font = FONT_BLACK; cc.fill = FILL_FORECAST_GREEN
                cc.number_format = FMT_CURRENCY_M
            r["eq_aff"] = row; row += 1

            # (+) Other Income / (Loss) — 明细外的残差
            ws.cell(row, 1, "(+) Other Income / (Loss) / 加: 其他非营业净收益"); ws.cell(row, 2, "M")
            for i, col in enumerate(self.HIST_COLS):
                v = self.d["hist_other_income"][self.hist_offset + i]
                cc = ws.cell(row, col, v); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
                cc.number_format = FMT_CURRENCY_M
                add_source_comment(cc, "Plug", "hist EBT - EBIT - FinInc + FinCost - EqAff (余量)")
            for col in self.FCST_COLS:
                cL = get_column_letter(col)
                f = f"={cL}{r['revenue']}*Assumptions!{cL}{assump['other_income_pct']}"
                cc = ws.cell(row, col, f); cc.font = FONT_BLACK; cc.fill = FILL_FORECAST_GREEN
                cc.number_format = FMT_CURRENCY_M
                add_comment(cc, "Other Income = Revenue x Other Income % (from Assumptions)")
            r["other_income"] = row; row += 1

            # ---- EBT (港股: EBIT + Fin Inc - Fin Cost + Eq Aff + Other Income) ----
            c = ws.cell(row, 1, "EBT / 税前利润"); c.font = FONT_BOLD
            ws.cell(row, 2, "M")
            for col in self.HIST_COLS + self.FCST_COLS:
                cL = get_column_letter(col)
                f = (f"={cL}{r['ebit']}+{cL}{r['fin_income']}+{cL}{r['interest']}"
                     f"+{cL}{r['eq_aff']}+{cL}{r['other_income']}")
                cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.number_format = FMT_CURRENCY_M
                add_comment(cc, "EBT = EBIT + Finance Income - Finance Cost + Equity Affiliates + Other Income")
            r["ebt"] = row; row += 1
        else:
            # ---- 美股/A股: 合并单行 "Non-Operating Items (Net)" ----
            # 富途 US/CN 利润表在 EBIT 与 EBT 之间无明细拆分 (利息费用/投资净收益/汇兑损益等
            # 全部收纳入"营业外收支净额"或直接从"营业利润→税前利润"跳跃), 故本模型:
            #   历史列 = hist_ebt - hist_ebit (即 hist_other_income, 直接读为 plug)
            #   预测列 = Revenue x Other Income % - Debt Schedule Interest
            #     (显式扣除新增有息负债的利息费用, 保证 Debt Schedule 与 IS 联动)
            ws.cell(row, 1, "(+/-) Non-Operating Items (Net) / 营业外净收支 (含利息/投资收益/汇兑等)")
            ws.cell(row, 2, "M")
            for i, col in enumerate(self.HIST_COLS):
                v = self.d["hist_other_income"][self.hist_offset + i]
                cc = ws.cell(row, col, v); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
                cc.number_format = FMT_CURRENCY_M
                add_source_comment(cc, "富途利润表 (合并口径)",
                                   f"hist EBT - EBIT = {v:.0f} "
                                   f"(涵盖 利息费用 / 投资净收益 / 汇兑 / 公允价值变动 等全部非营业项; "
                                   f"{market_type.upper()} 利润表无独立子项字段, 故合并为单行展示)")
            for col in self.FCST_COLS:
                cL = get_column_letter(col)
                f = (f"={cL}{r['revenue']}*Assumptions!{cL}{assump['other_income_pct']}"
                     f"-'Debt Schedule'!{cL}{debt['interest']}")
                cc = ws.cell(row, col, f); cc.font = FONT_BLACK; cc.fill = FILL_FORECAST_GREEN
                cc.number_format = FMT_CURRENCY_M
                add_comment(cc, "预测: Revenue x Other Income % (from Assumptions) - Debt Schedule Interest")
            # 为下游 EBT 公式统一, 把这一行同时登记为 fin_income / interest / eq_aff / other_income
            # (4 个 key 都指向本行, EBT 公式聚合时相当于只加一次)
            r["non_op"] = row
            row += 1

            # ---- EBT (美股/A股: 直接 = EBIT + Non-Operating Items) ----
            c = ws.cell(row, 1, "EBT / 税前利润"); c.font = FONT_BOLD
            ws.cell(row, 2, "M")
            for col in self.HIST_COLS + self.FCST_COLS:
                cL = get_column_letter(col)
                f = f"={cL}{r['ebit']}+{cL}{r['non_op']}"
                cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.number_format = FMT_CURRENCY_M
                add_comment(cc, "EBT = EBIT + Non-Operating Items (Net)")
            r["ebt"] = row; row += 1


        # ---- Taxes ----
        # 历史列: 直接引用富途"所得税" (蓝色输入, 负号展示), 与富途原始数据一致
        # 预测列: 公式 = -MAX(0, EBT) x Assumptions!Tax Rate (亏损时不计税)
        ws.cell(row, 1, "Less: Taxes / 减: 所得税 (亏损不缴)"); ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_tax"][self.hist_offset + i]
            cc = ws.cell(row, col, -abs(v) if v else 0)
            cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "富途利润表", "所得税 (以负数展示)")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"=-MAX(0,{cL}{r['ebt']})*Assumptions!{cL}{assump['tax_rate']}"
            c = ws.cell(row, col, f); c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
            c.fill = FILL_FORECAST_GREEN
            add_comment(c, "Tax = -MAX(0, EBT) x Tax Rate")
        r["tax"] = row; row += 1

        # ---- Net Income (归母口径) ----
        # 历史列: 直接引用 hist_ni (归母净利润, 蓝色输入), 避免 IS 自算传导误差; 保证与 BS Equity 口径一致
        # 预测列: 公式 = EBT + Tax
        c = ws.cell(row, 1, "Net Income / 净利润"); c.font = FONT_BOLD
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_ni"][self.hist_offset + i]
            cc = ws.cell(row, col, v); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_MEDIUM_BOTTOM
            add_source_comment(cc, "富途利润表", "归属母公司股东净利润 (归母口径, 与 BS Equity 一致)")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['ebt']}+{cL}{r['tax']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.fill = FILL_MEDIUM_BLUE
            cc.number_format = FMT_CURRENCY_M; cc.border = BORDER_MEDIUM_BOTTOM
            add_comment(cc, "预测: NI = EBT + Tax")
        r["ni"] = row; row += 1

        # ---- Net Margin % ----
        ws.cell(row, 1, "  Net Margin % / 净利润率").font = FONT_ITALIC_GREY
        ws.cell(row, 1).alignment = ALIGN_INDENT
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(row, col, f"={cL}{r['ni']}/{cL}{r['revenue']}")
            c.font = FONT_ITALIC_GREY; c.number_format = FMT_PERCENT
        row += 1

        self._fmt_column_widths(ws)

    # ==================== Tab 4: Cash Flow Statement ====================
    def _cash_flow(self):
        """三段式 CF (OCF / CFI / CFF) + 净现金变动 + 现金滚动.
        L1 (OCF/CFI/CFF Net) 从 API 直接抽取 (历史严格勾稽实际报告);
        L2 (二级指标) 精确映射 API fid; 每段以 plug 桶保证 L2 sum = L1.
        预测期用公式驱动 (NI+D&A+ΔWC / -CapEx / Debt+Div+Repurchase)。
        """
        ws = self.wb.create_sheet("Cash Flow")
        d = self.d; r = self.rows["cf"]
        assump = self.rows["assump"]
        is_rows = self.rows["is"]; wc_rows = self.rows["wc"]
        da_rows = self.rows["da"]; debt_rows = self.rows["debt"]

        self._write_section_header(ws, 1, "CASH FLOW STATEMENT -- 现金流量表 (L1/L2 结构)", span=12)
        row = 2
        self._write_column_headers(ws, row); row += 1

        hist_offset = self.hist_offset

        def _hist_write(row_idx, series_key, src, ref):
            """历史列蓝色输入 + 来源 comment"""
            for i, col in enumerate(self.HIST_COLS):
                v = self.d[series_key][hist_offset + i]
                cc = ws.cell(row_idx, col, v)
                cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
                cc.number_format = FMT_CURRENCY_M
                add_source_comment(cc, src, ref)

        # ==============================================================
        # ---- OPERATING CASH FLOW ----
        # ==============================================================
        c = ws.cell(row, 1, "OPERATING CASH FLOW -- 经营活动现金流 (L1)")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12); row += 1

        # L2 #1: Net Income (from IS, cross-sheet ref)
        ws.cell(row, 1, "  Net Income -- 净利润")
        ws.cell(row, 2, "M")
        # 历史用 API 的 CF NI (US 8017 持续经营净收入 / HK 5003 除税前利润)
        # 但为保持 CF NI 与 IS NI 一致, 我们用 IS Net Income 跨表引用 (US 场景)
        # HK 场景 CF NI 是税前利润, 与 IS NI 差 tax, plug 桶会吸收
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(row, col, f"='Income Statement'!{cL}{is_rows['ni']}")
            c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"跨表引用: 'Income Statement'!{cL}{is_rows['ni']}")
        r["ni"] = row; row += 1

        # L2 #2: D&A (from D&A Schedule)
        ws.cell(row, 1, "  (+) D&A -- 加: 折旧摊销")
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(row, col, f"='D&A Schedule'!{cL}{da_rows['da']}")
            c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "跨表引用: D&A Schedule (CF 广口径, 含无形/使用权/减值)")
        r["da"] = row; row += 1

        # L2 #3: (+/-) ΔWorking Capital (aggregated for hist; broken out for fcst)
        # 历史列: 用 API "营运资金变化" (US 8028 / HK 5034) 作合并 blue input
        # 预测列: 保持 Δ AR + Δ Inv + Δ AP 分项 (WC schedule)
        ws.cell(row, 1, "  (+/-) Δ Working Capital -- 营运资金变动")
        ws.cell(row, 2, "M")
        _hist_write(row, "hist_cf_wc_change", "富途 CF",
                    "美股 8028 营运资金变化 (含 AR/Inv/AP/其他 WC); 港股 5034 营运资金变动项目")
        # 预测: WC schedule 的 Δ AR + Δ Inv + Δ AP 三项之和
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = (f"='Working Capital'!{cL}{wc_rows['ar_chg']}"
                 f"+'Working Capital'!{cL}{wc_rows['inv_chg']}"
                 f"+'Working Capital'!{cL}{wc_rows['ap_chg']}")
            c = ws.cell(row, col, f); c.font = FONT_GREEN
            c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: Working Capital Δ AR + Δ Inv + Δ AP (Days-driven)")
        r["wc_change"] = row; r["ar_chg"] = row; r["inv_chg"] = row; r["ap_chg"] = row  # legacy alias
        row += 1

        # L2 #4: (+/-) Other Non-Cash Adjustments (plug)
        # 历史列 = OCF Total - NI - D&A - ΔWC (吸收递延所得税/公允价值变动/汇兑等)
        ws.cell(row, 1, "  (+/-) Other Non-Cash Adjustments -- 其他非现金调整 (plug)")
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            # 历史 plug = OCF Total - (CF NI + D&A + WC Change)
            # 注意: 我们 IS NI 与 CF NI 可能有差, 需用 API CF NI 计算 plug
            # 但为在 Excel 中让 plug + NI + D&A + WC = OCF Total, 使用 IS NI 会引入 NI 差异
            # 简化: hist plug = OCF Total - (IS NI + D&A + WC Change)
            #   -> 差异被 plug 吸收, 保证 OCF Total = 各项 sum
            ocf = self.d["hist_ocf_total"][hist_offset + i]
            ni  = self.d["hist_ni"][hist_offset + i]
            da  = self.d["hist_da_cf"][hist_offset + i]
            wc  = self.d["hist_cf_wc_change"][hist_offset + i]
            plug = ocf - ni - da - wc
            cc = ws.cell(row, col, plug); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "Plug (OCF−NI−D&A−ΔWC)",
                               "吸收 API 未拆解的非现金调整: 递延所得税/公允价值变动/汇兑/减值/CF与IS NI差 等")
        # 预测列: 简化为 0 (预测期假设无特殊非现金项)
        for col in self.FCST_COLS:
            c = ws.cell(row, col, 0); c.font = FONT_BLACK
            c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: 简化为 0 (无非常规非现金项假设)")
        r["ocf_plug"] = row; row += 1

        # Net OCF (L1)
        c = ws.cell(row, 1, "Net Operating Cash Flow -- 经营活动现金流净额 (L1)")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['ni']}+{cL}{r['da']}+{cL}{r['wc_change']}+{cL}{r['ocf_plug']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD
            cc.fill = FILL_MEDIUM_BLUE; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_MEDIUM_BOTTOM
            add_comment(cc, "L1 OCF = NI + D&A + ΔWC + Other Non-Cash Adj")
        r["ocf"] = row; row += 2

        # ==============================================================
        # ---- INVESTING CASH FLOW ----
        # ==============================================================
        c = ws.cell(row, 1, "INVESTING CASH FLOW -- 投资活动现金流 (L1)")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12); row += 1

        # L2 #1: CapEx (from D&A Schedule for hist/fcst)
        ws.cell(row, 1, "  (-) Capital Expenditure -- 减: 资本开支")
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(row, col, f"=-'D&A Schedule'!{cL}{da_rows['capex']}")
            c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "跨表引用: -D&A Schedule CapEx (美股 8046+8047 / 港股 5071+5073)")
        r["capex"] = row; row += 1

        # L2 #2: Other Investing (plug)
        ws.cell(row, 1, "  (+/-) Other Investing Activities -- 其他投资活动 (plug)")
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            cfi_total = self.d["hist_cfi_total"][hist_offset + i]
            capex_val = -abs(self.d["hist_capex"][hist_offset + i])  # 负号
            plug = cfi_total - capex_val
            cc = ws.cell(row, col, plug); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "Plug (CFI−CapEx)",
                               "美股 8049 业务交易+8051 投资产品+8054 其他; 港股 5070 出售+5074 出售附属+5076-5083 投资交易/利息/股息")
        for col in self.FCST_COLS:
            c = ws.cell(row, col, 0); c.font = FONT_BLACK
            c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: 简化为 0 (无非常规投资活动假设)")
        r["cfi_other"] = row; row += 1

        # Net CFI (L1)
        c = ws.cell(row, 1, "Net Investing Cash Flow -- 投资活动现金流净额 (L1)")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['capex']}+{cL}{r['cfi_other']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD
            cc.fill = FILL_MEDIUM_BLUE; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_MEDIUM_BOTTOM
            add_comment(cc, "L1 CFI = -CapEx + Other Investing")
        r["cfi"] = row; row += 2

        # ==============================================================
        # ---- FINANCING CASH FLOW ----
        # ==============================================================
        c = ws.cell(row, 1, "FINANCING CASH FLOW -- 融资活动现金流 (L1)")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12); row += 1

        # L2 #1: Debt Issuance
        ws.cell(row, 1, "  (+) Debt Issuance -- 加: 新增借款")
        ws.cell(row, 2, "M")
        # 历史: API "债务发行/偿还净额" 分离为 issue (正) / repay (负); 简化用 hist_cf_debt_net (含符号)
        # 我们把净额都放到 Debt Issuance 行, Debt Repayment 用 Debt Schedule 预测
        # 更精确: 历史列独立求正/负部分. 简化: 用净额, comment 说明
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_cf_debt_net"][hist_offset + i]
            issue_val = max(0.0, v)  # 只取正部分
            cc = ws.cell(row, col, issue_val); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "富途 CF",
                               "美股 8058 债务发行/偿还净额 (取正部分); 港股 5087 新增借款")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(row, col, f"='Debt Schedule'!{cL}{debt_rows['issuance']}")
            c.font = FONT_GREEN; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["issuance"] = row; row += 1

        # L2 #2: Debt Repayment
        ws.cell(row, 1, "  (-) Debt Repayment -- 减: 债务偿还 (强制+超额)")
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_cf_debt_net"][hist_offset + i]
            repay_val = min(0.0, v)  # 只取负部分
            cc = ws.cell(row, col, repay_val); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "富途 CF",
                               "美股 8058 债务发行/偿还净额 (取负部分); 港股 5088 偿还借款")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"=-('Debt Schedule'!{cL}{debt_rows['mandatory']}+'Debt Schedule'!{cL}{debt_rows['sweep']})"
            c = ws.cell(row, col, f); c.font = FONT_GREEN
            c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["repayment"] = row; row += 1

        # L2 #3: Dividends Paid
        ws.cell(row, 1, "  (-) Dividends Paid -- 减: 股利支付")
        ws.cell(row, 2, "M")
        _hist_write(row, "hist_cf_dividends", "富途 CF",
                    "美股 8061 现金股利支付; 港股 5094 已付股息-融资")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"=-MAX(0,{cL}{r['ni']})*Assumptions!$B${assump['div_payout']}"
            c = ws.cell(row, col, f); c.font = FONT_BLACK
            c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: Dividends = -MAX(0, NI) x Payout Ratio (from Assumptions)")
        r["dividends"] = row; row += 1

        # L2 #4: Share Repurchases (from equity_net)
        ws.cell(row, 1, "  (-) Share Repurchases / Issuance -- 减: 股份回购/发行")
        ws.cell(row, 2, "M")
        _hist_write(row, "hist_cf_equity_net", "富途 CF",
                    "美股 8059 普通股发行/回购净额 (含符号: 回购为负); 港股 5089 发行股份")
        for col in self.FCST_COLS:
            c = ws.cell(row, col, f"=-Assumptions!$B${assump['repurchase']}")
            c.font = FONT_GREEN; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["repurchase"] = row; row += 1

        # L2 #5: Other Financing (plug)
        ws.cell(row, 1, "  (+/-) Other Financing Activities -- 其他融资活动 (plug)")
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            cff_total = self.d["hist_cff_total"][hist_offset + i]
            debt_net = self.d["hist_cf_debt_net"][hist_offset + i]
            eq_net = self.d["hist_cf_equity_net"][hist_offset + i]
            div = self.d["hist_cf_dividends"][hist_offset + i]
            plug = cff_total - debt_net - eq_net - div
            cc = ws.cell(row, col, plug); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "Plug (CFF−Debt−Equity−Div)",
                               "美股 8065 其他融资活动; 港股 5090 发行债券+5093 已付利息+5091 发行费用+5096 融资其他")
        for col in self.FCST_COLS:
            c = ws.cell(row, col, 0); c.font = FONT_BLACK
            c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["cff_other"] = row; row += 1

        # Net CFF (L1)
        c = ws.cell(row, 1, "Net Financing Cash Flow -- 融资活动现金流净额 (L1)")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = (f"={cL}{r['issuance']}+{cL}{r['repayment']}+{cL}{r['dividends']}"
                 f"+{cL}{r['repurchase']}+{cL}{r['cff_other']}")
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD
            cc.fill = FILL_MEDIUM_BLUE; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_MEDIUM_BOTTOM
            add_comment(cc, "L1 CFF = Debt Iss + Debt Repay + Dividends + Repurchase + Other Fin")
        r["cff"] = row; row += 2

        # ---- NET CHANGE IN CASH ----
        ws.cell(row, 1, "Net Change in Cash -- 现金净变动"); ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['ocf']}+{cL}{r['cfi']}+{cL}{r['cff']}"
            c = ws.cell(row, col, f); c.font = FONT_BLACK_BOLD; c.number_format = FMT_CURRENCY_M
        r["net_change"] = row; row += 1

        ws.cell(row, 1, "Beginning Cash -- 期初现金"); ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            if i == 0:
                cL = get_column_letter(col)
                f = f"={cL}{row+1}-{cL}{r['net_change']}"
                c = ws.cell(row, col, f); c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
            else:
                prev_cL = get_column_letter(col - 1)
                c = ws.cell(row, col, f"={prev_cL}{row+1}")
                c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row+1}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["cash_beg"] = row; row += 1

        # Ending Cash
        c = ws.cell(row, 1, "Ending Cash -- 期末现金"); c.font = FONT_BOLD
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_cash"][hist_offset + i]
            cc = ws.cell(row, col, v); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "富途 BS", "现金及等价物 + 短期投资 + 定期存款 (与 BS 保持一致)")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['cash_beg']}+{cL}{r['net_change']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.fill = FILL_MEDIUM_BLUE
            cc.number_format = FMT_CURRENCY_M; cc.border = BORDER_MEDIUM_BOTTOM
            add_comment(cc, "Ending Cash = Beg + Net Change (=> BS Cash 核心勾稽)")
        r["cash_end"] = row; row += 1

        self._fmt_column_widths(ws)

    # ==================== Tab 3: Balance Sheet (3-tier: L1 → L2 → L3) ====================
    def _balance_sheet(self):
        """三级结构 BS: L1 (Assets/Liab/Equity) → L2 (Current/NC × Assets/Liab, Parent/Minority Equity)
        → L3 (3-5 aggregated buckets per group, 按变现难度/偿还优先级 排序, plug bucket 兜底).

        历史列: L2 直接从 API 抽取 (蓝色输入), L3 从 API + plug 兜底, 保证 L3 sum = L2 exact,
                Balance Check = 0.
        预测列: L2 用 简单驱动 (Cash/AR/AP/PPE/Debt 与其他 Schedule 联动; 其余项保持 prev year);
                L3 中 "Other" 桶作为 plug 强制 L3 sum = L2.
        """
        ws = self.wb.create_sheet("Balance Sheet")
        d = self.d; r = self.rows["bs"]
        wc_rows = self.rows["wc"]; da_rows = self.rows["da"]
        debt_rows = self.rows["debt"]; is_rows = self.rows["is"]
        cf_rows = self.rows["cf"]; assump = self.rows["assump"]

        self._write_section_header(ws, 1, "BALANCE SHEET -- 资产负债表 (三级结构 L1/L2/L3)", span=12)
        row = 2
        self._write_column_headers(ws, row); row += 1

        # 提前保存 hist_offset 以便简化下方引用
        hist_offset = self.hist_offset

        def _write_hist_l3(row_idx, hist_key, source_note, aggregation_note):
            """写入 L3 桶的历史列 (蓝色输入, 附带来源与合并说明)."""
            for i, col in enumerate(self.HIST_COLS):
                v = self.d[hist_key][hist_offset + i]
                cc = ws.cell(row_idx, col, v)
                cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
                cc.number_format = FMT_CURRENCY_M
                add_source_comment(cc, source_note, aggregation_note)

        # ==============================================================
        # ==== ASSETS (L1) ====
        # ==============================================================
        c = ws.cell(row, 1, "ASSETS -- 资产 (L1)"); c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12); row += 1

        # ---- Current Assets (L2) ----
        c = ws.cell(row, 1, "  Current Assets -- 流动资产 (L2)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12); row += 1

        # L3-CA #1: Cash & ST Investments (from CF Ending Cash for consistency)
        ws.cell(row, 1, "    Cash & ST Investments -- 现金及短期投资")
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(row, col, f"='Cash Flow'!{cL}{cf_rows['cash_end']}")
            c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "CF Ending Cash (含 现金/短期投资/定期存款流动) — 与 CF 勾稽")
        r["cash"] = row; r["l3_cash_sti"] = row; row += 1

        # L3-CA #2: AR & Prepayments (Days-driven forecast; hist from API)
        ws.cell(row, 1, "    AR & Prepayments -- 应收账款及预付款")
        ws.cell(row, 2, "M")
        # Hist: 用 API L3 桶值 (含应收款项父项 + 预付, 与 L2 CA total 严格勾稽)
        _write_hist_l3(row, "hist_l3_ar_prepaid", "富途 BS",
                       "美股 8006 应收款项+8016 预付款项; 港股 5007 应收+5014 预付按金")
        # Fcst: 用 Working Capital AR (Days-driven, 预付款保持 prev)
        for col in self.FCST_COLS:
            cL = get_column_letter(col); prev_cL = get_column_letter(col - 1)
            f = f"='Working Capital'!{cL}{wc_rows['ar_bal']}"
            c = ws.cell(row, col, f); c.font = FONT_GREEN
            c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: Working Capital AR (Days-driven)\n预付款忽略, 简化模型")
        r["ar"] = row; r["l3_ar_prepaid"] = row; row += 1

        # L3-CA #3: Inventory
        ws.cell(row, 1, "    Inventory -- 存货")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_inventory", "富途 BS", "存货 (美股 8017 / 港股 5019)")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"='Working Capital'!{cL}{wc_rows['inv_bal']}"
            c = ws.cell(row, col, f); c.font = FONT_GREEN
            c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["inv"] = row; r["l3_inventory"] = row; row += 1

        # L3-CA #4: Other Current Assets (plug 兜底: L2 CA - 前 3 项)
        ws.cell(row, 1, "    Other Current Assets -- 其他流动资产 (plug)")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_other_ca", "富途 BS (合并 plug)",
                       "美股 受限制现金/递延资产/其他流动; 港股 已抵押存款/流动特殊项; "
                       "计算为: L2 流动资产合计 − 前 3 桶, 保证 L3 sum = L2")
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: 保持 prev year (Other CA 不随主假设驱动)")
        r["other_ca"] = row; r["l3_other_ca"] = row; row += 1

        # Total Current Assets (L2, 直接 = L3 sum)
        c = ws.cell(row, 1, "  Total Current Assets -- 流动资产合计 (L2)")
        c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['l3_cash_sti']}+{cL}{r['l3_ar_prepaid']}+{cL}{r['l3_inventory']}+{cL}{r['l3_other_ca']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_THIN_BOTTOM
            add_comment(cc, "L2 = ΣL3 (4 桶合计)")
        r["ca"] = row; row += 1

        # ---- Non-Current Assets (L2) ----
        c = ws.cell(row, 1, "  Non-Current Assets -- 非流动资产 (L2)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12); row += 1

        # L3-NCA #1: PP&E & Property (from D&A Schedule for hist; Schedule for fcst)
        ws.cell(row, 1, "    PP&E & Property -- 固定资产及物业")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_ppe_property", "富途 BS",
                       "美股 8024 固定资产净额; 港股 5031 物业厂房+5032 在建工程+5033 投资物业+5034 土地使用权")
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(row, col, f"='D&A Schedule'!{cL}{da_rows['ppe_end']}")
            c.font = FONT_GREEN; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: D&A Schedule PP&E End (Roll-forward)")
        r["ppe"] = row; r["l3_ppe_property"] = row; row += 1

        # L3-NCA #2: Long-term Investments
        ws.cell(row, 1, "    Long-term Investments -- 长期投资")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_lt_investments", "富途 BS",
                       "美股 8028 总投资+8035 金融资产; 港股 5037 FVTPL+5039 长期投资+5050 联营+5053 合营+5054 定期存款(非流)")
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: 保持 prev year")
        r["l3_lt_investments"] = row; row += 1

        # L3-NCA #3: Goodwill & Intangibles
        ws.cell(row, 1, "    Goodwill & Intangibles -- 商誉及无形资产")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_goodwill_intangibles", "富途 BS",
                       "美股 8039 商誉及其他无形资产; 港股 5046 无形资产")
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: 保持 prev year (无收购摊销假设)")
        r["intangible"] = row; r["l3_goodwill_intangibles"] = row; row += 1

        # L3-NCA #4: Other Non-Current Assets (plug)
        ws.cell(row, 1, "    Other Non-Current Assets -- 其他非流动资产 (plug)")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_other_nca", "富途 BS (合并 plug)",
                       "美股 长期应收/预付/递延税项/其他; 港股 预付款项(非流)/递延税项/特殊; "
                       "计算为: L2 非流动资产合计 − 前 3 桶")
        # 预测列的 plug 需要引用下方 TL 与 TE, 稍后回填
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["other_nca"] = row; r["l3_other_nca"] = row; row += 1

        # Total Non-Current Assets (L2 = ΣL3)
        c = ws.cell(row, 1, "  Total Non-Current Assets -- 非流动资产合计 (L2)")
        c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['l3_ppe_property']}+{cL}{r['l3_lt_investments']}+{cL}{r['l3_goodwill_intangibles']}+{cL}{r['l3_other_nca']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_THIN_BOTTOM
        r["nca"] = row; row += 1

        # Total Assets (L1)
        c = ws.cell(row, 1, "TOTAL ASSETS -- 资产总计 (L1)")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['ca']}+{cL}{r['nca']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD
            cc.fill = FILL_MEDIUM_BLUE; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_MEDIUM_BOTTOM
        r["ta"] = row; row += 2

        # ==============================================================
        # ==== LIABILITIES (L1) ====
        # ==============================================================
        c = ws.cell(row, 1, "LIABILITIES -- 负债 (L1)"); c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12); row += 1

        # ---- Current Liabilities (L2) ----
        c = ws.cell(row, 1, "  Current Liabilities -- 流动负债 (L2)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12); row += 1

        # L3-CL #1: Short-term Debt (计息负债, 偿还优先级最高)
        ws.cell(row, 1, "    Short-term Debt -- 短期借款")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_st_debt", "富途 BS",
                       "美股 8057 短期借款与融资租赁负债 (父项); 港股 5070 银行贷款及透支+5072 短期融资租赁")
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: 保持 prev year (Debt Schedule 追踪长期债务变动)")
        r["l3_st_debt"] = row; row += 1

        # L3-CL #2: AP & Accrued Expenses
        ws.cell(row, 1, "    AP & Accrued Expenses -- 应付账款及应计费用")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_ap_accrued", "富途 BS",
                       "美股 8050 应付款项(除税项)+8056 应计费用; 港股 5062 应付账款+5066 应付票据+5067 其他应付+应计+5068 预收")
        # Fcst: 用 Working Capital AP (Days-driven)
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"='Working Capital'!{cL}{wc_rows['ap_bal']}"
            c = ws.cell(row, col, f); c.font = FONT_GREEN
            c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: Working Capital AP (Days-driven)")
        r["ap"] = row; r["l3_ap_accrued"] = row; row += 1

        # L3-CL #3: Taxes Payable
        ws.cell(row, 1, "    Taxes & Dividends Payable -- 应交税费及应付股利")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_taxes_payable", "富途 BS",
                       "美股 8052 -应交税费 (从应付款项 8050 中分离); 港股 5063 应交税费+5064 应付股利")
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["l3_taxes_payable"] = row; row += 1

        # L3-CL #4: Other Current Liab (plug)
        ws.cell(row, 1, "    Other Current Liabilities -- 其他流动负债 (plug)")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_other_cl", "富途 BS (合并 plug)",
                       "美股 8063 递延负债(流)+8064 其他流动负债; 港股 5083 递延收入(流); "
                       "计算为: L2 流动负债合计 − 前 3 桶")
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["ocl"] = row; r["l3_other_cl"] = row; row += 1

        # Total Current Liab (L2)
        c = ws.cell(row, 1, "  Total Current Liabilities -- 流动负债合计 (L2)")
        c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['l3_st_debt']}+{cL}{r['l3_ap_accrued']}+{cL}{r['l3_taxes_payable']}+{cL}{r['l3_other_cl']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_THIN_BOTTOM
        r["cl"] = row; row += 1

        # ---- Non-Current Liab (L2) ----
        c = ws.cell(row, 1, "  Non-Current Liabilities -- 非流动负债 (L2)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12); row += 1

        # L3-NCL #1: Long-term Debt
        ws.cell(row, 1, "    Long-term Debt -- 长期借款")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_lt_debt", "富途 BS",
                       "美股 8068 长期借款与租赁负债 (父项); 港股 5091 长期银行贷款+5093 长期融资租赁+5104 可转换票据")
        # 预测: 引用 Debt Schedule (端账)
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(row, col, f"='Debt Schedule'!{cL}{debt_rows['debt_end']}")
            c.font = FONT_GREEN; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: Debt Schedule 期末债务余额 (Beg + Iss - Repay - Sweep)")
        r["debt"] = row; r["l3_lt_debt"] = row; row += 1

        # L3-NCL #2: Deferred Liab
        ws.cell(row, 1, "    Deferred Liabilities -- 递延负债")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_deferred_liab", "富途 BS",
                       "美股 8074 递延负债(非流动); 港股 5101 递延税项+5102 递延收入(非流)")
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["l3_deferred_liab"] = row; row += 1

        # L3-NCL #3: Other Non-Current Liab (BS plug of last resort)
        ws.cell(row, 1, "    Other Non-Current Liabilities -- 其他非流动负债 (BS plug)")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_other_ncl", "富途 BS (合并 plug)",
                       "美股 8066 长期应付+8067 长期应计+8080 其他; 港股 5089 长期应付+5106 其他; "
                       "计算为: L2 非流动负债合计 − 前 2 桶")
        # 预测列: 稍后作为 BS 平衡 plug 回填, 先占位
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["oncl"] = row; r["l3_other_ncl"] = row; row += 1

        # Total Non-Current Liab (L2)
        c = ws.cell(row, 1, "  Total Non-Current Liabilities -- 非流动负债合计 (L2)")
        c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['l3_lt_debt']}+{cL}{r['l3_deferred_liab']}+{cL}{r['l3_other_ncl']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_THIN_BOTTOM
        r["ncl"] = row; row += 1

        # Total Liabilities (L1)
        c = ws.cell(row, 1, "TOTAL LIABILITIES -- 负债合计 (L1)")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['cl']}+{cL}{r['ncl']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD
            cc.fill = FILL_MEDIUM_BLUE; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_MEDIUM_BOTTOM
        r["tl"] = row; row += 2

        # ==============================================================
        # ==== EQUITY (L1) ====
        # ==============================================================
        c = ws.cell(row, 1, "EQUITY -- 股东权益 (L1)"); c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12); row += 1

        # ---- Parent Equity (L2, 归属于母公司) ----
        c = ws.cell(row, 1, "  Parent Equity -- 归属于母公司股东权益 (L2)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12); row += 1

        # L3-EQ #1: Common Stock + APIC
        ws.cell(row, 1, "    Common Stock + APIC -- 股本及资本公积")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_common_stock_apic", "富途 BS",
                       "美股 8086 股本+8090 资本公积; 港股 5111 股本+5112 股本溢价")
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: 保持 prev year (无增发/回购简化)")
        r["cs"] = row; r["l3_common_stock_apic"] = row; row += 1

        # L3-EQ #2: Retained Earnings (rolls with NI - Dividends)
        ws.cell(row, 1, "    Retained Earnings -- 留存收益")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_retained_earnings", "富途 BS",
                       "美股 8091 留存收益; 港股 5115 保留溢利")
        for col in self.FCST_COLS:
            cL = get_column_letter(col); prev_cL = get_column_letter(col - 1)
            f = f"={prev_cL}{row}+'Income Statement'!{cL}{is_rows['ni']}+'Cash Flow'!{cL}{cf_rows['dividends']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK; cc.fill = FILL_FORECAST_GREEN
            cc.number_format = FMT_CURRENCY_M
            add_comment(cc, "RE = Prior + NI - Dividends (Dividends 在 CF 已带负号)")
        r["re"] = row; r["l3_retained_earnings"] = row; row += 1

        # L3-EQ #3: Other Equity (库存股/其他储备/特殊项)
        ws.cell(row, 1, "    Other Equity -- 其他权益 (库存股/储备/plug)")
        ws.cell(row, 2, "M")
        _write_hist_l3(row, "hist_l3_other_equity", "富途 BS (合并 plug)",
                       "美股 8092 库存股+8093 不影响RE损益+8094 其他股本权益; 港股 5121 其他储备+5123 特殊项; "
                       "计算为: L2 归属母公司权益 − CS − RE")
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        r["l3_other_equity"] = row; row += 1

        # Parent Equity Total (L2)
        c = ws.cell(row, 1, "  Total Parent Equity -- 归属母公司权益合计 (L2)")
        c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['l3_common_stock_apic']}+{cL}{r['l3_retained_earnings']}+{cL}{r['l3_other_equity']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_THIN_BOTTOM
        r["parent_eq"] = row; row += 1

        # ---- Minority Interest (L2) ----
        ws.cell(row, 1, "  Minority Interest -- 少数股东权益 (L2)").font = FONT_BOLD
        ws.cell(row, 1).fill = FILL_LIGHT_BLUE
        ws.cell(row, 2, "M")
        for i, col in enumerate(self.HIST_COLS):
            v = self.d["hist_minority_eq"][hist_offset + i]
            cc = ws.cell(row, col, v); cc.font = FONT_BLUE; cc.fill = FILL_INPUT_GREY
            cc.number_format = FMT_CURRENCY_M
            add_source_comment(cc, "富途 BS", "美股 8095 少数股东权益 / 港股 5125 少数股东权益 (无则用 股东权益合计 - 归属母公司)")
        for col in self.FCST_COLS:
            prev_cL = get_column_letter(col - 1)
            c = ws.cell(row, col, f"={prev_cL}{row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, "预测: 保持 prev year")
        r["minority_eq"] = row; row += 1

        # Total Equity (L1 = Parent + Minority)
        c = ws.cell(row, 1, "TOTAL EQUITY -- 股东权益合计 (L1)")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['parent_eq']}+{cL}{r['minority_eq']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD
            cc.fill = FILL_MEDIUM_BLUE; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_MEDIUM_BOTTOM
        r["te"] = row; row += 2

        # ==============================================================
        # ==== 回填: 预测期 plug (强制 BS 平衡) ====
        # 由于历史列 L3 已 sum = L2 (extraction 时已 plug 调整), 历史 Balance Check 自动 = 0
        # 预测列需要一个 last-resort plug: 用 Other Non-Current Liab 吸收剩余不平衡
        # 公式: ONCL_fcst = TA - CL - LT Debt - Deferred Liab - Total Equity
        # ==============================================================
        for col in self.FCST_COLS:
            cL = get_column_letter(col)
            # 让 TL = TA - TE, 故 ONCL = TA - TE - CL - LT_debt - Deferred_Liab
            f = (f"={cL}{r['ta']}-{cL}{r['te']}-{cL}{r['cl']}"
                 f"-{cL}{r['l3_lt_debt']}-{cL}{r['l3_deferred_liab']}")
            cc = ws.cell(r["oncl"], col, f)
            cc.font = FONT_BLACK; cc.fill = FILL_FORECAST_GREEN; cc.number_format = FMT_CURRENCY_M
            add_comment(cc, "预测 plug: ONCL = TA − TE − CL − LT Debt − Deferred Liab (强制 BS 平衡)")

        # 同时预测期 Other CA/NCA/CL/Other Equity 保持 prev; 由于 CS/RE 会变化,
        # Other Equity 也可能需要吸收剩余. 简化: RE 独立滚动, CS/Other 保持 prev.

        # Total Liabilities & Equity
        c = ws.cell(row, 1, "Total Liabilities & Equity -- 负债及股东权益合计")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['tl']}+{cL}{r['te']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BLACK_BOLD
            cc.fill = FILL_MEDIUM_BLUE; cc.number_format = FMT_CURRENCY_M
            cc.border = BORDER_MEDIUM_BOTTOM
        r["tle"] = row; row += 2

        # ==== CHECKS ====
        c = ws.cell(row, 1, "Balance Check (TA − TL&E) -- 勾稽校验")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['ta']}-{cL}{r['tle']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BOLD; cc.fill = FILL_MEDIUM_BLUE
            cc.number_format = FMT_CHECK
            add_comment(cc, "必须为 0. 历史列由 L3 plug 自动保证; 预测列由 ONCL plug 强制平衡")
        r["balance_check"] = row; row += 1

        c = ws.cell(row, 1, "Cash Tie-Out (BS Cash − CF End) -- 现金勾稽")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE
        ws.cell(row, 2, "M")
        for col in self.HIST_COLS + self.FCST_COLS:
            cL = get_column_letter(col)
            f = f"={cL}{r['cash']}-'Cash Flow'!{cL}{cf_rows['cash_end']}"
            cc = ws.cell(row, col, f); cc.font = FONT_BOLD; cc.fill = FILL_MEDIUM_BLUE
            cc.number_format = FMT_CHECK
        r["cash_tieout"] = row; row += 1

        self._fmt_column_widths(ws)

    # ==================== Tab 8: DCF Valuation ====================
    def _dcf(self):
        """基于 3-Statement 已有历史 + 预测数据构建 DCF 估值 sheet.

        与 dcf-model skill 计算逻辑一致 (Header/Case Selector/Market Data/3 情景块/
        Selected Consolidation/Financials/FCF/Discounting/Valuation Summary/3 张 5x5 敏感性表).

        关键差异: 历史列 Revenue/EBIT/D&A/CapEx/ΔNWC 全部通过 **跨 Sheet 公式**
        引用 Income Statement / D&A Schedule / Working Capital 的已有行 (绿色字体),
        避免二次抽取产生数据分叉.

        Net Debt / Gross Debt 从 Balance Sheet L3 (Short-term Debt + Long-term Debt - Cash) 引用,
        Shares/Stock Price/FX 从 Assumptions 引用 (与 3-Statement 主体同源).

        列布局 (与 3-Statement 保持一致):
          A=标签, B..F=5 期历史 (右对齐), G..K=FY1..FY5 预测
        (与独立 dcf-model 的 B..F/G..K 相同, 便于用户对照)
        """
        ws = self.wb.create_sheet("DCF")
        d = self.d
        r = self.rows["dcf"]
        is_rows   = self.rows["is"]
        cf_rows   = self.rows["cf"]
        wc_rows   = self.rows["wc"]
        da_rows   = self.rows["da"]
        debt_rows = self.rows["debt"]
        bs_rows   = self.rows["bs"]
        assump    = self.rows["assump"]
        wacc_r    = self.rows["wacc"]

        # 币种与列布局
        rep_ccy = d.get("reporting_currency") or "N/A"
        trd_ccy = d.get("trading_currency") or "N/A"
        # 3-Statement 内部使用 HIST_COLS (C..G, 右对齐) + FCST_COLS (H..L)
        # 为便于用户查看, DCF sheet 保持相同列布局
        # DCF 内部历史列 (可能 <5 期): 与 3-Statement Sheets 严格对齐, 直接跨表引用
        HIST_COLS = self.HIST_COLS[:]   # C..G (右对齐)
        FCST_COLS = self.FCST_COLS[:]   # H..L
        latest_hist_col_letter = get_column_letter(HIST_COLS[-1]) if HIST_COLS else "B"
        SEC_MERGE_END = FCST_COLS[-1]   # 合并到 L
        # 3-Statement 各 Sheet 的历史列与 DCF sheet 同为 self.HIST_COLS/FCST_COLS
        # 故跨表引用直接使用相同 col letter (例如 D..G, H..L)

        # ============ Section 1: Header (Row 1-2) ============
        c = ws.cell(1, 1, f"{d['ticker']} DCF Valuation (from 3-Statement) / 基于 3-Statement 的 DCF 估值")
        c.fill = FILL_DARK_BLUE; c.font = FONT_WHITE_BOLD
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=SEC_MERGE_END)
        c = ws.cell(2, 1, f"Ticker: {d['ticker']}  |  Date: {date.today().isoformat()}  |  "
                          f"Reporting: Million {rep_ccy}  |  Trading Ccy: {trd_ccy}")
        c.font = FONT_ITALIC_GREY
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=SEC_MERGE_END)

        # ============ Section 2: Case Selector (Row 4-5) ============
        ws.cell(4, 1, "Case Selector (1=Bear 悲观 / 2=Base 基准 / 3=Bull 乐观)")
        self._apply_input(ws.cell(4, 2), 2, source_system="User input")
        add_comment(ws.cell(4, 2),
                    "Case Selector:\n  1 = Bear (悲观)\n  2 = Base (基准, 默认)\n  3 = Bull (乐观)\n\n"
                    "所有下游预测通过 CHOOSE($B$4, Bear, Base, Bull) 引用选中情景.")
        ws.cell(5, 1, "Selected Case Name")
        c = ws.cell(5, 2, '=CHOOSE($B$4,"Bear","Base","Bull")'); c.font = FONT_BLACK
        add_comment(c, '公式: =CHOOSE($B$4, "Bear", "Base", "Bull")')

        # 数据来源标注
        c = ws.cell(6, 1,
                    f"Data Source 数据来源: {d.get('shares_source', 'N/A')}  |  "
                    f"FX: {d.get('fx_source', 'N/A')}  |  Beta: {d.get('beta_source', 'N/A')}")
        c.font = FONT_GREEN
        ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=SEC_MERGE_END)

        # ============ Section 3: Market Data (Row 8-15) ============
        c = ws.cell(8, 1, "MARKET DATA -- 市场数据 (跨表引用 Assumptions)")
        c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=SEC_MERGE_END)

        # Row 9: Trading Stock Price (from Assumptions)
        ws.cell(9, 1, f"Current Stock Price ({trd_ccy}) -- 当前股价(交易币种)")
        c = ws.cell(9, 2, f"=Assumptions!B{assump['stock_price']}")
        c.font = FONT_GREEN; c.number_format = FMT_PRICE
        add_comment(c, f"跨 Sheet 引用: Assumptions!B{assump['stock_price']}")

        # Row 10: FX Rate (from Assumptions)
        ws.cell(10, 1, f"FX Rate: 1 {trd_ccy} = X {rep_ccy} -- 汇率")
        c = ws.cell(10, 2, f"=Assumptions!B{assump['fx_rate']}")
        c.font = FONT_GREEN; c.number_format = FMT_DECIMAL4
        add_comment(c, f"跨 Sheet 引用: Assumptions!B{assump['fx_rate']}\n"
                       f"来源: {d.get('fx_source', 'N/A')}")

        # Row 11: Reporting Stock Price
        ws.cell(11, 1, f"Current Stock Price ({rep_ccy}) -- 当前股价(报表币种)")
        c = ws.cell(11, 2, "=B9*B10"); c.font = FONT_BLACK; c.number_format = FMT_PRICE
        add_comment(c, f"Stock Price ({rep_ccy}) = Stock Price ({trd_ccy}) × FX Rate = B9 × B10")

        # Row 12: Diluted Shares (from Assumptions)
        ws.cell(12, 1, "Diluted Shares Outstanding (M) -- 稀释后总股本(百万)")
        c = ws.cell(12, 2, f"=Assumptions!B{assump['shares']}")
        c.font = FONT_GREEN; c.number_format = FMT_SHARES
        add_comment(c, f"跨 Sheet 引用: Assumptions!B{assump['shares']}")

        # Row 13: Market Cap
        ws.cell(13, 1, f"Market Capitalization ({rep_ccy} M) -- 市值(报表币种,百万)")
        c = ws.cell(13, 2, "=B11*B12"); c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        add_comment(c, f"Market Cap = Stock Price × Shares = B11 × B12")

        # Row 14: Net Debt (from Balance Sheet L3)
        # 用最新历史列的 BS L3: (Short-term Debt + Long-term Debt - Cash&STI)
        last_hist_L = latest_hist_col_letter
        ws.cell(14, 1, f"Net Debt / (Net Cash) ({rep_ccy} M) -- 净债务/(净现金)")
        c = ws.cell(14, 2,
                    f"='Balance Sheet'!{last_hist_L}{bs_rows['l3_st_debt']}"
                    f"+'Balance Sheet'!{last_hist_L}{bs_rows['l3_lt_debt']}"
                    f"-'Balance Sheet'!{last_hist_L}{bs_rows['l3_cash_sti']}")
        c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
        add_comment(c, f"跨 Sheet 引用: BS L3 最新历史列\n"
                       f"= 'Balance Sheet'!{last_hist_L}{bs_rows['l3_st_debt']} (Short-term Debt)\n"
                       f"+ 'Balance Sheet'!{last_hist_L}{bs_rows['l3_lt_debt']} (Long-term Debt)\n"
                       f"- 'Balance Sheet'!{last_hist_L}{bs_rows['l3_cash_sti']} (Cash & ST Investments)")

        # Row 15: Total Gross Debt (WACC 权重专用)
        ws.cell(15, 1, f"Total Gross Debt ({rep_ccy} M) -- 债务总额 (WACC 权重用)")
        c = ws.cell(15, 2,
                    f"='Balance Sheet'!{last_hist_L}{bs_rows['l3_st_debt']}"
                    f"+'Balance Sheet'!{last_hist_L}{bs_rows['l3_lt_debt']}")
        c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
        add_comment(c, "Short-term Debt + Long-term Debt (不减现金) — WACC 权重计算专用")

        # 关键行号索引
        row_index = {
            "case_selector": 4,
            "stock_price_trading": 9,
            "fx_rate": 10,
            "stock_price": 11,
            "shares": 12,
            "market_cap": 13,
            "net_debt": 14,
            "gross_debt": 15,
        }

        # ============ Section 4: 3 Scenario Blocks + Consolidation ============
        row_ptr = 17

        # 计算历史比率 (用于历史列展示; 从 d 字典的 hist series 中取)
        hist_data_by_key = {
            "rev_growth":  d.get("hist_rev_growth", []),
            "ebit_margin": d.get("hist_ebit_margin", []),
            # D&A% / CapEx% / Tax Rate 从原始 hist series 反推
            "da_pct":      [safe_divide(x, y) for x, y in zip(d.get("hist_da", []),    d.get("hist_revenue", []))],
            "capex_pct":   [safe_divide(x, y) for x, y in zip(d.get("hist_capex", []), d.get("hist_revenue", []))],
            "tax_rate":    [safe_divide(t, e) if e > 0 else 0.0
                            for t, e in zip(d.get("hist_tax", []), d.get("hist_ebt", []))],
        }

        # 三情景假设默认值 — Base 用最近一年实际值 (不加人为下限), Bear/Bull 相对 Base 情景化调整
        # D&A%/CapEx%/NWC% 都取自最近一年实际经营数据 (extract_financial_data 已归一化)
        base_growth = d.get("growth", 0.05)
        base_margin = d.get("ebit_margin", 0.10)
        base_da     = d.get("da_pct", 0.03)       # 最近一年 D&A / Revenue (实际)
        base_capex  = d.get("capex_pct", 0.05)    # 最近一年 CapEx / Revenue (实际)
        base_nwc    = d.get("nwc_pct", 0.01)      # 最近 3 年 ΔWC / ΔRevenue 均值 (实际, 可为负)
        base_tax    = d.get("tax_rate", 0.25)

        def _decay(base, delta_start, decay_step):
            return [base + delta_start - decay_step * i for i in range(5)]

        scenarios = {
            "Bear": {
                "rev_growth":      _decay(base_growth, -0.03, 0.005),
                "ebit_margin":     _decay(base_margin, -0.03, 0.002),
                # Bear: 折旧摊销 & 资本开支比重上升 (成本压力/资本效率下降)
                "da_pct":          [base_da] * 5,
                "capex_pct":       [base_capex + 0.01] * 5,
                # Bear: NWC 占用增加 (回款变差 / 库存积压), 相对 Base +1%
                "nwc_pct":         [base_nwc + 0.01] * 5,
                "tax_rate":        [base_tax] * 5,
                "terminal_growth": [0.020] + [None] * 4,
                "wacc":            [0.100] + [None] * 4,
            },
            "Base": {
                "rev_growth":      _decay(base_growth, 0.00, 0.003),
                "ebit_margin":     _decay(base_margin, 0.00, 0.000),
                # Base: 直接用最近一年实际比率 (投行 baseline 惯例)
                "da_pct":          [base_da] * 5,
                "capex_pct":       [base_capex] * 5,
                "nwc_pct":         [base_nwc] * 5,
                "tax_rate":        [base_tax] * 5,
                "terminal_growth": [0.025] + [None] * 4,
                "wacc":            [None] + [None] * 4,   # 由 WACC Sheet 引用
            },
            "Bull": {
                "rev_growth":      _decay(base_growth, 0.03, 0.005),
                "ebit_margin":     _decay(base_margin, 0.03, -0.002),   # 递增
                # Bull: 折旧摊销保持, 资本开支效率提升 (相对 Base -0.5%)
                "da_pct":          [base_da] * 5,
                "capex_pct":       [max(base_capex - 0.005, 0.0)] * 5,
                # Bull: NWC 效率提升 (相对 Base -0.5%)
                "nwc_pct":         [base_nwc - 0.005] * 5,
                "tax_rate":        [base_tax] * 5,
                "terminal_growth": [0.030] + [None] * 4,
                "wacc":            [0.080] + [None] * 4,
            },
        }

        ASSUMPTION_LABELS = [
            ("Revenue Growth % -- 营收增长率",           "rev_growth",      FMT_PERCENT),
            ("EBIT Margin % -- 息税前利润率",             "ebit_margin",     FMT_PERCENT),
            ("D&A % of Revenue -- 折旧摊销占营收比",      "da_pct",          FMT_PERCENT),
            ("CapEx % of Revenue -- 资本开支占营收比",    "capex_pct",       FMT_PERCENT),
            ("NWC % of Delta Revenue -- 营运资本占营收变动比", "nwc_pct",     FMT_PERCENT),
            ("Tax Rate -- 税率",                          "tax_rate",        FMT_PERCENT),
            ("Terminal Growth -- 永续增长率",             "terminal_growth", FMT_PERCENT),
            ("WACC -- 加权平均资本成本",                  "wacc",            FMT_PERCENT),
        ]

        HIST_KEYS = {"rev_growth", "ebit_margin", "da_pct", "capex_pct", "tax_rate"}
        scenario_rows = {}

        for scenario in ["Bear", "Base", "Bull"]:
            c = ws.cell(row_ptr, 1,
                        f"{scenario.upper()} CASE ASSUMPTIONS -- "
                        f"{'悲观' if scenario=='Bear' else ('基准' if scenario=='Base' else '乐观')}情景假设")
            c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
            ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=SEC_MERGE_END)
            row_ptr += 1

            ws.cell(row_ptr, 1, "Assumption -- 假设项").font = FONT_BOLD
            ws.cell(row_ptr, 1).fill = FILL_LIGHT_BLUE
            for i, col in enumerate(HIST_COLS):
                fy = self.hist_fys[self.hist_offset + i]
                c = ws.cell(row_ptr, col, f"{fy} (A)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
            for i, col in enumerate(FCST_COLS):
                c = ws.cell(row_ptr, col, f"FY{i+1} (E)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
            row_ptr += 1

            scenario_rows[scenario] = {}
            for label, key, fmt in ASSUMPTION_LABELS:
                ws.cell(row_ptr, 1, label)
                # 历史列: 展示实际比率 (3 个情景块显示相同)
                if key in HIST_KEYS:
                    hist_data = hist_data_by_key.get(key, [])
                    for i, col in enumerate(HIST_COLS):
                        idx = self.hist_offset + i
                        if idx < len(hist_data):
                            hv = hist_data[idx]
                            if hv is not None:
                                c = ws.cell(row_ptr, col, hv); c.font = FONT_BLACK; c.number_format = fmt
                # 预测列: 情景假设
                vals = scenarios[scenario][key]
                for i, col in enumerate(FCST_COLS):
                    v = vals[i]
                    if v is None:
                        if scenario == "Base" and key == "wacc" and i == 0:
                            c = ws.cell(row_ptr, col, "=WACC!B18"); c.font = FONT_GREEN
                            c.number_format = fmt
                            add_comment(c, "跨 Sheet 引用: WACC!B18 (由 WACC 表 CAPM 计算)")
                        continue
                    c = ws.cell(row_ptr, col, v); c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY
                    c.number_format = fmt
                scenario_rows[scenario][key] = row_ptr
                row_ptr += 1
            row_ptr += 1

        # ---- Selected Case Consolidation Block ----
        c = ws.cell(row_ptr, 1, "SELECTED CASE ASSUMPTIONS -- 选中情景 (由 Case Selector 驱动)")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=SEC_MERGE_END)
        row_ptr += 1
        ws.cell(row_ptr, 1, "Assumption -- 假设项").font = FONT_BOLD
        ws.cell(row_ptr, 1).fill = FILL_LIGHT_BLUE
        for i, col in enumerate(HIST_COLS):
            fy = self.hist_fys[self.hist_offset + i]
            c = ws.cell(row_ptr, col, f"{fy} (A)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        for i, col in enumerate(FCST_COLS):
            c = ws.cell(row_ptr, col, f"FY{i+1} (E)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        row_ptr += 1

        selected_rows = {}
        case_selector_ref = f"$B${row_index['case_selector']}"
        for label, key, fmt in ASSUMPTION_LABELS:
            ws.cell(row_ptr, 1, label)
            bear_row = scenario_rows["Bear"][key]
            base_row = scenario_rows["Base"][key]
            bull_row = scenario_rows["Bull"][key]
            if key in HIST_KEYS:
                for col in HIST_COLS:
                    col_letter = get_column_letter(col)
                    c = ws.cell(row_ptr, col, f"={col_letter}{base_row}")
                    c.font = FONT_PURPLE; c.number_format = fmt
            for i, col in enumerate(FCST_COLS):
                col_letter = get_column_letter(col)
                if key in ("wacc", "terminal_growth") and i > 0:
                    continue
                formula = (f"=CHOOSE({case_selector_ref},"
                           f"{col_letter}{bear_row},{col_letter}{base_row},{col_letter}{bull_row})")
                c = ws.cell(row_ptr, col, formula); c.font = FONT_BLACK
                c.fill = FILL_FORECAST_GREEN; c.number_format = fmt
                add_comment(c, f"=CHOOSE({case_selector_ref}, Bear, Base, Bull)")
            selected_rows[key] = row_ptr
            row_ptr += 1
        row_ptr += 1

        # ============ Section 5: Historical & Projected Financials ============
        # 历史列 (C-G): 跨 Sheet 引用真实数据 (IS/CF/BS/D&A/WC)
        # 预测列 (H-L): DCF sheet 内部公式, 直接由 Selected Case Assumptions 驱动
        #   → 用户切换 Case Selector (B4) 后, DCF 预测数据立即联动 (Revenue/EBIT/D&A/CapEx/ΔWC 全部重算)
        c = ws.cell(row_ptr, 1,
                    "HISTORICAL & PROJECTED FINANCIALS -- 历史 (跨表) + 预测 (由 Selected Case 内部驱动)")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=SEC_MERGE_END)
        row_ptr += 1

        ws.cell(row_ptr, 1, "Line Item (M) -- 项目").font = FONT_BOLD
        ws.cell(row_ptr, 1).fill = FILL_LIGHT_BLUE
        for i, col in enumerate(HIST_COLS):
            fy = self.hist_fys[self.hist_offset + i]
            c = ws.cell(row_ptr, col, f"{fy} (A)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        for i, col in enumerate(FCST_COLS):
            c = ws.cell(row_ptr, col, f"FY{i+1} (E)"); c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        row_ptr += 1

        # ---- Revenue ----
        # 历史列: 跨表引用 IS!revenue (真实数据)
        # 预测列: FY_n Revenue = FY_{n-1} Revenue × (1 + Selected.rev_growth FY_n)
        #        FY1 的前一列是 G (latest hist, 跨表引用 IS 最新历史 Revenue) → 天然锚定实际起点
        rev_row = row_ptr
        ws.cell(rev_row, 1, "Revenue -- 营业收入")
        for col in HIST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(rev_row, col, f"='Income Statement'!{cL}{is_rows['revenue']}")
            c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"历史 (跨表引用): 'Income Statement'!{cL}{is_rows['revenue']}")
        all_cols = HIST_COLS + FCST_COLS
        for i, col in enumerate(FCST_COLS):
            cL = get_column_letter(col)
            # 前一列: FY1 → G (latest hist); FY2 → H (FY1); FY3 → I; ...
            prev_col_letter = get_column_letter(col - 1)
            growth_ref = f"{cL}{selected_rows['rev_growth']}"
            f = f"={prev_col_letter}{rev_row}*(1+{growth_ref})"
            c = ws.cell(rev_row, col, f)
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"预测 (内部驱动):\n  Revenue(FY{i+1}) = Revenue(前年) × (1 + Growth%)\n"
                           f"  = {prev_col_letter}{rev_row} × (1 + {growth_ref})")
        row_ptr += 1

        # ---- % Revenue Growth (计算显示行) ----
        growth_disp_row = row_ptr
        ws.cell(growth_disp_row, 1, "  % Revenue Growth -- 营收增速")
        for pos, col in enumerate(all_cols):
            cL = get_column_letter(col)
            if pos == 0:
                # 最老列: 从 hist_data 读取 (需要 prior-to-hist, 直接放数值)
                hist_gr = hist_data_by_key.get("rev_growth", [])
                idx = self.hist_offset
                if idx < len(hist_gr) and hist_gr[idx] is not None:
                    c = ws.cell(growth_disp_row, col, hist_gr[idx])
                    c.font = FONT_BLACK; c.number_format = FMT_PERCENT
            else:
                prev_col_letter = get_column_letter(all_cols[pos - 1])
                c = ws.cell(growth_disp_row, col,
                            f"={cL}{rev_row}/{prev_col_letter}{rev_row}-1")
                c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        row_ptr += 1

        # ---- EBIT ----
        # 历史列: 跨表引用 IS!ebit (真实数据)
        # 预测列: EBIT = Revenue × Selected.ebit_margin FY_n
        ebit_row = row_ptr
        ws.cell(ebit_row, 1, "EBIT -- 息税前利润")
        for col in HIST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(ebit_row, col, f"='Income Statement'!{cL}{is_rows['ebit']}")
            c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"历史 (跨表引用): 'Income Statement'!{cL}{is_rows['ebit']}")
        for i, col in enumerate(FCST_COLS):
            cL = get_column_letter(col)
            margin_ref = f"{cL}{selected_rows['ebit_margin']}"
            f = f"={cL}{rev_row}*{margin_ref}"
            c = ws.cell(ebit_row, col, f)
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"预测 (内部驱动):\n  EBIT(FY{i+1}) = Revenue × EBIT Margin\n"
                           f"  = {cL}{rev_row} × {margin_ref}")
        row_ptr += 1

        # ---- % EBIT Margin (计算显示行) ----
        margin_disp_row = row_ptr
        ws.cell(margin_disp_row, 1, "  % EBIT Margin -- 息税前利润率")
        for col in HIST_COLS + FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(margin_disp_row, col, f"={cL}{ebit_row}/{cL}{rev_row}")
            c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        row_ptr += 1

        # ---- Tax Rate (Selected consolidation FCST + 历史反算) ----
        tax_disp_row = row_ptr
        ws.cell(tax_disp_row, 1, "Tax Rate -- 税率")
        for i, col in enumerate(HIST_COLS):
            hist_tr = hist_data_by_key.get("tax_rate", [])
            idx = self.hist_offset + i
            if idx < len(hist_tr):
                c = ws.cell(tax_disp_row, col, hist_tr[idx])
                c.font = FONT_BLUE; c.fill = FILL_INPUT_GREY; c.number_format = FMT_PERCENT
        for col in FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(tax_disp_row, col, f"={cL}{selected_rows['tax_rate']}")
            c.font = FONT_PURPLE; c.number_format = FMT_PERCENT
        row_ptr += 1

        # ---- Cash Taxes = -MAX(0, EBIT) × Tax Rate (内部公式, 不变) ----
        cash_tax_row = row_ptr
        ws.cell(cash_tax_row, 1, "Less: Cash Taxes -- 现金税额 (亏损不缴)")
        for col in HIST_COLS + FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(cash_tax_row, col,
                        f"=-MAX(0,{cL}{ebit_row})*{cL}{tax_disp_row}")
            c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
            if col in FCST_COLS: c.fill = FILL_FORECAST_GREEN
            add_comment(c, "Cash Taxes = -MAX(0, EBIT) × Tax Rate")
        row_ptr += 1

        # ---- NOPAT = EBIT + Cash Taxes (内部公式, 不变) ----
        nopat_row = row_ptr
        ws.cell(nopat_row, 1, "NOPAT -- 税后净营业利润").font = FONT_BOLD
        for col in HIST_COLS + FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(nopat_row, col, f"={cL}{ebit_row}+{cL}{cash_tax_row}")
            c.font = FONT_BLACK_BOLD; c.number_format = FMT_CURRENCY_M
            if col in FCST_COLS: c.fill = FILL_FORECAST_GREEN
        row_ptr += 2

        # ============ Section 6: Free Cash Flow Build ============
        c = ws.cell(row_ptr, 1, "FREE CASH FLOW BUILD -- 自由现金流构建"); c.font = FONT_BOLD
        c.fill = FILL_LIGHT_BLUE
        row_ptr += 1

        # NOPAT ref (内部引用)
        nopat_ref_row = row_ptr
        ws.cell(nopat_ref_row, 1, "NOPAT (from above)")
        for col in HIST_COLS + FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(nopat_ref_row, col, f"={cL}{nopat_row}")
            c.font = FONT_PURPLE; c.number_format = FMT_CURRENCY_M
        row_ptr += 1

        # ---- D&A ----
        # 历史列: 跨表引用 D&A Schedule (CF-side 广口径, 真实数据)
        # 预测列: D&A = Revenue × Selected.da_pct FY_n
        da_ref_row = row_ptr
        ws.cell(da_ref_row, 1, "Plus: D&A -- 折旧摊销 (加回)")
        for col in HIST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(da_ref_row, col, f"='D&A Schedule'!{cL}{da_rows['da']}")
            c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"历史 (跨表引用): 'D&A Schedule'!{cL}{da_rows['da']} (CF-side 广口径)")
        for i, col in enumerate(FCST_COLS):
            cL = get_column_letter(col)
            da_pct_ref = f"{cL}{selected_rows['da_pct']}"
            f = f"={cL}{rev_row}*{da_pct_ref}"
            c = ws.cell(da_ref_row, col, f)
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"预测 (内部驱动):\n  D&A(FY{i+1}) = Revenue × D&A %\n"
                           f"  = {cL}{rev_row} × {da_pct_ref}")
        row_ptr += 1

        # ---- CapEx ----
        # 历史列: 跨表引用 D&A Schedule capex (负号展示, 真实数据)
        # 预测列: CapEx = -Revenue × Selected.capex_pct FY_n (负号: 现金流出)
        capex_ref_row = row_ptr
        ws.cell(capex_ref_row, 1, "Less: CapEx -- 资本开支")
        for col in HIST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(capex_ref_row, col, f"=-'D&A Schedule'!{cL}{da_rows['capex']}")
            c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"历史 (跨表引用): -'D&A Schedule'!{cL}{da_rows['capex']}")
        for i, col in enumerate(FCST_COLS):
            cL = get_column_letter(col)
            capex_pct_ref = f"{cL}{selected_rows['capex_pct']}"
            f = f"=-{cL}{rev_row}*{capex_pct_ref}"
            c = ws.cell(capex_ref_row, col, f)
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"预测 (内部驱动):\n  CapEx(FY{i+1}) = -Revenue × CapEx %\n"
                           f"  = -{cL}{rev_row} × {capex_pct_ref}\n"
                           "  (负号: 表示现金流出)")
        row_ptr += 1

        # ---- ΔWorking Capital ----
        # 历史列: 跨表引用 Working Capital dNWC (真实数据, 已按 CF 符号约定)
        # 预测列: ΔWC = -(Revenue(FY_n) − Revenue(FY_{n-1})) × Selected.nwc_pct FY_n
        #   NWC% 定义: 每单位新增营收对应的营运资本占用 (ΔWC/ΔRev)
        #   负号约定: 营运资本增加 → 资产端占用现金 → CF 视角为流出 (-)
        nwc_ref_row = row_ptr
        ws.cell(nwc_ref_row, 1, "Less: Δ Working Capital -- 营运资本变动")
        for col in HIST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(nwc_ref_row, col, f"='Working Capital'!{cL}{wc_rows['dNWC']}")
            c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"历史 (跨表引用): 'Working Capital'!{cL}{wc_rows['dNWC']}\n"
                           "(Δ AR + Δ Inv + Δ AP, 按 CF 符号约定)")
        for i, col in enumerate(FCST_COLS):
            cL = get_column_letter(col)
            prev_col_letter = get_column_letter(col - 1)
            nwc_pct_ref = f"{cL}{selected_rows['nwc_pct']}"
            f = f"=-({cL}{rev_row}-{prev_col_letter}{rev_row})*{nwc_pct_ref}"
            c = ws.cell(nwc_ref_row, col, f)
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
            add_comment(c, f"预测 (内部驱动):\n"
                           f"  ΔWC(FY{i+1}) = -(Revenue(FY{i+1}) - Revenue(前年)) × NWC %\n"
                           f"  = -({cL}{rev_row} - {prev_col_letter}{rev_row}) × {nwc_pct_ref}\n"
                           "  (负号: 营运资本占用增加视为现金流出)")
        row_ptr += 1

        # ---- Unlevered FCF (内部公式, 不变) ----
        fcf_row = row_ptr
        ws.cell(fcf_row, 1, "Unlevered FCF -- 无杠杆自由现金流").font = FONT_BOLD
        for col in HIST_COLS + FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(fcf_row, col,
                        f"={cL}{nopat_ref_row}+{cL}{da_ref_row}+{cL}{capex_ref_row}+{cL}{nwc_ref_row}")
            c.font = FONT_BOLD; c.number_format = FMT_CURRENCY_M; c.border = BORDER_THIN_BOTTOM
            if col in FCST_COLS: c.fill = FILL_FORECAST_GREEN
            add_comment(c, "Unlevered FCF = NOPAT + D&A + CapEx + ΔWC")
        row_ptr += 2

        # ============ Section 7: Discounting & Terminal Value ============
        c = ws.cell(row_ptr, 1, "DISCOUNTING & TERMINAL VALUE -- 折现与终值")
        c.font = FONT_BOLD; c.fill = FILL_LIGHT_BLUE
        row_ptr += 1

        first_fcst_letter = get_column_letter(FCST_COLS[0])
        last_fcst_letter  = get_column_letter(FCST_COLS[-1])

        wacc_ref_row = row_ptr
        ws.cell(wacc_ref_row, 1, "WACC (from Selected Case)")
        c = ws.cell(wacc_ref_row, 2, f"={first_fcst_letter}{selected_rows['wacc']}")
        c.font = FONT_PURPLE; c.number_format = FMT_PERCENT
        row_ptr += 1

        tgr_ref_row = row_ptr
        ws.cell(tgr_ref_row, 1, "Terminal Growth Rate (from Selected Case)")
        c = ws.cell(tgr_ref_row, 2, f"={first_fcst_letter}{selected_rows['terminal_growth']}")
        c.font = FONT_PURPLE; c.number_format = FMT_PERCENT
        row_ptr += 1

        # Discount Period (Mid-year)
        period_row = row_ptr
        ws.cell(period_row, 1, "Discount Period (Mid-year)")
        for i, col in enumerate(FCST_COLS):
            c = ws.cell(period_row, col, i + 0.5); c.font = FONT_BLUE
            c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_DECIMAL4
        row_ptr += 1

        # Discount Factor
        df_row = row_ptr
        ws.cell(df_row, 1, "Discount Factor")
        wacc_abs = f"$B${wacc_ref_row}"
        for col in FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(df_row, col, f"=1/(1+{wacc_abs})^{cL}{period_row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_DECIMAL4
        row_ptr += 1

        # PV of FCF
        pv_fcf_row = row_ptr
        ws.cell(pv_fcf_row, 1, "PV of FCF -- 自由现金流现值")
        for col in FCST_COLS:
            cL = get_column_letter(col)
            c = ws.cell(pv_fcf_row, col, f"={cL}{fcf_row}*{cL}{df_row}")
            c.font = FONT_BLACK; c.fill = FILL_FORECAST_GREEN; c.number_format = FMT_CURRENCY_M
        row_ptr += 2

        # ============ Section 8: Valuation Summary ============
        c = ws.cell(row_ptr, 1, "VALUATION SUMMARY -- 估值汇总")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=SEC_MERGE_END)
        row_ptr += 1
        val_start = row_ptr

        pv_sum_row = row_ptr
        ws.cell(pv_sum_row, 1, "(+) PV of Explicit FCFs -- 显性期 FCF 现值合计")
        c = ws.cell(pv_sum_row, 2,
                    f"=SUM({first_fcst_letter}{pv_fcf_row}:{last_fcst_letter}{pv_fcf_row})")
        c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        row_ptr += 1

        term_fcf_row = row_ptr
        ws.cell(term_fcf_row, 1, "Terminal FCF (Year 5+1)")
        c = ws.cell(term_fcf_row, 2, f"={last_fcst_letter}{fcf_row}*(1+B{tgr_ref_row})")
        c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        row_ptr += 1

        tv_row = row_ptr
        ws.cell(tv_row, 1, "Terminal Value -- 终值 (Gordon Growth Model)")
        c = ws.cell(tv_row, 2, f"=B{term_fcf_row}/(B{wacc_ref_row}-B{tgr_ref_row})")
        c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        add_comment(c, "TV = Terminal FCF / (WACC - TGR)\n约束: TGR < WACC")
        row_ptr += 1

        pv_tv_row = row_ptr
        ws.cell(pv_tv_row, 1, "(+) PV of Terminal Value -- 终值现值")
        c = ws.cell(pv_tv_row, 2, f"=B{tv_row}*{last_fcst_letter}{df_row}")
        c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        row_ptr += 1

        ev_row = row_ptr
        ws.cell(ev_row, 1, f"  Enterprise Value ({rep_ccy} M) -- 企业价值 (=)").font = FONT_BOLD
        c = ws.cell(ev_row, 2, f"=B{pv_sum_row}+B{pv_tv_row}")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = FMT_CURRENCY_M
        row_ptr += 1

        tv_ev_row = row_ptr
        ws.cell(tv_ev_row, 1, "  Terminal Value / EV -- 终值占比 (合理区间 50–70%)")
        c = ws.cell(tv_ev_row, 2, f"=B{pv_tv_row}/B{ev_row}")
        c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        add_comment(c, "50-70% 正常; >75% 过度依赖远期; <40% 假设过于保守")
        row_ptr += 1

        nd_row = row_ptr
        ws.cell(nd_row, 1, "(-) Net Debt / (Net Cash)")
        c = ws.cell(nd_row, 2, f"=B{row_index['net_debt']}"); c.font = FONT_PURPLE
        c.number_format = FMT_CURRENCY_M
        row_ptr += 1

        eq_row = row_ptr
        ws.cell(eq_row, 1, f"  Equity Value ({rep_ccy} M) -- 股权价值 (=)").font = FONT_BOLD
        c = ws.cell(eq_row, 2, f"=B{ev_row}-B{nd_row}")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = FMT_CURRENCY_M
        row_ptr += 1

        shares_row = row_ptr
        ws.cell(shares_row, 1, "÷ Diluted Shares Outstanding (M)")
        c = ws.cell(shares_row, 2, f"=B{row_index['shares']}"); c.font = FONT_PURPLE
        c.number_format = FMT_SHARES
        row_ptr += 1

        implied_row = row_ptr
        ws.cell(implied_row, 1,
                f"  Implied Price per Share ({rep_ccy}) -- 每股内在价值 (=)").font = FONT_BOLD
        c = ws.cell(implied_row, 2, f"=B{eq_row}/B{shares_row}")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = FMT_PRICE
        row_ptr += 1

        cur_row = row_ptr
        ws.cell(cur_row, 1, f"Current Stock Price ({rep_ccy}) -- 当前股价(报表币种)")
        c = ws.cell(cur_row, 2, f"=B{row_index['stock_price']}"); c.font = FONT_PURPLE
        c.number_format = FMT_PRICE
        row_ptr += 1

        upside_row = row_ptr
        ws.cell(upside_row, 1, "Implied Upside / (Downside) -- 隐含涨跌空间").font = FONT_BOLD
        c = ws.cell(upside_row, 2, f"=B{implied_row}/B{cur_row}-1")
        c.font = FONT_BOLD; c.fill = FILL_MEDIUM_BLUE; c.number_format = FMT_PERCENT
        row_ptr += 1

        # 底纹
        for rr in range(val_start, upside_row + 1):
            for col in (1, 2):
                cell = ws.cell(rr, col)
                if cell.fill.fgColor is None or cell.fill.patternType is None:
                    cell.fill = FILL_VALUATION_ORANGE
        for rr in (ev_row, eq_row, implied_row, upside_row):
            ws.cell(rr, 2).fill = FILL_VALUATION_ORANGE_DARK

        row_ptr += 2

        # ---- 保存关键行号供敏感性表引用 ----
        self._dcf_refs = {
            "sheet": ws,
            "row_index": row_index,
            "selected_rows": selected_rows,
            "rev_row": rev_row,
            "fcf_row": fcf_row,
            "pv_fcf_row": pv_fcf_row,
            "df_row": df_row,
            "period_row": period_row,
            "wacc_ref_row": wacc_ref_row,
            "tgr_ref_row": tgr_ref_row,
            "implied_row": implied_row,
            "fcst_cols": FCST_COLS,
            "hist_cols": HIST_COLS,
            "latest_hist_col_letter": latest_hist_col_letter,
            "sec_merge_end": SEC_MERGE_END,
        }

        # 保存 DCF sheet 的行索引 (供外部/测试用)
        r.update({
            "case_selector": row_index["case_selector"],
            "stock_price_trading": row_index["stock_price_trading"],
            "fx_rate": row_index["fx_rate"],
            "stock_price": row_index["stock_price"],
            "shares": row_index["shares"],
            "market_cap": row_index["market_cap"],
            "net_debt": row_index["net_debt"],
            "gross_debt": row_index["gross_debt"],
            "rev_row": rev_row, "ebit_row": ebit_row, "fcf_row": fcf_row,
            "ev_row": ev_row, "eq_row": eq_row, "implied_row": implied_row,
        })

        # ============ Section 9: Sensitivity Tables (3× 5×5) ============
        self._dcf_sensitivity_tables(ws, row_ptr)

        # 列宽
        ws.column_dimensions["A"].width = 48
        for col in HIST_COLS + FCST_COLS:
            ws.column_dimensions[get_column_letter(col)].width = 13

    def _dcf_sensitivity_tables(self, ws, start_row: int):
        """在 DCF sheet 底部生成 3 张 5×5 敏感性表, 75 个闭式重算公式.

        Table 1: WACC × Terminal Growth → Implied Price
        Table 2: Revenue Growth (FY1) × EBIT Margin (FY1) → Implied Price
        Table 3: Beta × Risk-Free Rate → Implied Price (通过 WACC 传导)
        """
        refs = self._dcf_refs
        SEC_MERGE_END = refs["sec_merge_end"]

        row_ptr = start_row

        # ---- Table 1: WACC × TGR ----
        c = ws.cell(row_ptr, 1, "SENSITIVITY TABLE 1: WACC × Terminal Growth → Implied Share Price")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=7)
        row_ptr += 1
        row_ptr = self._dcf_sens_wacc_tgr(ws, row_ptr, 0.005, 0.005)
        row_ptr += 2

        # ---- Table 2: Growth × Margin ----
        c = ws.cell(row_ptr, 1,
                    "SENSITIVITY TABLE 2: Revenue Growth (FY1) × EBIT Margin (FY1) → Implied Share Price")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=7)
        row_ptr += 1
        row_ptr = self._dcf_sens_growth_margin(ws, row_ptr, 0.02, 0.02)
        row_ptr += 2

        # ---- Table 3: Beta × Rf ----
        c = ws.cell(row_ptr, 1, "SENSITIVITY TABLE 3: Beta × Risk-Free Rate → Implied Share Price")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=7)
        row_ptr += 1
        self._dcf_sens_beta_rf(ws, row_ptr, 0.15, 0.005)

    def _dcf_sens_wacc_tgr(self, ws, start_row: int, wacc_step: float, tgr_step: float) -> int:
        refs = self._dcf_refs
        row_index = refs["row_index"]
        fcf_row = refs["fcf_row"]
        wacc_ref_row = refs["wacc_ref_row"]; tgr_ref_row = refs["tgr_ref_row"]
        FCST_COLS = refs["fcst_cols"]
        fcf_cells = [f"{get_column_letter(c)}{fcf_row}" for c in FCST_COLS]
        periods = [0.5, 1.5, 2.5, 3.5, 4.5]
        net_debt = f"$B${row_index['net_debt']}"
        shares = f"$B${row_index['shares']}"
        base_wacc_cell = f"$B${wacc_ref_row}"
        base_tgr_cell = f"$B${tgr_ref_row}"

        hdr_row = start_row
        ws.cell(hdr_row, 1, "WACC \\ TGR").font = FONT_BOLD
        ws.cell(hdr_row, 1).fill = FILL_LIGHT_BLUE
        for j in range(5):
            offset = j - 2
            tgr_f = f"={base_tgr_cell}+{offset}*{tgr_step}"
            c = ws.cell(hdr_row, 3 + j, tgr_f); c.font = FONT_BLUE
            c.fill = FILL_LIGHT_BLUE; c.number_format = FMT_PERCENT
        for i in range(5):
            offset = i - 2
            r = hdr_row + 1 + i
            wacc_f = f"={base_wacc_cell}+{offset}*{wacc_step}"
            c = ws.cell(r, 2, wacc_f); c.font = FONT_BLUE
            c.fill = FILL_LIGHT_BLUE; c.number_format = FMT_PERCENT
            for j in range(5):
                cc = 3 + j
                col_letter = get_column_letter(cc)
                wacc_ref = f"$B{r}"
                tgr_ref = f"{col_letter}${hdr_row}"
                df_terms = [f"({fcf_cells[k]}/(1+{wacc_ref})^{periods[k]})" for k in range(5)]
                pv_sum = "+".join(df_terms)
                tv = f"{fcf_cells[-1]}*(1+{tgr_ref})/({wacc_ref}-{tgr_ref})"
                pv_tv = f"({tv})/(1+{wacc_ref})^{periods[-1]}"
                ev = f"({pv_sum})+({pv_tv})"
                equity = f"({ev})-{net_debt}"
                implied = f"IFERROR(({equity})/{shares},NA())"
                c = ws.cell(r, cc, f"={implied}"); c.font = FONT_BLACK; c.number_format = FMT_PRICE
                if i == 2 and j == 2:
                    c.font = FONT_BLACK_BOLD; c.fill = FILL_MEDIUM_BLUE
                    add_comment(c, "★ 中心格 = Base case Implied Price\n应等于 Valuation Summary 输出")
        return hdr_row + 5

    def _dcf_sens_growth_margin(self, ws, start_row: int, growth_step: float, margin_step: float) -> int:
        refs = self._dcf_refs
        row_index = refs["row_index"]
        rev_row = refs["rev_row"]; fcf_row = refs["fcf_row"]
        wacc_ref_row = refs["wacc_ref_row"]; tgr_ref_row = refs["tgr_ref_row"]
        selected_rows = refs["selected_rows"]
        FCST_COLS = refs["fcst_cols"]
        latest_hist_L = refs["latest_hist_col_letter"]

        first_fcst_letter = get_column_letter(FCST_COLS[0])
        base_growth_cell = f"${first_fcst_letter}${selected_rows['rev_growth']}"
        base_margin_cell = f"${first_fcst_letter}${selected_rows['ebit_margin']}"
        rev_ltm = f"${latest_hist_L}${rev_row}"
        base_wacc_cell = f"$B${wacc_ref_row}"
        base_tgr_cell = f"$B${tgr_ref_row}"
        net_debt = f"$B${row_index['net_debt']}"
        shares = f"$B${row_index['shares']}"
        tax_c = f"${first_fcst_letter}${selected_rows['tax_rate']}"
        da_c = f"${first_fcst_letter}${selected_rows['da_pct']}"
        capex_c = f"${first_fcst_letter}${selected_rows['capex_pct']}"
        nwc_c = f"${first_fcst_letter}${selected_rows['nwc_pct']}"
        fcf_2_to_5 = [f"{get_column_letter(c)}${fcf_row}" for c in FCST_COLS[1:]]
        periods = [0.5, 1.5, 2.5, 3.5, 4.5]

        hdr_row = start_row
        ws.cell(hdr_row, 1, "Growth \\ Margin").font = FONT_BOLD
        ws.cell(hdr_row, 1).fill = FILL_LIGHT_BLUE
        for j in range(5):
            offset = j - 2
            m_f = f"={base_margin_cell}+{offset}*{margin_step}"
            c = ws.cell(hdr_row, 3 + j, m_f); c.font = FONT_BLUE
            c.fill = FILL_LIGHT_BLUE; c.number_format = FMT_PERCENT
        for i in range(5):
            offset = i - 2
            r = hdr_row + 1 + i
            g_f = f"={base_growth_cell}+{offset}*{growth_step}"
            c = ws.cell(r, 2, g_f); c.font = FONT_BLUE
            c.fill = FILL_LIGHT_BLUE; c.number_format = FMT_PERCENT
            for j in range(5):
                cc = 3 + j
                col_letter = get_column_letter(cc)
                g_ref = f"$B{r}"; m_ref = f"{col_letter}${hdr_row}"
                rev1 = f"({rev_ltm}*(1+{g_ref}))"
                ebit1 = f"({rev1}*{m_ref})"
                tax1 = f"(-MAX(0,{ebit1})*{tax_c})"
                nopat1 = f"({ebit1}+{tax1})"
                da1 = f"({rev1}*{da_c})"
                capex1 = f"(-{rev1}*{capex_c})"
                nwc1 = f"(-({rev1}-{rev_ltm})*{nwc_c})"
                fcf1 = f"({nopat1}+{da1}+{capex1}+{nwc1})"
                pv_terms = [f"({fcf1}/(1+{base_wacc_cell})^{periods[0]})"] + \
                           [f"({fcf_2_to_5[k]}/(1+{base_wacc_cell})^{periods[k+1]})" for k in range(4)]
                pv_sum = "+".join(pv_terms)
                tv = f"{fcf_2_to_5[-1]}*(1+{base_tgr_cell})/({base_wacc_cell}-{base_tgr_cell})"
                pv_tv = f"({tv})/(1+{base_wacc_cell})^{periods[-1]}"
                ev = f"({pv_sum})+({pv_tv})"
                equity = f"({ev})-{net_debt}"
                implied = f"IFERROR(({equity})/{shares},NA())"
                c = ws.cell(r, cc, f"={implied}"); c.font = FONT_BLACK; c.number_format = FMT_PRICE
                if i == 2 and j == 2:
                    c.font = FONT_BLACK_BOLD; c.fill = FILL_MEDIUM_BLUE
                    add_comment(c, "★ 中心格 ≈ Base case Implied Price (仅 FY1 参数变化)")
        return hdr_row + 5

    def _dcf_sens_beta_rf(self, ws, start_row: int, beta_step: float, rf_step: float) -> int:
        refs = self._dcf_refs
        row_index = refs["row_index"]
        fcf_row = refs["fcf_row"]
        tgr_ref_row = refs["tgr_ref_row"]
        FCST_COLS = refs["fcst_cols"]
        fcf_cells = [f"{get_column_letter(c)}${fcf_row}" for c in FCST_COLS]
        periods = [0.5, 1.5, 2.5, 3.5, 4.5]
        net_debt = f"$B${row_index['net_debt']}"
        shares = f"$B${row_index['shares']}"
        base_tgr_cell = f"$B${tgr_ref_row}"
        we_ref = "WACC!$B$16"; wd_ref = "WACC!$B$17"
        kd_after_tax = "WACC!$B$10"
        erp_ref = "WACC!$B$4"
        base_beta_cell = "WACC!$B$3"; base_rf_cell = "WACC!$B$2"

        hdr_row = start_row
        ws.cell(hdr_row, 1, "Beta \\ Rf").font = FONT_BOLD
        ws.cell(hdr_row, 1).fill = FILL_LIGHT_BLUE
        for j in range(5):
            offset = j - 2
            rf_f = f"={base_rf_cell}+{offset}*{rf_step}"
            c = ws.cell(hdr_row, 3 + j, rf_f); c.font = FONT_BLUE
            c.fill = FILL_LIGHT_BLUE; c.number_format = FMT_PERCENT
        for i in range(5):
            offset = i - 2
            r = hdr_row + 1 + i
            beta_f = f"={base_beta_cell}+{offset}*{beta_step}"
            c = ws.cell(r, 2, beta_f); c.font = FONT_BLUE
            c.fill = FILL_LIGHT_BLUE; c.number_format = "0.00"
            for j in range(5):
                cc = 3 + j
                col_letter = get_column_letter(cc)
                beta_ref = f"$B{r}"; rf_ref = f"{col_letter}${hdr_row}"
                ke = f"({rf_ref}+{beta_ref}*{erp_ref})"
                wacc_f = f"({we_ref}*{ke}+{wd_ref}*{kd_after_tax})"
                df_terms = [f"({fcf_cells[k]}/(1+{wacc_f})^{periods[k]})" for k in range(5)]
                pv_sum = "+".join(df_terms)
                tv = f"{fcf_cells[-1]}*(1+{base_tgr_cell})/({wacc_f}-{base_tgr_cell})"
                pv_tv = f"({tv})/(1+{wacc_f})^{periods[-1]}"
                ev = f"({pv_sum})+({pv_tv})"
                equity = f"({ev})-{net_debt}"
                implied = f"IFERROR(({equity})/{shares},NA())"
                c = ws.cell(r, cc, f"={implied}"); c.font = FONT_BLACK; c.number_format = FMT_PRICE
                if i == 2 and j == 2:
                    c.font = FONT_BLACK_BOLD; c.fill = FILL_MEDIUM_BLUE
                    add_comment(c, "★ 中心格 = Base Beta × Base Rf (通过 WACC 传导)")
        return hdr_row + 5

    # ==================== Tab 9: WACC (CAPM) ====================
    def _wacc(self):
        """CAPM 股权成本 + 债务成本 + 资本结构权重 + WACC.

        与 dcf-model skill 保持一致:
          Ke = Rf + β × ERP        (CAPM)
          Kd_after_tax = Kd × (1 - Tax)
          Enterprise Capital = MAX(MC + Gross Debt, MC × 0.5)   (下限保护)
          We = MC / EC, Wd = Gross Debt / EC
          WACC = IF(Gross Debt ≤ 0.1% × MC, Ke, We×Ke + Wd×Kd_at)  (无债退化为 Ke)

        输入来源:
          Rf   — 按报表币种查常量表 (可覆盖)
          Beta — Futu 5Y monthly (vs benchmark 指数)
          ERP  — Damodaran country ERP
          Market Cap / Gross Debt — 引用 DCF!B13 / DCF!B15
          Tax Rate — 引用 Assumptions!B{tax_rate row 首列}
        """
        ws = self.wb.create_sheet("WACC")
        d = self.d
        assump = self.rows["assump"]
        r = self.rows["wacc"]
        rep_ccy = d.get("reporting_currency") or "N/A"

        # ============ Cost of Equity (CAPM) ============
        c = ws.cell(1, 1, "COST OF EQUITY (CAPM) -- 股权成本")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

        rf_src = d.get("rf_source", "10Y sovereign yield")
        beta_src = d.get("beta_source", "5Y Monthly Beta")
        erp_src = d.get("erp_source", "Damodaran ERP")
        benchmark = d.get("benchmark", "N/A")

        self._apply_input(ws.cell(2, 2), d.get("rf_rate", 0.043), FMT_PERCENT,
                          f"{rf_src} ({rep_ccy})", "10Y 主权债券收益率")
        ws.cell(2, 1, f"Risk-Free Rate ({rep_ccy} 10Y) -- 无风险利率")
        r["rf"] = 2

        self._apply_input(ws.cell(3, 2), d.get("beta", _BETA_FALLBACK), "0.00",
                          beta_src, f"基准: {benchmark}")
        ws.cell(3, 1, f"Beta (5Y Monthly, vs {benchmark}) -- 贝塔系数")
        r["beta"] = 3

        self._apply_input(ws.cell(4, 2), d.get("erp", 0.055), FMT_PERCENT,
                          f"{erp_src} ({rep_ccy})", "股权风险溢价 (国家/地区口径)")
        ws.cell(4, 1, f"Equity Risk Premium ({rep_ccy} country) -- 股权风险溢价")
        r["erp"] = 4

        ws.cell(5, 1, "Cost of Equity -- 股权成本")
        c = ws.cell(5, 2, "=B2+B3*B4"); c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        add_comment(c, "CAPM: Ke = Rf + β × ERP = B2 + B3 × B4")
        r["ke"] = 5

        # ============ Cost of Debt ============
        c = ws.cell(7, 1, "COST OF DEBT -- 债务成本"); c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2)

        self._apply_input(ws.cell(8, 2), 0.045, FMT_PERCENT,
                          "10-K / bond spread", "Pre-Tax Cost of Debt")
        ws.cell(8, 1, "Pre-Tax Cost of Debt -- 税前债务成本")
        r["kd_pre"] = 8

        # Tax Rate: 引用 Assumptions FY1 Tax Rate (与 IS/DCF 保持一致)
        first_fcst_letter = get_column_letter(self.FCST_COLS[0])
        c = ws.cell(9, 2, f"=Assumptions!{first_fcst_letter}{assump['tax_rate']}")
        c.font = FONT_GREEN; c.number_format = FMT_PERCENT
        add_comment(c, f"跨 Sheet 引用: Assumptions!{first_fcst_letter}{assump['tax_rate']} (FY1 Tax Rate)")
        ws.cell(9, 1, "Tax Rate -- 税率")
        r["tax"] = 9

        c = ws.cell(10, 2, "=B8*(1-B9)"); c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        add_comment(c, "Kd(after-tax) = Kd × (1 - Tax) = B8 × (1 - B9)")
        ws.cell(10, 1, "After-Tax Cost of Debt -- 税后债务成本")
        r["kd_after"] = 10

        # ============ Capital Structure ============
        c = ws.cell(12, 1, "CAPITAL STRUCTURE -- 资本结构")
        c.font = FONT_WHITE_BOLD; c.fill = FILL_DARK_BLUE
        ws.merge_cells(start_row=12, start_column=1, end_row=12, end_column=2)

        c = ws.cell(13, 2, "=DCF!B13"); c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
        add_comment(c, "跨 Sheet 引用: DCF!B13 (Market Cap, 报表币种)")
        ws.cell(13, 1, "Market Capitalization (M) -- 市值")
        r["mc"] = 13

        c = ws.cell(14, 2, "=DCF!B15"); c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
        add_comment(c, "跨 Sheet 引用: DCF!B15 (Gross Debt, 不减现金)\n"
                       "\n为什么用 Gross Debt: 净现金公司 Net Debt<0 会导致 Wd<0/We>100%, WACC 数值失真。")
        ws.cell(14, 1, "Total Gross Debt (M) -- 债务总额 (含现金前)")
        r["gd"] = 14

        c = ws.cell(15, 2, "=MAX(B13+B14, B13*0.5)"); c.font = FONT_BLACK; c.number_format = FMT_CURRENCY_M
        add_comment(c, "Enterprise Capital = MAX(MC + Gross Debt, MC × 0.5)\n下限保护避免 Wd>67%")
        ws.cell(15, 1, "Enterprise Capital (M) -- 企业资本 (含下限保护)")
        r["ec"] = 15

        c = ws.cell(16, 2, "=B13/B15"); c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        add_comment(c, "We = Market Cap / Enterprise Capital\n范围: [33%, 100%]")
        ws.cell(16, 1, "Equity Weight (We) -- 股权权重")
        r["we"] = 16

        c = ws.cell(17, 2, "=B14/B15"); c.font = FONT_BLACK; c.number_format = FMT_PERCENT
        add_comment(c, "Wd = Gross Debt / Enterprise Capital\n范围: [0, 67%]")
        ws.cell(17, 1, "Debt Weight (Wd) -- 债务权重")
        r["wd"] = 17

        c = ws.cell(18, 2, "=IF(B14<=0.001*B13, B5, B16*B5+B17*B10)")
        c.font = FONT_BLACK_BOLD; c.fill = FILL_VALUATION_ORANGE_DARK
        c.number_format = FMT_PERCENT
        add_comment(c,
                    "WACC:\n"
                    "  IF Gross Debt ≤ 0.1% × Market Cap: WACC = Ke (unlevered, 无债)\n"
                    "  ELSE: WACC = We × Ke + Wd × Kd(after-tax) = B16×B5 + B17×B10")
        ws.cell(18, 1, "WACC -- 加权平均资本成本")
        ws.cell(18, 1).font = FONT_BOLD
        r["wacc"] = 18

        # Net Debt 显示行 (仅显示, 不参与 WACC 计算)
        c = ws.cell(19, 2, "=DCF!B14"); c.font = FONT_GREEN; c.number_format = FMT_CURRENCY_M
        add_comment(c, "跨 Sheet 引用: DCF!B14 (Net Debt / Net Cash)\n"
                       "此行仅显示, Net Debt 在 EV → Equity Value bridge 中扣除")
        ws.cell(19, 1, "Net Debt / (Net Cash) (M) -- 净债务 (仅显示, EV→Equity 用)")
        r["nd_display"] = 19

        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 16


# ==================== Main ====================
def find_recent_report(excels_dir: Path, ticker: str, max_age_days: int = 7) -> Optional[Path]:
    """在 excels 目录下查找 {ticker}_3Statement_YYYY-MM-DD.xlsx 中最近 max_age_days 天内生成的文件.

    命名规范: `{ticker}_3Statement_{YYYY-MM-DD}.xlsx`, 日期由脚本落盘时的 `date.today()` 生成.
    优先按**文件名日期**判断 (可靠, 不受文件系统 mtime 修改影响); 无法解析时兜底为文件 mtime.

    Returns:
        Path — 命中的最新一份报告 (若存在多份都在窗口内, 返回文件名日期最新的一份)
        None — 目录不存在, 或无任何命中的报告
    """
    if not excels_dir.is_dir():
        return None
    pattern = re.compile(rf'^{re.escape(ticker)}_3Statement_(\d{{4}}-\d{{2}}-\d{{2}})\.xlsx$', re.IGNORECASE)
    today = date.today()
    candidates: List[Tuple[date, Path]] = []
    for fp in excels_dir.iterdir():
        if not fp.is_file(): continue
        if fp.name.startswith('~$'): continue   # 跳过 Excel 打开时的锁文件
        m = pattern.match(fp.name)
        if not m: continue
        try:
            file_date = date.fromisoformat(m.group(1))
        except ValueError:
            # 文件名日期不合法, 兜底 mtime
            try:
                file_date = date.fromtimestamp(fp.stat().st_mtime)
            except OSError:
                continue
        age_days = (today - file_date).days
        if 0 <= age_days <= max_age_days:
            candidates.append((file_date, fp))
    if not candidates:
        return None
    # 命中多份时选文件名日期最新的
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--ticker", required=True)
    parser.add_argument("--workspace", required=True)
    parser.add_argument("--force", action="store_true",
                        help="强制重新生成, 忽略 7 天内已有报告的缓存")
    parser.add_argument("--max-age-days", type=int, default=7,
                        help="缓存有效期(天), 默认 7 天")
    args = parser.parse_args()

    excels_dir = Path(args.workspace) / "excels"

    # ---- 生成前先检查缓存: 若 7 天内已有报告则直接复用, 无需重跑 ----
    if not args.force:
        cached = find_recent_report(excels_dir, args.ticker, args.max_age_days)
        if cached is not None:
            age = (date.today() - date.fromisoformat(cached.stem.split('_')[-1])).days \
                  if cached.stem.split('_')[-1].count('-') == 2 else 0
            logger.info(f"检测到 {args.max_age_days} 天内已存在报告 (age={age}天), 直接复用: {cached}")
            logger.info("如需强制重新生成, 请添加 --force 参数")
            print(str(cached))   # 打印路径到 stdout, 便于调用方拿到复用的文件路径
            _sys.exit(0)

    data = extract_financial_data(Path(args.workspace), args.ticker)
    output = excels_dir / f"{args.ticker}_3Statement_{date.today().isoformat()}.xlsx"
    ThreeStatementBuilder(data).build(output)
    print(str(output))
