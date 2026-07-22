#!/usr/bin/env python3
"""
生成包含实时公式的机构级可比公司分析 Excel 模型。
优先从本地 Excel 提取利润表/资产负债表数据，缺失的市场数据留空并标记 TODO 供 Web Search 填补。
"""
import argparse
import re
import zipfile
import tempfile
import os
from pathlib import Path
from typing import Any, Tuple

try:
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise ImportError("openpyxl required: pip install openpyxl") from exc

# ==================== Constants & Styles ====================
FONT_BLUE = Font(color="0000FF", name="Times New Roman", size=11)
FONT_BLACK = Font(color="000000", name="Times New Roman", size=11)
FONT_WHITE_BOLD = Font(color="FFFFFF", bold=True, name="Times New Roman", size=12)
FONT_BOLD = Font(bold=True, name="Times New Roman", size=11)

FILL_DARK_BLUE = PatternFill("solid", fgColor="1F4E79")
FILL_LIGHT_BLUE = PatternFill("solid", fgColor="D9E1F2")
FILL_LIGHT_GREY = PatternFill("solid", fgColor="F2F2F2")

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# ==================== Data Extraction Logic ====================
def safe_divide(n, d): return n / d if d != 0 else 0.0

def _ensure_shared_strings(file_path: Path) -> Tuple[Path, bool]:
    file_path = Path(file_path)
    try:
        with zipfile.ZipFile(str(file_path), 'r') as zf:
            namelist = zf.namelist()
            if 'xl/sharedStrings.xml' in namelist: return file_path, False
            ct = zf.read('[Content_Types].xml').decode('utf-8', errors='ignore')
            if 'sharedStrings' not in ct: return file_path, False
    except: return file_path, False

    fd, tmp = tempfile.mkstemp(suffix='.xlsx', dir=str(file_path.parent))
    os.close(fd)
    try:
        with zipfile.ZipFile(str(file_path), 'r') as src, zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as dst:
            for item in src.namelist(): dst.writestr(item, src.read(item))
            dst.writestr('xl/sharedStrings.xml', '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"/>')
        return Path(tmp), True
    except: return file_path, False

def _read_excel_map(file_path: Path) -> dict:
    fp, is_tmp = _ensure_shared_strings(file_path)
    try:
        wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
        sheet = wb.active
        data = {}
        header_row = 1
        for r in range(1, 5):
            if "FY" in " ".join(str(sheet.cell(r, c).value) for c in range(1, sheet.max_column+1) if sheet.cell(r,c).value):
                header_row = r; break

        periods = [str(sheet.cell(header_row, c).value) for c in range(3, sheet.max_column + 1)]
        for r in range(header_row + 1, sheet.max_row + 1):
            ind = sheet.cell(r, 1).value
            if not ind: continue
            ind = str(ind).strip()
            vals = {}
            for c, p in enumerate(periods):
                v = sheet.cell(r, c + 3).value
                num = 0.0
                if isinstance(v, (int, float)): num = float(v)
                elif isinstance(v, str) and v:
                    cl = v.replace(",", "").replace("-", "").strip()
                    if cl.replace(".", "", 1).isdigit(): num = float(cl) * (-1 if v.strip().startswith("-") else 1)
                vals[p] = num
            data[ind] = vals
        wb.close()
        return data
    finally:
        if is_temp and fp.exists(): os.remove(str(fp))

def find_local_file(workspace: Path, ticker: str, suffix: str) -> Path:
    pattern = re.compile(rf'^.*_{re.escape(ticker)}_{re.escape(suffix)}_.*\.(xlsx|xls)$', re.IGNORECASE)
    files = [f for f in (workspace / "excels").iterdir() if f.is_file() and pattern.match(f.name)]
    return sorted(files)[-1] if files else None

def extract_company_data(ticker: str, workspace: Path) -> dict:
    data = {"ticker": ticker, "revenue": None, "ebit": None, "net_income": None, "market_cap": None, "ev": None}
    inc_file = find_local_file(workspace, ticker, "income")

    if inc_file:
        inc_data = _read_excel_map(inc_file)
        fy_cols = sorted([c for c in inc_data.get("总收入", {}).keys() if "FY" in c], reverse=True)
        if fy_cols:
            latest = fy_cols[0]
            data["revenue"] = inc_data.get("总收入", {}).get(latest, 0.0)
            data["ebit"] = inc_data.get("营业利润", {}).get(latest, 0.0)
            data["net_income"] = inc_data.get("净利润", {}).get(latest, 0.0)
            data["source_inc"] = f"Local file: {inc_file.name}"
    return data

# ==================== Excel Builder Class ====================
class CompsModelBuilder:
    def __init__(self, tickers: list, workspace: Path):
        self.tickers = tickers
        self.workspace = workspace
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Comps Analysis"
        self.companies_data = []

    def build(self, output_path) -> Path:
        for t in self.tickers:
            self.companies_data.append(extract_company_data(t, self.workspace))

        self._build_header()
        self._build_operating_metrics()
        self._build_valuation_multiples()

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(path)
        print(f"Comps Analysis model saved to {path}")
        return path

    def _build_header(self) -> None:
        ws = self.ws
        ws.merge_cells("A1:G1")
        ws["A1"] = "COMPARABLE COMPANY ANALYSIS"
        ws["A1"].fill = FILL_DARK_BLUE; ws["A1"].font = FONT_WHITE_BOLD; ws["A1"].alignment = CENTER_ALIGN

        ws.merge_cells("A2:G2")
        ws["A2"] = " • ".join(self.tickers); ws["A2"].alignment = CENTER_ALIGN

        ws.merge_cells("A3:G3")
        ws["A3"] = "All figures in USD Millions | Market Data requires Web Search fill"; ws["A3"].alignment = CENTER_ALIGN

    def _build_operating_metrics(self) -> None:
        ws = self.ws
        ws.merge_cells("A5:G5")
        ws["A5"] = "OPERATING STATISTICS"; ws["A5"].fill = FILL_DARK_BLUE; ws["A5"].font = FONT_WHITE_BOLD

        headers = ["Company", "Revenue (LTM)", "Growth (YoY)", "EBIT", "EBIT Margin", "Net Income", "Net Margin"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(6, i, h); c.fill = FILL_LIGHT_BLUE; c.font = FONT_BOLD; c.alignment = CENTER_ALIGN

        start_row = 7
        for i, data in enumerate(self.companies_data):
            r = start_row + i
            ws.cell(r, 1, data["ticker"]).font = FONT_BLACK

            # Inputs with source comments
            if data["revenue"] is not None:
                ws.cell(r, 2, data["revenue"]).font = FONT_BLUE
                ws.cell(r, 2).comment = Comment(data.get("source_inc", "Source missing"), "Agent")
            else:
                ws.cell(r, 2, 0).font = FONT_BLUE
                ws.cell(r, 2).comment = Comment("TODO: Web Search (Eastmoney/Futu) for LTM Revenue", "Agent")

            # Growth requires prior year, setting as TODO for simplicity in this demo
            ws.cell(r, 3, 0).font = FONT_BLUE
            ws.cell(r, 3).comment = Comment("TODO: Web Search or Local File for YoY Growth", "Agent")

            if data["ebit"] is not None:
                ws.cell(r, 4, data["ebit"]).font = FONT_BLUE
                ws.cell(r, 4).comment = Comment(data.get("source_inc", "Source missing"), "Agent")
            else:
                ws.cell(r, 4, 0).font = FONT_BLUE
                ws.cell(r, 4).comment = Comment("TODO: Web Search for EBIT", "Agent")

            if data["net_income"] is not None:
                ws.cell(r, 6, data["net_income"]).font = FONT_BLUE
                ws.cell(r, 6).comment = Comment(data.get("source_inc", "Source missing"), "Agent")
            else:
                ws.cell(r, 6, 0).font = FONT_BLUE
                ws.cell(r, 6).comment = Comment("TODO: Web Search for Net Income", "Agent")

            # Formulas
            ws.cell(r, 5, f"=D{r}/B{r}").font = FONT_BLACK
            ws.cell(r, 7, f"=F{r}/B{r}").font = FONT_BLACK

            for c in range(2, 8):
                ws.cell(r, c).alignment = CENTER_ALIGN
                if c in [3, 5, 7]: ws.cell(r, c).number_format = "0.0%"
                else: ws.cell(r, c).number_format = "#,##0"

        self._add_stats_block(start_row, len(self.companies_data), [3, 5, 7], "0.0%")

    def _build_valuation_multiples(self) -> None:
        ws = self.ws
        val_start = 7 + len(self.companies_data) + 6 + 2
        ws.merge_cells(f"A{val_start}:G{val_start}")
        ws.cell(val_start, 1, "VALUATION MULTIPLES").fill = FILL_DARK_BLUE
        ws.cell(val_start, 1).font = FONT_WHITE_BOLD

        headers = ["Company", "Market Cap", "Enterprise Value", "EV/Revenue", "EV/EBIT", "P/E Ratio", "Price"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(val_start + 1, i, h); c.fill = FILL_LIGHT_BLUE; c.font = FONT_BOLD; c.alignment = CENTER_ALIGN

        data_start = val_start + 2
        op_data_start = 7
        for i, data in enumerate(self.companies_data):
            r = data_start + i
            op_r = op_data_start + i

            ws.cell(r, 1, data["ticker"]).font = FONT_BLACK

            # Market Cap and EV are typically missing from local financials, mark for Web Search
            ws.cell(r, 2, 0).font = FONT_BLUE
            ws.cell(r, 2).comment = Comment("TODO: Web Search (Futu/Yahoo) for Market Cap", "Agent")

            ws.cell(r, 3, 0).font = FONT_BLUE
            ws.cell(r, 3).comment = Comment("TODO: Calculate EV = Mkt Cap + Debt - Cash (fetch from BS or Web)", "Agent")

            ws.cell(r, 7, 0).font = FONT_BLUE
            ws.cell(r, 7).comment = Comment("TODO: Web Search for current Stock Price", "Agent")

            # Formulas referencing Operating Metrics
            # Note: Using IFERROR to handle 0s gracefully before web search fill
            ws.cell(r, 4, f"=IFERROR(C{r}/B{op_r}, 0)").font = FONT_BLACK
            ws.cell(r, 5, f"=IFERROR(C{r}/D{op_r}, 0)").font = FONT_BLACK
            ws.cell(r, 6, f"=IFERROR(B{r}/F{op_r}, 0)").font = FONT_BLACK

            for c in range(2, 8):
                ws.cell(r, c).alignment = CENTER_ALIGN
                if c in [4, 5, 6]: ws.cell(r, c).number_format = "0.0\"x\""
                else: ws.cell(r, c).number_format = "#,##0"

        self._add_stats_block(data_start, len(self.companies_data), [4, 5, 6], "0.0\"x\"")

        for col in range(1, 8): ws.column_dimensions[get_column_letter(col)].width = 18

    def _add_stats_block(self, start_row: int, count: int, cols: list, fmt: str):
        ws = self.ws
        stats_start = start_row + count + 1
        labels = ["Maximum", "75th Percentile", "Median", "25th Percentile", "Minimum"]

        for i, label in enumerate(labels):
            r = stats_start + i
            ws.cell(r, 1, label).font = FONT_BOLD
            ws.cell(r, 1).fill = FILL_LIGHT_GREY

            for col_idx in cols:
                col_l = get_column_letter(col_idx)
                rng = f"{col_l}{start_row}:{col_l}{start_row + count - 1}"

                if label == "Maximum": f = f"=MAX({rng})"
                elif label == "75th Percentile": f = f"=QUARTILE({rng}, 3)"
                elif label == "Median": f = f"=MEDIAN({rng})"
                elif label == "25th Percentile": f = f"=QUARTILE({rng}, 1)"
                elif label == "Minimum": f = f"=MIN({rng})"

                c = ws.cell(r, col_idx, f)
                c.font = FONT_BLACK; c.fill = FILL_LIGHT_GREY; c.alignment = CENTER_ALIGN; c.number_format = fmt

def main() -> None:
    parser = argparse.ArgumentParser(description="Build a Comps Analysis Excel model")
    parser.add_argument("--tickers", required=True, help="Comma-separated tickers (e.g., MSFT,GOOGL)")
    parser.add_argument("--workspace", required=True, help="Path to workspace directory containing 'excels' folder")
    parser.add_argument("--output", required=True, help="Path to output xlsx file")
    args = parser.parse_args()

    builder = CompsModelBuilder(args.tickers.split(','), Path(args.workspace))
    builder.build(args.output)

if __name__ == "__main__":
    main()
