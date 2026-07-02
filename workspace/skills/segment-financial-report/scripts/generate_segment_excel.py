#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
上市公司分部业务财务报表Excel生成工具

支持多层级分部展示（如BABA的三层业务结构），体现层级关系。
包含收入、EBITA等指标，支持YoY同比高亮显示。

用法：
    python generate_segment_excel.py <ticker> [--output <path>]
    python generate_segment_excel.py --json <json_file> [--output <path>]

示例：
    python generate_segment_excel.py BABA
    python generate_segment_excel.py BABA --output ./baba_segments.xlsx
"""

import json
import subprocess
import sys
import os
import argparse
import logging
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─────────────────────────────────────────────────────────────
# 颜色配置 - 按层级区分
# ─────────────────────────────────────────────────────────────

# 一级业务分部的背景色组合（用于区分不同的一级业务）
LEVEL1_COLORS = [
    {'bg': 'D9E1F2', 'font': '2F5496'},    # 浅蓝-深蓝字体
    {'bg': 'E2EFDA', 'font': '548235'},    # 浅绿-深绿字体
    {'bg': 'FFF2CC', 'font': 'BF8F00'},    # 浅黄-深黄字体
    {'bg': 'FCE4D6', 'font': 'C65911'},    # 浅橙-深橙字体
    {'bg': 'E2DAF0', 'font': '7030A0'},    # 浅紫-深紫字体
    {'bg': 'F8CBAD', 'font': 'B45F06'},    # 橙黄-深橙字体
]

COLORS = {
    'header': '4472C4',       # 表头-深蓝
    'header_font': 'FFFFFF',  # 表头字体-白色
    'revenue': 'DEEBF7',      # 收入-浅蓝
    'ebita': 'E2EFDA',        # EBITA-浅绿
    'yoy_alert': 'FF0000',    # YoY异常-红色（下跌超过5%或上涨超过30%）
    'ratio': 'FFF2CC',        # 比率-浅黄
    'margin_bg': 'F4B084',    # EBITA利润率-深橙色背景
    'default': None,
    # 层级颜色（保留兼容）
    'level1': 'D9E1F2',
    'level1_font': '1F3864',
    'level2': 'E2EFDA',
    'level2_font': '548235',
    'level3': 'F2F2F2',
    'level3_font': '000000',
}

# YoY阈值配置
YOY_DECLINE_THRESHOLD = -5.0   # 下降超过此阈值红色高亮
YOY_GROWTH_THRESHOLD = 30.0    # 增长超过此阈值红色高亮（异常高增长）

# ─────────────────────────────────────────────────────────────
# 辅助函数
# ─────────────────────────────────────────────────────────────

def setup_logging(log_file):
    """设置日志系统"""
    if not os.path.exists(os.path.dirname(log_file)):
        os.makedirs(os.path.dirname(log_file))

    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s',
        filename=log_file,
        filemode='w',
        encoding='utf-8'
    )
    return logging.getLogger(__name__)


def log_stdout(msg):
    """输出到 stdout（极简）"""
    print(msg, flush=True)


def get_all_periods(flat_metrics):
    """从扁平的指标数据中提取所有周期

    排序规则：年份逆序（新→旧），同一年内 FY 在最前面，然后是 Q4/Q3/Q2/Q1、H2/H1
    例如：[2025FY, 2025Q3, 2025Q2, 2025Q1, 2024FY, 2024Q4, 2024Q3, ...]
    """
    periods = set()
    for metric in flat_metrics:
        period = metric.get('period')
        if period:
            periods.add(period)

    def period_sort_key(p):
        """构造排序键：(年份逆序, 周期类型优先级)
        年份越大排前面（用负号实现），同年内 FY=0 > Q4=1 > Q3=2 > Q2=3 > Q1=4 > H2=5 > H1=6
        """
        import re
        # 提取年份
        year_match = re.search(r'(\d{4})', p)
        year = int(year_match.group(1)) if year_match else 0

        # 提取周期类型
        upper = p.upper()
        if 'FY' in upper:
            type_order = 0
        elif 'Q4' in upper:
            type_order = 1
        elif 'Q3' in upper:
            type_order = 2
        elif 'Q2' in upper:
            type_order = 3
        elif 'Q1' in upper:
            type_order = 4
        elif 'H2' in upper:
            type_order = 5
        elif 'H1' in upper:
            type_order = 6
        else:
            type_order = 9

        # 年份越大越靠前 → 用 -year；周期类型按上述顺序
        return (-year, type_order)

    return sorted(list(periods), key=period_sort_key)


def get_all_level1_segments(flat_metrics):
    """获取所有一级业务分部及其颜色映射"""
    level1_segments = {}
    for metric in flat_metrics:
        if metric.get('level') == 1:
            seg_code = metric.get('segmentCode')
            seg_name = metric.get('segmentName')
            if seg_code and seg_code not in level1_segments:
                level1_segments[seg_code] = seg_name

    # 按出现顺序返回一级分部编码列表
    return list(level1_segments.keys())


def get_level1_segment_code(flat_metrics, segment_code):
    """查找指定分部对应的一级分部编码"""
    # 构建父子关系映射
    parent_map = {}
    for metric in flat_metrics:
        seg_code = metric.get('segmentCode')
        parent_code = metric.get('parentSegmentCode')
        level = metric.get('level', 1)
        if seg_code and level == 1:
            parent_map[seg_code] = seg_code  # 一级分部的父级是自己
        elif seg_code and parent_code:
            parent_map[seg_code] = parent_code

    # 向上查找直到找到一级分部
    current = segment_code
    visited = set()
    while current in parent_map and current not in visited:
        visited.add(current)
        if parent_map[current] == current:  # 找到一级分部
            return current
        current = parent_map[current]

    return segment_code  # 如果找不到，返回自己


def build_segment_metric_map(flat_metrics, periods):
    """构建扁平数据的索引映射。

    以 (parentSegmentCode, segmentCode) 为节点唯一键 —— 因为部分公司（如美团 83690）
    多个 L1 下会共用同一个 L2 segmentCode（"佣金"既在核心本地商業下也在新業務下），
    单靠 segmentCode 会把两个 L1 的 L2 数据合并成一份，还会因 period 冲突互相覆盖。

    L1 节点用 (None, segmentCode) 作 key；每个节点独立存放它的 metrics/periods。

    返回: {
        (parent_code_or_None, segment_code): {
            'segmentName': '...',
            'segmentCode': '...',
            'level': 1,
            'parentSegmentCode': '...',
            'metrics': {
                'REVENUE': {period1: value, ...},
                'REVENUE_YOY': {period1: value, ...},
                'ADJUSTED_EBITA': {...},
                'COST': {...},
                'OPERATING_INCOME': {...},
                'OPERATING_EXPENSES': {...},
                ...
            }
        }
    }
    """
    seg_metric_map = {}

    for item in flat_metrics:
        seg_code = item.get('segmentCode')
        if not seg_code:
            continue
        parent_code = item.get('parentSegmentCode')
        # L1 节点的 parentSegmentCode 应为 None；防御性地把等于自己的 parent 也归为 L1
        if parent_code == seg_code:
            parent_code = None
        node_key = (parent_code, seg_code)

        if node_key not in seg_metric_map:
            seg_metric_map[node_key] = {
                'segmentName': item.get('segmentName', 'Unknown'),
                'segmentCode': seg_code,
                'level': item.get('level', 1),
                'parentSegmentCode': parent_code,
                'metrics': {}
            }

        metric_code = item.get('metricCode')
        if metric_code:
            period = item.get('period')
            value = item.get('value')
            yoy = item.get('yoyGrowth')

            # 存储指标值
            if metric_code not in seg_metric_map[node_key]['metrics']:
                seg_metric_map[node_key]['metrics'][metric_code] = {}
            seg_metric_map[node_key]['metrics'][metric_code][period] = value

            # 存储YoY值
            yoy_key = f"{metric_code}_YOY"
            if yoy_key not in seg_metric_map[node_key]['metrics']:
                seg_metric_map[node_key]['metrics'][yoy_key] = {}
            seg_metric_map[node_key]['metrics'][yoy_key][period] = yoy

    return seg_metric_map


def has_any_data(values):
    """检查是否有任何有效数据"""
    return any(v is not None for v in values)


def get_display_name(segment_info, level=None):
    """生成分部显示名称，包含层级缩进"""
    name = segment_info.get('segmentName', 'Unknown')
    level = level or segment_info.get('level', 1)

    # 根据层级添加缩进
    indent = '  ' * (level - 1)
    prefix = '├─ ' if level > 1 else ''
    return f"{indent}{prefix}{name}"


def build_data_rows_from_flat(flat_metrics, periods, logger=None):
    """从扁平数据构建数据行

    展示逻辑：
    1. 按一级业务分组，组内按层级深度优先遍历
    2. 每个分部依次展示：收入 → 成本 → 营业费用 → 营业利润 → EBITA
       每个数值指标后紧跟对应的 YoY 行；收入/EBITA 后额外附利润率行
    3. 空数据行过滤：指标值全为空则不展示
    4. L1 分部若没有直接的收入/成本/费用/利润数据，则用子分部同期数据累加填充
       —— 港股 SUBSEGMENT_MATRIX 布局里 L2 只披露 REVENUE、L1 只披露 COST/OP_INCOME，
       没有这一步，Excel 上 L1 就看不到收入行。
    """
    rows = []

    # 构建指标映射（key = (parentCode, segmentCode)）
    seg_metric_map = build_segment_metric_map(flat_metrics, periods)

    # 获取一级业务分部（用于分组和颜色）
    level1_codes = get_all_level1_segments(flat_metrics)

    if logger:
        logger.info(f"识别到 {len(level1_codes)} 个一级业务分部: {level1_codes}")
        logger.info(f"共 {len(seg_metric_map)} 个分部数据，{len(periods)} 个周期")

    # 构建父子索引：parent_code -> [(parent_code_of_child, child_code)]
    # 由于 seg_metric_map 已按 (parent, code) 唯一存储，同一 code 挂在不同 parent 下的会各占一个节点
    children_map = {}
    for node_key, seg_info in seg_metric_map.items():
        parent_code = seg_info.get('parentSegmentCode')
        if parent_code:
            children_map.setdefault(parent_code, []).append(node_key)

    # ---- 汇总辅助：把子节点某指标同期的值累加到 L1（仅当 L1 自己该指标该期为空时才填） ----
    AGGREGATABLE_METRICS = ('REVENUE', 'COST', 'GROSS_PROFIT', 'OPERATING_EXPENSES',
                            'OPERATING_INCOME', 'ADJUSTED_EBITA')

    def _sum_child_metric(node_key, metric_code, period):
        """深度累加：把 node_key 下所有后代该指标该期的值加起来；任一后代无值则跳过它。"""
        total = None
        for child_key in children_map.get(node_key[1], []):
            child_info = seg_metric_map.get(child_key)
            if not child_info:
                continue
            child_metrics = child_info.get('metrics', {}) or {}
            v = child_metrics.get(metric_code, {}).get(period)
            if v is None:
                # 递归看孙辈
                v = _sum_child_metric(child_key, metric_code, period)
            if v is not None:
                total = (total or 0) + v
        return total

    def _fill_parent_from_children(node_key):
        """若 L1（或任何有子节点的父节点）的某个可累加指标同期缺数，用后代同期累加值补上。"""
        seg_info = seg_metric_map.get(node_key)
        if not seg_info or not children_map.get(node_key[1]):
            return
        metrics = seg_info.setdefault('metrics', {})
        for m in AGGREGATABLE_METRICS:
            m_map = metrics.setdefault(m, {})
            for period in periods:
                if m_map.get(period) is not None:
                    continue
                agg = _sum_child_metric(node_key, m, period)
                if agg is not None:
                    m_map[period] = agg

    # 按一级业务分组深度遍历
    for level1_idx, level1_code in enumerate(level1_codes):
        level1_key = (None, level1_code)
        level1_info = seg_metric_map.get(level1_key)
        if not level1_info:
            continue

        # 先把 L1 缺的可累加指标从子节点补齐（子节点自己也做一遍，以支持 3+ 层）
        for node_key in list(seg_metric_map.keys()):
            # 只处理属于当前 L1 树的节点
            if node_key == level1_key or _belongs_to_level1(node_key, level1_code, seg_metric_map):
                _fill_parent_from_children(node_key)

        # 深度优先遍历当前一级业务的整棵树
        def traverse_tree(node_key):
            seg_info = seg_metric_map.get(node_key)
            if not seg_info:
                return

            level = seg_info.get('level', 1)
            display_name = get_display_name(seg_info, level)
            metrics = seg_info.get('metrics', {})

            row_level1_info = {
                'level1_code': level1_code,
                'level1_index': level1_idx
            }

            # ---- 按顺序输出：收入、成本、毛利、营业费用、营业利润、EBITA ----
            # 每个数值指标后跟 YoY 行；毛利/营业利润/EBITA 后各附一个"利润率(%)"行
            metric_specs = [
                # (row_type, display_label, metric_code, margin_label_or_None)
                ('revenue',            '收入',       'REVENUE',            None),
                ('cost',               '成本',       'COST',               None),
                ('gross_profit',       '毛利',       'GROSS_PROFIT',       '毛利率(%)'),
                ('operating_expenses', '营业费用',   'OPERATING_EXPENSES', None),
                ('operating_income',   '营业利润',   'OPERATING_INCOME',   '营业利润率(%)'),
                ('ebita',              'EBITA',      'ADJUSTED_EBITA',     'EBITA利润率(%)'),
            ]

            # 收入值用于计算利润率
            revenue_values = [metrics.get('REVENUE', {}).get(p) for p in periods]

            for row_type, label, metric_code, margin_label in metric_specs:
                values = [metrics.get(metric_code, {}).get(p) for p in periods]
                if not has_any_data(values):
                    continue

                rows.append({
                    'type': row_type,
                    'level': level,
                    'name': f"{display_name} - {label}",
                    'values': values,
                    'yoy_values': [],
                    'level1': row_level1_info
                })

                yoy_values = [metrics.get(f"{metric_code}_YOY", {}).get(p) for p in periods]
                if has_any_data(yoy_values):
                    rows.append({
                        'type': f'{row_type}_yoy',
                        'level': level,
                        'name': f"{display_name} - {label}YoY(%)",
                        'values': yoy_values,
                        'yoy_values': yoy_values,
                        'level1': row_level1_info
                    })

                # 营业利润 / EBITA 后附利润率行
                if margin_label:
                    margin_values = _calc_margin(revenue_values, values)
                    if has_any_data(margin_values):
                        rows.append({
                            'type': f'{row_type}_margin',
                            'level': level,
                            'name': f"{display_name} - {margin_label}",
                            'values': margin_values,
                            'yoy_values': [],
                            'level1': row_level1_info
                        })

            # 递归处理子分部
            for child_key in sorted(children_map.get(node_key[1], []), key=lambda k: k[1]):
                traverse_tree(child_key)

        traverse_tree(level1_key)

    if logger:
        logger.info(f"过滤空数据行后，共生成 {len(rows)} 行数据")

    return rows


def _calc_margin(revenue_values, numerator_values):
    """按周期计算 numerator / revenue * 100；分母为 0 或缺任一值时返回 None。"""
    result = []
    for rev, num in zip(revenue_values, numerator_values):
        if rev and rev != 0 and num is not None:
            result.append((num / rev) * 100)
        else:
            result.append(None)
    return result


def _belongs_to_level1(node_key, level1_code, seg_metric_map):
    """判断某个 node 是否属于指定 L1 的子树（沿 parentSegmentCode 向上查）。"""
    parent_code = seg_metric_map.get(node_key, {}).get('parentSegmentCode')
    visited = set()
    while parent_code and parent_code not in visited:
        if parent_code == level1_code:
            return True
        visited.add(parent_code)
        # 找该 parent 对应的节点（有多份时任取一份即可，因为 parent 关系向上唯一）
        parent_key = next((k for k in seg_metric_map.keys() if k[1] == parent_code), None)
        if not parent_key:
            break
        parent_code = seg_metric_map[parent_key].get('parentSegmentCode')
    return False


def format_value(val):
    """格式化数值 - 直接展示原始数据，不做单位转换"""
    if val is None:
        return '-'
    return val


def generate_excel_with_styling(flat_metrics, periods, output_file, ticker, logger=None):
    """生成带样式的Excel，体现层级关系和一级业务分组颜色

    每个一级业务的所有相关行（包括子业务的所有指标行）使用相同的背景色
    """
    if not HAS_OPENPYXL:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "分部财务数据"

    # 基础样式
    header_font = Font(bold=True, name="微软雅黑", color=COLORS['header_font'])
    normal_font = Font(name="微软雅黑")
    red_font = Font(name="微软雅黑", color=COLORS['yoy_alert'], bold=True)

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 构建数据行（从扁平数据）
    data_rows = build_data_rows_from_flat(flat_metrics, periods, logger)

    # 写入表头
    # 结构：业务分部 | 指标 | [各期数值]
    headers = ['业务分部', '指标'] + periods
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color=COLORS['header'],
                                end_color=COLORS['header'],
                                fill_type="solid")
        cell.alignment = center_align
        cell.border = thin_border

    # 写入数据
    row_idx = 2
    for data_row in data_rows:
        level = data_row['level']
        row_type = data_row['type']
        name = data_row['name']
        values = data_row['values']
        level1_info = data_row.get('level1', {})
        level1_index = level1_info.get('level1_index', 0)

        # 获取当前一级业务对应的背景色（循环使用 LEVEL1_COLORS）
        color_config = LEVEL1_COLORS[level1_index % len(LEVEL1_COLORS)]
        bg_color = color_config['bg']
        font_color = color_config['font']

        # 根据层级选择字体
        if level == 1:
            row_font = Font(bold=True, name="微软雅黑", color=font_color)
        elif level == 2:
            row_font = Font(bold=True, name="微软雅黑", color=font_color)
        else:
            row_font = Font(name="微软雅黑", color=font_color)

        # 拆分名称：分部名称 和 指标名称（如 "L1 - 集团总览 - 收入" → "L1 - 集团总览", "收入"）
        name_parts = name.rsplit(' - ', 1)
        if len(name_parts) == 2:
            segment_name = name_parts[0]
            metric_name = name_parts[1]
        else:
            segment_name = name
            metric_name = ''

        # ========== 确定当前行的背景色 ==========
        # 利润率行（*_margin）使用特殊的深色背景，其他行使用一级业务颜色
        is_margin_row = row_type.endswith('_margin')
        is_yoy_row = row_type.endswith('_yoy')
        is_pct_row = is_margin_row or is_yoy_row
        if is_margin_row:
            row_bg_color = COLORS['margin_bg']
        else:
            row_bg_color = bg_color

        # ========== 分部名称列 ==========
        name_cell = ws.cell(row=row_idx, column=1)
        name_cell.value = segment_name
        name_cell.border = thin_border
        name_cell.font = row_font
        name_cell.fill = PatternFill(start_color=row_bg_color,
                                    end_color=row_bg_color,
                                    fill_type="solid")
        name_cell.alignment = left_align

        # ========== 指标类型列 ==========
        metric_cell = ws.cell(row=row_idx, column=2)
        metric_cell.value = metric_name
        metric_cell.border = thin_border
        metric_cell.alignment = center_align
        metric_cell.font = row_font
        metric_cell.fill = PatternFill(start_color=row_bg_color,
                                      end_color=row_bg_color,
                                      fill_type="solid")

        # ========== 数据列 ==========
        for col_offset, val in enumerate(values, 0):
            val_cell = ws.cell(row=row_idx, column=3 + col_offset)
            val_cell.border = thin_border
            val_cell.alignment = center_align

            # 设置数值和格式
            if val is not None and isinstance(val, (int, float)):
                val_cell.value = round(val, 2) if is_pct_row else format_value(val)

                # YoY行的颜色高亮：下跌超过5%或上涨超过30%都高亮红色
                if is_yoy_row:
                    if val <= YOY_DECLINE_THRESHOLD or val >= YOY_GROWTH_THRESHOLD:
                        val_cell.font = red_font
                    else:
                        val_cell.font = normal_font
                else:
                    val_cell.font = normal_font

                # 数值格式
                if is_pct_row:
                    # 百分比类保留2位小数
                    val_cell.number_format = '0.00'
                elif val == int(val):
                    # 整数不带小数
                    val_cell.number_format = '#,##0'
                else:
                    # 小数保留2位
                    val_cell.number_format = '#,##0.00'
            else:
                val_cell.value = '-'
                val_cell.font = normal_font

            # 背景色：利润率行使用深色突出，其余行继承一级业务颜色
            val_cell.fill = PatternFill(start_color=row_bg_color,
                                       end_color=row_bg_color,
                                       fill_type="solid")

        # 移动到下一行
        row_idx += 1

    # ========== 设置列宽 ==========
    # 第1列：业务分部 - 宽度40（适应长名称，如 "├─ L2 - 平台业务 - 国内业务"）
    ws.column_dimensions[get_column_letter(1)].width = 40
    # 第2列：指标 - 宽度18（适应 "收入"、"收入YoY(%)"、"EBITA利润率(%)" 等）
    ws.column_dimensions[get_column_letter(2)].width = 18
    # 第3列及以后：数据列 - 统一宽度15，便于数值展示
    for col in range(3, len(periods) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # 保存工作簿
    wb.save(output_file)

    if logger:
        logger.info(f"Excel saved: {output_file}")

    return output_file


def main():
    parser = argparse.ArgumentParser(description='生成分部业务财务报表Excel')
    parser.add_argument('ticker', nargs='?', help='股票代码，如 BABA, AAPL')
    parser.add_argument('--json', '-j', required=True, help='包含Segment数据的JSON文件路径（由Java Tool传入）')
    parser.add_argument('--output', '-o', help='输出Excel文件路径')
    parser.add_argument('--workspace', '-w', help='项目工作空间路径')
    args = parser.parse_args()

    # 确定项目根目录
    if args.workspace:
        # workspace已经是workspace目录了，不需要再加一层workspace
        project_root = args.workspace
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.abspath(os.path.join(script_dir, '..', '..'))

    # 确保日志和输出目录存在
    logs_dir = os.path.join(project_root, 'excels', 'logs')
    excels_dir = os.path.join(project_root, 'excels')
    for d in [logs_dir, excels_dir]:
        if not os.path.exists(d):
            os.makedirs(d)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    ticker = args.ticker or 'UNKNOWN'

    log_file = os.path.join(logs_dir, f"segment_{ticker}_{timestamp}.log")
    logger = setup_logging(log_file)

    log_stdout(f">>> Processing {ticker} 分部财务数据")

    try:
        # 从JSON文件加载数据（由Java Tool传入）
        if args.json and os.path.exists(args.json):
            with open(args.json, 'r', encoding='utf-8') as f:
                segments = json.load(f)
            logger.info(f"从JSON文件加载数据: {args.json}")
        else:
            log_stdout(f"ERROR: JSON文件不存在: {args.json}")
            return 1

        if not segments:
            log_stdout("ERROR: 分部数据为空")
            return 1

        logger.info(f"获取到 {len(segments)} 条扁平指标数据")

        # 从扁平数据中获取所有周期（已按逆序排序：最新在前）
        periods = get_all_periods(segments)
        logger.info(f"包含 {len(periods)} 个财报周期: {periods}")

        if not periods:
            periods = ['Latest']

        # 确定输出路径
        if args.output:
            output_file = args.output
        else:
            output_file = os.path.join(excels_dir, f"{ticker}_segments_{timestamp}.xlsx")

        output_file = os.path.abspath(output_file)

        # 生成Excel
        generate_excel_with_styling(segments, periods, output_file, ticker, logger)

        # 最终输出
        result = {
            "status": "ok",
            "ticker": ticker,
            "segments_count": len(segments),
            "periods": periods,
            "excel_path": output_file,
            "log_path": os.path.abspath(log_file),
        }

        log_stdout("=" * 60)
        log_stdout(json.dumps(result, ensure_ascii=False))
        log_stdout("=" * 60)

        return 0

    except Exception as e:
        logger.exception("生成Excel失败")
        log_stdout(f"ERROR: {str(e)}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
