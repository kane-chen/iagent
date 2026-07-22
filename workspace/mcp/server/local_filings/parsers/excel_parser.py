# -*- coding: utf-8 -*-
"""
Excel 财务报表解析器——解析 workspace/excels/ 下的生成式财务报表工作簿。

工作簿格式约定（与 workspace/skills/financial-metrics-query 一致）：
  Row 1: period 列（'2026Q2', '2026Q1', '2025FY', ...）
  Row 2: 报告日期
  Col 1: 指标名称
  Col 2: 单位
  Col 3+: 数值
"""

from pathlib import Path
from typing import Any, Optional

import openpyxl

from ..storage import PERIOD_RE, period_is_annual, period_is_quarterly, period_sort_key, safe_float


def parse_financials_excel(
    file_map: dict[str, Optional[Path]],
    period: str = "quarterly",
    num_periods: int = 8,
) -> tuple[dict, dict, dict, dict]:
    """从 Excel 文件映射解析三大表 + 财务指标。

    file_map: {income|balance|cashflow: Path|None}
    period：quarterly | annual
    num_periods：最多返回的期间数
    返回 (income, balance, cashflow, indicators)
    """
    stmt_types = {
        "income": "income_statement",
        "balance": "balance_sheet",
        "cashflow": "cash_flow",
    }

    result = {}
    for stmt_key, stmt_name in stmt_types.items():
        path = file_map.get(stmt_key)
        if path is not None and path.is_file():
            result[stmt_name] = _parse_excel(path, period, num_periods)
        else:
            result[stmt_name] = {}

    # financial_indicators 暂不单独解析
    result["financial_indicators"] = {}

    return (
        result["income_statement"],
        result["balance_sheet"],
        result["cash_flow"],
        result["financial_indicators"],
    )


def _parse_excel(path: Path, period_filter: str, num_periods: int) -> dict:
    """解析单个工作簿文件，返回 {period: {metric: value}}。"""
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        ws = wb.active
        if ws is None:
            return {}

        # 第 1 行：period 列
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        # 第 3 行起：数据
        data_start = 3

        # 收集期间列索引
        period_cols: list[tuple[int, str]] = []
        for col_idx, label in enumerate(headers, start=1):
            if label is None:
                continue
            label_str = str(label).strip()
            if not PERIOD_RE.match(label_str):
                continue
            if period_filter == "quarterly" and not period_is_quarterly(label_str):
                continue
            if period_filter == "annual" and not period_is_annual(label_str):
                continue
            period_cols.append((col_idx, label_str))

        # 按期间倒序，截取 num_periods
        period_cols.sort(key=lambda x: period_sort_key(x[1]), reverse=True)
        period_cols = period_cols[:num_periods]

        if not period_cols:
            return {}

        # 逐行读取指标
        result: dict[str, dict] = {}
        for row in range(data_start, ws.max_row + 1):
            metric_name = ws.cell(row, 1).value
            if metric_name is None:
                continue
            metric_name = str(metric_name).strip()
            if not metric_name:
                continue

            for col_idx, label in period_cols:
                if label not in result:
                    result[label] = {}
                val = safe_float(ws.cell(row, col_idx).value)
                if val is not None:
                    result[label][metric_name] = val

        return result
    finally:
        wb.close()